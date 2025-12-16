#!/usr/bin/env python3
"""
Generate a nicely formatted Excel file showing all 9 dimensions
of the Technological Sovereignty concept analysis.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
dimension_font = Font(bold=True, size=11, color="FFFFFF")
dimension_fills = {
    "Quinean": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
    "Sellarsian": PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid"),
    "Brandomian": PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
    "Deleuzian": PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid"),
    "Bachelardian": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
    "Canguilhem": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
    "Davidson": PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
    "Blumenberg": PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid"),
    "Carey": PatternFill(start_color="16A085", end_color="16A085", fill_type="solid"),
}
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# Sheet 1: Overview
# ============================================================================
ws_overview = wb.active
ws_overview.title = "Overview"

overview_data = [
    ["TECHNOLOGICAL SOVEREIGNTY - 9-Dimensional Concept Analysis"],
    [""],
    ["Concept", "Technological Sovereignty"],
    ["Definition", "A state's capacity to control its own technological development and infrastructure"],
    ["Category", "sovereignty, technology, geopolitics"],
    [""],
    ["This analysis applies 9 philosophical frameworks to understand the concept's structure:"],
    [""],
    ["#", "Framework", "Focus", "Key Question"],
    ["1", "Quinean", "Web of Belief", "Where does this concept sit in inferential web?"],
    ["2", "Sellarsian", "Inferential Roles", "Is this concept falsely treated as 'given'?"],
    ["3", "Brandomian", "Social Practices", "What commitments/entitlements does using this concept involve?"],
    ["4", "Deleuzian", "Problems & Plane", "What problems does this concept address? What assumptions underlie it?"],
    ["5", "Bachelardian", "Obstacles", "Does this concept block understanding? Is it an epistemological obstacle?"],
    ["6", "Canguilhem", "Life History", "How was this concept born? How is it evolving? Is it healthy?"],
    ["7", "Davidson", "Reasoning Styles", "What reasoning style does this concept require/enable?"],
    ["8", "Blumenberg", "Metaphorology", "What root metaphors structure this concept?"],
    ["9", "Carey", "Bootstrapping", "What hierarchy level? What is it built from?"],
]

for row_idx, row_data in enumerate(overview_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif row_idx == 9:
            cell.font = header_font
            cell.fill = header_fill
        elif row_idx > 9 and col_idx == 1:
            cell.font = Font(bold=True)

ws_overview.column_dimensions['A'].width = 5
ws_overview.column_dimensions['B'].width = 25
ws_overview.column_dimensions['C'].width = 40
ws_overview.column_dimensions['D'].width = 60

# ============================================================================
# Sheet 2: Quinean - Web of Belief
# ============================================================================
ws_quinean = wb.create_sheet("1. Quinean")

quinean_data = [
    ["QUINEAN ANALYSIS - Web of Belief"],
    [""],
    ["Attribute", "Value", "Notes"],
    ["Centrality", "intermediate", "Not core (like 'sovereignty'), not peripheral - sits between"],
    ["Web Coherence Impact", "Changing this concept creates tension with both 'sovereignty' and 'technology dependency'", ""],
    [""],
    ["FORWARD INFERENCES (If tech sovereignty, then...)"],
    ["", "If tech sovereignty, then indigenous chip development possible", ""],
    ["", "If tech sovereignty, then reduced foreign dependency", ""],
    ["", "If tech sovereignty, then autonomous AI development feasible", ""],
    [""],
    ["BACKWARD INFERENCES (Tech sovereignty because...)"],
    ["", "Tech sovereignty because sufficient R&D investment", ""],
    ["", "Tech sovereignty because protected domestic markets", ""],
    [""],
    ["LATERAL CONNECTIONS"],
    ["", "Related to: economic sovereignty, data sovereignty, digital sovereignty", ""],
    ["", "Tension with: globalization, supply chain integration, technology transfer", ""],
    [""],
    ["CONTRADICTIONS (Evidence that breaks inferences)"],
    ["", "Material dependency evidence contradicts forward inferences", "Gulf $2T but can't invest at Series B"],
    ["", "Equipment imports contradict autonomy claims", "China depends on ASML despite huge investment"],
    ["", "Knowledge dependencies contradict independence narrative", "Requires ongoing Western expertise"],
]

for row_idx, row_data in enumerate(quinean_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_quinean.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Quinean"]
        elif row_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
        elif "INFERENCES" in str(value) or "CONNECTIONS" in str(value) or "CONTRADICTIONS" in str(value):
            cell.font = Font(bold=True, italic=True)

ws_quinean.column_dimensions['A'].width = 15
ws_quinean.column_dimensions['B'].width = 60
ws_quinean.column_dimensions['C'].width = 50

# ============================================================================
# Sheet 3: Sellarsian - Myths of the Given
# ============================================================================
ws_sellarsian = wb.create_sheet("2. Sellarsian")

sellarsian_data = [
    ["SELLARSIAN ANALYSIS - Inferential Roles & Myths of the Given"],
    [""],
    ["Attribute", "Value"],
    ["Is Myth of Given?", "YES - falsely treated as foundational/self-evident"],
    [""],
    ["GIVENNESS MARKERS (language that treats concept as obvious)"],
    ["", '"naturally" - as if sovereignty naturally extends to technology'],
    ["", '"clearly possible" - assumes achievability without evidence'],
    ["", '"inevitable" - treats progression as given'],
    ["", '"obvious strategic necessity" - forecloses questioning'],
    [""],
    ["SHOULD BE INFERRED FROM (what evidence/argument should support it)"],
    ["", "Analysis of actual material capacity"],
    ["", "Mapping of supply chain dependencies"],
    ["", "Assessment of knowledge transfer mechanisms"],
    ["", "Comparison with successful sovereignty projects"],
    [""],
    ["THEORETICAL COMMITMENTS EMBEDDED (hidden assumptions)"],
    ["", "States can achieve technological autonomy through will + investment"],
    ["", "Technology behaves like territory (can be 'owned' and 'controlled')"],
    ["", "Sovereignty is binary rather than a spectrum"],
    ["", "Domestic capacity can substitute for global integration"],
    [""],
    ["WHAT GIVENNESS ENABLES"],
    ["", "Mobilizes resources for technology projects"],
    ["", "Justifies protectionist policies"],
    ["", "Creates political legitimacy for tech investments"],
    [""],
    ["WHAT GIVENNESS BLOCKS"],
    ["", "Questions about structural impossibility under current conditions"],
    ["", "Analysis of why sovereignty might be unachievable"],
    ["", "Recognition of necessary interdependence"],
    ["", "Honest assessment of dependency management strategies"],
]

for row_idx, row_data in enumerate(sellarsian_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_sellarsian.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Sellarsian"]
        elif row_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
        elif any(x in str(value) for x in ["MARKERS", "INFERRED FROM", "COMMITMENTS", "ENABLES", "BLOCKS"]):
            cell.font = Font(bold=True, italic=True)

ws_sellarsian.column_dimensions['A'].width = 15
ws_sellarsian.column_dimensions['B'].width = 80

# ============================================================================
# Sheet 4: Brandomian - Commitments & Entitlements
# ============================================================================
ws_brandomian = wb.create_sheet("3. Brandomian")

brandomian_data = [
    ["BRANDOMIAN ANALYSIS - Commitments, Entitlements & Social Practices"],
    [""],
    ["Type", "Statement", "Honored?", "Evidence"],
    ["COMMITMENT", "Autonomy is achievable through investment", "NO", "Gulf $2T can't invest at Series B level"],
    ["COMMITMENT", "State can control tech development trajectory", "NO", "Dependent on foreign equipment, expertise"],
    ["COMMITMENT", "Domestic capacity can replace foreign dependency", "NO", "Supply chains remain global"],
    ["ENTITLEMENT", "Can claim sovereignty through rhetoric alone", "EXCEEDED", "Claims without material substance"],
    ["ENTITLEMENT", "Can define success on own terms", "EXCEEDED", "Moving goalposts when milestones missed"],
    ["INCOMPATIBILITY", "Sovereignty AND deep integration in global supply chains", "", "Logical tension unresolved"],
    ["INCOMPATIBILITY", "Autonomy AND reliance on foreign IP/equipment", "", "Material contradiction"],
    [""],
    ["PRACTICE BREAKDOWNS"],
    ["", "Concept used to mobilize resources but not to assess outcomes honestly", ""],
    ["", "Success criteria shifted to maintain sovereignty narrative despite evidence", ""],
    ["", "Gap between stated commitments and actual dependencies not addressed", ""],
    [""],
    ["WHO BENEFITS FROM CURRENT USAGE"],
    ["", "Political leaders (legitimacy for tech spending)", ""],
    ["", "Domestic tech firms (protectionist policies)", ""],
    ["", "Military-industrial complex (national security framing)", ""],
]

for row_idx, row_data in enumerate(brandomian_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_brandomian.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Brandomian"]
        elif row_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
        elif "BREAKDOWNS" in str(value) or "BENEFITS" in str(value):
            cell.font = Font(bold=True, italic=True)
        elif value == "NO":
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif value == "EXCEEDED":
            cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

ws_brandomian.column_dimensions['A'].width = 20
ws_brandomian.column_dimensions['B'].width = 55
ws_brandomian.column_dimensions['C'].width = 12
ws_brandomian.column_dimensions['D'].width = 45

# ============================================================================
# Sheet 5: Deleuzian - Problems & Plane
# ============================================================================
ws_deleuzian = wb.create_sheet("4. Deleuzian")

deleuzian_data = [
    ["DELEUZIAN ANALYSIS - Problems, Plane of Immanence & Becomings"],
    [""],
    ["PROBLEM ADDRESSED"],
    ["", "How to claim autonomy while remaining materially dependent"],
    [""],
    ["TENSION POLES"],
    ["Pole A", "Political need for sovereignty rhetoric (legitimacy, mobilization)"],
    ["Pole B", "Material reality of structural dependency (supply chains, knowledge, equipment)"],
    [""],
    ["CREATIVE RESPONSES (how concept navigates the problem)"],
    ["", "Temporal displacement: 'we will achieve sovereignty' (future promise)"],
    ["", "Definitional shifting: redefining sovereignty as 'partial' or 'strategic'"],
    ["", "Domain restriction: sovereignty in 'key areas' only"],
    ["", "Rhetorical emphasis: celebrating inputs (investment) over outputs (capability)"],
    [""],
    ["BECOMINGS ENABLED (what transformations concept allows)"],
    ["", "Becoming-investor: states transformed into tech funders"],
    ["", "Becoming-protectionist: legitimate barriers to foreign tech"],
    ["", "Becoming-strategic: ordinary tech decisions elevated to national security"],
    [""],
    ["BECOMINGS BLOCKED (what transformations concept prevents)"],
    ["", "Becoming-honest: cannot admit structural impossibility"],
    ["", "Becoming-interdependent: cannot embrace strategic dependency"],
    ["", "Becoming-collaborative: global cooperation framed as threat"],
    [""],
    ["PLANE OF IMMANENCE (unquestioned background assumptions)"],
    ["Assumption", "Makes Possible", "Makes Impossible"],
    ["States are primary tech actors", "National tech policy", "Recognizing corporate/network primacy"],
    ["Technology can be 'owned' like territory", "Sovereignty framing", "Understanding tech as relational"],
    ["Autonomy is achievable through investment", "Resource mobilization", "Structural critique of dependency"],
    ["Competition is natural state", "Strategic framings", "Genuine cooperation models"],
]

for row_idx, row_data in enumerate(deleuzian_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_deleuzian.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Deleuzian"]
        elif row_idx == 22:
            cell.font = header_font
            cell.fill = header_fill
        elif any(x in str(value) for x in ["PROBLEM", "TENSION", "RESPONSES", "ENABLED", "BLOCKED", "PLANE"]):
            cell.font = Font(bold=True, italic=True)

ws_deleuzian.column_dimensions['A'].width = 35
ws_deleuzian.column_dimensions['B'].width = 45
ws_deleuzian.column_dimensions['C'].width = 40

# ============================================================================
# Sheet 6: Bachelardian - Epistemological Obstacles
# ============================================================================
ws_bachelardian = wb.create_sheet("5. Bachelardian")

bachelardian_data = [
    ["BACHELARDIAN ANALYSIS - Epistemological Obstacles"],
    [""],
    ["Attribute", "Value"],
    ["Is Obstacle?", "YES - this concept blocks understanding"],
    ["Obstacle Type", "VERBAL - uses familiar word (sovereignty) to colonize new domain (technology)"],
    [""],
    ["WHAT IT BLOCKS"],
    ["", "Recognition of structural dependency as permanent feature"],
    ["", "Analysis of power asymmetries in global tech system"],
    ["", "Questions about whether autonomy is possible or desirable"],
    ["", "Understanding technology as inherently relational/networked"],
    ["", "Class analysis of who benefits from sovereignty discourse"],
    [""],
    ["EVIDENCE OF INADEQUACY"],
    ["", "Gulf states: $2T wealth but cannot invest autonomously at Series B"],
    ["", "China: Massive investment but still depends on ASML for advanced chips"],
    ["", "Russia: Sanctions reveal depth of tech dependencies"],
    ["", "India: 'Digital sovereignty' claims while dependent on US cloud providers"],
    [""],
    ["WHY IT PERSISTS (ideological/class function)"],
    ["", "Legitimizes massive state tech spending"],
    ["", "Creates protected markets for domestic firms"],
    ["", "Provides nationalist narrative for political mobilization"],
    ["", "Justifies surveillance under 'data sovereignty' framing"],
    [""],
    ["RUPTURE WOULD ENABLE"],
    ["", "Honest analysis of what kinds of autonomy are achievable"],
    ["", "Strategies for dependency management rather than elimination"],
    ["", "Recognition of necessary interdependence"],
    ["", "New concepts: 'strategic positioning', 'dependency management', 'leverage points'"],
    [""],
    ["RUPTURE TRIGGER (what would force abandonment)"],
    ["", "Sustained demonstration that investment doesn't produce autonomy"],
    ["", "Political change that devalues nationalist framing"],
    ["", "Alternative concepts that better serve mobilization function"],
]

for row_idx, row_data in enumerate(bachelardian_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_bachelardian.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Bachelardian"]
        elif row_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
        elif any(x in str(value) for x in ["BLOCKS", "EVIDENCE", "PERSISTS", "ENABLE", "TRIGGER"]):
            cell.font = Font(bold=True, italic=True)
        elif value == "YES - this concept blocks understanding":
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

ws_bachelardian.column_dimensions['A'].width = 20
ws_bachelardian.column_dimensions['B'].width = 80

# ============================================================================
# Sheet 7: Canguilhem - Life History
# ============================================================================
ws_canguilhem = wb.create_sheet("6. Canguilhem")

canguilhem_data = [
    ["CANGUILHEM ANALYSIS - Life History of Concepts"],
    [""],
    ["Attribute", "Value"],
    ["Health Status", "STRAINED - concept under pressure, not yet dying"],
    ["Birth Period", "2010s (US-China tech tensions, Snowden revelations)"],
    ["Birth Problem", "How to mobilize resources for tech independence in era of perceived vulnerability"],
    [""],
    ["CONCEPT EVOLUTION"],
    ["Period", "Form", "Problem Driving Change", "Who Transformed"],
    ["1648 (Westphalia)", "Territorial sovereignty", "Ending religious wars, establishing state system", "European diplomats"],
    ["1950s-70s", "Economic sovereignty", "Decolonization, control of resources", "Newly independent states"],
    ["1990s-2000s", "Data sovereignty", "Internet governance, privacy concerns", "EU regulators, privacy advocates"],
    ["2010s-present", "Technological sovereignty", "US-China rivalry, Snowden, supply chain fears", "Tech nationalists, security hawks"],
    [""],
    ["TRAJECTORY (where concept is heading)"],
    ["", "EITHER: Dies - exposed as impossible, replaced by 'strategic positioning'"],
    ["", "OR: Transforms - 'sovereignty-within-subordination', 'managed dependency'"],
    ["", "OR: Fragments - different 'sovereignties' for different domains (AI, chips, cloud)"],
    [""],
    ["NORMATIVE DIMENSIONS"],
    ["Value Embedded", "Whose Values", "What's Excluded"],
    ["State control is good", "State elites, security apparatus", "Individual autonomy, corporate freedom"],
    ["Autonomy > interdependence", "Nationalists, protectionists", "Cosmopolitans, free traders"],
    ["Technology is strategic", "Military-industrial complex", "Technology as commons/public good"],
]

for row_idx, row_data in enumerate(canguilhem_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_canguilhem.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Canguilhem"]
        elif row_idx in [3, 8, 20]:
            cell.font = header_font
            cell.fill = header_fill
        elif any(x in str(value) for x in ["EVOLUTION", "TRAJECTORY", "NORMATIVE"]):
            cell.font = Font(bold=True, italic=True)
        elif value == "STRAINED - concept under pressure, not yet dying":
            cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

ws_canguilhem.column_dimensions['A'].width = 25
ws_canguilhem.column_dimensions['B'].width = 35
ws_canguilhem.column_dimensions['C'].width = 45
ws_canguilhem.column_dimensions['D'].width = 35

# ============================================================================
# Sheet 8: Davidson - Reasoning Styles
# ============================================================================
ws_davidson = wb.create_sheet("7. Davidson")

davidson_data = [
    ["DAVIDSON ANALYSIS - Reasoning Styles"],
    [""],
    ["Attribute", "Value"],
    ["Style Required", "GEOPOLITICAL REALISM - frames everything through state competition lens"],
    [""],
    ["WHAT THIS STYLE MAKES VISIBLE"],
    ["", "State competition and rivalry"],
    ["", "Strategic sectors and chokepoints"],
    ["", "National interest calculations"],
    ["", "Security vulnerabilities"],
    ["", "Balance of power dynamics"],
    [""],
    ["WHAT THIS STYLE MAKES INVISIBLE"],
    ["", "Class dynamics within states (who benefits from sovereignty discourse?)"],
    ["", "Corporate interests driving state policy"],
    ["", "Impossibility of autonomy under current global production"],
    ["", "Technology as inherently networked/relational"],
    ["", "Labor and knowledge flows that underpin tech development"],
    ["", "Environmental costs of tech nationalism (duplicate supply chains)"],
    [""],
    ["EVIDENCE TYPES PRIVILEGED"],
    ["", "State investments and policies"],
    ["", "Trade statistics and dependencies"],
    ["", "Military applications of technology"],
    ["", "Government statements and strategies"],
    [""],
    ["EVIDENCE TYPES MARGINALIZED"],
    ["", "Corporate profit motives"],
    ["", "Worker mobility and knowledge transfer"],
    ["", "Academic collaboration networks"],
    ["", "Actual capability assessments (vs. aspirational claims)"],
    [""],
    ["CHARACTERISTIC INFERENCE PATTERNS"],
    ["", "'If dependent, then vulnerable' (assumes dependency = weakness)"],
    ["", "'If strategic sector, then must be controlled' (control imperative)"],
    ["", "'If rival invests, we must too' (competitive logic)"],
    ["", "'If we invest enough, we will achieve autonomy' (investment → capability)"],
]

for row_idx, row_data in enumerate(davidson_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_davidson.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Davidson"]
        elif row_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
        elif any(x in str(value) for x in ["VISIBLE", "INVISIBLE", "PRIVILEGED", "MARGINALIZED", "INFERENCE"]):
            cell.font = Font(bold=True, italic=True)

ws_davidson.column_dimensions['A'].width = 15
ws_davidson.column_dimensions['B'].width = 80

# ============================================================================
# Sheet 9: Blumenberg - Metaphorology
# ============================================================================
ws_blumenberg = wb.create_sheet("8. Blumenberg")

blumenberg_data = [
    ["BLUMENBERG ANALYSIS - Metaphorology"],
    [""],
    ["ROOT METAPHOR"],
    ["Metaphor", "Sovereignty as territorial possession"],
    ["Source Domain", "Territorial/military control"],
    [""],
    ["WHAT METAPHOR ENABLES"],
    ["", "Thinking technology can be 'owned' like territory"],
    ["", "Applying border/control concepts to data and code"],
    ["", "Framing tech competition as territorial contest"],
    ["", "Justifying exclusion (like border control)"],
    ["", "'Defending' national technology from foreign 'invasion'"],
    [""],
    ["WHAT METAPHOR HIDES"],
    ["", "Technology is networked - not bounded like territory"],
    ["", "Technology requires ongoing relationships (maintenance, updates, expertise)"],
    ["", "Knowledge flows through people, not just products"],
    ["", "Digital goods are non-rivalrous (copying ≠ taking)"],
    ["", "Innovation requires openness more than control"],
    [""],
    ["RESISTS CONCEPTUALIZATION?", "YES - this is an 'absolute metaphor'"],
    ["Why It Resists", "Territory and technology have fundamentally different autonomy conditions:"],
    ["", "- Territory: bounded, defensible, excludable"],
    ["", "- Technology: networked, distributed, requires ongoing relationships"],
    ["", "The metaphor cannot be made literal without exposing its inadequacy"],
    [""],
    ["CONCEPTUAL WORK IN PROGRESS"],
    ["Original Meaning", "Sovereignty: supreme authority over territory"],
    ["Current Work", "Attempting to extend territorial sovereignty logic to technological domain"],
    ["Who's Doing Work", "Tech nationalists, security establishment, some academics"],
    ["Work Status", "FAILING - evidence accumulates against possibility of tech sovereignty"],
    [""],
    ["ALTERNATIVE METAPHORS EMERGING"],
    ["", "'Strategic positioning' - chess/game metaphor (position, not control)"],
    ["", "'Dependency management' - systems metaphor (flows, not borders)"],
    ["", "'Leverage points' - network metaphor (influence, not ownership)"],
]

for row_idx, row_data in enumerate(blumenberg_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_blumenberg.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Blumenberg"]
        elif any(x in str(value) for x in ["ROOT METAPHOR", "ENABLES", "HIDES", "RESISTS", "WORK IN", "ALTERNATIVE"]):
            cell.font = Font(bold=True, italic=True)
        elif value == "FAILING - evidence accumulates against possibility of tech sovereignty":
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

ws_blumenberg.column_dimensions['A'].width = 25
ws_blumenberg.column_dimensions['B'].width = 75

# ============================================================================
# Sheet 10: Carey - Bootstrapping
# ============================================================================
ws_carey = wb.create_sheet("9. Carey")

carey_data = [
    ["CAREY ANALYSIS - Conceptual Bootstrapping & Hierarchy"],
    [""],
    ["Attribute", "Value"],
    ["Hierarchy Level", "3 (bootstrapped from multiple complex concepts)"],
    ["Bootstrap Status", "FAILED - the bootstrap didn't achieve required qualitative leap"],
    [""],
    ["BUILT FROM (component concepts)"],
    ["Level 0 (Primitives)", "Territory, control, authority, boundary"],
    ["Level 1 (Simple)", "State, technology, autonomy"],
    ["Level 2 (Complex)", "Sovereignty (state + territory + supreme authority)"],
    ["Level 3 (Bootstrapped)", "Technological Sovereignty (sovereignty + technology domain)"],
    [""],
    ["COMBINATION TYPE", "QUALITATIVE LEAP attempted - tried to extend sovereignty logic to new domain"],
    ["Transparency", "LOW - components don't obviously combine to produce claimed result"],
    [""],
    ["WHY BOOTSTRAP FAILED"],
    ["", "Territorial sovereignty logic doesn't transfer to technology domain:"],
    ["", "- Territory is bounded; technology is networked"],
    ["", "- Territory is static; technology requires continuous flows"],
    ["", "- Territorial control is binary; tech 'control' is always partial"],
    ["", "- Territory defended by military; tech defended by... what?"],
    [""],
    ["EVIDENCE OF BOOTSTRAP FAILURE"],
    ["", "No country has achieved tech sovereignty despite massive investment"],
    ["", "Even US (most capable) depends on TSMC for advanced chips"],
    ["", "Concept keeps being redefined to accommodate failures"],
    ["", "Success criteria shift from 'autonomy' to 'reduced dependency'"],
    [""],
    ["WHAT WOULD FIX THE BOOTSTRAP"],
    ["", "New Level 2 concept: 'Strategic Technology Position' (not sovereignty)"],
    ["", "Built from: leverage, network position, dependencies (managed)"],
    ["", "Acknowledges: partial control, necessary interdependence, strategic choices"],
    ["", "Enables: honest assessment, practical strategies, realistic goals"],
    [""],
    ["ALTERNATIVE BOOTSTRAPS"],
    ["Concept", "Built From", "What It Enables"],
    ["'Strategic Positioning'", "Network position + leverage + influence", "Realistic assessment of what's achievable"],
    ["'Dependency Management'", "Flows + risks + diversification", "Practical strategies for resilience"],
    ["'Tech Statecraft'", "State capacity + international relations + tech", "Focus on influence rather than control"],
]

for row_idx, row_data in enumerate(carey_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_carey.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = dimension_font
            cell.fill = dimension_fills["Carey"]
        elif row_idx in [3, 33]:
            cell.font = header_font
            cell.fill = header_fill
        elif any(x in str(value) for x in ["BUILT FROM", "COMBINATION", "FAILED", "EVIDENCE", "FIX", "ALTERNATIVE"]):
            cell.font = Font(bold=True, italic=True)
        elif value == "FAILED - the bootstrap didn't achieve required qualitative leap":
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

ws_carey.column_dimensions['A'].width = 25
ws_carey.column_dimensions['B'].width = 50
ws_carey.column_dimensions['C'].width = 45

# ============================================================================
# Sheet 11: Summary - Impact Analysis Template
# ============================================================================
ws_summary = wb.create_sheet("Impact Summary")

summary_data = [
    ["EVIDENCE CLUSTER IMPACT ANALYSIS - Technological Sovereignty"],
    [""],
    ["When evidence challenges this concept, analyze impact across all 9 dimensions:"],
    [""],
    ["Dimension", "Key Question", "Example Evidence Finding", "Impact Assessment"],
    ["1. Quinean", "Does evidence contradict inferences?", "Gulf has $2T but can't invest at Series B", "Breaks 'investment → autonomy' inference"],
    ["2. Sellarsian", "Does evidence expose givenness?", "Shows sovereignty assumed not proven", "Reveals myth of given - should be inferred from capacity analysis"],
    ["3. Brandomian", "Are commitments violated?", "Claims autonomy, evidence shows dependency", "Major violation - entitlements exceeded"],
    ["4. Deleuzian", "New problems revealed?", "How to maintain legitimacy without capacity", "Yes - new tension between rhetoric and reality"],
    ["5. Bachelardian", "Evidence of obstacle?", "Concept prevents seeing structural dependency", "Yes - blocking honest analysis"],
    ["6. Canguilhem", "Health impact?", "Evidence shows concept failing its function", "Worsening - trajectory toward death or transformation"],
    ["7. Davidson", "Style inadequacy?", "Geopolitical frame misses class/capital dynamics", "Yes - alternative style needed"],
    ["8. Blumenberg", "Metaphor strained?", "Territory metaphor breaking down", "Yes - technology ≠ territory"],
    ["9. Carey", "Bootstrap failing?", "Can't achieve Level 3 qualitative leap", "Yes - need alternative bootstrap"],
    [""],
    ["OVERALL ASSESSMENT"],
    ["", "Impact Severity", "TRANSFORMATIVE - challenges concept at structural level"],
    ["", "Recommended Response", "RUPTURE - concept needs replacement, not refinement"],
    ["", "Alternative Concepts", "'Strategic Positioning', 'Dependency Management', 'Tech Statecraft'"],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_idx == 5:
            cell.font = header_font
            cell.fill = header_fill
        elif "OVERALL" in str(value):
            cell.font = Font(bold=True, italic=True)
        elif value == "RUPTURE - concept needs replacement, not refinement":
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cell.font = Font(bold=True)

ws_summary.column_dimensions['A'].width = 15
ws_summary.column_dimensions['B'].width = 35
ws_summary.column_dimensions['C'].width = 45
ws_summary.column_dimensions['D'].width = 45

# Save the workbook
output_path = Path("/home/evgeny/projects/theory-service/documentation/Technological_Sovereignty_9D_Analysis.xlsx")
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
