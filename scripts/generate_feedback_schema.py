#!/usr/bin/env python3
"""
Generate Excel schema for Evidence Testing Feedback System

This captures:
1. How evidence clusters test concept values
2. Structured feedback from LLM analysis
3. Editorial decisions and revisions
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
new_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Colors for different sheets
colors = {
    "overview": "3498DB",     # Blue
    "feedback": "E74C3C",     # Red
    "decision": "27AE60",     # Green
    "revision": "9B59B6",     # Purple
    "workflow": "F39C12",     # Orange
}

def add_table(ws, start_row, table_name, fields, note=None, color="34495E"):
    row = start_row
    if note:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9)
        row += 1

    ws.merge_cells(f'A{row}:D{row}')
    title_cell = ws.cell(row=row, column=1, value=table_name)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    row += 1

    headers = ["Field", "Type", "Constraints", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for f in fields:
        ws.cell(row=row, column=1, value=f["name"]).font = field_font
        ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
        c3 = ws.cell(row=row, column=3, value=f.get("constraints", ""))
        if "FK" in f.get("constraints", ""):
            c3.fill = fk_fill
        ws.cell(row=row, column=4, value=f["desc"])
        if f.get("new"):
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = new_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1
    return row + 1

# ============================================================================
# SHEET 1: Overview
# ============================================================================
ws = wb.active
ws.title = "1. Overview"

ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="EVIDENCE TESTING FEEDBACK SYSTEM")
cell.font = title_font
cell.fill = PatternFill(start_color=colors["overview"], end_color=colors["overview"], fill_type="solid")

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Purpose: Capture how evidence clusters test concept values and drive schema refinement").font = Font(bold=True)

# Workflow explanation
ws.merge_cells('A5:D5')
ws.cell(row=5, column=1, value="WORKFLOW:").font = Font(bold=True, size=12)

workflow_steps = [
    "1. EVIDENCE TESTING: Essay-flow sends evidence clusters to test against concept schema",
    "2. LLM ANALYSIS: Claude analyzes each cluster against relevant concept dimensions",
    "3. FEEDBACK GENERATION: Structured feedback created for tensions, confirmations, extensions",
    "4. EDITORIAL REVIEW: Human reviews feedback, creates decisions",
    "5. REVISION EXECUTION: Approved changes applied to concept schema",
    "6. AUDIT TRAIL: All changes tracked with full provenance",
]
row = 6
for step in workflow_steps:
    ws.cell(row=row, column=1, value=step)
    row += 1

# Schema overview
row += 1
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="TABLES IN THIS SCHEMA:").font = Font(bold=True, size=12)
row += 1

tables = [
    ("evidence_test_sessions", "Groups evidence testing into sessions"),
    ("evidence_tests", "Individual evidence cluster tests against concept"),
    ("test_feedback", "Structured feedback from LLM analysis"),
    ("feedback_tensions", "Specific tensions identified between evidence and concept"),
    ("feedback_confirmations", "Evidence that confirms concept values"),
    ("feedback_extensions", "Proposed extensions to concept from evidence"),
    ("editorial_decisions", "Human decisions on feedback items"),
    ("concept_revisions", "Actual changes made to concept schema"),
    ("revision_audit_log", "Complete audit trail of all changes"),
]

for table, desc in tables:
    ws.cell(row=row, column=1, value=f"• {table}").font = Font(bold=True)
    ws.cell(row=row, column=2, value=desc)
    row += 1

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 50

# ============================================================================
# SHEET 2: Evidence Testing Tables
# ============================================================================
ws = wb.create_sheet("2. Evidence Testing")

ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="EVIDENCE TESTING TABLES")
cell.font = title_font
cell.fill = PatternFill(start_color=colors["feedback"], end_color=colors["feedback"], fill_type="solid")

row = 3

# Table: evidence_test_sessions
row = add_table(ws, row, "evidence_test_sessions", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "project_id", "type": "INTEGER", "constraints": "NOT NULL", "desc": "Essay-flow project ID"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Concept being tested"},
    {"name": "session_name", "type": "VARCHAR(200)", "constraints": "", "desc": "Human-readable session name"},
    {"name": "llm_model", "type": "VARCHAR(50)", "constraints": "NOT NULL", "desc": "Model used (e.g., claude-opus-4-5-20251101)"},
    {"name": "dimensions_tested", "type": "TEXT[]", "constraints": "", "desc": "Which dimensions were tested"},
    {"name": "cluster_count", "type": "INTEGER", "constraints": "", "desc": "Number of clusters tested"},
    {"name": "status", "type": "VARCHAR(20)", "constraints": "DEFAULT 'pending'", "desc": "pending/running/completed/failed"},
    {"name": "started_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "completed_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": ""},
], note="Groups evidence testing into sessions for batch processing", color=colors["feedback"])

# Table: evidence_tests
row = add_table(ws, row, "evidence_tests", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "session_id", "type": "INTEGER", "constraints": "FK → evidence_test_sessions", "desc": "Parent session"},
    {"name": "cluster_id", "type": "INTEGER", "constraints": "NOT NULL", "desc": "Essay-flow cluster_id"},
    {"name": "cluster_name", "type": "TEXT", "constraints": "", "desc": "Cluster name for reference"},
    {"name": "cluster_summary", "type": "TEXT", "constraints": "", "desc": "AI summary of cluster content"},
    {"name": "dimension_tested", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "quinean/sellarsian/brandomian/etc"},
    {"name": "table_tested", "type": "VARCHAR(100)", "constraints": "", "desc": "Specific table tested"},
    {"name": "llm_prompt", "type": "TEXT", "constraints": "", "desc": "Prompt sent to LLM"},
    {"name": "llm_response", "type": "TEXT", "constraints": "", "desc": "Raw LLM response"},
    {"name": "tokens_used", "type": "INTEGER", "constraints": "", "desc": "API tokens consumed"},
    {"name": "processing_time_ms", "type": "INTEGER", "constraints": "", "desc": "LLM response time"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Individual evidence cluster tests", color=colors["feedback"])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 55

# ============================================================================
# SHEET 3: Feedback Tables
# ============================================================================
ws = wb.create_sheet("3. Feedback")

ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="STRUCTURED FEEDBACK TABLES")
cell.font = title_font
cell.fill = PatternFill(start_color=colors["feedback"], end_color=colors["feedback"], fill_type="solid")

row = 3

# Table: test_feedback
row = add_table(ws, row, "test_feedback", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "evidence_test_id", "type": "INTEGER", "constraints": "FK → evidence_tests, NOT NULL", "desc": "Parent test"},
    {"name": "feedback_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "tension/confirmation/extension/question"},
    {"name": "severity", "type": "VARCHAR(20)", "constraints": "", "desc": "critical/major/minor/informational"},
    {"name": "summary", "type": "TEXT", "constraints": "NOT NULL", "desc": "Brief description of finding"},
    {"name": "details", "type": "TEXT", "constraints": "", "desc": "Full explanation"},
    {"name": "affected_field", "type": "VARCHAR(100)", "constraints": "", "desc": "Specific field affected (e.g., inference_claim)"},
    {"name": "affected_value_id", "type": "INTEGER", "constraints": "", "desc": "ID of specific row affected"},
    {"name": "confidence", "type": "FLOAT", "constraints": "", "desc": "LLM's confidence in this finding"},
    {"name": "requires_human_review", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Needs editorial decision?"},
    {"name": "status", "type": "VARCHAR(20)", "constraints": "DEFAULT 'pending'", "desc": "pending/reviewed/actioned/dismissed"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Master feedback table - one row per finding", color=colors["feedback"])

# Table: feedback_tensions
row = add_table(ws, row, "feedback_tensions", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "feedback_id", "type": "INTEGER", "constraints": "FK → test_feedback, NOT NULL", "desc": "Parent feedback"},
    {"name": "tension_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "contradiction/limitation/over_claim/misfit"},
    {"name": "existing_claim", "type": "TEXT", "constraints": "", "desc": "What the concept schema currently says"},
    {"name": "evidence_claim", "type": "TEXT", "constraints": "", "desc": "What the evidence suggests"},
    {"name": "nature_of_tension", "type": "TEXT", "constraints": "", "desc": "Explanation of the incompatibility"},
    {"name": "suggested_resolution", "type": "TEXT", "constraints": "", "desc": "LLM's proposed resolution"},
    {"name": "resolution_type", "type": "VARCHAR(30)", "constraints": "", "desc": "revise/delete/contextualize/split"},
    {"name": "philosophical_significance", "type": "TEXT", "constraints": "", "desc": "Why this matters theoretically"},
], note="Tensions = when evidence challenges existing concept values", color=colors["feedback"])

# Table: feedback_confirmations
row = add_table(ws, row, "feedback_confirmations", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "feedback_id", "type": "INTEGER", "constraints": "FK → test_feedback, NOT NULL", "desc": "Parent feedback"},
    {"name": "confirmation_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "direct/indirect/extended"},
    {"name": "confirmed_claim", "type": "TEXT", "constraints": "", "desc": "Which claim is confirmed"},
    {"name": "supporting_evidence", "type": "TEXT", "constraints": "", "desc": "How evidence supports it"},
    {"name": "strength_increase", "type": "FLOAT", "constraints": "", "desc": "How much confidence should increase"},
    {"name": "new_example", "type": "TEXT", "constraints": "", "desc": "Example from evidence to add"},
], note="Confirmations = when evidence supports existing concept values", color=colors["feedback"])

# Table: feedback_extensions
row = add_table(ws, row, "feedback_extensions", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "feedback_id", "type": "INTEGER", "constraints": "FK → test_feedback, NOT NULL", "desc": "Parent feedback"},
    {"name": "extension_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "new_claim/new_component/new_relation/refinement"},
    {"name": "target_table", "type": "VARCHAR(100)", "constraints": "", "desc": "Which table to extend"},
    {"name": "proposed_value", "type": "TEXT", "constraints": "", "desc": "The new value to add"},
    {"name": "rationale", "type": "TEXT", "constraints": "", "desc": "Why this should be added"},
    {"name": "evidence_basis", "type": "TEXT", "constraints": "", "desc": "Evidence supporting this extension"},
    {"name": "related_existing_ids", "type": "INTEGER[]", "constraints": "", "desc": "IDs of related existing values"},
], note="Extensions = evidence suggests adding something new to concept", color=colors["feedback"])

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 55

# ============================================================================
# SHEET 4: Editorial Decisions
# ============================================================================
ws = wb.create_sheet("4. Editorial Decisions")

ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="EDITORIAL DECISION TABLES")
cell.font = title_font
cell.fill = PatternFill(start_color=colors["decision"], end_color=colors["decision"], fill_type="solid")

row = 3

# Table: editorial_decisions
row = add_table(ws, row, "editorial_decisions", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "feedback_id", "type": "INTEGER", "constraints": "FK → test_feedback, NOT NULL", "desc": "Feedback being decided on"},
    {"name": "decision", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "accept/reject/defer/modify"},
    {"name": "decision_rationale", "type": "TEXT", "constraints": "", "desc": "Why this decision was made"},
    {"name": "modifications", "type": "JSONB", "constraints": "", "desc": "If modified, what changes were made"},
    {"name": "priority", "type": "VARCHAR(20)", "constraints": "", "desc": "immediate/standard/low"},
    {"name": "assigned_to", "type": "VARCHAR(100)", "constraints": "", "desc": "Who will implement revision"},
    {"name": "decided_by", "type": "VARCHAR(100)", "constraints": "", "desc": "Who made the decision"},
    {"name": "decided_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "execution_status", "type": "VARCHAR(20)", "constraints": "DEFAULT 'pending'", "desc": "pending/in_progress/completed"},
], note="Human decisions on LLM feedback items", color=colors["decision"])

# Table: decision_batch
row = add_table(ws, row, "decision_batches", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "session_id", "type": "INTEGER", "constraints": "FK → evidence_test_sessions", "desc": "Evidence session this batch is for"},
    {"name": "batch_name", "type": "VARCHAR(200)", "constraints": "", "desc": "Human-readable name"},
    {"name": "total_items", "type": "INTEGER", "constraints": "", "desc": "Total feedback items in batch"},
    {"name": "accepted", "type": "INTEGER", "constraints": "", "desc": "Count of accepted"},
    {"name": "rejected", "type": "INTEGER", "constraints": "", "desc": "Count of rejected"},
    {"name": "deferred", "type": "INTEGER", "constraints": "", "desc": "Count of deferred"},
    {"name": "notes", "type": "TEXT", "constraints": "", "desc": "Batch-level notes"},
    {"name": "created_by", "type": "VARCHAR(100)", "constraints": "", "desc": "Who processed this batch"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="For batch processing multiple decisions at once", color=colors["decision"])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 55

# ============================================================================
# SHEET 5: Revision Tables
# ============================================================================
ws = wb.create_sheet("5. Revisions")

ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="REVISION & AUDIT TABLES")
cell.font = title_font
cell.fill = PatternFill(start_color=colors["revision"], end_color=colors["revision"], fill_type="solid")

row = 3

# Table: concept_revisions
row = add_table(ws, row, "concept_revisions", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "decision_id", "type": "INTEGER", "constraints": "FK → editorial_decisions, NOT NULL", "desc": "Decision that triggered this"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Concept being revised"},
    {"name": "target_table", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Table being changed"},
    {"name": "target_row_id", "type": "INTEGER", "constraints": "", "desc": "Specific row (for updates/deletes)"},
    {"name": "operation", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "INSERT/UPDATE/DELETE"},
    {"name": "old_value", "type": "JSONB", "constraints": "", "desc": "Previous state (for rollback)"},
    {"name": "new_value", "type": "JSONB", "constraints": "", "desc": "New state"},
    {"name": "field_changed", "type": "VARCHAR(100)", "constraints": "", "desc": "Specific field (for updates)"},
    {"name": "revision_note", "type": "TEXT", "constraints": "", "desc": "Explanation of change"},
    {"name": "executed_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "executed_by", "type": "VARCHAR(100)", "constraints": "", "desc": "System or user who applied"},
    {"name": "can_rollback", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Is rollback possible?"},
], note="Actual changes made to concept schema", color=colors["revision"])

# Table: revision_audit_log
row = add_table(ws, row, "revision_audit_log", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "revision_id", "type": "INTEGER", "constraints": "FK → concept_revisions", "desc": "Linked revision (if applicable)"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Concept affected"},
    {"name": "action", "type": "VARCHAR(50)", "constraints": "NOT NULL", "desc": "What happened"},
    {"name": "actor", "type": "VARCHAR(100)", "constraints": "", "desc": "Who/what performed action"},
    {"name": "actor_type", "type": "VARCHAR(20)", "constraints": "", "desc": "user/llm/system/api"},
    {"name": "evidence_session_id", "type": "INTEGER", "constraints": "", "desc": "If from evidence testing"},
    {"name": "project_id", "type": "INTEGER", "constraints": "", "desc": "If from specific project"},
    {"name": "details", "type": "JSONB", "constraints": "", "desc": "Full details of action"},
    {"name": "ip_address", "type": "INET", "constraints": "", "desc": "For security audit"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Complete audit trail of all concept changes", color=colors["revision"])

# Table: revision_rollback
row = add_table(ws, row, "revision_rollbacks", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "revision_id", "type": "INTEGER", "constraints": "FK → concept_revisions, NOT NULL", "desc": "Revision being rolled back"},
    {"name": "reason", "type": "TEXT", "constraints": "", "desc": "Why rolling back"},
    {"name": "rolled_back_by", "type": "VARCHAR(100)", "constraints": "", "desc": "Who initiated rollback"},
    {"name": "rolled_back_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "success", "type": "BOOLEAN", "constraints": "", "desc": "Did rollback succeed?"},
    {"name": "error_message", "type": "TEXT", "constraints": "", "desc": "If failed, why"},
], note="Track rollback operations for audit", color=colors["revision"])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 55

# ============================================================================
# SHEET 6: Example Workflow
# ============================================================================
ws = wb.create_sheet("6. Example Workflow")

ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="EXAMPLE: Testing 'Technological Sovereignty' Against Evidence")
cell.font = title_font
cell.fill = PatternFill(start_color=colors["workflow"], end_color=colors["workflow"], fill_type="solid")

row = 3

# Step 1
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="STEP 1: Evidence Test Session Created").font = Font(bold=True, size=12)
row += 1

example_session = [
    ("project_id", "2"),
    ("concept_id", "15 (Tech Sovereignty)"),
    ("dimensions_tested", "['quinean', 'brandomian', 'deleuzian']"),
    ("cluster_count", "12"),
    ("llm_model", "claude-opus-4-5-20251101"),
]
for field, value in example_session:
    ws.cell(row=row, column=1, value=field).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)
    row += 1

row += 1

# Step 2
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="STEP 2: LLM Tests Cluster Against Brandomian Dimension").font = Font(bold=True, size=12)
row += 1

ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="Cluster: 'EU Chips Act funding mechanisms'").font = Font(italic=True)
row += 1
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="Testing against: concept_inferential_roles.inference_claim").font = Font(italic=True)
row += 2

# Step 3
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="STEP 3: Feedback Generated").font = Font(bold=True, size=12)
row += 1

ws.cell(row=row, column=1, value="Type:").font = Font(bold=True)
ws.cell(row=row, column=2, value="TENSION")
row += 1
ws.cell(row=row, column=1, value="Severity:").font = Font(bold=True)
ws.cell(row=row, column=2, value="major")
row += 1
ws.cell(row=row, column=1, value="Existing claim:").font = Font(bold=True)
ws.cell(row=row, column=2, value="'Tech sovereignty → right to exclude foreign tech'")
row += 1
ws.cell(row=row, column=1, value="Evidence suggests:").font = Font(bold=True)
ws.cell(row=row, column=2, value="'EU Chips Act creates dependency on US firms, not exclusion'")
row += 1
ws.cell(row=row, column=1, value="Suggested resolution:").font = Font(bold=True)
ws.cell(row=row, column=2, value="'Contextualize: add limitation_type = strategic_sectors_only'")
row += 2

# Step 4
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="STEP 4: Editorial Decision").font = Font(bold=True, size=12)
row += 1

ws.cell(row=row, column=1, value="Decision:").font = Font(bold=True)
ws.cell(row=row, column=2, value="ACCEPT (with modification)")
row += 1
ws.cell(row=row, column=1, value="Rationale:").font = Font(bold=True)
ws.cell(row=row, column=2, value="Evidence is compelling; need to add scope limitation")
row += 1
ws.cell(row=row, column=1, value="Priority:").font = Font(bold=True)
ws.cell(row=row, column=2, value="standard")
row += 2

# Step 5
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="STEP 5: Revision Executed").font = Font(bold=True, size=12)
row += 1

ws.cell(row=row, column=1, value="Table:").font = Font(bold=True)
ws.cell(row=row, column=2, value="concept_inferential_roles")
row += 1
ws.cell(row=row, column=1, value="Operation:").font = Font(bold=True)
ws.cell(row=row, column=2, value="UPDATE")
row += 1
ws.cell(row=row, column=1, value="Old value:").font = Font(bold=True)
ws.cell(row=row, column=2, value="limitation_type: NULL")
row += 1
ws.cell(row=row, column=1, value="New value:").font = Font(bold=True)
ws.cell(row=row, column=2, value="limitation_type: 'strategic_sectors_only'")
row += 1
ws.cell(row=row, column=1, value="Source tracking:").font = Font(bold=True)
ws.cell(row=row, column=2, value="source_type='evidence_testing', source_reference='cluster_42, project_2'")

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 40

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Evidence_Feedback_Schema.xlsx")
wb.save(output_path)
print(f"Feedback schema saved to: {output_path}")
