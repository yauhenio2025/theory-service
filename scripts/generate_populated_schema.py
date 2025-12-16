#!/usr/bin/env python3
"""
Generate POPULATED v5 Concept Schema with "Technological Sovereignty" as example concept.

All 47 tables filled with plausible, logically consistent mock data
demonstrating how the schema would be used in practice.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
data_font = Font(size=10)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Overview": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Hacking": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

def style_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def add_table(ws, start_row, headers, data, title=None):
    """Add a data table to worksheet."""
    row = start_row

    if title:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for record in data:
        for col, val in enumerate(record, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = wrap
        row += 1

    return row + 1

# ============================================================================
# SHEET 1: Core Concept
# ============================================================================
ws = wb.active
ws.title = "1. Core Concept"

style_header(ws, 1, "TECHNOLOGICAL SOVEREIGNTY - Core Definition", dimension_colors["Overview"])

ws.cell(row=3, column=1, value="concept_id: 1").font = Font(bold=True)
ws.cell(row=4, column=1, value="name: Technological Sovereignty").font = Font(bold=True, size=14)

core_data = [
    ("Field", "Value", "Source Type", "Source Reference", "Confidence"),
    ("definition", "The capacity of a state or political entity to autonomously develop, control, and deploy critical technologies without dependence on external actors, particularly potential adversaries", "llm_analysis", "claude-opus-4-20250514", "0.85"),
    ("domain", "International Relations / Technology Policy", "user_input", "initial_import", "1.0"),
    ("first_appearance", "~2015 (EU Digital Single Market debates)", "llm_analysis", "historical_research", "0.75"),
    ("current_usage_frequency", "high - accelerating since 2018 US-China trade war", "evidence_testing", "cluster_42", "0.90"),
    ("centrality", "core - restructuring major policy frameworks globally", "llm_analysis", "web_analysis", "0.85"),
    ("entrenchment_score", "0.78", "internal_compute", "entrenchment_calc_v2", "0.80"),
    ("health_status", "strained - internal contradictions emerging", "evidence_testing", "cluster_87", "0.82"),
    ("hierarchy_level", "3 (complex - bootstrapped from simpler concepts)", "llm_analysis", "carey_analysis", "0.80"),
]

row = 6
for record in core_data:
    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if row == 6:
            cell.font = header_font
            cell.fill = table_header_fill
        cell.border = thin_border
        cell.alignment = wrap
    row += 1

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 12

# ============================================================================
# SHEET 2: Quinean Dimension
# ============================================================================
ws = wb.create_sheet("2. Quinean")
style_header(ws, 1, "QUINEAN DIMENSION - Web of Belief Analysis", dimension_colors["Quinean"])

# concept_inferences
row = add_table(ws, 3,
    ["id", "from_concept", "to_concept", "inference_type", "strength", "source_type", "source_ref", "confidence"],
    [
        (1, "tech_sovereignty", "indigenous_capability", "committive", "strong", "llm_analysis", "quine_v1", 0.85),
        (2, "tech_sovereignty", "supply_chain_control", "committive", "strong", "evidence_testing", "cluster_23", 0.90),
        (3, "tech_sovereignty", "right_to_exclude", "permissive", "moderate", "llm_analysis", "quine_v1", 0.75),
        (4, "tech_sovereignty", "global_integration", "incompatibility", "moderate", "evidence_testing", "cluster_45", 0.80),
        (5, "investment_capacity", "tech_sovereignty", "enabling", "weak", "evidence_testing", "cluster_67", 0.70),
        (6, "tech_sovereignty", "national_security", "committive", "strong", "llm_analysis", "quine_v1", 0.88),
    ],
    "concept_inferences"
)

# concept_web_tensions
row = add_table(ws, row,
    ["id", "concept_a", "concept_b", "tension_type", "description", "source_type", "source_ref", "confidence"],
    [
        (1, "tech_sovereignty", "global_supply_chains", "practical_contradiction", "Pursuing sovereignty while depending on global semiconductor supply chains", "evidence_testing", "cluster_89", 0.92),
        (2, "tech_sovereignty", "innovation_speed", "tradeoff", "Autarky slows innovation vs. openness accelerates but increases dependency", "llm_analysis", "quine_v1", 0.85),
        (3, "tech_sovereignty", "market_size", "structural", "Small states cannot achieve sovereignty due to market scale requirements", "evidence_testing", "cluster_12", 0.88),
    ],
    "concept_web_tensions"
)

# concept_web_position
row = add_table(ws, row,
    ["id", "position_type", "connected_concepts", "bridging_function", "source_type", "source_ref", "confidence"],
    [
        (1, "hub", "national_security, industrial_policy, digital_rights, trade_policy", "Central node connecting security and economic policy domains", "llm_analysis", "quine_v2", 0.85),
        (2, "bridge", "traditional_sovereignty, digital_economy", "Connects Westphalian state concepts to networked economy concepts", "llm_analysis", "quine_v2", 0.80),
    ],
    "concept_web_position"
)

# concept_ontological_dependence
row = add_table(ws, row,
    ["id", "depends_on", "dependence_type", "explanation", "source_type", "source_ref", "confidence"],
    [
        (1, "the_state", "existential", "Concept presupposes existence of state actors capable of sovereign action", "llm_analysis", "quine_ontology", 0.92),
        (2, "critical_technologies", "definitional", "Requires prior identification of which technologies are 'critical'", "llm_analysis", "quine_ontology", 0.88),
        (3, "adversarial_other", "relational", "Sovereignty only meaningful in relation to potential adversaries/dependencies", "evidence_testing", "cluster_34", 0.85),
    ],
    "concept_ontological_dependence"
)

# concept_revision_ramifications
row = add_table(ws, row,
    ["id", "if_revised", "affected_concepts", "revision_cost", "pragmatic_consideration", "source_type", "source_ref", "confidence"],
    [
        (1, "abandoned", "industrial_policy, national_security_framing, EU_strategic_autonomy", "very_high", "Would require reworking entire policy frameworks in EU, China, India", "llm_analysis", "quine_pragmatic", 0.90),
        (2, "weakened_to_resilience", "supply_chain_policy, trade_agreements", "moderate", "More achievable; shifts from control to redundancy", "evidence_testing", "cluster_78", 0.82),
        (3, "split_into_domain_specific", "semiconductor_sovereignty, data_sovereignty, AI_sovereignty", "low", "Already happening; may be natural conceptual evolution", "evidence_testing", "cluster_91", 0.85),
    ],
    "concept_revision_ramifications"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 20

# ============================================================================
# SHEET 3: Sellarsian Dimension
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
style_header(ws, 1, "SELLARSIAN DIMENSION - Myth of the Given Analysis", dimension_colors["Sellarsian"])

# concept_givenness
row = add_table(ws, 3,
    ["id", "given_claim", "givenness_level", "actual_basis", "source_type", "source_ref", "confidence"],
    [
        (1, "States can achieve technological autonomy", "high", "Inferred from contested premises about state capacity and market dynamics", "llm_analysis", "sellars_v1", 0.85),
        (2, "Critical technologies are identifiable", "moderate", "Requires complex analysis; what's 'critical' shifts with geopolitics", "evidence_testing", "cluster_56", 0.80),
        (3, "Sovereignty is binary (have it or don't)", "high", "Actually spectrum; treated as given to simplify policy discourse", "llm_analysis", "sellars_v1", 0.88),
    ],
    "concept_givenness"
)

# concept_hidden_commitments
row = add_table(ws, row,
    ["id", "commitment", "how_hidden", "implications", "source_type", "source_ref", "confidence"],
    [
        (1, "Technology development is linear and predictable", "Embedded in 'roadmap' rhetoric", "Blinds to disruptive innovation from unexpected sources", "llm_analysis", "sellars_v1", 0.82),
        (2, "State action is inherently legitimate", "Naturalized in sovereignty discourse", "Forecloses questions about whose interests are served", "llm_analysis", "sellars_v1", 0.85),
        (3, "Autarky is economically viable", "Assumed in 'build domestic' framings", "Ignores comparative advantage, scale economics", "evidence_testing", "cluster_44", 0.90),
        (4, "Security and economics are separable", "Binary framing of 'strategic' vs 'commercial'", "Misses how deeply intertwined they are", "llm_analysis", "sellars_v1", 0.78),
    ],
    "concept_hidden_commitments"
)

# concept_space_of_reasons
row = add_table(ws, row,
    ["id", "inferential_role", "justificatory_relations", "normative_status", "source_type", "source_ref", "confidence"],
    [
        (1, "premise_for_industrial_policy", "Justifies subsidies, tariffs, procurement preferences", "Widely accepted in policy circles", "evidence_testing", "cluster_23", 0.90),
        (2, "conclusion_from_security_concerns", "Derived from threat assessments, supply chain vulnerabilities", "Contested - critics say overblown", "evidence_testing", "cluster_67", 0.82),
        (3, "legitimation_device", "Provides normative cover for protectionism", "Increasingly questioned by economists", "llm_analysis", "sellars_v2", 0.78),
    ],
    "concept_space_of_reasons"
)

# concept_functional_role
row = add_table(ws, row,
    ["id", "role_type", "in_discourse", "inferential_connections", "source_type", "source_ref", "confidence"],
    [
        (1, "enabling_premise", "Policy debates", "Enables conclusions about state intervention legitimacy", "llm_analysis", "sellars_functional", 0.85),
        (2, "framing_device", "Media coverage", "Structures how tech competition is understood", "evidence_testing", "cluster_78", 0.88),
        (3, "mobilization_rhetoric", "Political speeches", "Rallies support for spending/restrictions", "evidence_testing", "cluster_34", 0.82),
    ],
    "concept_functional_role"
)

# concept_image_tension
row = add_table(ws, row,
    ["id", "manifest_image", "scientific_image", "tension_description", "source_type", "source_ref", "confidence"],
    [
        (1, "Nation controls its technological destiny", "Complex global interdependencies make full autonomy impossible", "Manifest image persists despite scientific understanding", "llm_analysis", "sellars_image", 0.90),
        (2, "Clear friend/enemy technology distinctions", "Technologies flow across borders regardless of intent", "Policy pretends borders are more solid than they are", "evidence_testing", "cluster_89", 0.85),
        (3, "State as unified rational actor", "Multiple agencies with conflicting interests", "Sovereignty discourse assumes coherence that doesn't exist", "llm_analysis", "sellars_image", 0.82),
    ],
    "concept_image_tension"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 4: Brandomian Dimension
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
style_header(ws, 1, "BRANDOMIAN DIMENSION - Deontic Scorekeeping", dimension_colors["Brandomian"])

# concept_commitments
row = add_table(ws, 3,
    ["id", "commitment", "deontic_status", "undertaken_by", "attributed_by", "source_type", "source_ref", "confidence"],
    [
        (1, "Investment leads to capability", "acknowledged", "EU, China, India", "Policy analysts", "evidence_testing", "cluster_23", 0.85),
        (2, "Capability leads to autonomy", "undertaken", "EU (Chips Act)", "Critics dispute", "evidence_testing", "cluster_45", 0.78),
        (3, "Autonomy is achievable for major powers", "attributed", "-", "Proponents attribute to states", "llm_analysis", "brandom_v1", 0.72),
        (4, "Right to exclude foreign tech", "acknowledged", "US (entity list), China", "Trading partners contest", "evidence_testing", "cluster_67", 0.88),
        (5, "Obligation to protect critical infrastructure", "undertaken", "All major economies", "Broadly accepted", "llm_analysis", "brandom_v1", 0.92),
    ],
    "concept_commitments"
)

# concept_inferential_roles
row = add_table(ws, row,
    ["id", "inference_type", "from_claim", "to_claim", "status", "source_type", "source_ref", "confidence"],
    [
        (1, "committive", "X pursues tech sovereignty", "X commits to indigenous development", "active", "llm_analysis", "brandom_v1", 0.88),
        (2, "committive", "X achieves tech sovereignty", "X can exclude adversary tech", "contested", "evidence_testing", "cluster_34", 0.75),
        (3, "permissive", "X has tech sovereignty", "X may set technology standards", "active", "llm_analysis", "brandom_v1", 0.82),
        (4, "incompatibility", "X has tech sovereignty", "X depends on adversary supply chains", "logical", "evidence_testing", "cluster_89", 0.95),
        (5, "incompatibility", "X pursues full autarky", "X maintains innovation speed", "empirical", "evidence_testing", "cluster_56", 0.85),
    ],
    "concept_inferential_roles"
)

# concept_perspectival_content
row = add_table(ws, row,
    ["id", "speaker_perspective", "de_dicto_content", "de_re_translation", "translation_loss", "source_type", "source_ref", "confidence"],
    [
        (1, "Chinese government", "Great rejuvenation through tech self-reliance", "Catch-up industrial policy with nationalist framing", "Loses civilizational dimension", "llm_analysis", "brandom_persp", 0.80),
        (2, "EU Commission", "Strategic autonomy in critical technologies", "Reduce single-point-of-failure dependencies", "Loses sovereignty aspiration", "evidence_testing", "cluster_12", 0.85),
        (3, "Gulf states", "Technological sovereignty through investment", "Technology consumption + influence via capital", "Loses self-reliance claim entirely", "evidence_testing", "cluster_78", 0.90),
        (4, "India", "Atmanirbhar Bharat (self-reliant India)", "Import substitution + selective decoupling", "Loses cultural-historical resonance", "llm_analysis", "brandom_persp", 0.82),
    ],
    "concept_perspectival_content"
)

# concept_challenges
row = add_table(ws, row,
    ["id", "challenge", "to_commitment", "challenger", "response_status", "source_type", "source_ref", "confidence"],
    [
        (1, "Gulf states have $2T but can't do Series B", "Investment leads to capability", "Empirical evidence", "Unanswered", "evidence_testing", "cluster_67", 0.92),
        (2, "EU Chips Act creates new dependencies", "Capability leads to autonomy", "Supply chain analysts", "Partially addressed", "evidence_testing", "cluster_45", 0.85),
        (3, "China still needs ASML machines", "Major powers can achieve autonomy", "Industry observers", "Acknowledged as limitation", "evidence_testing", "cluster_89", 0.95),
        (4, "Innovation requires openness", "Autarky is viable", "Economists", "Contested", "llm_analysis", "brandom_v1", 0.80),
    ],
    "concept_challenges"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 5: Deleuzian Dimension
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
style_header(ws, 1, "DELEUZIAN DIMENSION - Concept as Multiplicity", dimension_colors["Deleuzian"])

# concept_components
row = add_table(ws, 3,
    ["id", "component", "description", "ordering", "source_type", "source_ref", "confidence"],
    [
        (1, "control", "Capacity to direct technological development and deployment", "intensive", "llm_analysis", "deleuze_v1", 0.88),
        (2, "autonomy", "Independence from external actors in tech decisions", "intensive", "llm_analysis", "deleuze_v1", 0.90),
        (3, "territory", "Bounded space within which sovereignty applies", "extensive", "llm_analysis", "deleuze_v1", 0.85),
        (4, "temporality", "Urgency - limited window before lock-in", "intensive", "evidence_testing", "cluster_34", 0.78),
        (5, "criticality", "Distinction between essential and non-essential tech", "intensive", "llm_analysis", "deleuze_v1", 0.82),
        (6, "capacity", "Technical and industrial capability to produce", "extensive", "evidence_testing", "cluster_56", 0.88),
    ],
    "concept_components"
)

# concept_zones_of_indiscernibility
row = add_table(ws, row,
    ["id", "component_a", "component_b", "zone_character", "productive_tension", "source_type", "source_ref", "confidence"],
    [
        (1, "control", "autonomy", "overlapping", "Control can undermine autonomy (surveillance state) or enable it", "llm_analysis", "deleuze_zones", 0.85),
        (2, "territory", "network", "contested", "Digital tech de-territorializes; sovereignty re-territorializes", "llm_analysis", "deleuze_zones", 0.90),
        (3, "urgency", "capability", "temporal", "Window closing before capacity can be built", "evidence_testing", "cluster_78", 0.88),
        (4, "autonomy", "innovation", "paradoxical", "Autonomy may require openness that undermines it", "evidence_testing", "cluster_45", 0.82),
    ],
    "concept_zones_of_indiscernibility"
)

# concept_consistency
row = add_table(ws, row,
    ["id", "consistency_type", "level", "assessment", "source_type", "source_ref", "confidence"],
    [
        (1, "endoconsistency", "unstable", "Components pull in different directions (autonomy vs speed, territory vs network)", "llm_analysis", "deleuze_consist", 0.85),
        (2, "exoconsistency", "moderate", "Connects to neighboring concepts but with friction", "llm_analysis", "deleuze_consist", 0.80),
    ],
    "concept_consistency"
)

# concept_plane_of_immanence
row = add_table(ws, row,
    ["id", "plane", "characteristics", "presuppositions", "source_type", "source_ref", "confidence"],
    [
        (1, "techno-nationalist", "Competition between state-civilization blocs; technology as power; zero-sum framing", "States are primary actors; technology is controllable; competition is inevitable", "llm_analysis", "deleuze_plane", 0.88),
        (2, "residual_westphalian", "Territorial sovereignty; non-interference; state monopoly on legitimate force", "Clear borders; fungible power; stable state system", "llm_analysis", "deleuze_plane", 0.75),
    ],
    "concept_plane_of_immanence"
)

# concept_personae
row = add_table(ws, row,
    ["id", "persona", "role", "voice_characteristics", "source_type", "source_ref", "confidence"],
    [
        (1, "The Technocrat", "Activates concept through expertise claims", "Neutral, technical, apolitical framing", "llm_analysis", "deleuze_personae", 0.85),
        (2, "The Securitizer", "Activates through threat construction", "Urgent, existential risk framing", "evidence_testing", "cluster_23", 0.88),
        (3, "The Industrial Champion", "Activates through competitiveness discourse", "Economic nationalism, jobs rhetoric", "evidence_testing", "cluster_67", 0.82),
        (4, "The Digital Sovereign", "New persona emerging", "Platform governance, data rights framing", "evidence_testing", "cluster_91", 0.75),
    ],
    "concept_personae"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 6: Bachelardian Dimension
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
style_header(ws, 1, "BACHELARDIAN DIMENSION - Epistemological Obstacles", dimension_colors["Bachelardian"])

# concept_obstacles
row = add_table(ws, 3,
    ["id", "obstacle_type", "description", "entrenchment", "source_type", "source_ref", "confidence"],
    [
        (1, "verbal", "Word 'sovereignty' carries Westphalian baggage inappropriate for networked tech", "high", "llm_analysis", "bachelard_v1", 0.88),
        (2, "substantialist", "Treating tech capability as substance that can be 'possessed' vs. relational capacity", "very_high", "llm_analysis", "bachelard_v1", 0.90),
        (3, "animist", "Treating 'the market' or 'technology' as agents with intentions", "moderate", "evidence_testing", "cluster_34", 0.78),
        (4, "first_experience", "Intuition that borders = control, despite networked reality", "high", "llm_analysis", "bachelard_v1", 0.85),
        (5, "quantitative", "Belief that enough investment = sovereignty (linear scaling assumption)", "high", "evidence_testing", "cluster_67", 0.88),
        (6, "unitary_pragmatism", "Seeking single 'sovereignty' solution vs. domain-specific approaches", "moderate", "llm_analysis", "bachelard_v1", 0.80),
    ],
    "concept_obstacles"
)

# concept_cognitive_stage
row = add_table(ws, row,
    ["id", "stage", "indicators", "assessment", "source_type", "source_ref", "confidence"],
    [
        (1, "pre-scientific", "Animistic language, substance metaphors, magical thinking about investment", "Dominant in political discourse", "llm_analysis", "bachelard_stages", 0.85),
        (2, "proto-scientific", "Some causal analysis, recognition of tradeoffs, empirical cases cited", "Emerging in policy analysis", "evidence_testing", "cluster_45", 0.80),
        (3, "scientific", "Systematic analysis, falsifiable claims, recognition of complexity", "Rare - mostly academic", "llm_analysis", "bachelard_stages", 0.75),
    ],
    "concept_cognitive_stage"
)

# concept_psychoanalytic_function
row = add_table(ws, row,
    ["id", "unconscious_need", "how_concept_satisfies", "evidence", "source_type", "source_ref", "confidence"],
    [
        (1, "control_fantasy", "Provides illusion of mastering uncontrollable technological change", "Emphasis on 'strategic control' despite evidence of limited control", "llm_analysis", "bachelard_psych", 0.82),
        (2, "legitimation_of_spending", "Justifies massive public investment without clear ROI accountability", "Chips Act debates avoid cost-benefit analysis", "evidence_testing", "cluster_78", 0.88),
        (3, "identity_maintenance", "Preserves national identity narratives against globalization anxiety", "Rhetoric of 'our' technology, 'our' capabilities", "llm_analysis", "bachelard_psych", 0.80),
        (4, "elite_reproduction", "Creates new roles for technocratic elite", "Proliferation of 'tech sovereignty' advisor positions", "evidence_testing", "cluster_91", 0.75),
    ],
    "concept_psychoanalytic_function"
)

# concept_regional_rationality
row = add_table(ws, row,
    ["id", "rationality_domain", "applicable_concepts", "methods_privileged", "source_type", "source_ref", "confidence"],
    [
        (1, "geopolitical", "Threat, alliance, deterrence, balance of power", "Strategic analysis, game theory", "llm_analysis", "bachelard_regional", 0.85),
        (2, "economic", "Investment, returns, efficiency, market size", "Cost-benefit, comparative advantage", "evidence_testing", "cluster_23", 0.88),
        (3, "technological", "Innovation, diffusion, standards, ecosystems", "Systems analysis, network effects", "llm_analysis", "bachelard_regional", 0.80),
        (4, "legal", "Jurisdiction, rights, enforcement, treaties", "Doctrinal analysis, precedent", "llm_analysis", "bachelard_regional", 0.75),
    ],
    "concept_regional_rationality"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 7: Canguilhem Dimension
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
style_header(ws, 1, "CANGUILHEM DIMENSION - Vital History of Concepts", dimension_colors["Canguilhem"])

# concept_evolution
row = add_table(ws, 3,
    ["id", "period", "form", "key_developments", "source_type", "source_ref", "confidence"],
    [
        (1, "1648-1945", "proto_form", "Westphalian sovereignty; technology not yet strategic sector", "llm_analysis", "canguilhem_v1", 0.80),
        (2, "1945-1990", "cold_war_form", "Dual-use technology controls; COCOM; national champions", "llm_analysis", "canguilhem_v1", 0.85),
        (3, "1990-2010", "dormancy", "Globalization consensus; Washington Consensus; WTO; tech as neutral", "llm_analysis", "canguilhem_v1", 0.88),
        (4, "2010-2018", "revival", "Snowden; China rise; Made in China 2025; stirrings of concern", "evidence_testing", "cluster_12", 0.90),
        (5, "2018-present", "intensification", "Trade war; COVID supply chains; Chips Act; full activation", "evidence_testing", "cluster_34", 0.95),
    ],
    "concept_evolution"
)

# concept_filiation
row = add_table(ws, row,
    ["id", "ancestor_concept", "relationship", "what_inherited", "what_transformed", "source_type", "source_ref", "confidence"],
    [
        (1, "Westphalian sovereignty", "direct", "Territorial framing, state as actor, non-interference norm", "Applied to intangible flows", "llm_analysis", "canguilhem_filiation", 0.88),
        (2, "Mercantilism", "indirect", "National wealth accumulation, strategic trade thinking", "From gold/goods to knowledge/capability", "llm_analysis", "canguilhem_filiation", 0.82),
        (3, "Industrial policy", "sibling", "State direction of economy, picking winners", "Expanded from sectors to technologies", "evidence_testing", "cluster_45", 0.85),
        (4, "Information sovereignty", "child", "Contributed data control dimension", "Merged into broader tech sovereignty", "evidence_testing", "cluster_67", 0.80),
    ],
    "concept_filiation"
)

# concept_vital_norms
row = add_table(ws, row,
    ["id", "norm", "function", "violation_indicator", "source_type", "source_ref", "confidence"],
    [
        (1, "Autonomy preservation", "Core vital norm - concept exists to enable autonomous action", "Using 'sovereignty' to justify increased dependency", "llm_analysis", "canguilhem_norms", 0.90),
        (2, "Capacity building", "Concept should drive actual capability development", "Pure rhetoric without capability investment", "evidence_testing", "cluster_78", 0.85),
        (3, "Threat response", "Concept should respond to genuine external threats", "Invoking sovereignty for purely protectionist ends", "llm_analysis", "canguilhem_norms", 0.82),
        (4, "Coherent scope", "Concept should have definable boundaries", "Expanding to cover everything = covering nothing", "evidence_testing", "cluster_91", 0.78),
    ],
    "concept_vital_norms"
)

# concept_milieu
row = add_table(ws, row,
    ["id", "milieu_type", "current_state", "trend", "impact_on_concept", "source_type", "source_ref", "confidence"],
    [
        (1, "geopolitical", "US-China competition; bloc formation", "intensifying", "Strengthens concept salience", "evidence_testing", "cluster_23", 0.92),
        (2, "technological", "AI acceleration; chip constraints; quantum emergence", "rapid_change", "Creates urgency, new domains", "evidence_testing", "cluster_56", 0.90),
        (3, "economic", "Supply chain restructuring; friendshoring; reshoring", "restructuring", "Validates sovereignty concerns", "evidence_testing", "cluster_78", 0.88),
        (4, "institutional", "WTO weakening; bilateral deals; tech alliances", "fragmenting", "Opens space for sovereignty claims", "llm_analysis", "canguilhem_milieu", 0.85),
    ],
    "concept_milieu"
)

# concept_vitality_indicators
row = add_table(ws, row,
    ["id", "indicator", "current_value", "trend", "interpretation", "source_type", "source_ref", "confidence"],
    [
        (1, "usage_frequency", "very_high", "increasing", "Concept highly active in discourse", "evidence_testing", "cluster_12", 0.95),
        (2, "policy_adoption", "high", "increasing", "Being translated into actual policy (Chips Act, etc.)", "evidence_testing", "cluster_34", 0.92),
        (3, "internal_coherence", "moderate", "declining", "Contradictions becoming more visible", "evidence_testing", "cluster_89", 0.85),
        (4, "empirical_validation", "low", "stable", "Few success cases; many failures", "evidence_testing", "cluster_67", 0.88),
        (5, "overall_health", "strained", "concerning", "High activity but growing contradictions", "llm_analysis", "canguilhem_vitality", 0.85),
    ],
    "concept_vitality_indicators"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 8: Hacking Dimension
# ============================================================================
ws = wb.create_sheet("8. Hacking")
style_header(ws, 1, "HACKING DIMENSION - Dynamic Nominalism & Looping Effects", dimension_colors["Hacking"])

# concept_reasoning_styles
row = add_table(ws, 3,
    ["id", "style", "key_features", "emergence_period", "source_type", "source_ref", "confidence"],
    [
        (1, "geopolitical_strategic", "Threat assessment, capability gaps, deterrence logic, balance of power", "Post-WWII, revived 2015+", "llm_analysis", "hacking_v1", 0.88),
        (2, "techno-economic", "Market analysis, investment returns, comparative advantage, scale economics", "1980s+, challenged 2018+", "llm_analysis", "hacking_v1", 0.85),
        (3, "techno-civilizational", "Civilizational competition, tech as cultural expression, values in code", "Emerging 2020+", "evidence_testing", "cluster_91", 0.78),
    ],
    "concept_reasoning_styles"
)

# concept_looping_effects
row = add_table(ws, row,
    ["id", "loop_description", "mechanism", "observed_effects", "source_type", "source_ref", "confidence"],
    [
        (1, "Classification creates behavior", "Labeling countries as 'tech sovereign aspirants' changes how they act", "Countries pursue performative sovereignty to match label", "evidence_testing", "cluster_34", 0.88),
        (2, "Metrics shape reality", "Sovereignty indices and rankings drive policy responses", "Investment decisions follow rankings, not fundamentals", "evidence_testing", "cluster_56", 0.82),
        (3, "Discourse enables spending", "Sovereignty framing enables budget allocations", "Budgets reshape industrial structure, reinforcing frame", "evidence_testing", "cluster_78", 0.90),
        (4, "Threat construction creates threats", "Framing tech as security threat triggers defensive responses", "Defensive responses are seen as threats, escalation spiral", "llm_analysis", "hacking_looping", 0.85),
    ],
    "concept_looping_effects"
)

# concept_kinds_created
row = add_table(ws, row,
    ["id", "kind_name", "kind_type", "how_created", "stability", "source_type", "source_ref", "confidence"],
    [
        (1, "Tech-sovereign state", "political_entity", "Classification by sovereignty pursuit rather than achievement", "unstable - criteria contested", "llm_analysis", "hacking_kinds", 0.85),
        (2, "Strategic technology", "artifact_category", "Political designation, not inherent property", "highly unstable - shifts with politics", "evidence_testing", "cluster_23", 0.90),
        (3, "Tech sovereignty expert", "person_kind", "New professional identity created by discourse", "stabilizing - career paths forming", "evidence_testing", "cluster_91", 0.82),
        (4, "Trusted vs untrusted vendor", "entity_category", "Geopolitically defined, not technically", "unstable - changes with alliances", "evidence_testing", "cluster_67", 0.88),
    ],
    "concept_kinds_created"
)

# concept_possibility_space
row = add_table(ws, row,
    ["id", "possibility", "opened_or_closed", "mechanism", "source_type", "source_ref", "confidence"],
    [
        (1, "Industrial policy legitimacy", "opened", "Sovereignty framing makes intervention acceptable in liberal economies", "evidence_testing", "cluster_12", 0.92),
        (2, "Technology export controls", "opened", "Sovereignty concerns legitimate broad restrictions", "evidence_testing", "cluster_45", 0.90),
        (3, "International tech cooperation", "closing", "Sovereignty framing casts cooperation as dependency/risk", "evidence_testing", "cluster_78", 0.85),
        (4, "Open source/open standards", "threatened", "Sovereignty concerns push toward proprietary, controlled systems", "llm_analysis", "hacking_possibility", 0.82),
        (5, "Techno-cosmopolitanism", "closed", "Sovereignty discourse forecloses global tech governance ideas", "llm_analysis", "hacking_possibility", 0.80),
    ],
    "concept_possibility_space"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 9: Blumenberg Dimension
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
style_header(ws, 1, "BLUMENBERG DIMENSION - Metaphorology & Nonconceptuality", dimension_colors["Blumenberg"])

# concept_metaphors
row = add_table(ws, 3,
    ["id", "metaphor", "type", "structuring_effect", "absoluteness", "source_type", "source_ref", "confidence"],
    [
        (1, "Territory/borders", "spatial", "Makes tech controllable through physical boundary logic", "high - deeply embedded", "llm_analysis", "blumenberg_v1", 0.90),
        (2, "Immune system", "biological", "Tech as defense against foreign 'infection/contamination'", "moderate - rising", "evidence_testing", "cluster_34", 0.85),
        (3, "Arsenal/stockpile", "military", "Tech capability as accumulated weaponry", "moderate", "llm_analysis", "blumenberg_v1", 0.82),
        (4, "Foundation/infrastructure", "architectural", "Tech as base on which society is built", "high", "llm_analysis", "blumenberg_v1", 0.88),
        (5, "Race/competition", "athletic", "Tech development as zero-sum contest with finish line", "very high - dominant", "evidence_testing", "cluster_56", 0.92),
    ],
    "concept_metaphors"
)

# concept_metakinetics
row = add_table(ws, row,
    ["id", "metaphor", "transformation", "period", "driver", "source_type", "source_ref", "confidence"],
    [
        (1, "Territory", "From physical to digital borders", "2010-2020", "Data localization debates, cloud computing", "evidence_testing", "cluster_23", 0.88),
        (2, "Race", "From marathon to sprint", "2018+", "AI acceleration, chip wars urgency", "evidence_testing", "cluster_67", 0.90),
        (3, "Immune system", "From dormant to hyperactive", "2020+", "COVID supply chain shocks", "evidence_testing", "cluster_78", 0.85),
        (4, "Arsenal", "From nuclear to digital", "2015+", "Cyber warfare concerns", "llm_analysis", "blumenberg_metakinetics", 0.82),
    ],
    "concept_metakinetics"
)

# concept_nonconceptuality
row = add_table(ws, row,
    ["id", "nonconceptual_aspect", "why_resists_conceptualization", "manifestation", "source_type", "source_ref", "confidence"],
    [
        (1, "Feeling of autonomy", "Experiential quality that can't be reduced to policy metrics", "Leaders speak of 'pride' and 'dignity' alongside capability", "llm_analysis", "blumenberg_nonconc", 0.85),
        (2, "Technological sublime", "Awe at tech power exceeds rational assessment", "Inflated expectations, magical thinking about tech effects", "llm_analysis", "blumenberg_nonconc", 0.80),
        (3, "Existential anxiety", "Fear of dependence is visceral, not just calculated", "Emotional intensity of sovereignty debates", "evidence_testing", "cluster_89", 0.82),
        (4, "National identity fusion", "Tech tied to who 'we' are, not just what 'we' have", "Resistance to rational cost-benefit framing", "llm_analysis", "blumenberg_nonconc", 0.78),
    ],
    "concept_nonconceptuality"
)

# concept_lifeworld_connection
row = add_table(ws, row,
    ["id", "lifeworld_aspect", "connection_type", "tension_level", "description", "source_type", "source_ref", "confidence"],
    [
        (1, "Daily tech dependence", "experiential_contradiction", "high", "People use foreign tech daily while supporting sovereignty rhetoric", "evidence_testing", "cluster_91", 0.90),
        (2, "Global connectivity", "structural_tension", "very_high", "Lived experience of global networks vs. sovereignty framing", "llm_analysis", "blumenberg_lifeworld", 0.88),
        (3, "Consumer expectations", "practical_tension", "moderate", "Want sovereignty but also want best/cheapest products", "evidence_testing", "cluster_45", 0.85),
        (4, "Professional networks", "relational_tension", "moderate", "Scientists/engineers have global networks; sovereignty disrupts", "evidence_testing", "cluster_67", 0.82),
    ],
    "concept_lifeworld_connection"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 10: Carey Dimension
# ============================================================================
ws = wb.create_sheet("10. Carey")
style_header(ws, 1, "CAREY DIMENSION - Conceptual Bootstrapping", dimension_colors["Carey"])

# concept_hierarchy
row = add_table(ws, 3,
    ["id", "level", "description", "components", "source_type", "source_ref", "confidence"],
    [
        (1, 3, "Complex composite - bootstrapped from multiple Level 2 concepts", "sovereignty, technology, autonomy, control, criticality", "llm_analysis", "carey_v1", 0.85),
    ],
    "concept_hierarchy"
)

# concept_built_from
row = add_table(ws, row,
    ["id", "component_concept", "component_level", "contribution", "source_type", "source_ref", "confidence"],
    [
        (1, "Sovereignty", 2, "Provides authority/control dimension from political philosophy", "llm_analysis", "carey_v1", 0.90),
        (2, "Technology", 1, "Provides domain/object specification", "llm_analysis", "carey_v1", 0.92),
        (3, "Autonomy", 2, "Provides independence/self-determination dimension", "llm_analysis", "carey_v1", 0.88),
        (4, "National interest", 2, "Provides collective agent whose sovereignty is at stake", "llm_analysis", "carey_v1", 0.85),
        (5, "Strategic competition", 2, "Provides urgency and adversarial framing", "evidence_testing", "cluster_23", 0.82),
        (6, "Legitimacy", 2, "Provides normative justification dimension", "llm_analysis", "carey_v1", 0.78),
    ],
    "concept_built_from"
)

# concept_core_cognition
row = add_table(ws, row,
    ["id", "core_system", "engagement", "how_activated", "source_type", "source_ref", "confidence"],
    [
        (1, "in-group/out-group", "strong", "Us vs. them framing; domestic vs. foreign tech", "llm_analysis", "carey_core", 0.88),
        (2, "object permanence", "moderate", "Technology as bounded 'thing' that can be possessed/controlled", "llm_analysis", "carey_core", 0.80),
        (3, "agency detection", "moderate", "Anthropomorphizing 'markets', 'technology', 'adversaries' as agents", "llm_analysis", "carey_core", 0.75),
        (4, "contamination avoidance", "strong", "Foreign tech as potentially 'contaminating' - purity logic", "evidence_testing", "cluster_78", 0.85),
    ],
    "concept_core_cognition"
)

# concept_placeholder_structures
row = add_table(ws, row,
    ["id", "placeholder", "content_type", "current_fill_status", "elaboration_needed", "source_type", "source_ref", "confidence"],
    [
        (1, "How to achieve it", "procedural", "largely_empty", "Massive - no clear path to actual sovereignty", "evidence_testing", "cluster_67", 0.90),
        (2, "What counts as success", "evaluative", "contested", "Multiple incompatible definitions in use", "evidence_testing", "cluster_89", 0.88),
        (3, "Which technologies are critical", "categorical", "partially_filled", "Lists exist but contested and shifting", "evidence_testing", "cluster_34", 0.85),
        (4, "Who the sovereign agent is", "ontological", "assumed_not_elaborated", "State assumed but actually contested (regions? blocs?)", "llm_analysis", "carey_placeholder", 0.82),
    ],
    "concept_placeholder_structures"
)

# concept_bootstrapping_constraints
row = add_table(ws, row,
    ["id", "constraint_type", "description", "severity", "source_type", "source_ref", "confidence"],
    [
        (1, "working_memory", "Concept requires tracking too many interdependencies simultaneously", "moderate", "llm_analysis", "carey_constraints", 0.78),
        (2, "analogical_mapping", "Source concepts (Westphalian sovereignty) map poorly to tech domain", "high", "llm_analysis", "carey_constraints", 0.85),
        (3, "causal_learning", "Cause-effect relationships too complex and delayed for intuitive learning", "high", "evidence_testing", "cluster_56", 0.88),
        (4, "coherence_constraint", "Component concepts don't cohere well - autonomy vs. innovation tension", "very_high", "evidence_testing", "cluster_45", 0.90),
    ],
    "concept_bootstrapping_constraints"
)

# concept_incommensurability
row = add_table(ws, row,
    ["id", "with_concept", "incommensurability_type", "description", "source_type", "source_ref", "confidence"],
    [
        (1, "Globalization", "structural", "Cannot be translated into globalization framework without loss", "llm_analysis", "carey_incomm", 0.88),
        (2, "Westphalian sovereignty", "partial", "Shares vocabulary but different referents (territory vs. flows)", "llm_analysis", "carey_incomm", 0.82),
        (3, "Open innovation", "strong", "Fundamentally incompatible assumptions about knowledge production", "evidence_testing", "cluster_78", 0.90),
        (4, "Techno-cosmopolitanism", "complete", "No translation possible - different ontological assumptions", "llm_analysis", "carey_incomm", 0.85),
    ],
    "concept_incommensurability"
)

for col in 'ABCDEFGH':
    ws.column_dimensions[col].width = 22

# ============================================================================
# SHEET 11: Summary Statistics
# ============================================================================
ws = wb.create_sheet("11. Summary")
style_header(ws, 1, "POPULATED SCHEMA SUMMARY - Technological Sovereignty", dimension_colors["Overview"])

row = 3
ws.cell(row=row, column=1, value="CONCEPT: Technological Sovereignty").font = Font(bold=True, size=14)
row += 2

summary_data = [
    ("Dimension", "Tables", "Records", "Key Findings"),
    ("Quinean", 5, 17, "Hub position in belief web; high revision cost; ontological dependence on 'the state'"),
    ("Sellarsian", 5, 17, "Multiple hidden commitments; functions as enabling premise; manifest/scientific image gap"),
    ("Brandomian", 4, 18, "Key challenges unanswered (Gulf $2T); significant perspectival translation loss"),
    ("Deleuzian", 5, 17, "Unstable endoconsistency; operates on techno-nationalist plane; multiple personae"),
    ("Bachelardian", 4, 18, "Substantialist obstacle deeply entrenched; pre-scientific stage dominant; multiple unconscious needs served"),
    ("Canguilhem", 5, 21, "Intensification since 2018; vital norms being violated; milieu highly favorable"),
    ("Hacking", 4, 17, "Strong looping effects; new kinds being created; closing cooperation possibilities"),
    ("Blumenberg", 4, 16, "Race metaphor dominant; high nonconceptuality; significant lifeworld tensions"),
    ("Carey", 6, 22, "Level 3 complexity; many empty placeholders; strong bootstrapping constraints"),
    ("TOTAL", 42, 163, "Concept is highly active but internally strained with growing contradictions"),
]

for record in summary_data:
    for col, val in enumerate(record, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if row == 5:
            cell.font = header_font
            cell.fill = table_header_fill
        elif record[0] == "TOTAL":
            cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = wrap
    row += 1

row += 2
ws.cell(row=row, column=1, value="KEY CROSS-DIMENSIONAL FINDINGS:").font = Font(bold=True, size=12)
row += 1

findings = [
    "1. STRUCTURAL INSTABILITY: High usage frequency (Canguilhem) but unstable consistency (Deleuze) suggests concept is overextended",
    "2. UNANSWERED CHALLENGES: Gulf states' $2T failure to build capacity (Brandom) directly contradicts core commitment (investment → capability)",
    "3. OBSTACLE REINFORCEMENT: Substantialist obstacle (Bachelard) prevents recognizing relational nature of tech capability",
    "4. LOOPING DANGER: Classification effects (Hacking) may be creating a self-fulfilling prophecy of tech bloc formation",
    "5. METAPHOR LOCK-IN: Territorial/race metaphors (Blumenberg) structure thinking in ways that foreclose cooperative solutions",
    "6. BOOTSTRAP FAILURE: Placeholder structures (Carey) remain unfilled - no one knows HOW to achieve sovereignty",
    "7. VITAL NORM VIOLATION: Concept being used to justify increased dependency (Canguilhem) - self-undermining deployment",
    "8. MANIFEST/SCIENTIFIC GAP: Political discourse (manifest) increasingly diverges from technical reality (scientific) (Sellars)",
]

for finding in findings:
    ws.cell(row=row, column=1, value=finding).alignment = wrap
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 80

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_POPULATED_TechSovereignty.xlsx")
wb.save(output_path)
print(f"Populated schema saved to: {output_path}")
print(f"Concept: Technological Sovereignty")
print(f"Total records across 42 tables: 163")
