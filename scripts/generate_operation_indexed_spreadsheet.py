"""
Generate Operation-Indexed Concept Analysis Spreadsheet

Demonstrates the new schema where dimensions are organized by analytical operation
rather than by thinker. Applies comprehensive analysis to "Technological Sovereignty".
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Color palette for analytical dimensions
DIMENSION_COLORS = {
    'positional': 'E3F2FD',      # Blue - where it sits
    'genealogical': 'FFF3E0',    # Orange - where it comes from
    'presuppositional': 'FCE4EC', # Pink - what it assumes
    'commitment': 'F3E5F5',      # Purple - what it commits to
    'affordance': 'E8F5E9',      # Green - what it enables
    'normalization': 'FFEBEE',   # Red - what it normalizes
    'boundary': 'E0F7FA',        # Cyan - its limits
    'dynamic': 'FFF8E1',         # Amber - how it changes
}

# The new schema structure
ANALYTICAL_DIMENSIONS = {
    '1. Positional Analysis': {
        'core_question': 'Where does this concept SIT in various networks?',
        'description': 'Analyzes the concept\'s position in belief webs, paradigms, and discursive formations. Maps its relational structure.',
        'color': 'positional',
        'operations': [
            {
                'name': 'Inferential Mapping',
                'description': 'What follows from this concept? What contradicts it? What does it connect to?',
                'influences': ['Quine'],
                'output_type': 'inference_map',
                'key_questions': [
                    'What beliefs must you hold if you accept this concept?',
                    'What does accepting this concept rule out?',
                    'What lateral connections does it have to adjacent concepts?'
                ]
            },
            {
                'name': 'Paradigm Positioning',
                'description': 'Which paradigm does this belong to? Is it normal science or revolutionary?',
                'influences': ['Kuhn'],
                'output_type': 'paradigm_assessment',
                'key_questions': [
                    'What paradigm does this concept operate within?',
                    'What are the exemplary instances?',
                    'Is this concept extending or challenging the paradigm?'
                ]
            },
            {
                'name': 'Discursive Positioning',
                'description': 'What statements does this concept make possible or impossible?',
                'influences': ['Foucault'],
                'output_type': 'discourse_map',
                'key_questions': [
                    'What can be said using this concept that couldn\'t be said before?',
                    'What statements does it foreclose?',
                    'What subject positions does it create?'
                ]
            },
            {
                'name': 'Web Centrality Assessment',
                'description': 'How central or peripheral is this concept? What is the cost of revision?',
                'influences': ['Quine'],
                'output_type': 'centrality_score',
                'key_questions': [
                    'How many other beliefs depend on this concept?',
                    'What would have to change if this concept were abandoned?',
                    'Is this core or peripheral to your framework?'
                ]
            }
        ]
    },
    '2. Genealogical Analysis': {
        'core_question': 'Where does this concept COME FROM?',
        'description': 'Traces the historical emergence, conceptual predecessors, and conditions of possibility for the concept.',
        'color': 'genealogical',
        'operations': [
            {
                'name': 'Conceptual Bootstrapping',
                'description': 'What simpler concepts combine to produce this more complex one?',
                'influences': ['Carey'],
                'output_type': 'bootstrap_tree',
                'key_questions': [
                    'What simpler concepts is this built from?',
                    'What emergent properties arise from their combination?',
                    'What conceptual development was required?'
                ]
            },
            {
                'name': 'Historical Emergence',
                'description': 'What historical/discursive conditions made this concept possible?',
                'influences': ['Foucault'],
                'output_type': 'emergence_narrative',
                'key_questions': [
                    'What had to be in place for this concept to emerge?',
                    'What historical moment does it belong to?',
                    'What institutions or practices enabled it?'
                ]
            },
            {
                'name': 'Rupture Detection',
                'description': 'Does this concept mark an epistemological break? What does it break from?',
                'influences': ['Bachelard'],
                'output_type': 'rupture_assessment',
                'key_questions': [
                    'Does this concept represent a break from previous thinking?',
                    'What older concept or framework does it supersede?',
                    'What made the break possible or necessary?'
                ]
            },
            {
                'name': 'Metaphorical Archaeology',
                'description': 'What root metaphors structure this concept? What do they reveal/conceal?',
                'influences': ['Blumenberg'],
                'output_type': 'metaphor_analysis',
                'key_questions': [
                    'What is the governing metaphor?',
                    'What does this metaphor make visible?',
                    'What does it hide or distort?'
                ]
            },
            {
                'name': 'Predecessor Mapping',
                'description': 'What concepts did this replace or displace?',
                'influences': ['Foucault', 'Bachelard'],
                'output_type': 'predecessor_list',
                'key_questions': [
                    'What was being said before this concept existed?',
                    'What concepts did it make obsolete?',
                    'What problems did predecessors fail to solve?'
                ]
            }
        ]
    },
    '3. Presuppositional Analysis': {
        'core_question': 'What does this concept TAKE FOR GRANTED?',
        'description': 'Excavates hidden assumptions, unexamined givens, and background commitments embedded in the concept.',
        'color': 'presuppositional',
        'operations': [
            {
                'name': 'Givenness Detection',
                'description': 'What does this concept treat as self-evident that actually requires justification?',
                'influences': ['Sellars'],
                'output_type': 'givenness_list',
                'key_questions': [
                    'What does this concept treat as obvious?',
                    'What would someone demand justification for?',
                    'What epistemic work is being smuggled in?'
                ]
            },
            {
                'name': 'Assumption Excavation',
                'description': 'What hidden background beliefs are required for this concept to work?',
                'influences': ['Sellars'],
                'output_type': 'assumption_inventory',
                'key_questions': [
                    'What must be true for this concept to make sense?',
                    'What worldview does it presuppose?',
                    'What empirical claims does it smuggle in?'
                ]
            },
            {
                'name': 'Obstacle Detection',
                'description': 'Does this concept itself block understanding of something deeper?',
                'influences': ['Bachelard'],
                'output_type': 'obstacle_assessment',
                'key_questions': [
                    'Does using this concept prevent seeing something important?',
                    'What epistemological obstacles does it create?',
                    'What would you understand better without it?'
                ]
            },
            {
                'name': 'Manifest/Scientific Tension',
                'description': 'Is there tension between everyday and theoretical understandings?',
                'influences': ['Sellars'],
                'output_type': 'tension_analysis',
                'key_questions': [
                    'How does common sense view this differently from theory?',
                    'What folk concepts does this technical concept challenge?',
                    'Where do intuitions mislead?'
                ]
            }
        ]
    },
    '4. Commitment Analysis': {
        'core_question': 'What does accepting this concept COMMIT YOU TO?',
        'description': 'Maps the inferential, normative, and practical commitments that follow from adopting the concept.',
        'color': 'commitment',
        'operations': [
            {
                'name': 'Inferential Commitment Mapping',
                'description': 'What other beliefs must you accept if you accept this?',
                'influences': ['Brandom'],
                'output_type': 'commitment_list',
                'key_questions': [
                    'If you assert this, what else must you be prepared to assert?',
                    'What claims become unjustifiable if you hold this?',
                    'What inferential moves does this license?'
                ]
            },
            {
                'name': 'Normative Entailment Tracking',
                'description': 'What values, norms, or obligations follow from this concept?',
                'influences': ['Brandom'],
                'output_type': 'normative_entailments',
                'key_questions': [
                    'What should you value if you accept this concept?',
                    'What actions become obligatory or forbidden?',
                    'What normative stance does it encode?'
                ]
            },
            {
                'name': 'Incompatibility Detection',
                'description': 'What concepts or positions are materially incompatible with this one?',
                'influences': ['Brandom', 'Quine'],
                'output_type': 'incompatibility_map',
                'key_questions': [
                    'What can you NOT believe if you believe this?',
                    'What positions does this rule out?',
                    'Where are the hard incompatibilities vs. soft tensions?'
                ]
            },
            {
                'name': 'Entitlement Tracking',
                'description': 'What are you ENTITLED to claim if you accept this concept?',
                'influences': ['Brandom'],
                'output_type': 'entitlement_list',
                'key_questions': [
                    'What further claims does this justify?',
                    'What moves in the space of reasons does it enable?',
                    'What authority does it confer?'
                ]
            }
        ]
    },
    '5. Affordance Analysis': {
        'core_question': 'What does this concept ENABLE or BLOCK?',
        'description': 'Maps what becomes possible or impossible, visible or invisible, sayable or unsayable through this concept.',
        'color': 'affordance',
        'operations': [
            {
                'name': 'Transformation Mapping',
                'description': 'What changes, becomings, or movements does this concept enable?',
                'influences': ['Deleuze'],
                'output_type': 'transformation_vectors',
                'key_questions': [
                    'What transformations does this concept make thinkable?',
                    'What lines of flight does it open?',
                    'What becomings does it enable?'
                ]
            },
            {
                'name': 'Practical Effect Tracking',
                'description': 'What actions, interventions, or practices does this enable?',
                'influences': ['James', 'Dewey'],
                'output_type': 'practical_effects',
                'key_questions': [
                    'What can you DO with this concept?',
                    'What practical difference does it make?',
                    'What interventions does it enable?'
                ]
            },
            {
                'name': 'Visibility Mapping',
                'description': 'What does this concept make visible vs. invisible?',
                'influences': ['Hacking', 'Foucault'],
                'output_type': 'visibility_map',
                'key_questions': [
                    'What phenomena does this concept bring into focus?',
                    'What does it make harder to see?',
                    'What does naming this create?'
                ]
            },
            {
                'name': 'Vocabulary Extension',
                'description': 'What new things can you SAY with this concept?',
                'influences': ['Rorty'],
                'output_type': 'vocabulary_additions',
                'key_questions': [
                    'What conversations does this concept open?',
                    'What was unsayable before this concept existed?',
                    'What new descriptions does it enable?'
                ]
            },
            {
                'name': 'Inquiry Structuring',
                'description': 'How does this concept structure problem-solving and research?',
                'influences': ['Dewey'],
                'output_type': 'inquiry_structure',
                'key_questions': [
                    'What does this concept make into a "problem"?',
                    'What does it treat as "solved"?',
                    'How does it organize inquiry?'
                ]
            }
        ]
    },
    '6. Normalization Analysis': {
        'core_question': 'What does this concept NORMALIZE or make invisible?',
        'description': 'Exposes the norms, power relations, and governmentality embedded in and enabled by the concept.',
        'color': 'normalization',
        'operations': [
            {
                'name': 'Norm Embedding Detection',
                'description': 'What norms are baked into this concept?',
                'influences': ['Canguilhem'],
                'output_type': 'embedded_norms',
                'key_questions': [
                    'What does this concept treat as normal/healthy/proper?',
                    'What values are built into it?',
                    'What standards does it encode?'
                ]
            },
            {
                'name': 'Pathological Boundary Mapping',
                'description': 'What does this concept mark as deviant, abnormal, or pathological?',
                'influences': ['Canguilhem'],
                'output_type': 'pathological_boundaries',
                'key_questions': [
                    'What does this concept exclude as abnormal?',
                    'What gets pathologized?',
                    'Where is the normal/pathological boundary drawn?'
                ]
            },
            {
                'name': 'Power Relation Naturalization',
                'description': 'What power relations does this concept make seem natural or inevitable?',
                'influences': ['Foucault'],
                'output_type': 'naturalized_power',
                'key_questions': [
                    'What power arrangements does this concept make invisible?',
                    'Whose interests does it serve while appearing neutral?',
                    'What domination does it naturalize?'
                ]
            },
            {
                'name': 'Authority Legitimization',
                'description': 'Whose expertise or authority does this concept validate?',
                'influences': ['Foucault'],
                'output_type': 'authority_map',
                'key_questions': [
                    'Who gets to speak authoritatively about this?',
                    'What credentials does it require?',
                    'Whose knowledge counts?'
                ]
            },
            {
                'name': 'Governmentality Mapping',
                'description': 'What populations or phenomena does this concept make manageable?',
                'influences': ['Foucault'],
                'output_type': 'governmentality_analysis',
                'key_questions': [
                    'What does this concept make governable?',
                    'What techniques of management does it enable?',
                    'How does it render things calculable?'
                ]
            },
            {
                'name': 'Truth Regime Analysis',
                'description': 'What counts as true within the framework this concept establishes?',
                'influences': ['Foucault'],
                'output_type': 'truth_regime',
                'key_questions': [
                    'What counts as evidence within this framework?',
                    'What would falsify claims made with this concept?',
                    'Who arbitrates truth here?'
                ]
            }
        ]
    },
    '7. Boundary Analysis': {
        'core_question': 'What are the LIMITS of this concept?',
        'description': 'Maps anomalies, contradictions, incommensurabilities, and edge cases that test the concept\'s boundaries.',
        'color': 'boundary',
        'operations': [
            {
                'name': 'Anomaly Detection',
                'description': 'What cases can this concept not handle?',
                'influences': ['Kuhn'],
                'output_type': 'anomaly_list',
                'key_questions': [
                    'What phenomena does this concept fail to explain?',
                    'What are the stubborn counterexamples?',
                    'Where does it break down?'
                ]
            },
            {
                'name': 'Contradiction Mapping',
                'description': 'What does this concept logically rule out?',
                'influences': ['Quine', 'Brandom'],
                'output_type': 'contradiction_map',
                'key_questions': [
                    'What is strictly incompatible with this concept?',
                    'What would constitute a logical contradiction?',
                    'Where are the hard boundaries?'
                ]
            },
            {
                'name': 'Incommensurability Analysis',
                'description': 'Can this concept be translated into rival frameworks?',
                'influences': ['Kuhn'],
                'output_type': 'incommensurability_assessment',
                'key_questions': [
                    'What rival frameworks exist?',
                    'Can this concept be expressed in their terms?',
                    'Where is translation impossible?'
                ]
            },
            {
                'name': 'Gray Zone Identification',
                'description': 'What are the boundary cases where application is uncertain?',
                'influences': ['Hacking'],
                'output_type': 'gray_zones',
                'key_questions': [
                    'What cases are neither clearly in nor out?',
                    'Where do experts disagree about application?',
                    'What makes boundary cases hard?'
                ]
            },
            {
                'name': 'Crisis Indicator Tracking',
                'description': 'What would force abandonment of this concept?',
                'influences': ['Kuhn'],
                'output_type': 'crisis_indicators',
                'key_questions': [
                    'What accumulation of anomalies would be fatal?',
                    'What discovery would make this concept untenable?',
                    'How close are we to crisis?'
                ]
            }
        ]
    },
    '8. Dynamic Analysis': {
        'core_question': 'How does this concept CHANGE over time?',
        'description': 'Analyzes how the concept evolves, creates feedback loops, and shapes its own conditions.',
        'color': 'dynamic',
        'operations': [
            {
                'name': 'Looping Effect Detection',
                'description': 'Does naming this change the thing named?',
                'influences': ['Hacking'],
                'output_type': 'looping_effects',
                'key_questions': [
                    'Do the things classified by this concept change because of being classified?',
                    'Is there awareness of being categorized?',
                    'How does classification feed back?'
                ]
            },
            {
                'name': 'Habit Formation Tracking',
                'description': 'What habits of thought and practice does using this concept cultivate?',
                'influences': ['Dewey'],
                'output_type': 'habit_inventory',
                'key_questions': [
                    'What ways of thinking does this concept train?',
                    'What dispositions does repeated use create?',
                    'What becomes automatic?'
                ]
            },
            {
                'name': 'Revision Condition Mapping',
                'description': 'Under what conditions would this concept need revision?',
                'influences': ['Quine', 'Kuhn'],
                'output_type': 'revision_conditions',
                'key_questions': [
                    'What new evidence would force modification?',
                    'How revisable is this concept?',
                    'What would a revised version look like?'
                ]
            },
            {
                'name': 'Kind-Making Analysis',
                'description': 'Does this concept create new categories of being?',
                'influences': ['Hacking'],
                'output_type': 'kind_creation',
                'key_questions': [
                    'Does this concept bring a new kind of thing into existence?',
                    'What didn\'t exist (as a kind) before this concept?',
                    'Is this an interactive kind?'
                ]
            },
            {
                'name': 'Disciplinary Matrix Evolution',
                'description': 'How do shared practices around this concept shift over time?',
                'influences': ['Kuhn'],
                'output_type': 'matrix_evolution',
                'key_questions': [
                    'How has expert consensus shifted?',
                    'What exemplars have changed?',
                    'How has the community using this concept evolved?'
                ]
            }
        ]
    }
}

# Comprehensive analysis data for Technological Sovereignty
TECH_SOV_ANALYSIS = {
    '1. Positional Analysis': {
        'Inferential Mapping': {
            'canonical_statement': 'Technological Sovereignty entails the capacity for indigenous technology development, contradicts complete market liberalization of tech sectors, and connects to industrial policy, national security, and digital rights frameworks.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'forward_inferences': [
                    {'statement': 'States should invest in domestic semiconductor manufacturing', 'strength': 0.9},
                    {'statement': 'Critical infrastructure must have national ownership requirements', 'strength': 0.85},
                    {'statement': 'Data localization laws are justified for strategic sectors', 'strength': 0.8},
                    {'statement': 'Technology transfer requirements can be imposed on foreign firms', 'strength': 0.75},
                    {'statement': 'Public funding for R&D in strategic technologies is necessary', 'strength': 0.9},
                ],
                'backward_inferences': [
                    {'statement': 'States have legitimate interests in technological capacity', 'strength': 0.95},
                    {'statement': 'Technology is not politically neutral', 'strength': 0.9},
                    {'statement': 'Dependency creates vulnerability', 'strength': 0.85},
                    {'statement': 'Markets alone cannot secure strategic interests', 'strength': 0.8},
                ],
                'contradictions': [
                    {'statement': 'Complete free trade in all technology sectors', 'severity': 'hard'},
                    {'statement': 'Technology as purely private good', 'severity': 'hard'},
                    {'statement': 'National borders irrelevant for digital governance', 'severity': 'medium'},
                    {'statement': 'Efficiency as sole criterion for tech procurement', 'severity': 'hard'},
                ],
                'lateral_connections': [
                    {'concept': 'Food Sovereignty', 'relation': 'analogical', 'shared_structure': 'self-determination in strategic sector'},
                    {'concept': 'Energy Independence', 'relation': 'analogical', 'shared_structure': 'reducing foreign dependency'},
                    {'concept': 'Digital Rights', 'relation': 'complementary', 'shared_structure': 'citizen empowerment in digital sphere'},
                    {'concept': 'Industrial Policy', 'relation': 'instrumental', 'shared_structure': 'state intervention in strategic sectors'},
                ]
            }
        },
        'Paradigm Positioning': {
            'canonical_statement': 'Technological Sovereignty operates within the paradigm of strategic autonomy, challenging neoliberal assumptions about borderless technology flows while remaining within the state-centric international order.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'paradigm_name': 'Strategic Autonomy / Post-Neoliberal Statecraft',
                'paradigm_assumptions': [
                    'States remain primary actors in international system',
                    'Economic interdependence creates vulnerabilities',
                    'Technology is geopolitically strategic',
                    'Self-reliance reduces coercive leverage',
                ],
                'exemplary_instances': [
                    {'case': 'EU GDPR and Digital Markets Act', 'why_paradigmatic': 'Asserts regulatory sovereignty over foreign tech giants'},
                    {'case': 'China Semiconductor Self-Sufficiency Drive', 'why_paradigmatic': 'Massive state investment to escape dependency'},
                    {'case': 'India\'s UPI Payment System', 'why_paradigmatic': 'National alternative to foreign payment infrastructure'},
                    {'case': 'GAIA-X European Cloud Initiative', 'why_paradigmatic': 'Federated European alternative to US hyperscalers'},
                ],
                'normal_vs_revolutionary': 'NORMAL SCIENCE with revolutionary implications - extends existing state sovereignty concepts to new domain rather than challenging state-centrism itself',
                'shared_values': ['Strategic autonomy', 'Technological self-determination', 'Reduced dependency', 'Democratic control over technology'],
            }
        },
        'Discursive Positioning': {
            'canonical_statement': 'Technological Sovereignty makes it possible to discuss technology policy in terms of collective self-determination rather than mere market efficiency, while foreclosing purely techno-utopian or techno-libertarian framings.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'statements_enabled': [
                    '"We need to control our own data infrastructure"',
                    '"Foreign technology dependency is a security risk"',
                    '"Citizens should have a say in how algorithms govern them"',
                    '"Strategic sectors require public investment"',
                    '"Technology is too important to leave to markets alone"',
                ],
                'statements_foreclosed': [
                    '"The market will naturally optimize technology allocation"',
                    '"National borders are irrelevant in cyberspace"',
                    '"Technology policy should be purely technocratic"',
                    '"Consumer choice is sufficient for tech governance"',
                ],
                'subject_positions_created': [
                    {'position': 'The Sovereign Digital Citizen', 'description': 'Someone whose digital rights are protected by their state'},
                    {'position': 'The Strategic Technology Policymaker', 'description': 'Official who weighs technology through security/autonomy lens'},
                    {'position': 'The Dependent Technology User', 'description': 'Entity reliant on foreign technology - problematized subject'},
                ],
            }
        },
        'Web Centrality Assessment': {
            'canonical_statement': 'Technological Sovereignty is moderately central - abandoning it would require revising views on state legitimacy in tech governance, but does not threaten core metaphysical or epistemological commitments.',
            'version': '1.0',
            'confidence': 0.75,
            'analysis': {
                'centrality_score': 0.65,
                'centrality_rationale': 'High practical importance but not foundational to worldview',
                'dependent_beliefs': [
                    'Data localization policy positions',
                    'Industrial policy justifications',
                    'Critical infrastructure regulation frameworks',
                    'Technology transfer negotiation stances',
                ],
                'revision_cost': 'MEDIUM - Would require rethinking tech policy framework but not core political values',
                'core_or_peripheral': 'PERIPHERAL-TO-MIDDLING - Important policy concept but not identity-defining',
            }
        }
    },
    '2. Genealogical Analysis': {
        'Conceptual Bootstrapping': {
            'canonical_statement': 'Technological Sovereignty bootstraps from simpler concepts of sovereignty (political authority over territory), technology (applied knowledge systems), and dependency (reliance creating vulnerability), combining them to produce the emergent idea that political authority extends to technological systems.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'simpler_concepts': [
                    {
                        'concept': 'Sovereignty',
                        'definition': 'Supreme authority over a territory and its people',
                        'contribution': 'Provides the political-authority component'
                    },
                    {
                        'concept': 'Technology',
                        'definition': 'Applied knowledge systems, tools, and techniques',
                        'contribution': 'Provides the domain to which sovereignty is extended'
                    },
                    {
                        'concept': 'Dependency',
                        'definition': 'Reliance on external actors creating vulnerability',
                        'contribution': 'Provides the problem-structure that motivates the concept'
                    },
                    {
                        'concept': 'Strategic Sector',
                        'definition': 'Economic domain with national security implications',
                        'contribution': 'Provides justification for state intervention'
                    },
                    {
                        'concept': 'Self-Determination',
                        'definition': 'Capacity to make decisions about one\'s own affairs',
                        'contribution': 'Provides normative grounding'
                    },
                ],
                'emergent_properties': [
                    'Technology as domain of political contestation (not just economic)',
                    'Digital infrastructure as extension of territorial sovereignty',
                    'Algorithmic governance as subject to democratic accountability',
                    'Data as strategic resource comparable to natural resources',
                ],
                'developmental_stages': [
                    {'stage': 'Classical Sovereignty', 'period': 'pre-20th century', 'scope': 'Territory and population'},
                    {'stage': 'Economic Sovereignty', 'period': 'mid-20th century', 'scope': 'Added natural resources'},
                    {'stage': 'Information Sovereignty', 'period': 'late 20th century', 'scope': 'Added data and communications'},
                    {'stage': 'Technological Sovereignty', 'period': '21st century', 'scope': 'Full technology stack'},
                ],
            }
        },
        'Historical Emergence': {
            'canonical_statement': 'Technological Sovereignty emerged from the convergence of post-2008 skepticism of market fundamentalism, Snowden revelations about surveillance infrastructure, US-China tech rivalry, and COVID-19 supply chain disruptions.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'conditions_of_possibility': [
                    {
                        'condition': 'Concentration of tech power in few firms/states',
                        'how_enabling': 'Made dependency visible and threatening'
                    },
                    {
                        'condition': 'Snowden revelations (2013)',
                        'how_enabling': 'Demonstrated US surveillance via tech infrastructure'
                    },
                    {
                        'condition': 'US-China tech rivalry',
                        'how_enabling': 'Showed technology as geopolitical weapon'
                    },
                    {
                        'condition': 'COVID-19 supply chain disruptions',
                        'how_enabling': 'Revealed vulnerability of globalized tech production'
                    },
                    {
                        'condition': 'Platform monopolization',
                        'how_enabling': 'Created resentment of foreign platform power'
                    },
                ],
                'key_moments': [
                    {'year': 2013, 'event': 'Snowden revelations', 'impact': 'European data sovereignty concerns'},
                    {'year': 2016, 'event': 'Brexit/Trump', 'impact': 'Legitimized economic nationalism'},
                    {'year': 2018, 'event': 'US sanctions on ZTE/Huawei', 'impact': 'Demonstrated tech as coercive tool'},
                    {'year': 2020, 'event': 'COVID chip shortage', 'impact': 'Supply chain vulnerability became visceral'},
                    {'year': 2022, 'event': 'CHIPS Act passage', 'impact': 'US embraces tech sovereignty logic'},
                ],
                'enabling_institutions': [
                    'EU Commission (regulatory entrepreneurship)',
                    'National security establishments',
                    'Industrial policy think tanks',
                    'Digital rights movements',
                ],
            }
        },
        'Rupture Detection': {
            'canonical_statement': 'Technological Sovereignty represents an epistemological break from the 1990s-2000s consensus that technology naturally transcends borders and that market liberalization in tech is always beneficial.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'is_rupture': True,
                'rupture_from': 'Techno-globalist consensus / Cyber-libertarianism',
                'previous_framework': {
                    'name': 'Borderless Digital World',
                    'core_claims': [
                        'Information wants to be free',
                        'The internet routes around censorship',
                        'Technology transcends national boundaries',
                        'Global markets optimize technology allocation',
                    ],
                    'dominant_period': '1995-2015',
                },
                'nature_of_break': 'Rejection of inevitability and desirability of tech globalization',
                'what_forced_break': 'Accumulation of evidence that tech IS territorially bounded, CAN be weaponized, and DOES concentrate power',
                'residual_elements': 'Still accepts innovation as good; contest is over governance, not technology itself',
            }
        },
        'Metaphorical Archaeology': {
            'canonical_statement': 'Technological Sovereignty is governed by territorial metaphors (digital borders, data territory, cyber-space) that reveal technology as controllable domain but conceal the fundamental differences between bits and atoms.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'root_metaphors': [
                    {
                        'metaphor': 'Digital Territory',
                        'what_it_reveals': 'Technology can be bounded and governed',
                        'what_it_conceals': 'Data flows are fundamentally different from physical territory',
                        'source_domain': 'Westphalian territorial state',
                    },
                    {
                        'metaphor': 'Technology Stack as Infrastructure',
                        'what_it_reveals': 'Layered dependencies and chokepoints',
                        'what_it_conceals': 'Unlike physical infrastructure, software can be forked/replicated',
                        'source_domain': 'Physical infrastructure (roads, pipes)',
                    },
                    {
                        'metaphor': 'Data as Resource/Oil',
                        'what_it_reveals': 'Data has economic and strategic value',
                        'what_it_conceals': 'Data is non-rivalrous and infinitely copyable',
                        'source_domain': 'Natural resource extraction',
                    },
                ],
                'metaphor_tensions': 'Territorial metaphors struggle with borderless nature of networks',
            }
        },
        'Predecessor Mapping': {
            'canonical_statement': 'Technological Sovereignty displaces cyber-libertarianism and techno-utopianism while drawing on older traditions of economic nationalism and strategic trade theory.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'displaced_concepts': [
                    {
                        'concept': 'Cyber-libertarianism',
                        'what_it_claimed': 'Internet inherently resists government control',
                        'why_displaced': 'Governments proved very capable of controlling internet',
                    },
                    {
                        'concept': 'Techno-utopianism',
                        'what_it_claimed': 'Technology naturally democratizes and empowers',
                        'why_displaced': 'Platform monopolization concentrated power instead',
                    },
                    {
                        'concept': 'Digital Exceptionalism',
                        'what_it_claimed': 'Digital domain follows different rules than physical',
                        'why_displaced': 'Digital proved embedded in physical infrastructure and geopolitics',
                    },
                ],
                'ancestor_concepts': [
                    {
                        'concept': 'Economic Nationalism',
                        'relation': 'Extends to technology domain',
                    },
                    {
                        'concept': 'Strategic Trade Theory',
                        'relation': 'Provides economic justification',
                    },
                    {
                        'concept': 'Dependency Theory',
                        'relation': 'Provides North-South framing',
                    },
                ],
            }
        }
    },
    '3. Presuppositional Analysis': {
        'Givenness Detection': {
            'canonical_statement': 'Technological Sovereignty treats as given that states are appropriate technology governors, that technology dependency is problematic, and that technological capacity can be meaningfully "owned" by political entities.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'treated_as_given': [
                    {
                        'assumption': 'States are legitimate technology governors',
                        'justification_required': 'Why states rather than cities, regions, or international bodies?',
                        'hidden_work': 'Normalizes state-level governance in domain that might work better at other scales',
                    },
                    {
                        'assumption': 'Dependency is inherently problematic',
                        'justification_required': 'Interdependence can also be beneficial and peace-promoting',
                        'hidden_work': 'Frames interdependence negatively without argument',
                    },
                    {
                        'assumption': 'Technology can be "sovereign"',
                        'justification_required': 'Technology doesn\'t respect borders the way territory does',
                        'hidden_work': 'Imports territorial logic to non-territorial domain',
                    },
                    {
                        'assumption': 'Strategic sectors are identifiable',
                        'justification_required': 'Boundaries of "strategic" are contested and expandable',
                        'hidden_work': 'Naturalizes politically-determined boundaries',
                    },
                ],
            }
        },
        'Assumption Excavation': {
            'canonical_statement': 'Technological Sovereignty presupposes a world of competing states, technology as power multiplier, capacity for national technology development, and the desirability of reduced interdependence.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'background_assumptions': [
                    {
                        'assumption': 'Interstate competition is permanent',
                        'type': 'Worldview',
                        'alternatives_foreclosed': 'Global cooperation on technology governance',
                    },
                    {
                        'assumption': 'Technology = Power',
                        'type': 'Empirical',
                        'alternatives_foreclosed': 'Technology as pure welfare-enhancer',
                    },
                    {
                        'assumption': 'National development is possible',
                        'type': 'Empirical',
                        'alternatives_foreclosed': 'Technology too complex for national autarky',
                    },
                    {
                        'assumption': 'Self-reliance beats interdependence',
                        'type': 'Normative',
                        'alternatives_foreclosed': 'Managed interdependence as superior strategy',
                    },
                    {
                        'assumption': 'States can control technology flows',
                        'type': 'Empirical',
                        'alternatives_foreclosed': 'Technology inherently uncontrollable',
                    },
                ],
            }
        },
        'Obstacle Detection': {
            'canonical_statement': 'Technological Sovereignty can become an epistemological obstacle when it prevents seeing the benefits of technological interdependence, the legitimacy of non-state governance, or the ways sovereignty-talk serves incumbent interests.',
            'version': '1.0',
            'confidence': 0.75,
            'analysis': {
                'potential_obstacles': [
                    {
                        'obstacle': 'State-centrism',
                        'what_it_blocks': 'Seeing cities, regions, or communities as tech governance actors',
                        'severity': 'Medium',
                    },
                    {
                        'obstacle': 'Security framing',
                        'what_it_blocks': 'Seeing technology governance as welfare/rights issue',
                        'severity': 'Medium',
                    },
                    {
                        'obstacle': 'Sovereignty nostalgia',
                        'what_it_blocks': 'Genuinely new governance forms for networked technology',
                        'severity': 'High',
                    },
                ],
                'is_obstacle': 'PARTIALLY - Valuable corrective to cyber-utopianism but can overcorrect',
            }
        },
        'Manifest/Scientific Tension': {
            'canonical_statement': 'Tension exists between the manifest image of technology as controllable national asset and the scientific image of technology as globally networked, rapidly evolving, and boundary-transgressing.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'manifest_image': {
                    'description': 'Technology as national capability that can be developed, protected, and controlled like other strategic assets',
                    'intuitive_appeal': 'Fits with existing frameworks of national security and industrial policy',
                },
                'scientific_image': {
                    'description': 'Technology as globally networked, rapidly evolving ecosystem where innovation depends on openness and boundaries are porous',
                    'challenges_to_manifest': [
                        'Software can be copied infinitely across borders',
                        'Innovation depends on global knowledge flows',
                        'Supply chains are deeply intertwined',
                        'Talent moves globally',
                    ],
                },
                'tension_implications': 'Tech sovereignty policies may be less effective than manifest image suggests',
            }
        }
    },
    '4. Commitment Analysis': {
        'Inferential Commitment Mapping': {
            'canonical_statement': 'Accepting Technological Sovereignty commits you to: legitimacy of state tech intervention, technology as strategic domain, problematization of dependency, and some degree of techno-economic nationalism.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'you_must_also_accept': [
                    {
                        'commitment': 'State intervention in technology markets is legitimate',
                        'strength': 0.95,
                        'reason': 'Core premise of the concept'
                    },
                    {
                        'commitment': 'Some technologies are strategic (not just commercial)',
                        'strength': 0.9,
                        'reason': 'Justifies differential treatment'
                    },
                    {
                        'commitment': 'Foreign dependency creates vulnerability',
                        'strength': 0.85,
                        'reason': 'Motivates sovereignty pursuit'
                    },
                    {
                        'commitment': 'Nations can meaningfully develop technology capabilities',
                        'strength': 0.8,
                        'reason': 'Otherwise sovereignty pursuit is futile'
                    },
                    {
                        'commitment': 'Efficiency is not the only criterion for tech policy',
                        'strength': 0.85,
                        'reason': 'Otherwise pure market allocation would suffice'
                    },
                ],
                'inferential_licenses': [
                    'Can argue for subsidies to domestic tech firms',
                    'Can justify technology transfer requirements',
                    'Can support data localization mandates',
                    'Can advocate for public alternatives to foreign platforms',
                ],
            }
        },
        'Normative Entailment Tracking': {
            'canonical_statement': 'Technological Sovereignty encodes normative commitments to collective self-determination in technology, democratic accountability for algorithmic systems, and the priority of strategic security over pure market efficiency.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'normative_commitments': [
                    {
                        'norm': 'Collective self-determination',
                        'description': 'Peoples should control the technologies that govern them',
                        'implications': 'Democratic input into technology policy'
                    },
                    {
                        'norm': 'Accountability',
                        'description': 'Those deploying powerful technologies must answer to affected populations',
                        'implications': 'Regulation of foreign tech platforms'
                    },
                    {
                        'norm': 'Security priority',
                        'description': 'Strategic security can trump pure economic efficiency',
                        'implications': 'Accepting higher costs for reduced dependency'
                    },
                    {
                        'norm': 'Productive capacity matters',
                        'description': 'Ability to make things matters, not just ability to buy them',
                        'implications': 'Industrial policy, manufacturing focus'
                    },
                ],
            }
        },
        'Incompatibility Detection': {
            'canonical_statement': 'Technological Sovereignty is materially incompatible with pure free-trade liberalism in technology, techno-libertarianism, and strong versions of global governance that dissolve state authority over technology.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'hard_incompatibilities': [
                    {
                        'position': 'Pure free-trade liberalism in tech',
                        'why_incompatible': 'Treats state intervention as distortion',
                        'can_be_reconciled': False
                    },
                    {
                        'position': 'Techno-libertarianism',
                        'why_incompatible': 'Rejects state authority over technology',
                        'can_be_reconciled': False
                    },
                    {
                        'position': 'Strong tech globalism',
                        'why_incompatible': 'Treats borders as obsolete in digital domain',
                        'can_be_reconciled': False
                    },
                ],
                'soft_tensions': [
                    {
                        'position': 'Cosmopolitanism',
                        'nature_of_tension': 'Prioritizes universal over particular',
                        'reconciliation_possible': 'Cosmopolitan case for sovereignties in plural'
                    },
                    {
                        'position': 'Open source ideology',
                        'nature_of_tension': 'Favors borderless collaboration',
                        'reconciliation_possible': 'Open source as tool FOR sovereignty'
                    },
                ],
            }
        },
        'Entitlement Tracking': {
            'canonical_statement': 'Accepting Technological Sovereignty entitles you to claim that technology policy is a legitimate domain of political contestation, that dependency concerns are valid, and that alternatives to foreign platforms deserve support.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'you_become_entitled_to_claim': [
                    {
                        'claim': 'Technology policy is legitimate political domain',
                        'audience': 'Those who say "leave it to the market"',
                    },
                    {
                        'claim': 'Dependency concerns are valid, not protectionist pretexts',
                        'audience': 'Free trade advocates',
                    },
                    {
                        'claim': 'Public alternatives to foreign platforms deserve support',
                        'audience': 'Those who say "use what works best"',
                    },
                    {
                        'claim': 'Technology is not neutral - it embeds values',
                        'audience': 'Techno-neutrality advocates',
                    },
                ],
                'authority_conferred': 'Standing to participate in strategic technology policy debates',
            }
        }
    },
    '5. Affordance Analysis': {
        'Transformation Mapping': {
            'canonical_statement': 'Technological Sovereignty enables transformations from passive technology consumers to active technology shapers, from market-takers to strategic actors, from dependency to capability.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'transformations_enabled': [
                    {
                        'from': 'Passive technology consumer',
                        'to': 'Active technology shaper',
                        'mechanism': 'Legitimizes policy intervention in tech markets'
                    },
                    {
                        'from': 'Market-price-taker',
                        'to': 'Strategic actor',
                        'mechanism': 'Frames technology choices as geopolitical, not just economic'
                    },
                    {
                        'from': 'Dependency',
                        'to': 'Capability',
                        'mechanism': 'Provides framework for capability-building investment'
                    },
                    {
                        'from': 'Global efficiency optimization',
                        'to': 'Resilience optimization',
                        'mechanism': 'Changes what counts as "optimal"'
                    },
                ],
                'becomings': [
                    'A nation becoming a technology producer',
                    'Citizens becoming digital subjects of their own state',
                    'Firms becoming national champions',
                ],
            }
        },
        'Practical Effect Tracking': {
            'canonical_statement': 'Technological Sovereignty enables policy interventions including: industrial policy for tech sectors, data localization requirements, critical infrastructure regulation, technology transfer negotiations, and public platform alternatives.',
            'version': '1.0',
            'confidence': 0.9,
            'analysis': {
                'actions_enabled': [
                    {
                        'action': 'Industrial policy for strategic tech sectors',
                        'examples': ['CHIPS Act', 'EU Chips Act', 'Made in China 2025'],
                        'practical_difference': 'Billions in public investment in semiconductors'
                    },
                    {
                        'action': 'Data localization requirements',
                        'examples': ['GDPR data transfer rules', 'Russia data localization', 'India data protection bill'],
                        'practical_difference': 'Forces foreign firms to store data locally'
                    },
                    {
                        'action': 'Critical infrastructure designation',
                        'examples': ['5G vendor restrictions', 'Cloud security requirements'],
                        'practical_difference': 'Excludes certain foreign vendors from key systems'
                    },
                    {
                        'action': 'Public platform alternatives',
                        'examples': ['India UPI', 'GAIA-X', 'Brazil PIX'],
                        'practical_difference': 'State-backed alternatives to foreign platforms'
                    },
                ],
                'interventions_blocked': [
                    'Cannot easily argue for complete openness',
                    'Cannot dismiss security concerns as protectionism',
                ],
            }
        },
        'Visibility Mapping': {
            'canonical_statement': 'Technological Sovereignty makes visible: chokepoints in technology supply chains, asymmetric dependencies, and the political economy of platform power. It makes less visible: benefits of interdependence and costs of autarky.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'makes_visible': [
                    {
                        'phenomenon': 'Supply chain chokepoints',
                        'how': 'Frames dependency as vulnerability to be mapped',
                        'examples': ['Taiwan semiconductor concentration', 'Rare earth processing in China']
                    },
                    {
                        'phenomenon': 'Platform power',
                        'how': 'Frames foreign platforms as governance challenge',
                        'examples': ['Facebook/Meta content moderation', 'Google search dominance']
                    },
                    {
                        'phenomenon': 'Data flows as power flows',
                        'how': 'Frames data movement geopolitically',
                        'examples': ['Transatlantic data transfers', 'Chinese data access']
                    },
                ],
                'makes_invisible': [
                    {
                        'phenomenon': 'Benefits of interdependence',
                        'how': 'Security framing crowds out efficiency/welfare frame',
                    },
                    {
                        'phenomenon': 'Costs of autarky',
                        'how': 'Focus on dependency costs obscures self-reliance costs',
                    },
                    {
                        'phenomenon': 'Non-state governance possibilities',
                        'how': 'State-centric framing obscures alternatives',
                    },
                ],
            }
        },
        'Vocabulary Extension': {
            'canonical_statement': 'Technological Sovereignty extends vocabulary to discuss: digital colonialism, data sovereignty, platform imperialism, tech stack dependencies, and sovereign clouds - concepts unsayable in pure market framing.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'new_sayables': [
                    {
                        'phrase': 'Digital colonialism',
                        'what_it_allows': 'Critique of tech firm power in Global South in political terms',
                        'previously_unsayable': True
                    },
                    {
                        'phrase': 'Data sovereignty',
                        'what_it_allows': 'Claim national authority over data generated within borders',
                        'previously_unsayable': True
                    },
                    {
                        'phrase': 'Platform imperialism',
                        'what_it_allows': 'Critique of US platform dominance as geopolitical',
                        'previously_unsayable': True
                    },
                    {
                        'phrase': 'Sovereign cloud',
                        'what_it_allows': 'Demand for nationally-controlled cloud infrastructure',
                        'previously_unsayable': True
                    },
                    {
                        'phrase': 'Tech stack dependency',
                        'what_it_allows': 'Analyze vulnerability at each layer of technology',
                        'previously_unsayable': True
                    },
                ],
                'conversations_opened': [
                    'Should we build national alternatives to foreign platforms?',
                    'Which technologies are strategically critical?',
                    'How do we reduce technology dependency without autarky?',
                ],
            }
        },
        'Inquiry Structuring': {
            'canonical_statement': 'Technological Sovereignty structures inquiry around dependency mapping, capability gap analysis, and strategic sector identification - making these into "problems" while treating global efficiency as already "solved" (negatively).',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'what_becomes_a_problem': [
                    'Dependency on foreign technology',
                    'Capability gaps in strategic sectors',
                    'Foreign platform dominance',
                    'Data flows to adversarial jurisdictions',
                ],
                'what_is_treated_as_solved': [
                    'Pure market allocation (rejected)',
                    'Techno-globalist optimism (rejected)',
                    'Borderless digital governance (rejected)',
                ],
                'inquiry_organized_how': {
                    'typical_research_questions': [
                        'What are our technology dependencies?',
                        'Which sectors are strategically critical?',
                        'What capabilities do we lack?',
                        'How can we reduce vulnerability?',
                    ],
                    'methods_privileged': [
                        'Supply chain mapping',
                        'Capability assessment',
                        'Vulnerability analysis',
                        'Strategic technology roadmapping',
                    ],
                },
            }
        }
    },
    '6. Normalization Analysis': {
        'Norm Embedding Detection': {
            'canonical_statement': 'Technological Sovereignty embeds norms of: collective self-determination (as good), strategic autonomy (as necessary), and productive capacity (as valuable) - treating dependency as inherently problematic.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'embedded_norms': [
                    {
                        'norm': 'Self-determination is good',
                        'how_embedded': 'Built into the very concept of "sovereignty"',
                        'contested_by': 'Cosmopolitans who value universal over particular'
                    },
                    {
                        'norm': 'Strategic autonomy is necessary',
                        'how_embedded': 'Presupposed as goal worth pursuing',
                        'contested_by': 'Those who see interdependence as peace-promoting'
                    },
                    {
                        'norm': 'Productive capacity matters',
                        'how_embedded': 'Implicit preference for making over buying',
                        'contested_by': 'Pure comparative advantage arguments'
                    },
                    {
                        'norm': 'Security can trump efficiency',
                        'how_embedded': 'Justifies higher costs for reduced dependency',
                        'contested_by': 'Market efficiency advocates'
                    },
                ],
            }
        },
        'Pathological Boundary Mapping': {
            'canonical_statement': 'Technological Sovereignty marks as pathological: complete dependency on foreign technology, inability to produce strategic technologies domestically, and subordination of technology policy to pure market logic.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'marked_as_pathological': [
                    {
                        'condition': 'Complete foreign technology dependency',
                        'pathology_label': 'Strategic vulnerability',
                        'treatment_implied': 'Capability building investment'
                    },
                    {
                        'condition': 'No domestic production capacity',
                        'pathology_label': 'Deindustrialization',
                        'treatment_implied': 'Reindustrialization programs'
                    },
                    {
                        'condition': 'Pure market logic in tech policy',
                        'pathology_label': 'Naive liberalism',
                        'treatment_implied': 'Strategic industrial policy'
                    },
                    {
                        'condition': 'Data flowing freely to adversaries',
                        'pathology_label': 'Data colonialism',
                        'treatment_implied': 'Data localization'
                    },
                ],
                'normal_ideal': 'Strategic autonomy in critical technologies with managed interdependence in others',
            }
        },
        'Power Relation Naturalization': {
            'canonical_statement': 'Technological Sovereignty naturalizes: state authority over technology, expert determination of "strategic" sectors, and national framing of technology governance - potentially obscuring class, corporate, and bureaucratic power.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'power_relations_naturalized': [
                    {
                        'relation': 'State authority over technology',
                        'how_naturalized': 'Sovereignty implies state is natural technology governor',
                        'who_benefits': 'State bureaucracies, national security establishments',
                        'what_obscured': 'Corporate capture of state tech policy'
                    },
                    {
                        'relation': 'Expert authority to define "strategic"',
                        'how_naturalized': 'Technical complexity justifies expert judgment',
                        'who_benefits': 'Defense/intelligence agencies, industrial policy officials',
                        'what_obscured': 'Political nature of strategic sector boundaries'
                    },
                    {
                        'relation': 'National framing',
                        'how_naturalized': 'Sovereignty inherently national-scale',
                        'who_benefits': 'National governments over local/regional',
                        'what_obscured': 'City or regional tech governance possibilities'
                    },
                ],
            }
        },
        'Authority Legitimization': {
            'canonical_statement': 'Technological Sovereignty legitimizes authority of: national security officials, industrial policy experts, national-level regulators, and domestic technology champions over foreign firms, global institutions, and pure market signals.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'authorities_legitimized': [
                    {
                        'authority': 'National security officials',
                        'in_domain': 'Technology threat assessment',
                        'over_whom': 'Trade officials, foreign firms'
                    },
                    {
                        'authority': 'Industrial policy experts',
                        'in_domain': 'Strategic sector identification',
                        'over_whom': 'Free market advocates'
                    },
                    {
                        'authority': 'National regulators',
                        'in_domain': 'Technology governance',
                        'over_whom': 'Global platforms, international bodies'
                    },
                    {
                        'authority': 'Domestic tech firms',
                        'in_domain': 'Strategic technology provision',
                        'over_whom': 'Foreign competitors (even if more efficient)'
                    },
                ],
                'credentials_required': 'National security clearance, industrial policy expertise, regulatory authority',
            }
        },
        'Governmentality Mapping': {
            'canonical_statement': 'Technological Sovereignty makes governable: technology supply chains (through dependency mapping), digital populations (through data sovereignty), platform markets (through regulatory intervention), and foreign tech firms (through localization requirements).',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'what_made_governable': [
                    {
                        'domain': 'Technology supply chains',
                        'governance_technique': 'Dependency mapping, chokepoint identification',
                        'knowledge_produced': 'Strategic vulnerability assessments'
                    },
                    {
                        'domain': 'Digital populations',
                        'governance_technique': 'Data localization, digital identity',
                        'knowledge_produced': 'Data flow maps, digital census'
                    },
                    {
                        'domain': 'Platform markets',
                        'governance_technique': 'Antitrust, content regulation, interoperability',
                        'knowledge_produced': 'Market concentration metrics'
                    },
                    {
                        'domain': 'Foreign tech firms',
                        'governance_technique': 'Localization requirements, licensing',
                        'knowledge_produced': 'Foreign ownership registries'
                    },
                ],
                'new_calculabilities': [
                    'Dependency scores for technology sectors',
                    'Data sovereignty indices',
                    'Strategic autonomy metrics',
                ],
            }
        },
        'Truth Regime Analysis': {
            'canonical_statement': 'Within the Technological Sovereignty framework: claims about strategic vulnerability count as true when supported by supply chain analysis; expert security assessments outweigh pure market signals; and dependency reduction is presumptively good.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'what_counts_as_true': [
                    {
                        'claim_type': 'X is strategically critical',
                        'evidence_required': 'Supply chain analysis showing concentrated foreign production',
                        'arbiters': 'National security experts, industrial policy officials'
                    },
                    {
                        'claim_type': 'We are vulnerable to Y',
                        'evidence_required': 'Dependency mapping, scenario analysis',
                        'arbiters': 'Intelligence agencies, defense establishments'
                    },
                    {
                        'claim_type': 'Policy Z advances sovereignty',
                        'evidence_required': 'Capability building metrics, dependency reduction',
                        'arbiters': 'Industrial policy institutions'
                    },
                ],
                'excluded_knowledge': [
                    'Pure efficiency calculations',
                    'Consumer welfare analysis',
                    'Global optimization perspectives',
                ],
            }
        }
    },
    '7. Boundary Analysis': {
        'Anomaly Detection': {
            'canonical_statement': 'Anomalies for Technological Sovereignty include: genuinely borderless technologies (open source, decentralized systems), small states where autarky is impossible, and cases where sovereignty pursuit backfires (innovation slowdown).',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'anomalies': [
                    {
                        'anomaly': 'Decentralized/open source technologies',
                        'why_problematic': 'Cannot be "owned" by any sovereign',
                        'how_addressed': 'Argue these are exceptions or tools for sovereignty',
                        'seriousness': 'Moderate'
                    },
                    {
                        'anomaly': 'Small states',
                        'why_problematic': 'Cannot achieve meaningful tech sovereignty alone',
                        'how_addressed': 'Bloc-level sovereignty (EU) or managed dependency',
                        'seriousness': 'Moderate'
                    },
                    {
                        'anomaly': 'Sovereignty pursuit slowing innovation',
                        'why_problematic': 'Undermines the very capability it seeks',
                        'how_addressed': 'Balance sovereignty with openness; accept tradeoffs',
                        'seriousness': 'High'
                    },
                    {
                        'anomaly': 'Non-state tech governance successes',
                        'why_problematic': 'Shows states aren\'t only/best governors',
                        'how_addressed': 'Argue these are limited exceptions',
                        'seriousness': 'Low'
                    },
                ],
            }
        },
        'Contradiction Mapping': {
            'canonical_statement': 'Technological Sovereignty strictly contradicts: complete tech free trade advocacy, cyber-libertarian claims about ungovernable internet, and techno-utopian assumptions about automatically beneficial technology diffusion.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'logical_contradictions': [
                    {
                        'contradicting_position': 'Complete free trade in all technology',
                        'nature_of_contradiction': 'Denies legitimacy of intervention TS requires',
                        'boundary_type': 'Hard'
                    },
                    {
                        'contradicting_position': 'Internet inherently ungovernable',
                        'nature_of_contradiction': 'Denies state capacity TS assumes',
                        'boundary_type': 'Hard'
                    },
                    {
                        'contradicting_position': 'Technology automatically beneficial',
                        'nature_of_contradiction': 'Denies problematization TS requires',
                        'boundary_type': 'Hard'
                    },
                    {
                        'contradicting_position': 'Efficiency is only criterion',
                        'nature_of_contradiction': 'Denies strategic considerations',
                        'boundary_type': 'Hard'
                    },
                ],
            }
        },
        'Incommensurability Analysis': {
            'canonical_statement': 'Technological Sovereignty is partially incommensurable with techno-libertarianism (different views on state legitimacy) but translatable with economic nationalism and dependency theory (shared concern with power asymmetry).',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'rival_frameworks': [
                    {
                        'framework': 'Techno-libertarianism',
                        'translation_possible': False,
                        'fundamental_disagreement': 'Role of state in technology',
                        'incommensurability_type': 'Normative - different values'
                    },
                    {
                        'framework': 'Pure free trade liberalism',
                        'translation_possible': False,
                        'fundamental_disagreement': 'Legitimacy of intervention',
                        'incommensurability_type': 'Paradigmatic - different starting points'
                    },
                    {
                        'framework': 'Economic nationalism',
                        'translation_possible': True,
                        'shared_vocabulary': 'Strategic sectors, national champions, industrial policy',
                        'translation_note': 'TS is economic nationalism applied to tech'
                    },
                    {
                        'framework': 'Dependency theory',
                        'translation_possible': True,
                        'shared_vocabulary': 'Core-periphery, dependency, development',
                        'translation_note': 'TS shares concern with structural dependency'
                    },
                ],
            }
        },
        'Gray Zone Identification': {
            'canonical_statement': 'Gray zones for Technological Sovereignty include: dual-use technologies, gradually strategic sectors, technologies with distributed production, and sovereignty claims that conflict (national vs. EU).',
            'version': '1.0',
            'confidence': 0.75,
            'analysis': {
                'boundary_cases': [
                    {
                        'case': 'Dual-use technologies',
                        'why_gray': 'Strategic AND commercial simultaneously',
                        'experts_disagree_about': 'Where commercial ends and strategic begins'
                    },
                    {
                        'case': 'Consumer internet services',
                        'why_gray': 'Seem commercial but enable surveillance/influence',
                        'experts_disagree_about': 'Whether TikTok is a security threat'
                    },
                    {
                        'case': 'Globally distributed open source',
                        'why_gray': 'No clear national ownership possible',
                        'experts_disagree_about': 'Whether sovereignty applies'
                    },
                    {
                        'case': 'EU vs. member state sovereignty',
                        'why_gray': 'Competing sovereignty claims',
                        'experts_disagree_about': 'Which level should govern tech'
                    },
                ],
            }
        },
        'Crisis Indicator Tracking': {
            'canonical_statement': 'Crisis indicators for Technological Sovereignty would include: complete failure of sovereignty policies to build capability, massive costs exceeding benefits, or emergence of genuinely borderless technology systems that make sovereignty irrelevant.',
            'version': '1.0',
            'confidence': 0.75,
            'analysis': {
                'crisis_indicators': [
                    {
                        'indicator': 'Sovereignty policies fail to build capability',
                        'threshold': 'Major investments yield no competitive firms',
                        'current_status': 'Too early to tell',
                        'crisis_probability': 'Possible'
                    },
                    {
                        'indicator': 'Costs massively exceed benefits',
                        'threshold': 'Clear evidence of large deadweight loss',
                        'current_status': 'Costs mounting but benefits unclear',
                        'crisis_probability': 'Possible'
                    },
                    {
                        'indicator': 'Truly borderless systems emerge',
                        'threshold': 'Technology escapes all sovereign control',
                        'current_status': 'Cryptocurrency/Web3 are tests',
                        'crisis_probability': 'Unlikely'
                    },
                ],
                'overall_crisis_assessment': 'Concept is robust but faces testable challenges',
            }
        }
    },
    '8. Dynamic Analysis': {
        'Looping Effect Detection': {
            'canonical_statement': 'Technological Sovereignty creates looping effects: classifying technologies as "strategic" changes how they are developed; labeling dependencies as "vulnerabilities" changes how firms structure supply chains.',
            'version': '1.0',
            'confidence': 0.85,
            'analysis': {
                'looping_effects': [
                    {
                        'classification': 'Strategic technology',
                        'how_classified_changes': 'Attracts state investment, faces export controls, becomes geopoliticized',
                        'awareness': 'Firms are very aware of strategic designation',
                        'feedback': 'Firms position for strategic status to get subsidies'
                    },
                    {
                        'classification': 'Dependent state/firm',
                        'how_classified_changes': 'Triggers diversification efforts, changes procurement',
                        'awareness': 'High awareness of dependency discourse',
                        'feedback': 'Supply chains restructure in response'
                    },
                    {
                        'classification': 'Foreign technology threat',
                        'how_classified_changes': 'Faces restrictions, loses market access',
                        'awareness': 'Foreign firms intensely aware',
                        'feedback': 'May localize or exit; may lobby against'
                    },
                ],
                'is_interactive_kind': True,
                'awareness_level': 'High - all actors very aware of sovereignty discourse',
            }
        },
        'Habit Formation Tracking': {
            'canonical_statement': 'Using Technological Sovereignty cultivates habits of: mapping dependencies, thinking geopolitically about technology, prioritizing resilience over efficiency, and viewing foreign technology with suspicion.',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'habits_cultivated': [
                    {
                        'habit': 'Dependency mapping',
                        'description': 'Routinely identifying foreign dependencies in technology',
                        'domain': 'Policy analysis, corporate strategy',
                        'becomes_automatic': True
                    },
                    {
                        'habit': 'Geopolitical technology thinking',
                        'description': 'Viewing tech through power/security lens',
                        'domain': 'Technology policy, procurement',
                        'becomes_automatic': True
                    },
                    {
                        'habit': 'Resilience prioritization',
                        'description': 'Valuing redundancy and self-reliance over pure efficiency',
                        'domain': 'Supply chain management, infrastructure',
                        'becomes_automatic': True
                    },
                    {
                        'habit': 'Foreign tech suspicion',
                        'description': 'Default skepticism toward foreign technology',
                        'domain': 'Procurement, security review',
                        'becomes_automatic': True
                    },
                ],
            }
        },
        'Revision Condition Mapping': {
            'canonical_statement': 'Technological Sovereignty would require revision if: interdependence proved clearly peace-promoting, autarky proved clearly worse, or genuinely post-national technology governance emerged successfully.',
            'version': '1.0',
            'confidence': 0.75,
            'analysis': {
                'revision_conditions': [
                    {
                        'condition': 'Tech interdependence proves peace-promoting',
                        'evidence_needed': 'Clear cases where tech ties prevented conflict',
                        'revision_type': 'Soften toward managed interdependence',
                        'likelihood': 'Possible'
                    },
                    {
                        'condition': 'Autarky proves clearly worse',
                        'evidence_needed': 'Sovereignty-pursuing states fall behind',
                        'revision_type': 'Narrow scope of strategic sectors',
                        'likelihood': 'Possible'
                    },
                    {
                        'condition': 'Post-national governance works',
                        'evidence_needed': 'Successful global tech governance institution',
                        'revision_type': 'Reduce state-centrism',
                        'likelihood': 'Unlikely'
                    },
                ],
                'revisability': 'Moderate - core commitments are strong but scope is adjustable',
            }
        },
        'Kind-Making Analysis': {
            'canonical_statement': 'Technological Sovereignty creates new kinds including: "strategic technologies" (a category that didn\'t exist as such), "tech-sovereign states" (a new state-type), and "digital colonies" (new way of being subordinate).',
            'version': '1.0',
            'confidence': 0.8,
            'analysis': {
                'kinds_created': [
                    {
                        'kind': 'Strategic technology',
                        'existed_before': False,
                        'what_makes_it_new': 'Technology categorized by political rather than technical properties',
                        'interactive': True
                    },
                    {
                        'kind': 'Tech-sovereign state',
                        'existed_before': False,
                        'what_makes_it_new': 'State classified by technological capabilities',
                        'interactive': True
                    },
                    {
                        'kind': 'Digital colony',
                        'existed_before': False,
                        'what_makes_it_new': 'Entity subordinated through technology dependency',
                        'interactive': True
                    },
                    {
                        'kind': 'Technological dependency',
                        'existed_before': False,
                        'what_makes_it_new': 'Dependency as analytical category for technology relations',
                        'interactive': True
                    },
                ],
            }
        },
        'Disciplinary Matrix Evolution': {
            'canonical_statement': 'The disciplinary matrix around Technological Sovereignty is evolving rapidly: experts are splitting between security hawks and managed-interdependence moderates; exemplars are shifting from EU data protection to US CHIPS Act.',
            'version': '1.0',
            'confidence': 0.75,
            'analysis': {
                'expert_community_evolution': [
                    {
                        'period': '2015-2018',
                        'dominant_view': 'European data sovereignty focus',
                        'exemplars': ['GDPR', 'Schrems decisions'],
                        'key_institutions': 'EU Commission, data protection authorities'
                    },
                    {
                        'period': '2018-2022',
                        'dominant_view': 'US-China tech rivalry framing',
                        'exemplars': ['Huawei ban', 'Entity List expansion'],
                        'key_institutions': 'US Commerce, national security agencies'
                    },
                    {
                        'period': '2022-present',
                        'dominant_view': 'Industrial policy renaissance',
                        'exemplars': ['CHIPS Act', 'EU Chips Act', 'supply chain reshoring'],
                        'key_institutions': 'Industrial policy agencies, new tech offices'
                    },
                ],
                'emerging_fault_lines': [
                    'Security hawks vs. managed interdependence',
                    'National vs. bloc-level sovereignty',
                    'Defensive sovereignty vs. tech leadership ambition',
                ],
            }
        }
    }
}

# Theoretical influences reference
THEORETICAL_INFLUENCES = {
    'Quine': {
        'name': 'Willard Van Orman Quine',
        'years': '1908-2000',
        'key_works': ['Two Dogmas of Empiricism (1951)', 'Word and Object (1960)'],
        'core_insight': 'Beliefs face evidence as a corporate body; no belief is immune to revision; knowledge forms an interconnected web.',
        'operations_influenced': ['Inferential Mapping', 'Web Centrality Assessment', 'Contradiction Mapping', 'Revision Condition Mapping']
    },
    'Kuhn': {
        'name': 'Thomas Kuhn',
        'years': '1922-1996',
        'key_works': ['The Structure of Scientific Revolutions (1962)'],
        'core_insight': 'Science operates within paradigms; normal science solves puzzles within paradigms; anomalies accumulate until paradigm shifts occur.',
        'operations_influenced': ['Paradigm Positioning', 'Anomaly Detection', 'Incommensurability Analysis', 'Crisis Indicator Tracking', 'Disciplinary Matrix Evolution']
    },
    'Foucault': {
        'name': 'Michel Foucault',
        'years': '1926-1984',
        'key_works': ['Discipline and Punish (1975)', 'History of Sexuality Vol. 1 (1976)', 'The Archaeology of Knowledge (1969)'],
        'core_insight': 'Knowledge and power are intertwined; discourse determines what can be said; concepts make phenomena governable.',
        'operations_influenced': ['Discursive Positioning', 'Historical Emergence', 'Power Relation Naturalization', 'Authority Legitimization', 'Governmentality Mapping', 'Truth Regime Analysis']
    },
    'Sellars': {
        'name': 'Wilfrid Sellars',
        'years': '1912-1989',
        'key_works': ['Empiricism and the Philosophy of Mind (1956)'],
        'core_insight': 'Nothing is epistemically "given"; all knowledge requires conceptual mediation; the manifest and scientific images may conflict.',
        'operations_influenced': ['Givenness Detection', 'Assumption Excavation', 'Manifest/Scientific Tension']
    },
    'Brandom': {
        'name': 'Robert Brandom',
        'years': '1950-',
        'key_works': ['Making It Explicit (1994)', 'Articulating Reasons (2000)'],
        'core_insight': 'Concepts are defined by their inferential role; assertion involves undertaking commitments and entitlements.',
        'operations_influenced': ['Inferential Commitment Mapping', 'Normative Entailment Tracking', 'Incompatibility Detection', 'Entitlement Tracking']
    },
    'Deleuze': {
        'name': 'Gilles Deleuze',
        'years': '1925-1995',
        'key_works': ['Difference and Repetition (1968)', 'A Thousand Plateaus (1980, with Guattari)'],
        'core_insight': 'Concepts enable transformation and becoming; philosophy creates concepts that open new possibilities.',
        'operations_influenced': ['Transformation Mapping']
    },
    'Bachelard': {
        'name': 'Gaston Bachelard',
        'years': '1884-1962',
        'key_works': ['The Formation of the Scientific Mind (1938)', 'The Philosophy of No (1940)'],
        'core_insight': 'Scientific progress involves epistemological ruptures; prior concepts can become obstacles to knowledge.',
        'operations_influenced': ['Rupture Detection', 'Obstacle Detection']
    },
    'Blumenberg': {
        'name': 'Hans Blumenberg',
        'years': '1920-1996',
        'key_works': ['Paradigms for a Metaphorology (1960)', 'The Legitimacy of the Modern Age (1966)'],
        'core_insight': 'Absolute metaphors structure thought before explicit concepts; they reveal and conceal simultaneously.',
        'operations_influenced': ['Metaphorical Archaeology']
    },
    'Carey': {
        'name': 'Susan Carey',
        'years': '1942-',
        'key_works': ['The Origin of Concepts (2009)', 'Conceptual Change in Childhood (1985)'],
        'core_insight': 'Complex concepts bootstrap from simpler ones through developmental processes; conceptual change can be revolutionary.',
        'operations_influenced': ['Conceptual Bootstrapping']
    },
    'Canguilhem': {
        'name': 'Georges Canguilhem',
        'years': '1904-1995',
        'key_works': ['The Normal and the Pathological (1943/1966)'],
        'core_insight': 'Norms are not discovered but created; the normal/pathological distinction is value-laden and historically contingent.',
        'operations_influenced': ['Norm Embedding Detection', 'Pathological Boundary Mapping']
    },
    'Hacking': {
        'name': 'Ian Hacking',
        'years': '1936-2023',
        'key_works': ['The Social Construction of What? (1999)', 'Historical Ontology (2002)'],
        'core_insight': 'Classification of people creates "looping effects"; interactive kinds change because they are classified.',
        'operations_influenced': ['Visibility Mapping', 'Gray Zone Identification', 'Looping Effect Detection', 'Kind-Making Analysis']
    },
    'James': {
        'name': 'William James',
        'years': '1842-1910',
        'key_works': ['Pragmatism (1907)', 'The Meaning of Truth (1909)'],
        'core_insight': 'The meaning of a concept lies in its practical consequences; truth is what works.',
        'operations_influenced': ['Practical Effect Tracking']
    },
    'Dewey': {
        'name': 'John Dewey',
        'years': '1859-1952',
        'key_works': ['Experience and Nature (1925)', 'Logic: The Theory of Inquiry (1938)'],
        'core_insight': 'Inquiry is problem-solving; concepts are tools; habits structure thought and action.',
        'operations_influenced': ['Practical Effect Tracking', 'Inquiry Structuring', 'Habit Formation Tracking']
    },
    'Rorty': {
        'name': 'Richard Rorty',
        'years': '1931-2007',
        'key_works': ['Philosophy and the Mirror of Nature (1979)', 'Contingency, Irony, and Solidarity (1989)'],
        'core_insight': 'Progress comes through vocabulary change; new ways of speaking open new possibilities.',
        'operations_influenced': ['Vocabulary Extension']
    }
}


def create_operation_indexed_spreadsheet():
    """Create comprehensive operation-indexed concept analysis spreadsheet."""
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== OVERVIEW SHEET ==========
    ws = wb.active
    ws.title = 'Overview'

    ws['A1'] = 'TECHNOLOGICAL SOVEREIGNTY'
    ws['A1'].font = Font(bold=True, size=20)
    ws.merge_cells('A1:F1')

    ws['A2'] = 'Operation-Indexed Concept Analysis (Schema v2.0)'
    ws['A2'].font = Font(italic=True, size=14)
    ws.merge_cells('A2:F2')

    ws['A4'] = 'CONCEPT DEFINITION'
    ws['A4'].font = subheader_font

    ws['A5'] = 'The capacity of a state or political entity to control, develop, and govern critical technologies within its jurisdiction, reducing dependency on foreign actors for essential technological capabilities.'
    ws.merge_cells('A5:F5')
    ws['A5'].alignment = Alignment(wrap_text=True)

    ws['A7'] = 'ANALYSIS METADATA'
    ws['A7'].font = subheader_font

    metadata = [
        ('Analysis Date', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('Schema Version', '2.0 (Operation-Indexed)'),
        ('Analyst', 'Claude'),
        ('Dimensions Analyzed', '8'),
        ('Total Operations', '38'),
        ('Overall Confidence', '0.82'),
    ]

    for i, (key, value) in enumerate(metadata, start=8):
        ws[f'A{i}'] = key
        ws[f'A{i}'].font = bold_font
        ws[f'B{i}'] = value

    ws['A16'] = 'THE 8 ANALYTICAL DIMENSIONS'
    ws['A16'].font = subheader_font

    row = 17
    ws[f'A{row}'] = 'Dimension'
    ws[f'B{row}'] = 'Core Question'
    ws[f'C{row}'] = 'Operations'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'].font = bold_font
    ws[f'C{row}'].font = bold_font

    for dim_name, dim_info in ANALYTICAL_DIMENSIONS.items():
        row += 1
        ws[f'A{row}'] = dim_name.split('. ')[1]
        ws[f'B{row}'] = dim_info['core_question']
        ws[f'C{row}'] = len(dim_info['operations'])
        fill_color = DIMENSION_COLORS[dim_info['color']]
        ws[f'A{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15

    # ========== DIMENSION SHEETS ==========
    for dim_name, dim_info in ANALYTICAL_DIMENSIONS.items():
        ws = wb.create_sheet(dim_name.split('. ')[1][:20])
        fill_color = DIMENSION_COLORS[dim_info['color']]

        # Header
        ws['A1'] = dim_name.upper()
        ws['A1'].font = header_font
        ws['A1'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        ws.merge_cells('A1:F1')

        ws['A2'] = f'Core Question: {dim_info["core_question"]}'
        ws['A2'].font = Font(italic=True, size=11)
        ws.merge_cells('A2:F2')

        ws['A3'] = dim_info['description']
        ws.merge_cells('A3:F3')
        ws['A3'].alignment = Alignment(wrap_text=True)

        row = 5

        # Get analysis data for this dimension
        dim_analysis = TECH_SOV_ANALYSIS.get(dim_name, {})

        for op in dim_info['operations']:
            op_name = op['name']
            op_analysis = dim_analysis.get(op_name, {})

            # Operation header
            ws[f'A{row}'] = op_name
            ws[f'A{row}'].font = subheader_font
            ws[f'A{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Operation description
            ws[f'A{row}'] = f'Description: {op["description"]}'
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Theoretical influences
            ws[f'A{row}'] = f'Theoretical Influences: {", ".join(op["influences"])}'
            ws[f'A{row}'].font = Font(italic=True)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Canonical Statement
            if op_analysis:
                ws[f'A{row}'] = 'CANONICAL STATEMENT:'
                ws[f'A{row}'].font = bold_font
                row += 1
                ws[f'A{row}'] = op_analysis.get('canonical_statement', '')
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1

                # Version and confidence
                ws[f'A{row}'] = f'Version: {op_analysis.get("version", "1.0")} | Confidence: {op_analysis.get("confidence", 0.8)}'
                ws[f'A{row}'].font = Font(italic=True, size=9)
                row += 1

                # Analysis details
                analysis = op_analysis.get('analysis', {})
                if analysis:
                    ws[f'A{row}'] = 'DETAILED ANALYSIS:'
                    ws[f'A{row}'].font = bold_font
                    row += 1

                    for key, value in analysis.items():
                        ws[f'A{row}'] = key.replace('_', ' ').title()
                        ws[f'A{row}'].font = Font(bold=True, size=10)
                        row += 1

                        if isinstance(value, list):
                            for item in value[:8]:  # Limit to 8 items per list
                                if isinstance(item, dict):
                                    item_str = ' | '.join([f'{k}: {v}' for k, v in list(item.items())[:3]])
                                    ws[f'B{row}'] = item_str[:150]
                                else:
                                    ws[f'B{row}'] = str(item)[:150]
                                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                                row += 1
                        elif isinstance(value, dict):
                            for k, v in list(value.items())[:6]:
                                ws[f'B{row}'] = f'{k}: {str(v)[:100]}'
                                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                                row += 1
                        else:
                            ws[f'B{row}'] = str(value)[:200]
                            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                            row += 1

            row += 2  # Space between operations

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25

    # ========== THEORETICAL INFLUENCES SHEET ==========
    ws = wb.create_sheet('Theoretical Influences')

    ws['A1'] = 'THEORETICAL INFLUENCES'
    ws['A1'].font = header_font
    ws.merge_cells('A1:F1')

    ws['A3'] = 'These thinkers inform the analytical operations. The schema is organized by OPERATION, with thinkers as references within each.'
    ws.merge_cells('A3:F3')

    row = 5
    ws[f'A{row}'] = 'Thinker'
    ws[f'B{row}'] = 'Years'
    ws[f'C{row}'] = 'Core Insight'
    ws[f'D{row}'] = 'Operations Influenced'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'].font = bold_font
    ws[f'C{row}'].font = bold_font
    ws[f'D{row}'].font = bold_font

    for thinker_id, thinker in THEORETICAL_INFLUENCES.items():
        row += 1
        ws[f'A{row}'] = thinker['name']
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = thinker['years']
        ws[f'C{row}'] = thinker['core_insight']
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        ws[f'D{row}'] = ', '.join(thinker['operations_influenced'])
        ws[f'D{row}'].alignment = Alignment(wrap_text=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 50

    # ========== CANONICAL STATEMENTS SUMMARY SHEET ==========
    ws = wb.create_sheet('Canonical Statements')

    ws['A1'] = 'ALL CANONICAL STATEMENTS'
    ws['A1'].font = header_font
    ws.merge_cells('A1:D1')

    ws['A3'] = 'One-sentence summaries for each analytical operation, enabling quick reference and IE export.'
    ws.merge_cells('A3:D3')

    row = 5
    ws[f'A{row}'] = 'Dimension'
    ws[f'B{row}'] = 'Operation'
    ws[f'C{row}'] = 'Canonical Statement'
    ws[f'D{row}'] = 'Confidence'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'].font = bold_font
    ws[f'C{row}'].font = bold_font
    ws[f'D{row}'].font = bold_font

    for dim_name, dim_info in ANALYTICAL_DIMENSIONS.items():
        dim_analysis = TECH_SOV_ANALYSIS.get(dim_name, {})
        fill_color = DIMENSION_COLORS[dim_info['color']]

        for op in dim_info['operations']:
            op_analysis = dim_analysis.get(op['name'], {})
            row += 1
            ws[f'A{row}'] = dim_name.split('. ')[1]
            ws[f'A{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws[f'B{row}'] = op['name']
            ws[f'C{row}'] = op_analysis.get('canonical_statement', '[Not yet analyzed]')
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            ws[f'D{row}'] = op_analysis.get('confidence', 0.0)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 100
    ws.column_dimensions['D'].width = 12

    # ========== VERSION HISTORY SHEET ==========
    ws = wb.create_sheet('Version History')

    ws['A1'] = 'VERSION HISTORY'
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')

    ws['A3'] = 'Track changes to canonical statements and analysis over time.'
    ws.merge_cells('A3:E3')

    row = 5
    headers = ['Date', 'Operation', 'Version', 'Change Type', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = bold_font

    row += 1
    ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d'))
    ws.cell(row=row, column=2, value='ALL')
    ws.cell(row=row, column=3, value='1.0')
    ws.cell(row=row, column=4, value='Initial Analysis')
    ws.cell(row=row, column=5, value='First comprehensive operation-indexed analysis of Technological Sovereignty')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 60

    # ========== SCHEMA REFERENCE SHEET ==========
    ws = wb.create_sheet('Schema Reference')

    ws['A1'] = 'OPERATION-INDEXED SCHEMA REFERENCE'
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')

    ws['A3'] = 'This schema organizes concept analysis by WHAT YOU\'RE TRYING TO LEARN, not by WHICH THINKER INSPIRED IT.'
    ws.merge_cells('A3:E3')

    row = 5
    ws[f'A{row}'] = 'KEY DIFFERENCES FROM THINKER-INDEXED SCHEMA:'
    ws[f'A{row}'].font = subheader_font

    differences = [
        ('Old', 'Quinean dimension', 'New', 'Positional Analysis > Inferential Mapping'),
        ('Old', 'Sellarsian dimension', 'New', 'Presuppositional Analysis > Givenness Detection'),
        ('Old', 'Brandomian dimension', 'New', 'Commitment Analysis > Inferential Commitment Mapping'),
        ('Old', 'Deleuzian dimension', 'New', 'Affordance Analysis > Transformation Mapping'),
        ('Old', 'Bachelardian dimension', 'New', 'Genealogical Analysis > Rupture Detection'),
        ('Old', 'Canguilhem dimension', 'New', 'Normalization Analysis > Norm Embedding Detection'),
        ('Old', 'Hacking dimension', 'New', 'Dynamic Analysis > Looping Effect Detection'),
        ('Old', 'Blumenberg dimension', 'New', 'Genealogical Analysis > Metaphorical Archaeology'),
        ('Old', 'Carey dimension', 'New', 'Genealogical Analysis > Conceptual Bootstrapping'),
        ('Old', 'Kuhnian dimension', 'New', 'Positional Analysis > Paradigm Positioning'),
        ('Old', 'Pragmatist dimension', 'New', 'Affordance Analysis > Practical Effect Tracking'),
        ('Old', 'Foucauldian dimension', 'New', 'Multiple: Normalization, Genealogical, Positional'),
    ]

    for old, old_val, new, new_val in differences:
        row += 1
        ws[f'A{row}'] = old_val
        ws[f'B{row}'] = '→'
        ws[f'C{row}'] = new_val

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 50

    # Save
    output_path = '/home/evgeny/Downloads/TechSovereignty_OperationIndexed_Analysis.xlsx'
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


if __name__ == '__main__':
    create_operation_indexed_spreadsheet()
