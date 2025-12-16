#!/usr/bin/env python3
"""
Generate FULLY MODULAR concept schema Excel.
Every multi-valued field is a separate table with individually addressable rows.
Enhanced based on research into all 9 thinkers.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
new_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")  # Yellow for new from research
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Core": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Davidson": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

def style_header(ws, row, col, value, color="2C3E50"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return cell

def add_schema_table(ws, start_row, table_name, fields, new_from_research=False):
    """Add a table schema section to worksheet."""
    row = start_row

    # Table name
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value=table_name)
    cell.font = Font(bold=True, size=12)
    if new_from_research:
        cell.fill = new_fill
    row += 1

    # Headers
    for col, header in enumerate(["Field", "Type", "Constraints", "Description"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    # Fields
    for field in fields:
        ws.cell(row=row, column=1, value=field["name"]).font = field_font
        ws.cell(row=row, column=2, value=field["type"])
        ws.cell(row=row, column=2).fill = type_fill
        ws.cell(row=row, column=3, value=field.get("constraints", ""))
        if "FK" in field.get("constraints", ""):
            ws.cell(row=row, column=3).fill = fk_fill
        ws.cell(row=row, column=4, value=field["desc"])
        if field.get("new"):
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = new_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    return row + 1

def add_example_rows(ws, start_row, headers, rows):
    """Add example data rows."""
    row = start_row
    ws.cell(row=row, column=1, value="EXAMPLE DATA:").font = Font(bold=True)
    row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for data_row in rows:
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap
        row += 1

    return row + 1

# ============================================================================
# Sheet 1: Overview
# ============================================================================
ws = wb.active
ws.title = "Overview"

overview = [
    ["ENRICHED CONCEPT SCHEMA v2 - FULLY MODULAR"],
    [""],
    ["KEY CHANGES FROM v1:"],
    ["1. All multi-valued fields are now SEPARATE TABLES with individual rows"],
    ["2. Each row has its own ID, timestamps, confidence, and can be individually revised"],
    ["3. New fields added based on research into each thinker (marked in YELLOW)"],
    [""],
    ["RESEARCH ENHANCEMENTS BY DIMENSION:"],
    [""],
    ["Dimension", "Thinker", "New Insights Added"],
    ["Quinean", "W.V.O. Quine", "Ontological relativity, theory holism, entrenchment score"],
    ["Sellarsian", "Wilfrid Sellars", "Space of reasons, manifest/scientific image tension"],
    ["Brandomian", "Robert Brandom", "Deontic scorekeeping, inheritance chains"],
    ["Deleuzian", "Gilles Deleuze", "Lines of flight, deterritorialization, virtual/actual"],
    ["Bachelardian", "Gaston Bachelard", "Obstacle subtypes, psychoanalytic function, epistemological profile"],
    ["Canguilhem", "Georges Canguilhem", "Concept filiation, vitality indicators, normal vs normative"],
    ["Davidson/Hacking", "Ian Hacking", "Style emergence, objects created, dynamic nominalism"],
    ["Blumenberg", "Hans Blumenberg", "Nonconceptuality, metakinetics, absolutism of reality"],
    ["Carey", "Susan Carey", "Core cognition, incommensurability, fast/extended mapping"],
    [""],
    ["TABLE COUNT: 24 tables (was 13)"],
    [""],
    ["MODULARITY PRINCIPLE:"],
    ["Every claim about a concept is a separate row that can be:"],
    ["- Challenged by evidence"],
    ["- Assigned confidence scores"],
    ["- Revised or removed independently"],
    ["- Traced to its source"],
]

for row_idx, row_data in enumerate(overview, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif row_idx == 10:
            cell.font = header_font
            cell.fill = header_fill

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 60

# ============================================================================
# Sheet 2: Core concepts table
# ============================================================================
ws = wb.create_sheet("1. concepts")
style_header(ws, 1, 1, "CORE CONCEPT TABLE", "2C3E50")
ws.merge_cells('A1:D1')

row = add_schema_table(ws, 3, "concepts", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "term", "type": "VARCHAR(255)", "constraints": "NOT NULL, INDEX", "desc": "Concept name/term"},
    {"name": "definition", "type": "TEXT", "constraints": "NOT NULL", "desc": "Core definition"},
    {"name": "category", "type": "VARCHAR(100)", "constraints": "", "desc": "Domain category"},
    {"name": "source_id", "type": "INTEGER", "constraints": "FK → theory_sources", "desc": "Where concept originated"},
    {"name": "status", "type": "ENUM", "constraints": "draft/active/deprecated/challenged", "desc": "Current status"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 1.0", "desc": "Overall confidence 0-1"},
    {"name": "--- QUINEAN ---", "type": "", "constraints": "", "desc": ""},
    {"name": "centrality", "type": "VARCHAR(20)", "constraints": "core/intermediate/peripheral", "desc": "Position in inferential web"},
    {"name": "entrenchment_score", "type": "FLOAT", "constraints": "0-1", "desc": "How costly to revise (Quine)", "new": True},
    {"name": "--- CANGUILHEM ---", "type": "", "constraints": "", "desc": ""},
    {"name": "health_status", "type": "VARCHAR(20)", "constraints": "healthy/strained/dying/being_born", "desc": "Concept vitality"},
    {"name": "birth_period", "type": "VARCHAR(100)", "constraints": "", "desc": "When concept emerged"},
    {"name": "birth_problem", "type": "TEXT", "constraints": "", "desc": "Problem that birthed concept"},
    {"name": "--- CAREY ---", "type": "", "constraints": "", "desc": ""},
    {"name": "hierarchy_level", "type": "INTEGER", "constraints": "0=primitive, 1-3+=complex", "desc": "Conceptual hierarchy level"},
    {"name": "bootstrap_status", "type": "VARCHAR(20)", "constraints": "successful/partial/failed", "desc": "Did bootstrap achieve leap?"},
    {"name": "core_cognition_derived", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "From innate core systems? (Carey)", "new": True},
    {"name": "--- TIMESTAMPS ---", "type": "", "constraints": "", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "updated_at", "type": "TIMESTAMPTZ", "constraints": "ON UPDATE NOW()", "desc": ""},
])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 45

# ============================================================================
# Sheet 3: Quinean - Inferences (modular)
# ============================================================================
ws = wb.create_sheet("2. Quinean")
style_header(ws, 1, 1, "QUINEAN DIMENSION - Web of Belief", dimension_colors["Quinean"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on W.V.O. Quine: Confirmation holism, web of belief, ontological relativity")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_inferences (individual inferential connections)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "forward/backward/lateral/contradiction"},
    {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The inference itself"},
    {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept (if in DB)"},
    {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Inference strength 0-1"},
    {"name": "defeasible", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Can be defeated by evidence?", "new": True},
    {"name": "source_note", "type": "TEXT", "constraints": "", "desc": "Where this inference comes from"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Confidence in this specific inference"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When last challenged by evidence"},
])

row = add_example_rows(ws, row, ["id", "concept_id", "type", "statement", "strength"], [
    [1, 42, "forward", "If tech sovereignty → indigenous chip development possible", 0.7],
    [2, 42, "forward", "If tech sovereignty → reduced foreign dependency", 0.8],
    [3, 42, "backward", "Tech sovereignty ← sufficient R&D investment", 0.6],
    [4, 42, "lateral", "Related to economic sovereignty", 0.9],
    [5, 42, "contradiction", "Material dependency contradicts autonomy claims", 0.85],
])

row = add_schema_table(ws, row + 1, "concept_web_tensions (NEW - theory holism tensions)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "tension_with_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Concept it's in tension with"},
    {"name": "tension_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Nature of the tension"},
    {"name": "resolution_cost", "type": "VARCHAR(20)", "constraints": "low/medium/high/catastrophic", "desc": "Cost to resolve (Duhem-Quine)"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], new_from_research=True)

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 4: Sellarsian (modular)
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
style_header(ws, 1, 1, "SELLARSIAN DIMENSION - Myth of the Given", dimension_colors["Sellarsian"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Wilfrid Sellars: Space of reasons, myth of the given, manifest vs scientific image")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_givenness (1:1 with concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "is_myth_of_given", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Is concept falsely treated as foundational?"},
    {"name": "should_be_inferred_from", "type": "TEXT", "constraints": "", "desc": "What evidence should actually support this"},
    {"name": "space_of_reasons_role", "type": "TEXT", "constraints": "", "desc": "Role in justification/inference (NEW)", "new": True},
    {"name": "manifest_scientific_tension", "type": "TEXT", "constraints": "", "desc": "Tension between everyday and scientific views (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_givenness_markers (modular - one per row)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "marker_text", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Language marker: 'obviously', 'naturally', etc."},
    {"name": "example_usage", "type": "TEXT", "constraints": "", "desc": "Example of this marker in use"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Confidence this is a givenness marker"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "concept_id", "marker_text", "example_usage"], [
    [1, 42, "naturally", "'Tech naturally extends state power'"],
    [2, 42, "clearly possible", "'Sovereignty is clearly possible with investment'"],
    [3, 42, "inevitable", "'Technological independence is inevitable'"],
    [4, 42, "obvious strategic necessity", "'Obviously a strategic necessity'"],
])

row = add_schema_table(ws, row, "concept_hidden_commitments (modular - one per row)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "commitment_text", "type": "TEXT", "constraints": "NOT NULL", "desc": "Hidden assumption baked in"},
    {"name": "evidence_that_exposes", "type": "TEXT", "constraints": "", "desc": "What evidence could expose this"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_givenness_effects (modular - enables/blocks)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "effect_type", "type": "VARCHAR(20)", "constraints": "enables/blocks", "desc": "Does givenness enable or block?"},
    {"name": "effect_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is enabled or blocked"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "concept_id", "effect_type", "effect_description"], [
    [1, 42, "enables", "Mobilizes resources for technology projects"],
    [2, 42, "enables", "Justifies protectionist policies"],
    [3, 42, "enables", "Creates political legitimacy for tech investments"],
    [4, 42, "blocks", "Questions about structural impossibility"],
    [5, 42, "blocks", "Analysis of why sovereignty might be unachievable"],
    [6, 42, "blocks", "Recognition of necessary interdependence"],
])

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 5: Brandomian (modular)
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
style_header(ws, 1, 1, "BRANDOMIAN DIMENSION - Commitments & Entitlements", dimension_colors["Brandomian"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Robert Brandom: Deontic scorekeeping, game of giving and asking for reasons")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_commitments (one commitment per row)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "commitment_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "commitment/entitlement/incompatibility"},
    {"name": "statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "What you're committed to / entitled to"},
    {"name": "is_honored", "type": "BOOLEAN", "constraints": "", "desc": "Is this commitment honored in practice?"},
    {"name": "violation_evidence", "type": "TEXT", "constraints": "", "desc": "If violated, where/how"},
    {"name": "inherited_from_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Inherited from another concept? (NEW)", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When last challenged"},
])

row = add_example_rows(ws, row, ["id", "type", "statement", "honored?", "violation_evidence"], [
    [1, "commitment", "Autonomy is achievable through investment", "NO", "Gulf $2T can't invest at Series B"],
    [2, "commitment", "State can control tech development trajectory", "NO", "Dependent on foreign equipment"],
    [3, "entitlement", "Can claim sovereignty through rhetoric", "EXCEEDED", "Claims without substance"],
    [4, "incompatibility", "Sovereignty AND deep supply chain integration", "n/a", "Logical tension"],
])

row = add_schema_table(ws, row, "concept_scorekeeping (NEW - track deontic score changes)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "score_event", "type": "TEXT", "constraints": "NOT NULL", "desc": "What changed the score"},
    {"name": "commitment_id", "type": "INTEGER", "constraints": "FK → concept_commitments", "desc": "Which commitment affected"},
    {"name": "score_change", "type": "VARCHAR(20)", "constraints": "gained/lost/challenged", "desc": "Type of change"},
    {"name": "evidence_source", "type": "TEXT", "constraints": "", "desc": "What evidence triggered this"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], new_from_research=True)

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 6: Deleuzian (modular)
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
style_header(ws, 1, 1, "DELEUZIAN DIMENSION - Problems & Becomings", dimension_colors["Deleuzian"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Gilles Deleuze: Plane of immanence, becoming, deterritorialization, lines of flight")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_problems (problems/tensions addressed)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "problem_addressed", "type": "TEXT", "constraints": "NOT NULL", "desc": "What tension/problem does concept navigate"},
    {"name": "tension_pole_a", "type": "TEXT", "constraints": "", "desc": "One side of the tension"},
    {"name": "tension_pole_b", "type": "TEXT", "constraints": "", "desc": "Other side of the tension"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_creative_responses (modular - one per row)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "problem_id", "type": "INTEGER", "constraints": "FK → concept_problems, NOT NULL", "desc": "Which problem this responds to"},
    {"name": "response_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "How concept navigates the problem"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_becomings (modular - one per row)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "becoming_type", "type": "VARCHAR(20)", "constraints": "enabled/blocked", "desc": "Does concept enable or block this?"},
    {"name": "becoming_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "The transformation"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "concept_id", "type", "description"], [
    [1, 42, "enabled", "Becoming-investor: states transformed into tech funders"],
    [2, 42, "enabled", "Becoming-protectionist: legitimate barriers to foreign tech"],
    [3, 42, "blocked", "Becoming-honest: cannot admit structural impossibility"],
    [4, 42, "blocked", "Becoming-interdependent: cannot embrace strategic dependency"],
])

row = add_schema_table(ws, row, "concept_lines_of_flight (NEW - escape routes from concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "line_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Potential escape/transformation route"},
    {"name": "deterritorializes_toward", "type": "TEXT", "constraints": "", "desc": "What new territory does this open?"},
    {"name": "blocked_by", "type": "TEXT", "constraints": "", "desc": "What prevents this line of flight?"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], new_from_research=True)

row = add_schema_table(ws, row, "concept_plane_assumptions (background assumptions)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "assumption", "type": "TEXT", "constraints": "NOT NULL", "desc": "The unquestioned assumption"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_plane_effects (modular - what assumptions enable/foreclose)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "assumption_id", "type": "INTEGER", "constraints": "FK → concept_plane_assumptions", "desc": "Which assumption"},
    {"name": "effect_type", "type": "VARCHAR(20)", "constraints": "makes_possible/makes_impossible", "desc": ""},
    {"name": "effect_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is enabled or foreclosed"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 7: Bachelardian (modular)
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
style_header(ws, 1, 1, "BACHELARDIAN DIMENSION - Epistemological Obstacles", dimension_colors["Bachelardian"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Gaston Bachelard: Epistemological obstacles, rupture, psychoanalysis of knowledge")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_obstacles (1:1 - main obstacle analysis)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "is_obstacle", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Does concept block understanding?"},
    {"name": "obstacle_type", "type": "VARCHAR(30)", "constraints": "", "desc": "experience/verbal/pragmatic/quantitative/substantialist/animist"},
    {"name": "why_persists", "type": "TEXT", "constraints": "", "desc": "Ideological/class function"},
    {"name": "rupture_would_enable", "type": "TEXT", "constraints": "", "desc": "What becomes thinkable after rupture"},
    {"name": "rupture_trigger", "type": "TEXT", "constraints": "", "desc": "What would force abandonment"},
    {"name": "psychoanalytic_function", "type": "TEXT", "constraints": "", "desc": "Unconscious need it serves (NEW)", "new": True},
    {"name": "epistemological_profile", "type": "TEXT", "constraints": "", "desc": "Where concept sits on progress scale (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_obstacle_blocks (modular - what it blocks)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "blocked_understanding", "type": "TEXT", "constraints": "NOT NULL", "desc": "What understanding/question it prevents"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "concept_id", "blocked_understanding"], [
    [1, 42, "Recognition of structural dependency as permanent"],
    [2, 42, "Analysis of power asymmetries in global tech system"],
    [3, 42, "Questions about whether autonomy is possible or desirable"],
])

row = add_schema_table(ws, row, "concept_inadequacy_evidence (modular - empirical challenges)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "evidence_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Empirical challenge to the concept"},
    {"name": "source_cluster_id", "type": "INTEGER", "constraints": "FK → evidence_clusters", "desc": "Source evidence cluster"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "concept_id", "evidence_description"], [
    [1, 42, "Gulf states: $2T wealth but cannot invest at Series B"],
    [2, 42, "China: Massive investment but still depends on ASML"],
    [3, 42, "Russia: Sanctions reveal depth of tech dependencies"],
])

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 8: Canguilhem (modular)
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
style_header(ws, 1, 1, "CANGUILHEM DIMENSION - Life History of Concepts", dimension_colors["Canguilhem"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Georges Canguilhem: Normal vs normative, concept life cycles, filiation of concepts")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_evolution (historical transformations - one per period)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "period", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Time period of transformation"},
    {"name": "transformation_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What changed"},
    {"name": "problem_driving", "type": "TEXT", "constraints": "", "desc": "What problem drove this change"},
    {"name": "who_transformed", "type": "TEXT", "constraints": "", "desc": "Intellectual tradition, actors"},
    {"name": "predecessor_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Concept filiation (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "period", "transformation", "problem"], [
    [1, "1648 (Westphalia)", "Territorial sovereignty", "Ending religious wars"],
    [2, "1950s-70s", "Economic sovereignty", "Decolonization"],
    [3, "1990s-2000s", "Data sovereignty", "Internet, privacy"],
    [4, "2010s-present", "Technological sovereignty", "US-China rivalry, Snowden"],
])

row = add_schema_table(ws, row, "concept_normative_dimensions (modular - embedded values)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "value_embedded", "type": "TEXT", "constraints": "NOT NULL", "desc": "What value is embedded"},
    {"name": "whose_values", "type": "TEXT", "constraints": "", "desc": "Whose interests this serves"},
    {"name": "what_excluded", "type": "TEXT", "constraints": "", "desc": "What's marked as abnormal/excluded"},
    {"name": "normal_vs_normative", "type": "VARCHAR(20)", "constraints": "normal/normative", "desc": "Is this treated as norm or value? (NEW)", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_vitality_indicators (NEW - tracking concept health)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "indicator_type", "type": "VARCHAR(30)", "constraints": "", "desc": "growth/strain/crisis/revival"},
    {"name": "indicator_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What shows the concept's health"},
    {"name": "observed_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When this was observed"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], new_from_research=True)

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 9: Davidson/Hacking (modular)
# ============================================================================
ws = wb.create_sheet("8. Davidson-Hacking")
style_header(ws, 1, 1, "DAVIDSON/HACKING DIMENSION - Styles of Reasoning", dimension_colors["Davidson"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Ian Hacking & Arnold Davidson: Historical epistemology, styles of reasoning, dynamic nominalism")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_reasoning_styles (styles required by concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "style_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Name of reasoning style"},
    {"name": "style_emerged", "type": "VARCHAR(100)", "constraints": "", "desc": "When this style emerged (NEW)", "new": True},
    {"name": "objects_created", "type": "TEXT", "constraints": "", "desc": "What objects of inquiry does style create? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_style_visibility (modular - what style makes visible/invisible)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": "Parent style"},
    {"name": "visibility_type", "type": "VARCHAR(20)", "constraints": "visible/invisible", "desc": ""},
    {"name": "what_affected", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is made visible or invisible"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "style_id", "type", "what_affected"], [
    [1, 1, "visible", "State competition and rivalry"],
    [2, 1, "visible", "Strategic sectors and chokepoints"],
    [3, 1, "invisible", "Class dynamics within states"],
    [4, 1, "invisible", "Corporate interests driving policy"],
    [5, 1, "invisible", "Impossibility of autonomy under global production"],
])

row = add_schema_table(ws, row, "concept_style_evidence (modular - evidence types privileged/marginalized)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": "Parent style"},
    {"name": "evidence_status", "type": "VARCHAR(20)", "constraints": "privileged/marginalized", "desc": ""},
    {"name": "evidence_type", "type": "TEXT", "constraints": "NOT NULL", "desc": "Type of evidence"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_style_inferences (modular - characteristic inference patterns)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": "Parent style"},
    {"name": "inference_pattern", "type": "TEXT", "constraints": "NOT NULL", "desc": "Characteristic reasoning move"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 10: Blumenberg (modular)
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
style_header(ws, 1, 1, "BLUMENBERG DIMENSION - Metaphorology", dimension_colors["Blumenberg"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Hans Blumenberg: Absolute metaphors, nonconceptuality, metakinetics")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_metaphors (root metaphors - one per row)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "root_metaphor", "type": "TEXT", "constraints": "NOT NULL", "desc": "The metaphor itself"},
    {"name": "source_domain", "type": "TEXT", "constraints": "", "desc": "Where metaphor comes from"},
    {"name": "is_absolute", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Cannot be translated to concepts? (NEW)", "new": True},
    {"name": "nonconceptuality_aspect", "type": "TEXT", "constraints": "", "desc": "What resists conceptualization? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_metaphor_effects (modular - enables/hides)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "metaphor_id", "type": "INTEGER", "constraints": "FK → concept_metaphors, NOT NULL", "desc": "Parent metaphor"},
    {"name": "effect_type", "type": "VARCHAR(20)", "constraints": "enables/hides", "desc": ""},
    {"name": "effect_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is enabled or hidden"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "metaphor_id", "type", "description"], [
    [1, 1, "enables", "Thinking tech can be 'owned' like territory"],
    [2, 1, "enables", "Border concepts for data"],
    [3, 1, "hides", "Tech is networked not bounded"],
    [4, 1, "hides", "Requires ongoing relationships"],
])

row = add_schema_table(ws, row, "concept_metakinetics (NEW - how metaphors transform over time)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "metaphor_id", "type": "INTEGER", "constraints": "FK → concept_metaphors, NOT NULL", "desc": "Parent metaphor"},
    {"name": "period", "type": "VARCHAR(100)", "constraints": "", "desc": "Time period"},
    {"name": "transformation", "type": "TEXT", "constraints": "NOT NULL", "desc": "How metaphor meaning shifted"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], new_from_research=True)

row = add_schema_table(ws, row, "concept_work_in_progress (conceptual work being done)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "original_meaning", "type": "TEXT", "constraints": "", "desc": "What concept originally meant"},
    {"name": "current_work", "type": "TEXT", "constraints": "NOT NULL", "desc": "What transformation is being attempted"},
    {"name": "who_doing_work", "type": "TEXT", "constraints": "", "desc": "Intellectual tradition, actors"},
    {"name": "work_status", "type": "VARCHAR(30)", "constraints": "succeeding/failing/ongoing", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 11: Carey (modular)
# ============================================================================
ws = wb.create_sheet("10. Carey")
style_header(ws, 1, 1, "CAREY DIMENSION - Conceptual Bootstrapping", dimension_colors["Carey"])
ws.merge_cells('A1:D1')

ws.cell(row=3, column=1, value="Based on Susan Carey: Core cognition, bootstrapping, conceptual discontinuities, incommensurability")
ws.cell(row=3, column=1).font = Font(italic=True)

row = add_schema_table(ws, 5, "concept_hierarchy (1:1 - bootstrap structure)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "combination_type", "type": "VARCHAR(30)", "constraints": "", "desc": "simple_aggregation/interactive/qualitative_leap"},
    {"name": "transparency", "type": "VARCHAR(20)", "constraints": "high/medium/low", "desc": "How visible are components"},
    {"name": "bootstrap_failure_reason", "type": "TEXT", "constraints": "", "desc": "If failed, why"},
    {"name": "what_would_fix", "type": "TEXT", "constraints": "", "desc": "What would make bootstrap succeed"},
    {"name": "incommensurable_with", "type": "TEXT", "constraints": "", "desc": "Prior concepts this can't be reduced to (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_schema_table(ws, row, "concept_built_from (modular - component concepts)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "component_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Component concept (if in DB)"},
    {"name": "component_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Description of component"},
    {"name": "component_level", "type": "INTEGER", "constraints": "", "desc": "Hierarchy level of component (0-3)"},
    {"name": "how_combined", "type": "TEXT", "constraints": "", "desc": "How this component is used"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_example_rows(ws, row, ["id", "concept_id", "component_desc", "level", "how_combined"], [
    [1, 42, "Territory (bounded space)", 0, "Provides 'bounded control' aspect"],
    [2, 42, "Authority (legitimate power)", 0, "Provides 'supreme control' aspect"],
    [3, 42, "State (political entity)", 1, "Actor who exercises sovereignty"],
    [4, 42, "Sovereignty (state+territory+authority)", 2, "Base concept being extended"],
    [5, 42, "Technology domain", 1, "New domain sovereignty extends to"],
])

row = add_schema_table(ws, row, "concept_mapping_process (NEW - how concept is learned/acquired)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "mapping_type", "type": "VARCHAR(20)", "constraints": "fast/extended", "desc": "Fast mapping or extended? (Carey)"},
    {"name": "mapping_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "How concept gets acquired"},
    {"name": "executive_function_needed", "type": "TEXT", "constraints": "", "desc": "What cognitive effort needed for conceptual change"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], new_from_research=True)

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 12: Table Summary
# ============================================================================
ws = wb.create_sheet("Table Summary")
style_header(ws, 1, 1, "COMPLETE TABLE LIST", "2C3E50")
ws.merge_cells('A1:E1')

headers = ["#", "Table Name", "Dimension", "Relationship", "Purpose"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

tables = [
    [1, "concepts", "Core", "1 per concept", "Main concept record with Quinean/Canguilhem/Carey extensions"],
    [2, "concept_inferences", "Quinean", "Many:1", "Individual inferential connections"],
    [3, "concept_web_tensions", "Quinean", "Many:1", "NEW: Tensions with other concepts (Duhem-Quine)"],
    [4, "concept_givenness", "Sellarsian", "1:1", "Myth of given analysis"],
    [5, "concept_givenness_markers", "Sellarsian", "Many:1", "Individual givenness language markers"],
    [6, "concept_hidden_commitments", "Sellarsian", "Many:1", "Individual hidden assumptions"],
    [7, "concept_givenness_effects", "Sellarsian", "Many:1", "Individual enables/blocks effects"],
    [8, "concept_commitments", "Brandomian", "Many:1", "Individual commitments/entitlements"],
    [9, "concept_scorekeeping", "Brandomian", "Many:1", "NEW: Deontic score change events"],
    [10, "concept_problems", "Deleuzian", "Many:1", "Problems/tensions addressed"],
    [11, "concept_creative_responses", "Deleuzian", "Many:1", "Individual creative responses"],
    [12, "concept_becomings", "Deleuzian", "Many:1", "Individual enabled/blocked becomings"],
    [13, "concept_lines_of_flight", "Deleuzian", "Many:1", "NEW: Escape routes from concept"],
    [14, "concept_plane_assumptions", "Deleuzian", "Many:1", "Background assumptions"],
    [15, "concept_plane_effects", "Deleuzian", "Many:1", "Effects of each assumption"],
    [16, "concept_obstacles", "Bachelardian", "1:1", "Main obstacle analysis"],
    [17, "concept_obstacle_blocks", "Bachelardian", "Many:1", "Individual blocked understandings"],
    [18, "concept_inadequacy_evidence", "Bachelardian", "Many:1", "Individual empirical challenges"],
    [19, "concept_evolution", "Canguilhem", "Many:1", "Historical transformations"],
    [20, "concept_normative_dimensions", "Canguilhem", "Many:1", "Individual embedded values"],
    [21, "concept_vitality_indicators", "Canguilhem", "Many:1", "NEW: Health indicators over time"],
    [22, "concept_reasoning_styles", "Davidson", "Many:1", "Required reasoning styles"],
    [23, "concept_style_visibility", "Davidson", "Many:1", "What each style makes visible/invisible"],
    [24, "concept_style_evidence", "Davidson", "Many:1", "Evidence types privileged/marginalized"],
    [25, "concept_style_inferences", "Davidson", "Many:1", "Characteristic inference patterns"],
    [26, "concept_metaphors", "Blumenberg", "Many:1", "Root metaphors"],
    [27, "concept_metaphor_effects", "Blumenberg", "Many:1", "What each metaphor enables/hides"],
    [28, "concept_metakinetics", "Blumenberg", "Many:1", "NEW: How metaphors transform over time"],
    [29, "concept_work_in_progress", "Blumenberg", "Many:1", "Conceptual transformation attempts"],
    [30, "concept_hierarchy", "Carey", "1:1", "Bootstrap structure"],
    [31, "concept_built_from", "Carey", "Many:1", "Component concepts"],
    [32, "concept_mapping_process", "Carey", "Many:1", "NEW: How concept is acquired"],
]

for row_idx, row_data in enumerate(tables, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap
        if "NEW" in str(value):
            cell.fill = new_fill

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 55

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D_Modular_v2.xlsx")
wb.save(output_path)
print(f"Modular schema Excel saved to: {output_path}")
