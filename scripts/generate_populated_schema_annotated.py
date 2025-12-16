#!/usr/bin/env python3
"""
Generate ANNOTATED & POPULATED v5 Concept Schema with "Technological Sovereignty".

Each sheet includes:
1. Dimension explanation (philosophical background)
2. Table explanations (what each table captures)
3. Field explanations (what each column means)
4. Actual populated data with examples
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
section_font = Font(bold=True, size=12, color="2C3E50")
explain_font = Font(italic=True, size=10, color="555555")
field_explain_font = Font(size=9, color="666666")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
explain_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
field_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
data_font = Font(size=10)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Overview": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Hacking": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

def style_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return row + 1

def add_explanation(ws, row, text, merge_cols=8):
    """Add an explanation paragraph."""
    ws.merge_cells(f'A{row}:{chr(64+merge_cols)}{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = explain_font
    cell.fill = explain_fill
    cell.alignment = wrap
    ws.row_dimensions[row].height = max(30, len(text) // 2)
    return row + 1

def add_section_header(ws, row, text):
    """Add a section header."""
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    return row + 1

def add_field_explanations(ws, row, field_explanations):
    """Add field explanation rows."""
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value="FIELD EXPLANATIONS:").font = Font(bold=True, size=10)
    row += 1
    for field, explanation in field_explanations:
        ws.cell(row=row, column=1, value=f"  {field}:").font = Font(bold=True, size=9)
        ws.merge_cells(f'B{row}:H{row}')
        ws.cell(row=row, column=2, value=explanation).font = field_explain_font
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = field_fill
        row += 1
    return row + 1

def add_table(ws, start_row, headers, data, title=None):
    """Add a data table to worksheet."""
    row = start_row

    if title:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for record in data:
        for col, val in enumerate(record, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = wrap
        row += 1

    return row + 1

def set_column_widths(ws, widths=None):
    if widths is None:
        widths = {'A': 22, 'B': 25, 'C': 25, 'D': 20, 'E': 20, 'F': 15, 'G': 18, 'H': 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

# ============================================================================
# SHEET 1: Overview & Core Concept
# ============================================================================
ws = wb.active
ws.title = "1. Overview"

row = style_header(ws, 1, "CONCEPT SCHEMA v5 - TECHNOLOGICAL SOVEREIGNTY", dimension_colors["Overview"])
row += 1

row = add_section_header(ws, row, "WHAT IS THIS SCHEMA?")
row = add_explanation(ws, row,
    "This schema analyzes concepts through 9 philosophical dimensions. Each dimension reveals different aspects of how a concept works, "
    "what it presupposes, what it commits users to, and how it evolves over time. Every value is individually addressable and can be "
    "tested against evidence. The schema enables precise identification of which specific aspect of a concept is challenged by evidence.")
row += 1

row = add_section_header(ws, row, "THE 9 DIMENSIONS")
dimensions_explained = [
    ("Quinean", "How the concept connects to other beliefs in a 'web'. Based on Quine's confirmation holism - beliefs face evidence as a corporate body."),
    ("Sellarsian", "What the concept treats as 'given' (self-evident) that actually isn't. Exposes hidden assumptions. Based on Sellars' critique of the 'myth of the given'."),
    ("Brandomian", "What commitments using this concept involves. What you're obligated to accept if you use the concept. Based on Brandom's inferentialism."),
    ("Deleuzian", "The internal structure - what components make up the concept and how they hold together. Based on Deleuze's concept theory."),
    ("Bachelardian", "How the concept blocks understanding. What epistemological obstacles it creates. Based on Bachelard's philosophy of science."),
    ("Canguilhem", "The life history of the concept - its evolution, health, and vitality. Based on Canguilhem's historical epistemology."),
    ("Hacking", "How the concept creates what it describes through 'looping effects'. Based on Hacking's dynamic nominalism."),
    ("Blumenberg", "The metaphors that structure the concept, especially 'absolute' metaphors that can't be translated into literal terms."),
    ("Carey", "How the concept was 'bootstrapped' from simpler concepts. What cognitive constraints affect its learnability."),
]
for dim, expl in dimensions_explained:
    ws.cell(row=row, column=1, value=dim).font = Font(bold=True, size=10)
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row, column=2, value=expl).font = explain_font
    row += 1
row += 1

row = add_section_header(ws, row, "SOURCE TYPES - Where Values Come From")
row = add_explanation(ws, row,
    "Every value in this schema has a 'source_type' indicating its origin. This is critical for understanding how confident we are "
    "and what might challenge it.")
row += 1

sources_explained = [
    ("user_input", "Values entered directly by users during concept setup or manual editing. Confidence typically 1.0."),
    ("llm_analysis", "Values generated by LLM analysis of the concept definition and context. Confidence 0.6-0.9."),
    ("evidence_testing", "Values derived from testing the concept against evidence clusters (e.g., from essay-flow). Most valuable for refinement."),
    ("internal_compute", "Values computed from other values in the database (e.g., entrenchment score from inference counts)."),
    ("import", "Values imported from external sources like academic databases or ontologies."),
]
for src, expl in sources_explained:
    ws.cell(row=row, column=1, value=src).font = Font(bold=True, size=10, color="E74C3C")
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row, column=2, value=expl).font = explain_font
    row += 1
row += 2

row = add_section_header(ws, row, "CORE CONCEPT RECORD")
row = add_explanation(ws, row,
    "The master record for the concept. All dimension tables reference this via concept_id. This contains the basic identity and "
    "computed health metrics.")
row += 1

row = add_field_explanations(ws, row, [
    ("name", "The concept's name as it appears in discourse"),
    ("definition", "The core definition - what the concept means"),
    ("domain", "Primary field/discipline (e.g., International Relations, Philosophy)"),
    ("first_appearance", "When the concept first emerged in roughly its current form"),
    ("current_usage_frequency", "How often the concept is used now (low/moderate/high)"),
    ("centrality", "How central to the belief web: core (revision ripples widely), intermediate, or peripheral"),
    ("entrenchment_score", "0-1 score of how resistant to revision (computed from inference count, usage, institutional embedding)"),
    ("health_status", "Vital status: healthy, strained (tensions emerging), declining, or transforming"),
    ("hierarchy_level", "Carey's bootstrapping level: 1=primitive, 2=simple composite, 3=complex composite, 4=theoretical"),
])

core_headers = ["Field", "Value", "Source Type", "Source Reference", "Confidence"]
core_data = [
    ("name", "Technological Sovereignty", "user_input", "initial_import", 1.0),
    ("definition", "The capacity of a state to autonomously develop, control, and deploy critical technologies without dependence on external actors", "user_input", "initial_import", 1.0),
    ("domain", "International Relations / Technology Policy", "user_input", "initial_import", 1.0),
    ("first_appearance", "~2015 (EU Digital Single Market debates)", "llm_analysis", "historical_research", 0.75),
    ("current_usage_frequency", "high - accelerating since 2018 trade war", "evidence_testing", "cluster_42", 0.90),
    ("centrality", "core", "llm_analysis", "web_analysis", 0.85),
    ("entrenchment_score", "0.78", "internal_compute", "entrenchment_calc_v2", 0.80),
    ("health_status", "strained", "evidence_testing", "cluster_87", 0.82),
    ("hierarchy_level", "3 (complex composite)", "llm_analysis", "carey_analysis", 0.80),
]
row = add_table(ws, row, core_headers, core_data, "concepts (Master Record)")

set_column_widths(ws, {'A': 25, 'B': 50, 'C': 15, 'D': 20, 'E': 12, 'F': 15, 'G': 15, 'H': 15})

# ============================================================================
# SHEET 2: Quinean Dimension
# ============================================================================
ws = wb.create_sheet("2. Quinean")
row = style_header(ws, 1, "QUINEAN DIMENSION - Web of Belief Analysis", dimension_colors["Quinean"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on W.V.O. Quine's 'Two Dogmas of Empiricism' (1951). Key ideas: (1) CONFIRMATION HOLISM - beliefs don't face evidence "
    "individually but as a corporate body; (2) NO ANALYTIC/SYNTHETIC DISTINCTION - no beliefs are immune to revision; "
    "(3) WEB OF BELIEF - concepts are connected by inferential relations forming a web; (4) ONTOLOGICAL RELATIVITY - what exists "
    "is relative to a conceptual scheme; (5) PRAGMATIC REVISABILITY - we revise beliefs to maximize coherence and utility.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: How the concept is positioned in the web of belief - its inferential connections to other concepts, "
    "tensions with other parts of the web, and the consequences of revising it.")
row += 1

# Table 1: concept_inferences
row = add_section_header(ws, row, "TABLE: concept_inferences")
row = add_explanation(ws, row,
    "Maps the INFERENTIAL CONNECTIONS from/to this concept. If you accept this concept, what else are you committed to? "
    "What enables or defeats it? These are the 'edges' in the web of belief.")
row = add_field_explanations(ws, row, [
    ("from_concept", "The source concept in the inference (can be this concept or another)"),
    ("to_concept", "The target concept - what follows from the source"),
    ("inference_type", "COMMITTIVE (A commits you to B), PERMISSIVE (A entitles you to B), INCOMPATIBILITY (A precludes B), ENABLING (A is precondition for B)"),
    ("strength", "How strong the inferential connection is: strong, moderate, weak"),
])
row = add_table(ws, row,
    ["id", "from_concept", "to_concept", "inference_type", "strength", "source_type", "source_ref", "confidence"],
    [
        (1, "tech_sovereignty", "indigenous_capability", "committive", "strong", "llm_analysis", "quine_v1", 0.85),
        (2, "tech_sovereignty", "supply_chain_control", "committive", "strong", "evidence_testing", "cluster_23", 0.90),
        (3, "tech_sovereignty", "right_to_exclude", "permissive", "moderate", "llm_analysis", "quine_v1", 0.75),
        (4, "tech_sovereignty", "global_integration", "incompatibility", "moderate", "evidence_testing", "cluster_45", 0.80),
        (5, "investment_capacity", "tech_sovereignty", "enabling", "weak", "evidence_testing", "cluster_67", 0.70),
    ],
    "concept_inferences"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 says 'If you claim tech sovereignty, you're COMMITTED to indigenous capability' (strong connection, "
    "identified by LLM). Row 4 says 'tech sovereignty is INCOMPATIBLE with global integration' (discovered through evidence testing).")
row += 1

# Table 2: concept_web_tensions
row = add_section_header(ws, row, "TABLE: concept_web_tensions")
row = add_explanation(ws, row,
    "Captures TENSIONS and CONTRADICTIONS between concepts in the web. These are stress points where the web doesn't hang together well. "
    "Critical for identifying where concepts may need revision.")
row = add_field_explanations(ws, row, [
    ("concept_a / concept_b", "The two concepts in tension"),
    ("tension_type", "PRACTICAL_CONTRADICTION (can't pursue both), TRADEOFF (sacrifice one for other), STRUCTURAL (deep incompatibility)"),
    ("description", "Explanation of what the tension is and how it manifests"),
])
row = add_table(ws, row,
    ["id", "concept_a", "concept_b", "tension_type", "description", "source_type", "source_ref", "confidence"],
    [
        (1, "tech_sovereignty", "global_supply_chains", "practical_contradiction", "Pursuing sovereignty while depending on global semiconductor supply chains", "evidence_testing", "cluster_89", 0.92),
        (2, "tech_sovereignty", "innovation_speed", "tradeoff", "Autarky slows innovation vs. openness accelerates but increases dependency", "llm_analysis", "quine_v1", 0.85),
        (3, "tech_sovereignty", "market_size", "structural", "Small states cannot achieve sovereignty due to market scale requirements", "evidence_testing", "cluster_12", 0.88),
    ],
    "concept_web_tensions"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 is a PRACTICAL CONTRADICTION - you can't actually pursue sovereignty while relying on global chip supply chains. "
    "This was discovered through evidence testing (cluster_89) with high confidence (0.92).")
row += 1

# Table 3: concept_web_position
row = add_section_header(ws, row, "TABLE: concept_web_position")
row = add_explanation(ws, row,
    "Identifies the STRUCTURAL POSITION of this concept in the belief web. Is it a hub (many connections), a bridge (connects separate clusters), "
    "a leaf (peripheral), or a foundation (rarely questioned, supports others)?")
row = add_field_explanations(ws, row, [
    ("position_type", "HUB (central, many connections), BRIDGE (connects separate domains), LEAF (peripheral), FOUNDATION (supports others)"),
    ("connected_concepts", "List of concepts this one connects to"),
    ("bridging_function", "If a bridge, what domains does it connect?"),
])
row = add_table(ws, row,
    ["id", "position_type", "connected_concepts", "bridging_function", "source_type", "source_ref", "confidence"],
    [
        (1, "hub", "national_security, industrial_policy, digital_rights, trade_policy", "Central node connecting security and economic policy domains", "llm_analysis", "quine_v2", 0.85),
        (2, "bridge", "traditional_sovereignty, digital_economy", "Connects Westphalian state concepts to networked economy concepts", "llm_analysis", "quine_v2", 0.80),
    ],
    "concept_web_position"
)
row += 1

# Table 4: concept_ontological_dependence
row = add_section_header(ws, row, "TABLE: concept_ontological_dependence")
row = add_explanation(ws, row,
    "Captures what ENTITIES or CONCEPTS this concept PRESUPPOSES for its existence or meaning. From Quine's ontological relativity - "
    "what must exist for this concept to make sense?")
row = add_field_explanations(ws, row, [
    ("depends_on", "The entity or concept this one depends on"),
    ("dependence_type", "EXISTENTIAL (presupposes existence), DEFINITIONAL (definition requires it), RELATIONAL (only meaningful in relation to)"),
    ("explanation", "How the dependence works"),
])
row = add_table(ws, row,
    ["id", "depends_on", "dependence_type", "explanation", "source_type", "source_ref", "confidence"],
    [
        (1, "the_state", "existential", "Concept presupposes existence of state actors capable of sovereign action", "llm_analysis", "quine_ontology", 0.92),
        (2, "critical_technologies", "definitional", "Requires prior identification of which technologies are 'critical'", "llm_analysis", "quine_ontology", 0.88),
        (3, "adversarial_other", "relational", "Sovereignty only meaningful in relation to potential adversaries/dependencies", "evidence_testing", "cluster_34", 0.85),
    ],
    "concept_ontological_dependence"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 shows EXISTENTIAL dependence - 'tech sovereignty' makes no sense without states. If you don't believe in "
    "states as actors, the concept collapses. This is a deep presupposition.")
row += 1

# Table 5: concept_revision_ramifications
row = add_section_header(ws, row, "TABLE: concept_revision_ramifications")
row = add_explanation(ws, row,
    "Models what happens IF this concept is revised or abandoned. Quine's pragmatic revisability - we can revise any belief, but some "
    "revisions cost more than others. What's the ripple effect?")
row = add_field_explanations(ws, row, [
    ("if_revised", "Type of revision: abandoned, weakened, split into domain-specific variants, etc."),
    ("affected_concepts", "What other concepts would need to change"),
    ("revision_cost", "How costly: very_high, high, moderate, low"),
    ("pragmatic_consideration", "Practical real-world implications of this revision"),
])
row = add_table(ws, row,
    ["id", "if_revised", "affected_concepts", "revision_cost", "pragmatic_consideration", "source_type", "source_ref", "confidence"],
    [
        (1, "abandoned", "industrial_policy, national_security_framing, EU_strategic_autonomy", "very_high", "Would require reworking entire policy frameworks in EU, China, India", "llm_analysis", "quine_pragmatic", 0.90),
        (2, "weakened_to_resilience", "supply_chain_policy, trade_agreements", "moderate", "More achievable; shifts from control to redundancy", "evidence_testing", "cluster_78", 0.82),
        (3, "split_into_domains", "semiconductor_sovereignty, data_sovereignty, AI_sovereignty", "low", "Already happening; may be natural evolution", "evidence_testing", "cluster_91", 0.85),
    ],
    "concept_revision_ramifications"
)

set_column_widths(ws)

# ============================================================================
# SHEET 3: Sellarsian Dimension
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
row = style_header(ws, 1, "SELLARSIAN DIMENSION - Myth of the Given", dimension_colors["Sellarsian"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Wilfrid Sellars' 'Empiricism and the Philosophy of Mind' (1956). Key ideas: (1) MYTH OF THE GIVEN - nothing is simply "
    "'given' to consciousness without conceptual mediation; what seems self-evident is actually inferred; (2) SPACE OF REASONS - "
    "concepts get meaning from their role in justification/inference, not from pointing at things; (3) FUNCTIONAL ROLE SEMANTICS - "
    "meaning = inferential role in the network; (4) MANIFEST VS SCIENTIFIC IMAGE - tension between everyday and scientific pictures.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: What the concept treats as 'obviously true' that actually isn't, what hidden assumptions it carries, "
    "and how it functions in justificatory discourse. Exposes the foundations that aren't really foundational.")
row += 1

# Table 1: concept_givenness
row = add_section_header(ws, row, "TABLE: concept_givenness")
row = add_explanation(ws, row,
    "Identifies what the concept TREATS AS GIVEN (self-evident, foundational, obvious) that is ACTUALLY INFERRED from other premises. "
    "This is the core Sellarsian move - exposing false foundations.")
row = add_field_explanations(ws, row, [
    ("given_claim", "What's treated as obviously true, beyond question"),
    ("givenness_level", "How strongly it's treated as given: high (never questioned), moderate (sometimes), low (occasionally)"),
    ("actual_basis", "What it's ACTUALLY inferred from - the hidden premises"),
])
row = add_table(ws, row,
    ["id", "given_claim", "givenness_level", "actual_basis", "source_type", "source_ref", "confidence"],
    [
        (1, "States can achieve technological autonomy", "high", "Inferred from contested premises about state capacity and market dynamics", "llm_analysis", "sellars_v1", 0.85),
        (2, "Critical technologies are identifiable", "moderate", "Requires complex analysis; what's 'critical' shifts with geopolitics", "evidence_testing", "cluster_56", 0.80),
        (3, "Sovereignty is binary (have it or don't)", "high", "Actually a spectrum; binary framing simplifies policy discourse", "llm_analysis", "sellars_v1", 0.88),
    ],
    "concept_givenness"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 exposes that 'states CAN achieve tech autonomy' is treated as obvious, but it's actually based on debatable "
    "assumptions about what states can do. This is a FALSE GIVEN - it looks foundational but isn't.")
row += 1

# Table 2: concept_hidden_commitments
row = add_section_header(ws, row, "TABLE: concept_hidden_commitments")
row = add_explanation(ws, row,
    "Exposes ASSUMPTIONS BAKED INTO the concept that users may not realize they're making. These are the 'fine print' of the concept - "
    "accept the concept, accept these commitments (whether you know it or not).")
row = add_field_explanations(ws, row, [
    ("commitment", "The hidden assumption"),
    ("how_hidden", "Why it's not obvious - how the concept conceals it"),
    ("implications", "What following this assumption leads to - the downstream effects"),
])
row = add_table(ws, row,
    ["id", "commitment", "how_hidden", "implications", "source_type", "source_ref", "confidence"],
    [
        (1, "Technology development is linear and predictable", "Embedded in 'roadmap' rhetoric", "Blinds to disruptive innovation from unexpected sources", "llm_analysis", "sellars_v1", 0.82),
        (2, "State action is inherently legitimate", "Naturalized in sovereignty discourse", "Forecloses questions about whose interests are served", "llm_analysis", "sellars_v1", 0.85),
        (3, "Autarky is economically viable", "Assumed in 'build domestic' framings", "Ignores comparative advantage and scale economics", "evidence_testing", "cluster_44", 0.90),
        (4, "Security and economics are separable", "Binary framing of 'strategic' vs 'commercial'", "Misses how deeply intertwined they are", "llm_analysis", "sellars_v1", 0.78),
    ],
    "concept_hidden_commitments"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 3 shows that using 'tech sovereignty' commits you to believing autarky is viable - but evidence testing "
    "(cluster_44, confidence 0.90) suggests this hidden commitment is false. The concept smuggles in a bad assumption.")
row += 1

# Table 3: concept_functional_role
row = add_section_header(ws, row, "TABLE: concept_functional_role")
row = add_explanation(ws, row,
    "Specifies WHAT ROLE the concept plays in different discursive contexts. Per Sellars' functional role semantics - meaning IS "
    "the role in inference and justification. Same word can play different roles in different contexts.")
row = add_field_explanations(ws, row, [
    ("role_type", "What function: ENABLING_PREMISE, FRAMING_DEVICE, MOBILIZATION_RHETORIC, LEGITIMATION_DEVICE, etc."),
    ("in_discourse", "Which context: policy debates, media coverage, political speeches, academic analysis"),
    ("inferential_connections", "What inferences or conclusions this role enables"),
])
row = add_table(ws, row,
    ["id", "role_type", "in_discourse", "inferential_connections", "source_type", "source_ref", "confidence"],
    [
        (1, "enabling_premise", "Policy debates", "Enables conclusions about state intervention legitimacy", "llm_analysis", "sellars_functional", 0.85),
        (2, "framing_device", "Media coverage", "Structures how tech competition is understood", "evidence_testing", "cluster_78", 0.88),
        (3, "mobilization_rhetoric", "Political speeches", "Rallies support for spending/restrictions", "evidence_testing", "cluster_34", 0.82),
    ],
    "concept_functional_role"
)
row += 1

# Table 4: concept_image_tension
row = add_section_header(ws, row, "TABLE: concept_image_tension")
row = add_explanation(ws, row,
    "Captures tension between MANIFEST IMAGE (everyday understanding) and SCIENTIFIC IMAGE (expert/technical understanding). "
    "Sellars argued these often diverge significantly - common sense vs scientific reality.")
row = add_field_explanations(ws, row, [
    ("manifest_image", "How ordinary people/politicians understand it - the everyday picture"),
    ("scientific_image", "How experts/analysts understand it - the technical picture"),
    ("tension_description", "How these two images conflict"),
])
row = add_table(ws, row,
    ["id", "manifest_image", "scientific_image", "tension_description", "source_type", "source_ref", "confidence"],
    [
        (1, "Nation controls its technological destiny", "Complex global interdependencies make full autonomy impossible", "Manifest image persists despite scientific understanding", "llm_analysis", "sellars_image", 0.90),
        (2, "Clear friend/enemy technology distinctions", "Technologies flow across borders regardless of intent", "Policy pretends borders are more solid than they are", "evidence_testing", "cluster_89", 0.85),
        (3, "State as unified rational actor", "Multiple agencies with conflicting interests", "Sovereignty discourse assumes coherence that doesn't exist", "llm_analysis", "sellars_image", 0.82),
    ],
    "concept_image_tension"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 shows a major tension - ordinary understanding says nations control their destiny, but scientific "
    "understanding says interdependencies make this impossible. The manifest image 'persists despite' the scientific one.")

set_column_widths(ws)

# ============================================================================
# SHEET 4: Brandomian Dimension
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
row = style_header(ws, 1, "BRANDOMIAN DIMENSION - Deontic Scorekeeping", dimension_colors["Brandomian"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Robert Brandom's 'Making It Explicit' (1994). Key ideas: (1) INFERENTIALISM - meaning is constituted by inferential role, "
    "not by pointing at things; (2) DEONTIC SCOREKEEPING - using concepts involves tracking COMMITMENTS (what you're obligated to) and "
    "ENTITLEMENTS (what you're permitted to claim); (3) GAME OF GIVING AND ASKING FOR REASONS - discourse is challenging and justifying; "
    "(4) DE DICTO VS DE RE - difference between what someone says (their words) and what we translate it to (our interpretation).")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: What you commit yourself to when you use this concept, what challenges the concept faces, and how "
    "different speakers may mean different things by it. The normative 'score' of using the concept.")
row += 1

# Table 1: concept_commitments
row = add_section_header(ws, row, "TABLE: concept_commitments")
row = add_explanation(ws, row,
    "Tracks the DEONTIC STATUS of commitments associated with the concept. When you use this concept, what are you committed to? "
    "This is like tracking the 'score' in a game of reasons.")
row = add_field_explanations(ws, row, [
    ("commitment", "The claim or obligation"),
    ("deontic_status", "ACKNOWLEDGED (explicitly accepted), UNDERTAKEN (actions commit you, whether you admit it), ATTRIBUTED (others say you're committed)"),
    ("undertaken_by", "Who undertakes/acknowledges this commitment"),
    ("attributed_by", "Who attributes this commitment to others"),
])
row = add_table(ws, row,
    ["id", "commitment", "deontic_status", "undertaken_by", "attributed_by", "source_type", "source_ref", "confidence"],
    [
        (1, "Investment leads to capability", "acknowledged", "EU, China, India", "Policy analysts", "evidence_testing", "cluster_23", 0.85),
        (2, "Capability leads to autonomy", "undertaken", "EU (Chips Act)", "Critics dispute", "evidence_testing", "cluster_45", 0.78),
        (3, "Autonomy is achievable for major powers", "attributed", "-", "Proponents attribute to states", "llm_analysis", "brandom_v1", 0.72),
        (4, "Right to exclude foreign tech", "acknowledged", "US (entity list), China", "Trading partners contest", "evidence_testing", "cluster_67", 0.88),
    ],
    "concept_commitments"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 shows an ACKNOWLEDGED commitment - EU, China, India explicitly accept 'investment leads to capability'. "
    "Row 2 shows an UNDERTAKEN commitment - the EU's Chips Act commits them to 'capability leads to autonomy' whether they say so or not. "
    "Row 3 shows an ATTRIBUTED commitment - proponents attribute this to states, but states don't explicitly acknowledge it.")
row += 1

# Table 2: concept_challenges
row = add_section_header(ws, row, "TABLE: concept_challenges")
row = add_explanation(ws, row,
    "Tracks CHALLENGES to the concept's commitments and whether they've been ANSWERED. In Brandom's 'game of giving and asking for "
    "reasons,' claims can be challenged and must be justified. Unanswered challenges are serious problems.")
row = add_field_explanations(ws, row, [
    ("challenge", "The challenging evidence or argument"),
    ("to_commitment", "Which commitment is being challenged"),
    ("challenger", "Source of the challenge"),
    ("response_status", "UNANSWERED (no response), PARTIALLY_ADDRESSED (some response), ADDRESSED (answered), ACKNOWLEDGED (accepted as limitation)"),
])
row = add_table(ws, row,
    ["id", "challenge", "to_commitment", "challenger", "response_status", "source_type", "source_ref", "confidence"],
    [
        (1, "Gulf states have $2T but can't do Series B", "Investment leads to capability", "Empirical evidence", "unanswered", "evidence_testing", "cluster_67", 0.92),
        (2, "EU Chips Act creates new dependencies", "Capability leads to autonomy", "Supply chain analysts", "partially_addressed", "evidence_testing", "cluster_45", 0.85),
        (3, "China still needs ASML machines", "Major powers can achieve autonomy", "Industry observers", "acknowledged", "evidence_testing", "cluster_89", 0.95),
        (4, "Innovation requires openness", "Autarky is viable", "Economists", "contested", "llm_analysis", "brandom_v1", 0.80),
    ],
    "concept_challenges"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 is CRITICAL - Gulf states have $2 trillion but can't produce Series B startups. This DIRECTLY CHALLENGES "
    "the commitment that 'investment leads to capability'. The challenge is UNANSWERED with 0.92 confidence. This is a major problem for the concept.")
row += 1

# Table 3: concept_perspectival_content
row = add_section_header(ws, row, "TABLE: concept_perspectival_content")
row = add_explanation(ws, row,
    "Captures how DIFFERENT SPEAKERS understand the concept differently. Brandom's DE DICTO (what they say) vs DE RE (what we think "
    "they mean). Translation between perspectives can lose important content.")
row = add_field_explanations(ws, row, [
    ("speaker_perspective", "Whose perspective"),
    ("de_dicto_content", "What they say, in their own terms"),
    ("de_re_translation", "Our translation of what they mean"),
    ("translation_loss", "What's lost when we translate - important nuances that disappear"),
])
row = add_table(ws, row,
    ["id", "speaker_perspective", "de_dicto_content", "de_re_translation", "translation_loss", "source_type", "source_ref", "confidence"],
    [
        (1, "Chinese government", "Great rejuvenation through tech self-reliance", "Catch-up industrial policy with nationalist framing", "Loses civilizational dimension", "llm_analysis", "brandom_persp", 0.80),
        (2, "EU Commission", "Strategic autonomy in critical technologies", "Reduce single-point-of-failure dependencies", "Loses sovereignty aspiration", "evidence_testing", "cluster_12", 0.85),
        (3, "Gulf states", "Technological sovereignty through investment", "Technology consumption + influence via capital", "Loses self-reliance claim entirely", "evidence_testing", "cluster_78", 0.90),
        (4, "India", "Atmanirbhar Bharat (self-reliant India)", "Import substitution + selective decoupling", "Loses cultural-historical resonance", "llm_analysis", "brandom_persp", 0.82),
    ],
    "concept_perspectival_content"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 3 is striking - when Gulf states say 'tech sovereignty,' our translation (de re) is 'technology consumption + "
    "capital influence.' The translation LOSES THE SELF-RELIANCE CLAIM ENTIRELY. They're using the same words but meaning something fundamentally different.")

set_column_widths(ws)

# ============================================================================
# SHEET 5: Deleuzian Dimension
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
row = style_header(ws, 1, "DELEUZIAN DIMENSION - Concept as Multiplicity", dimension_colors["Deleuzian"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Deleuze & Guattari's 'What is Philosophy?' (1991). Key ideas: (1) CONCEPTS AS MULTIPLICITIES - concepts are not simple "
    "unities but composed of distinct COMPONENTS; (2) ZONES OF INDISCERNIBILITY - where components blur into each other, creating "
    "productive tensions; (3) ENDOCONSISTENCY/EXOCONSISTENCY - internal coherence vs external relations; (4) PLANE OF IMMANENCE - "
    "the pre-philosophical 'image of thought' that makes concepts thinkable; (5) CONCEPTUAL PERSONAE - the 'characters' who deploy concepts.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: The internal structure of the concept - what components make it up, how they hold together (or don't), "
    "what plane of thought it operates on, and who activates it.")
row += 1

# Table 1: concept_components
row = add_section_header(ws, row, "TABLE: concept_components")
row = add_explanation(ws, row,
    "Identifies the DISTINCT COMPONENTS that make up the concept's multiplicity. A concept isn't a simple unity - it's a composition "
    "of heterogeneous elements that somehow hold together.")
row = add_field_explanations(ws, row, [
    ("component", "Name of the component"),
    ("description", "What this component contributes to the concept"),
    ("ordering", "INTENSIVE (qualitative, non-divisible, like temperature) vs EXTENSIVE (quantitative, additive, like volume)"),
])
row = add_table(ws, row,
    ["id", "component", "description", "ordering", "source_type", "source_ref", "confidence"],
    [
        (1, "control", "Capacity to direct technological development and deployment", "intensive", "llm_analysis", "deleuze_v1", 0.88),
        (2, "autonomy", "Independence from external actors in tech decisions", "intensive", "llm_analysis", "deleuze_v1", 0.90),
        (3, "territory", "Bounded space within which sovereignty applies", "extensive", "llm_analysis", "deleuze_v1", 0.85),
        (4, "temporality", "Urgency - limited window before lock-in", "intensive", "evidence_testing", "cluster_34", 0.78),
        (5, "criticality", "Distinction between essential and non-essential tech", "intensive", "llm_analysis", "deleuze_v1", 0.82),
        (6, "capacity", "Technical and industrial capability to produce", "extensive", "evidence_testing", "cluster_56", 0.88),
    ],
    "concept_components"
)
row = add_explanation(ws, row,
    "READING THE DATA: Tech sovereignty is composed of 6 components. 'Control' and 'autonomy' are INTENSIVE (qualitative) - you can't "
    "add them up. 'Territory' and 'capacity' are EXTENSIVE (quantitative) - they can be measured and aggregated. The concept holds "
    "these heterogeneous elements together.")
row += 1

# Table 2: concept_zones_of_indiscernibility
row = add_section_header(ws, row, "TABLE: concept_zones_of_indiscernibility")
row = add_explanation(ws, row,
    "Maps WHERE COMPONENTS BLUR INTO EACH OTHER, creating zones of productive tension. These are the 'interesting' parts of the concept "
    "where things get complicated and generative.")
row = add_field_explanations(ws, row, [
    ("component_a / component_b", "The two components that form the zone"),
    ("zone_character", "OVERLAPPING (share territory), CONTESTED (compete), PARADOXICAL (require and undermine each other), TEMPORAL (shifts over time)"),
    ("productive_tension", "What the zone produces - the creative/problematic output"),
])
row = add_table(ws, row,
    ["id", "component_a", "component_b", "zone_character", "productive_tension", "source_type", "source_ref", "confidence"],
    [
        (1, "control", "autonomy", "overlapping", "Control can undermine autonomy (surveillance state) or enable it", "llm_analysis", "deleuze_zones", 0.85),
        (2, "territory", "network", "contested", "Digital tech de-territorializes; sovereignty re-territorializes", "llm_analysis", "deleuze_zones", 0.90),
        (3, "urgency", "capability", "temporal", "Window closing before capacity can be built", "evidence_testing", "cluster_78", 0.88),
        (4, "autonomy", "innovation", "paradoxical", "Autonomy may require openness that undermines it", "evidence_testing", "cluster_45", 0.82),
    ],
    "concept_zones_of_indiscernibility"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 4 shows a PARADOXICAL zone - autonomy and innovation require and undermine each other. To innovate, you need "
    "openness. But openness undermines autonomy. The concept lives in this paradox.")
row += 1

# Table 3: concept_plane_of_immanence
row = add_section_header(ws, row, "TABLE: concept_plane_of_immanence")
row = add_explanation(ws, row,
    "Identifies the PRE-PHILOSOPHICAL PLANE on which the concept operates. This is the 'image of thought' - the background assumptions "
    "that make thinking this concept possible in the first place. Different planes make different concepts thinkable.")
row = add_field_explanations(ws, row, [
    ("plane", "Name of the plane"),
    ("characteristics", "What defines this plane - its key features"),
    ("presuppositions", "What the plane takes for granted - the background assumptions"),
])
row = add_table(ws, row,
    ["id", "plane", "characteristics", "presuppositions", "source_type", "source_ref", "confidence"],
    [
        (1, "techno-nationalist", "Competition between state-civilization blocs; technology as power; zero-sum framing", "States are primary actors; technology is controllable; competition is inevitable", "llm_analysis", "deleuze_plane", 0.88),
        (2, "residual_westphalian", "Territorial sovereignty; non-interference; state monopoly on legitimate force", "Clear borders; fungible power; stable state system", "llm_analysis", "deleuze_plane", 0.75),
    ],
    "concept_plane_of_immanence"
)
row = add_explanation(ws, row,
    "READING THE DATA: The concept operates mainly on a 'techno-nationalist' plane. This plane PRESUPPOSES that states are primary actors, "
    "tech is controllable, and competition is inevitable. If you don't accept these presuppositions, the concept doesn't make sense.")
row += 1

# Table 4: concept_personae
row = add_section_header(ws, row, "TABLE: concept_personae")
row = add_explanation(ws, row,
    "Identifies the CONCEPTUAL PERSONAE that activate and deploy the concept. These are the 'characters' - not real people, but types "
    "of thinkers/speakers who give voice to the concept.")
row = add_field_explanations(ws, row, [
    ("persona", "Name of the conceptual persona"),
    ("role", "How this persona activates/deploys the concept"),
    ("voice_characteristics", "How this persona speaks - their rhetorical style"),
])
row = add_table(ws, row,
    ["id", "persona", "role", "voice_characteristics", "source_type", "source_ref", "confidence"],
    [
        (1, "The Technocrat", "Activates concept through expertise claims", "Neutral, technical, apolitical framing", "llm_analysis", "deleuze_personae", 0.85),
        (2, "The Securitizer", "Activates through threat construction", "Urgent, existential risk framing", "evidence_testing", "cluster_23", 0.88),
        (3, "The Industrial Champion", "Activates through competitiveness discourse", "Economic nationalism, jobs rhetoric", "evidence_testing", "cluster_67", 0.82),
        (4, "The Digital Sovereign", "New persona emerging", "Platform governance, data rights framing", "evidence_testing", "cluster_91", 0.75),
    ],
    "concept_personae"
)

set_column_widths(ws)

# ============================================================================
# SHEET 6: Bachelardian Dimension
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
row = style_header(ws, 1, "BACHELARDIAN DIMENSION - Epistemological Obstacles", dimension_colors["Bachelardian"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Gaston Bachelard's 'The Formation of the Scientific Mind' (1938). Key ideas: (1) EPISTEMOLOGICAL OBSTACLES - concepts can "
    "BLOCK rather than enable understanding; (2) OBSTACLE TAXONOMY - verbal, substantialist, animist, first experience, quantitative, "
    "unitary pragmatism; (3) THREE STAGES - pre-scientific, proto-scientific, scientific mind; (4) PSYCHOANALYSIS OF KNOWLEDGE - "
    "concepts serve UNCONSCIOUS NEEDS; (5) REGIONAL RATIONALISM - different domains require different rationalities.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: How the concept BLOCKS understanding, what stage of thinking it represents, and what psychological needs "
    "it serves. This is about the concept as an OBSTACLE, not just a tool.")
row += 1

# Table 1: concept_obstacles
row = add_section_header(ws, row, "TABLE: concept_obstacles")
row = add_explanation(ws, row,
    "Identifies how the concept functions as an EPISTEMOLOGICAL OBSTACLE - blocking understanding rather than enabling it. "
    "Bachelard's key insight: what we think with can prevent us from thinking further.")
row = add_field_explanations(ws, row, [
    ("obstacle_type", "VERBAL (word carries wrong associations), SUBSTANTIALIST (treats relations as substances), ANIMIST (attributes agency to non-agents), FIRST_EXPERIENCE (over-relies on intuition), QUANTITATIVE (more=better fallacy), UNITARY_PRAGMATISM (one solution for all)"),
    ("description", "How the obstacle blocks understanding"),
    ("entrenchment", "How deeply embedded: very_high, high, moderate, low"),
])
row = add_table(ws, row,
    ["id", "obstacle_type", "description", "entrenchment", "source_type", "source_ref", "confidence"],
    [
        (1, "verbal", "Word 'sovereignty' carries Westphalian baggage inappropriate for networked tech", "high", "llm_analysis", "bachelard_v1", 0.88),
        (2, "substantialist", "Treating tech capability as substance that can be 'possessed' vs relational capacity", "very_high", "llm_analysis", "bachelard_v1", 0.90),
        (3, "animist", "Treating 'the market' or 'technology' as agents with intentions", "moderate", "evidence_testing", "cluster_34", 0.78),
        (4, "first_experience", "Intuition that borders = control, despite networked reality", "high", "llm_analysis", "bachelard_v1", 0.85),
        (5, "quantitative", "Belief that enough investment = sovereignty (linear scaling assumption)", "high", "evidence_testing", "cluster_67", 0.88),
    ],
    "concept_obstacles"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 2 shows a SUBSTANTIALIST obstacle (very_high entrenchment) - treating capability as a 'substance' you can "
    "possess, rather than a relational capacity that depends on networks. This is deeply embedded and hard to overcome.")
row += 1

# Table 2: concept_psychoanalytic_function
row = add_section_header(ws, row, "TABLE: concept_psychoanalytic_function")
row = add_explanation(ws, row,
    "Identifies UNCONSCIOUS NEEDS the concept serves. Bachelard argued knowledge serves psychological functions beyond truth-seeking. "
    "What emotional/psychological work does this concept do for those who use it?")
row = add_field_explanations(ws, row, [
    ("unconscious_need", "What psychological need"),
    ("how_concept_satisfies", "How the concept meets this need"),
    ("evidence", "Evidence that this function is operating"),
])
row = add_table(ws, row,
    ["id", "unconscious_need", "how_concept_satisfies", "evidence", "source_type", "source_ref", "confidence"],
    [
        (1, "control_fantasy", "Provides illusion of mastering uncontrollable technological change", "Emphasis on 'strategic control' despite evidence of limited control", "llm_analysis", "bachelard_psych", 0.82),
        (2, "legitimation_of_spending", "Justifies massive public investment without clear ROI accountability", "Chips Act debates avoid cost-benefit analysis", "evidence_testing", "cluster_78", 0.88),
        (3, "identity_maintenance", "Preserves national identity narratives against globalization anxiety", "Rhetoric of 'our' technology, 'our' capabilities", "llm_analysis", "bachelard_psych", 0.80),
        (4, "elite_reproduction", "Creates new roles for technocratic elite", "Proliferation of 'tech sovereignty' advisor positions", "evidence_testing", "cluster_91", 0.75),
    ],
    "concept_psychoanalytic_function"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 2 shows that the concept serves to LEGITIMATE SPENDING - it lets politicians spend massively without being "
    "held to cost-benefit standards. This is why Chips Act debates 'avoid cost-benefit analysis' - the concept does psychological work "
    "that makes such analysis seem inappropriate.")

set_column_widths(ws)

# ============================================================================
# SHEET 7: Canguilhem Dimension
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
row = style_header(ws, 1, "CANGUILHEM DIMENSION - Vital History of Concepts", dimension_colors["Canguilhem"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Georges Canguilhem's 'The Normal and the Pathological' (1943). Key ideas: (1) VITAL NORMATIVITY - concepts have 'health' "
    "and can be vital, strained, or dying; (2) CONCEPT FILIATION - concepts have lineages, parents, offspring; (3) MILIEU - concepts "
    "exist in environments that shape them; (4) NORMAL VS NORMATIVE - distinction between descriptive average and prescriptive ideal.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: The LIFE HISTORY of the concept - how it evolved, what lineage it comes from, how healthy it is now, "
    "and what environment it operates in. Treats concepts as living things that can thrive or decline.")
row += 1

# Table 1: concept_evolution
row = add_section_header(ws, row, "TABLE: concept_evolution")
row = add_explanation(ws, row,
    "Traces the HISTORICAL DEVELOPMENT of the concept through different periods. Concepts aren't static - they emerge, transform, "
    "go dormant, and revive.")
row = add_field_explanations(ws, row, [
    ("period", "Time period"),
    ("form", "What form the concept took during this period"),
    ("key_developments", "What happened - major events, transformations"),
])
row = add_table(ws, row,
    ["id", "period", "form", "key_developments", "source_type", "source_ref", "confidence"],
    [
        (1, "1648-1945", "proto_form", "Westphalian sovereignty; technology not yet strategic sector", "llm_analysis", "canguilhem_v1", 0.80),
        (2, "1945-1990", "cold_war_form", "Dual-use technology controls; COCOM; national champions", "llm_analysis", "canguilhem_v1", 0.85),
        (3, "1990-2010", "dormancy", "Globalization consensus; Washington Consensus; WTO; tech as neutral", "llm_analysis", "canguilhem_v1", 0.88),
        (4, "2010-2018", "revival", "Snowden; China rise; Made in China 2025; stirrings of concern", "evidence_testing", "cluster_12", 0.90),
        (5, "2018-present", "intensification", "Trade war; COVID supply chains; Chips Act; full activation", "evidence_testing", "cluster_34", 0.95),
    ],
    "concept_evolution"
)
row = add_explanation(ws, row,
    "READING THE DATA: The concept was DORMANT from 1990-2010 during globalization consensus. It REVIVED 2010-2018 and has been "
    "INTENSIFYING since 2018. This historical trajectory matters - the concept carries its history with it.")
row += 1

# Table 2: concept_vital_norms
row = add_section_header(ws, row, "TABLE: concept_vital_norms")
row = add_explanation(ws, row,
    "Identifies the concept's OWN VITAL NORMS - what it 'needs' to remain healthy. These are the concept's internal standards. "
    "Violating them means using the concept against itself.")
row = add_field_explanations(ws, row, [
    ("norm", "The vital norm"),
    ("function", "What this norm does for the concept's health"),
    ("violation_indicator", "How to tell when this norm is being violated"),
])
row = add_table(ws, row,
    ["id", "norm", "function", "violation_indicator", "source_type", "source_ref", "confidence"],
    [
        (1, "Autonomy preservation", "Core vital norm - concept exists to enable autonomous action", "Using 'sovereignty' to justify increased dependency", "llm_analysis", "canguilhem_norms", 0.90),
        (2, "Capacity building", "Concept should drive actual capability development", "Pure rhetoric without capability investment", "evidence_testing", "cluster_78", 0.85),
        (3, "Threat response", "Concept should respond to genuine external threats", "Invoking sovereignty for purely protectionist ends", "llm_analysis", "canguilhem_norms", 0.82),
    ],
    "concept_vital_norms"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 is crucial - the concept's core vital norm is 'autonomy preservation'. If you use 'tech sovereignty' "
    "to justify things that INCREASE dependency, you're violating the concept's own norms - using it against itself.")
row += 1

# Table 3: concept_vitality_indicators
row = add_section_header(ws, row, "TABLE: concept_vitality_indicators")
row = add_explanation(ws, row,
    "Tracks METRICS of the concept's health. Like vital signs for a living thing. Is the concept thriving or declining?")
row = add_field_explanations(ws, row, [
    ("indicator", "What's being measured"),
    ("current_value", "Current level"),
    ("trend", "Direction: increasing, stable, declining"),
    ("interpretation", "What this means for the concept's health"),
])
row = add_table(ws, row,
    ["id", "indicator", "current_value", "trend", "interpretation", "source_type", "source_ref", "confidence"],
    [
        (1, "usage_frequency", "very_high", "increasing", "Concept highly active in discourse", "evidence_testing", "cluster_12", 0.95),
        (2, "policy_adoption", "high", "increasing", "Being translated into actual policy (Chips Act, etc.)", "evidence_testing", "cluster_34", 0.92),
        (3, "internal_coherence", "moderate", "declining", "Contradictions becoming more visible", "evidence_testing", "cluster_89", 0.85),
        (4, "empirical_validation", "low", "stable", "Few success cases; many failures", "evidence_testing", "cluster_67", 0.88),
        (5, "overall_health", "strained", "concerning", "High activity but growing contradictions", "llm_analysis", "canguilhem_vitality", 0.85),
    ],
    "concept_vitality_indicators"
)
row = add_explanation(ws, row,
    "READING THE DATA: The concept is 'STRAINED' - usage is very high and increasing, but coherence is declining and empirical validation "
    "is low. It's very active but increasingly contradictory. This is a concept under stress.")

set_column_widths(ws)

# ============================================================================
# SHEET 8: Hacking Dimension
# ============================================================================
ws = wb.create_sheet("8. Hacking")
row = style_header(ws, 1, "HACKING DIMENSION - Dynamic Nominalism", dimension_colors["Hacking"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Ian Hacking's 'Historical Ontology' (2002). Key ideas: (1) DYNAMIC NOMINALISM - categories and the things they classify "
    "EVOLVE TOGETHER; (2) LOOPING EFFECTS - classifications change the behavior of classified things (especially people); "
    "(3) MAKING UP PEOPLE - categories create new kinds of people/things that didn't exist before; (4) STYLES OF REASONING - "
    "different epochs have different standards of truth; (5) SPACE OF POSSIBILITIES - what's thinkable is historically contingent.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: How the concept CREATES WHAT IT DESCRIBES through looping effects, what new kinds of things it brings "
    "into being, and what possibilities it opens or closes. The concept isn't just describing reality - it's making reality.")
row += 1

# Table 1: concept_looping_effects
row = add_section_header(ws, row, "TABLE: concept_looping_effects")
row = add_explanation(ws, row,
    "Captures how the concept CHANGES WHAT IT CLASSIFIES. When you call something 'tech sovereign,' that classification changes "
    "how the thing behaves. The concept loops back on reality.")
row = add_field_explanations(ws, row, [
    ("loop_description", "What the loop is"),
    ("mechanism", "How the loop works - the causal pathway"),
    ("observed_effects", "What changes have actually been observed"),
])
row = add_table(ws, row,
    ["id", "loop_description", "mechanism", "observed_effects", "source_type", "source_ref", "confidence"],
    [
        (1, "Classification creates behavior", "Labeling countries as 'tech sovereign aspirants' changes how they act", "Countries pursue performative sovereignty to match label", "evidence_testing", "cluster_34", 0.88),
        (2, "Metrics shape reality", "Sovereignty indices and rankings drive policy responses", "Investment decisions follow rankings, not fundamentals", "evidence_testing", "cluster_56", 0.82),
        (3, "Discourse enables spending", "Sovereignty framing enables budget allocations", "Budgets reshape industrial structure, reinforcing frame", "evidence_testing", "cluster_78", 0.90),
        (4, "Threat construction creates threats", "Framing tech as security threat triggers defensive responses", "Defensive responses are seen as threats, escalation spiral", "llm_analysis", "hacking_looping", 0.85),
    ],
    "concept_looping_effects"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 shows CLASSIFICATION CREATES BEHAVIOR - when you label a country a 'tech sovereign aspirant,' they start "
    "acting like one, pursuing 'performative sovereignty' to match the label. The concept is making the reality it describes.")
row += 1

# Table 2: concept_kinds_created
row = add_section_header(ws, row, "TABLE: concept_kinds_created")
row = add_explanation(ws, row,
    "Identifies NEW KINDS of things/people the concept brings into being. Hacking's 'making up people' - categories create new kinds "
    "that didn't exist before the category existed.")
row = add_field_explanations(ws, row, [
    ("kind_name", "Name of the new kind"),
    ("kind_type", "What type: political_entity, artifact_category, person_kind, etc."),
    ("how_created", "How the concept creates this kind"),
    ("stability", "How stable is this kind - will it persist?"),
])
row = add_table(ws, row,
    ["id", "kind_name", "kind_type", "how_created", "stability", "source_type", "source_ref", "confidence"],
    [
        (1, "Tech-sovereign state", "political_entity", "Classification by sovereignty pursuit rather than achievement", "unstable - criteria contested", "llm_analysis", "hacking_kinds", 0.85),
        (2, "Strategic technology", "artifact_category", "Political designation, not inherent property", "highly unstable - shifts with politics", "evidence_testing", "cluster_23", 0.90),
        (3, "Tech sovereignty expert", "person_kind", "New professional identity created by discourse", "stabilizing - career paths forming", "evidence_testing", "cluster_91", 0.82),
        (4, "Trusted vs untrusted vendor", "entity_category", "Geopolitically defined, not technically", "unstable - changes with alliances", "evidence_testing", "cluster_67", 0.88),
    ],
    "concept_kinds_created"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 3 shows the concept has created a new PERSON KIND - 'tech sovereignty expert.' These people didn't exist "
    "before the discourse. Now there are career paths, job titles, consulting practices. The concept made them up.")
row += 1

# Table 3: concept_possibility_space
row = add_section_header(ws, row, "TABLE: concept_possibility_space")
row = add_explanation(ws, row,
    "Maps what the concept MAKES THINKABLE OR UNTHINKABLE. Concepts open some possibilities and close others. What can you think/do "
    "once you adopt this concept? What becomes unthinkable?")
row = add_field_explanations(ws, row, [
    ("possibility", "The possibility"),
    ("opened_or_closed", "OPENED (newly thinkable), CLOSING (becoming harder to think), CLOSED (now unthinkable), THREATENED (under pressure)"),
    ("mechanism", "How the concept affects this possibility"),
])
row = add_table(ws, row,
    ["id", "possibility", "opened_or_closed", "mechanism", "source_type", "source_ref", "confidence"],
    [
        (1, "Industrial policy legitimacy", "opened", "Sovereignty framing makes intervention acceptable in liberal economies", "evidence_testing", "cluster_12", 0.92),
        (2, "Technology export controls", "opened", "Sovereignty concerns legitimate broad restrictions", "evidence_testing", "cluster_45", 0.90),
        (3, "International tech cooperation", "closing", "Sovereignty framing casts cooperation as dependency/risk", "evidence_testing", "cluster_78", 0.85),
        (4, "Open source/open standards", "threatened", "Sovereignty concerns push toward proprietary, controlled systems", "llm_analysis", "hacking_possibility", 0.82),
    ],
    "concept_possibility_space"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 3 shows that the concept is CLOSING the possibility of international tech cooperation - it frames cooperation "
    "as creating dependency/risk. What was once normal (collaborating on tech) is becoming unthinkable.")

set_column_widths(ws)

# ============================================================================
# SHEET 9: Blumenberg Dimension
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
row = style_header(ws, 1, "BLUMENBERG DIMENSION - Metaphorology", dimension_colors["Blumenberg"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Hans Blumenberg's 'Paradigms for a Metaphorology' (1960). Key ideas: (1) ABSOLUTE METAPHORS - some metaphors CANNOT be "
    "translated into literal concepts; they're foundational to thought itself; (2) NONCONCEPTUALITY (Unbegrifflichkeit) - some aspects "
    "of understanding RESIST conceptualization; (3) METAKINETICS - metaphors transform historically; (4) LIFEWORLD BACKGROUND - "
    "pre-theoretical understanding that enables concepts.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: The METAPHORICAL FOUNDATIONS of the concept - what root metaphors structure it, which are 'absolute' "
    "(untranslatable), how they've transformed, and what resists being put into concepts at all.")
row += 1

# Table 1: concept_metaphors
row = add_section_header(ws, row, "TABLE: concept_metaphors")
row = add_explanation(ws, row,
    "Identifies the ROOT METAPHORS structuring the concept. These aren't just decorative comparisons - they're foundational to how "
    "we think. Some are 'absolute' - you can't translate them away without losing the concept.")
row = add_field_explanations(ws, row, [
    ("metaphor", "The root metaphor"),
    ("type", "Category: spatial, biological, military, architectural, athletic, etc."),
    ("structuring_effect", "How this metaphor shapes understanding of the concept"),
    ("absoluteness", "How irreplaceable: very_high (can't think without it), high, moderate, low (easily replaced)"),
])
row = add_table(ws, row,
    ["id", "metaphor", "type", "structuring_effect", "absoluteness", "source_type", "source_ref", "confidence"],
    [
        (1, "Territory/borders", "spatial", "Makes tech controllable through physical boundary logic", "high", "llm_analysis", "blumenberg_v1", 0.90),
        (2, "Immune system", "biological", "Tech as defense against foreign 'infection/contamination'", "moderate", "evidence_testing", "cluster_34", 0.85),
        (3, "Arsenal/stockpile", "military", "Tech capability as accumulated weaponry", "moderate", "llm_analysis", "blumenberg_v1", 0.82),
        (4, "Race/competition", "athletic", "Tech development as zero-sum contest with finish line", "very_high", "evidence_testing", "cluster_56", 0.92),
    ],
    "concept_metaphors"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 4 shows the RACE metaphor is VERY_HIGH absoluteness - it's nearly impossible to think 'tech sovereignty' "
    "without the race/competition framing. This metaphor is so deep you can't translate it away. It's 'absolute' in Blumenberg's sense.")
row += 1

# Table 2: concept_nonconceptuality
row = add_section_header(ws, row, "TABLE: concept_nonconceptuality")
row = add_explanation(ws, row,
    "Identifies what RESISTS CONCEPTUALIZATION. Blumenberg's key insight: not everything can be put into concepts. Some aspects of "
    "understanding remain nonconceptual - felt, imagined, but never fully articulated.")
row = add_field_explanations(ws, row, [
    ("nonconceptual_aspect", "What resists being conceptualized"),
    ("why_resists_conceptualization", "Why it can't be captured in concepts"),
    ("manifestation", "How this nonconceptual element shows up"),
])
row = add_table(ws, row,
    ["id", "nonconceptual_aspect", "why_resists_conceptualization", "manifestation", "source_type", "source_ref", "confidence"],
    [
        (1, "Feeling of autonomy", "Experiential quality that can't be reduced to policy metrics", "Leaders speak of 'pride' and 'dignity' alongside capability", "llm_analysis", "blumenberg_nonconc", 0.85),
        (2, "Technological sublime", "Awe at tech power exceeds rational assessment", "Inflated expectations, magical thinking about tech effects", "llm_analysis", "blumenberg_nonconc", 0.80),
        (3, "Existential anxiety", "Fear of dependence is visceral, not just calculated", "Emotional intensity of sovereignty debates", "evidence_testing", "cluster_89", 0.82),
    ],
    "concept_nonconceptuality"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 shows that 'feeling of autonomy' can't be reduced to metrics. When leaders talk about tech sovereignty, "
    "they invoke 'pride' and 'dignity' - affective qualities that resist conceptual capture. This is why purely rational policy "
    "analysis misses something important about the concept.")
row += 1

# Table 3: concept_lifeworld_connection
row = add_section_header(ws, row, "TABLE: concept_lifeworld_connection")
row = add_explanation(ws, row,
    "Maps TENSION between the concept and LIVED EXPERIENCE. Concepts operate against a background of pre-theoretical, everyday "
    "understanding. What happens when they diverge?")
row = add_field_explanations(ws, row, [
    ("lifeworld_aspect", "Which aspect of lived experience"),
    ("connection_type", "EXPERIENTIAL_CONTRADICTION, STRUCTURAL_TENSION, PRACTICAL_TENSION, RELATIONAL_TENSION"),
    ("tension_level", "How severe: very_high, high, moderate, low"),
    ("description", "What the tension is"),
])
row = add_table(ws, row,
    ["id", "lifeworld_aspect", "connection_type", "tension_level", "description", "source_type", "source_ref", "confidence"],
    [
        (1, "Daily tech dependence", "experiential_contradiction", "high", "People use foreign tech daily while supporting sovereignty rhetoric", "evidence_testing", "cluster_91", 0.90),
        (2, "Global connectivity", "structural_tension", "very_high", "Lived experience of global networks vs sovereignty framing", "llm_analysis", "blumenberg_lifeworld", 0.88),
        (3, "Consumer expectations", "practical_tension", "moderate", "Want sovereignty but also want best/cheapest products", "evidence_testing", "cluster_45", 0.85),
    ],
    "concept_lifeworld_connection"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 shows EXPERIENTIAL CONTRADICTION - people use iPhones and TikTok daily (foreign tech) while supporting "
    "sovereignty rhetoric. Their lived experience contradicts the concept they endorse. This is unsustainable.")

set_column_widths(ws)

# ============================================================================
# SHEET 10: Carey Dimension
# ============================================================================
ws = wb.create_sheet("10. Carey")
row = style_header(ws, 1, "CAREY DIMENSION - Conceptual Bootstrapping", dimension_colors["Carey"])
row += 1

row = add_section_header(ws, row, "PHILOSOPHICAL BACKGROUND")
row = add_explanation(ws, row,
    "Based on Susan Carey's 'The Origin of Concepts' (2009). Key ideas: (1) CORE COGNITION - innate cognitive systems (object, number, "
    "agency, etc.); (2) QUINIAN BOOTSTRAPPING - building complex concepts from simpler ones through analogy and combination; "
    "(3) PLACEHOLDER STRUCTURES - partial concepts awaiting elaboration; (4) CONCEPTUAL DISCONTINUITY - new concepts can be genuinely "
    "incommensurable with predecessors; (5) COMPUTATIONAL CONSTRAINTS - cognitive limitations affect what concepts are learnable.")
row = add_explanation(ws, row,
    "THIS DIMENSION CAPTURES: How the concept was BUILT UP from simpler components, what cognitive resources it engages, what parts "
    "are still 'placeholder' (undefined), and what constraints affect learning/using it.")
row += 1

# Table 1: concept_built_from
row = add_section_header(ws, row, "TABLE: concept_built_from")
row = add_explanation(ws, row,
    "Lists the COMPONENT CONCEPTS from which this one was 'bootstrapped'. Complex concepts are built from simpler ones through "
    "combination and analogy.")
row = add_field_explanations(ws, row, [
    ("component_concept", "Name of the component"),
    ("component_level", "That component's hierarchy level (1=primitive, 2=simple composite, 3=complex, 4=theoretical)"),
    ("contribution", "What this component contributes to the concept"),
])
row = add_table(ws, row,
    ["id", "component_concept", "component_level", "contribution", "source_type", "source_ref", "confidence"],
    [
        (1, "Sovereignty", 2, "Provides authority/control dimension from political philosophy", "llm_analysis", "carey_v1", 0.90),
        (2, "Technology", 1, "Provides domain/object specification", "llm_analysis", "carey_v1", 0.92),
        (3, "Autonomy", 2, "Provides independence/self-determination dimension", "llm_analysis", "carey_v1", 0.88),
        (4, "National interest", 2, "Provides collective agent whose sovereignty is at stake", "llm_analysis", "carey_v1", 0.85),
        (5, "Strategic competition", 2, "Provides urgency and adversarial framing", "evidence_testing", "cluster_23", 0.82),
    ],
    "concept_built_from"
)
row = add_explanation(ws, row,
    "READING THE DATA: 'Tech sovereignty' is Level 3 (complex composite) built from Level 1-2 components. 'Technology' is Level 1 "
    "(primitive). 'Sovereignty', 'Autonomy', 'National interest' are Level 2 (simple composites). The concept combines these.")
row += 1

# Table 2: concept_core_cognition
row = add_section_header(ws, row, "TABLE: concept_core_cognition")
row = add_explanation(ws, row,
    "Identifies engagement with INNATE COGNITIVE SYSTEMS. Carey argues we have 'core cognition' - evolved systems for objects, "
    "numbers, agents, in-group/out-group, etc. Complex concepts often recruit these systems.")
row = add_field_explanations(ws, row, [
    ("core_system", "Which innate system: object, number, agency, in-group/out-group, contamination"),
    ("engagement", "How strongly engaged: strong, moderate, weak, none"),
    ("how_activated", "How the concept activates this system"),
])
row = add_table(ws, row,
    ["id", "core_system", "engagement", "how_activated", "source_type", "source_ref", "confidence"],
    [
        (1, "in-group/out-group", "strong", "Us vs them framing; domestic vs foreign tech", "llm_analysis", "carey_core", 0.88),
        (2, "object", "moderate", "Technology as bounded 'thing' that can be possessed/controlled", "llm_analysis", "carey_core", 0.80),
        (3, "agency", "moderate", "Anthropomorphizing 'markets', 'technology', 'adversaries' as agents", "llm_analysis", "carey_core", 0.75),
        (4, "contamination", "strong", "Foreign tech as potentially 'contaminating' - purity logic", "evidence_testing", "cluster_78", 0.85),
    ],
    "concept_core_cognition"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 4 shows CONTAMINATION system is strongly engaged - foreign tech is framed as potentially 'contaminating' "
    "domestic purity. This is deep evolutionary psychology being recruited. It's why debates get so emotional.")
row += 1

# Table 3: concept_placeholder_structures
row = add_section_header(ws, row, "TABLE: concept_placeholder_structures")
row = add_explanation(ws, row,
    "Identifies PARTIAL STRUCTURES awaiting elaboration. Many concepts have 'slots' that are labeled but not filled in. "
    "People use the concept without having worked out crucial details.")
row = add_field_explanations(ws, row, [
    ("placeholder", "What's placeholder - the unfilled slot"),
    ("content_type", "What type of content is missing: procedural, evaluative, categorical, ontological"),
    ("current_fill_status", "LARGELY_EMPTY, PARTIALLY_FILLED, CONTESTED, ASSUMED_NOT_ELABORATED"),
    ("elaboration_needed", "What elaboration would be needed to fill the slot"),
])
row = add_table(ws, row,
    ["id", "placeholder", "content_type", "current_fill_status", "elaboration_needed", "source_type", "source_ref", "confidence"],
    [
        (1, "How to achieve it", "procedural", "largely_empty", "Massive - no clear path to actual sovereignty", "evidence_testing", "cluster_67", 0.90),
        (2, "What counts as success", "evaluative", "contested", "Multiple incompatible definitions in use", "evidence_testing", "cluster_89", 0.88),
        (3, "Which technologies are critical", "categorical", "partially_filled", "Lists exist but contested and shifting", "evidence_testing", "cluster_34", 0.85),
        (4, "Who the sovereign agent is", "ontological", "assumed_not_elaborated", "State assumed but actually contested (regions? blocs?)", "llm_analysis", "carey_placeholder", 0.82),
    ],
    "concept_placeholder_structures"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 1 is devastating - 'How to achieve it' is LARGELY EMPTY. Everyone uses the term but no one has worked out "
    "HOW to actually achieve tech sovereignty. The concept is all aspiration, no roadmap. This is a serious placeholder problem.")
row += 1

# Table 4: concept_bootstrapping_constraints
row = add_section_header(ws, row, "TABLE: concept_bootstrapping_constraints")
row = add_explanation(ws, row,
    "Identifies COGNITIVE CONSTRAINTS on learning/using the concept. Some concepts demand more than human cognition can easily deliver.")
row = add_field_explanations(ws, row, [
    ("constraint_type", "WORKING_MEMORY (too much to hold), ANALOGICAL_MAPPING (source doesn't map well), CAUSAL_LEARNING (too complex), COHERENCE (components don't fit)"),
    ("description", "What the constraint is"),
    ("severity", "How severe: very_high, high, moderate, low"),
])
row = add_table(ws, row,
    ["id", "constraint_type", "description", "severity", "source_type", "source_ref", "confidence"],
    [
        (1, "working_memory", "Concept requires tracking too many interdependencies simultaneously", "moderate", "llm_analysis", "carey_constraints", 0.78),
        (2, "analogical_mapping", "Source concepts (Westphalian sovereignty) map poorly to tech domain", "high", "llm_analysis", "carey_constraints", 0.85),
        (3, "causal_learning", "Cause-effect relationships too complex and delayed for intuitive learning", "high", "evidence_testing", "cluster_56", 0.88),
        (4, "coherence", "Component concepts don't cohere well - autonomy vs innovation tension", "very_high", "evidence_testing", "cluster_45", 0.90),
    ],
    "concept_bootstrapping_constraints"
)
row = add_explanation(ws, row,
    "READING THE DATA: Row 4 shows a VERY_HIGH coherence constraint - the components don't fit together. 'Autonomy' and 'innovation' "
    "are in tension. The concept is trying to combine incompatible elements. This explains why it feels unstable.")

set_column_widths(ws)

# ============================================================================
# SHEET 11: Summary
# ============================================================================
ws = wb.create_sheet("11. Summary")
row = style_header(ws, 1, "SUMMARY - Key Cross-Dimensional Findings", dimension_colors["Overview"])
row += 1

row = add_section_header(ws, row, "CONCEPT OVERVIEW")
ws.cell(row=row, column=1, value="Concept:").font = Font(bold=True)
ws.cell(row=row, column=2, value="Technological Sovereignty")
row += 1
ws.cell(row=row, column=1, value="Health Status:").font = Font(bold=True)
ws.cell(row=row, column=2, value="STRAINED - High usage but growing contradictions")
row += 1
ws.cell(row=row, column=1, value="Hierarchy Level:").font = Font(bold=True)
ws.cell(row=row, column=2, value="3 (Complex composite)")
row += 2

row = add_section_header(ws, row, "CRITICAL FINDINGS BY DIMENSION")
findings = [
    ("Quinean", "HIGH revision cost; hub position; existential dependence on 'the state'; practical contradiction with global supply chains"),
    ("Sellarsian", "Multiple false givens ('states CAN achieve autonomy'); hidden commitments to autarky viability; manifest/scientific image gap"),
    ("Brandomian", "UNANSWERED CHALLENGE: Gulf $2T proves investment ≠ capability; major perspectival translation losses between actors"),
    ("Deleuzian", "UNSTABLE endoconsistency; operates on techno-nationalist plane; paradoxical autonomy/innovation zone"),
    ("Bachelardian", "VERY HIGH substantialist obstacle; serves control fantasy and spending legitimation needs; pre-scientific stage dominant"),
    ("Canguilhem", "Currently in INTENSIFICATION phase but STRAINED health; vital norm violation (sovereignty justifying dependency)"),
    ("Hacking", "Strong LOOPING EFFECTS creating self-fulfilling prophecy; closing international cooperation possibilities; making up new 'kinds'"),
    ("Blumenberg", "ABSOLUTE race metaphor; nonconceptual 'feeling of autonomy'; high lifeworld contradiction (daily foreign tech use)"),
    ("Carey", "LARGELY EMPTY 'how to achieve it' placeholder; VERY HIGH coherence constraint; strong contamination cognition engaged"),
]
for dim, finding in findings:
    ws.cell(row=row, column=1, value=dim).font = Font(bold=True, color=dimension_colors.get(dim, "000000"))
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row, column=2, value=finding).alignment = wrap
    row += 1
row += 1

row = add_section_header(ws, row, "TOP 5 MOST CONCERNING ISSUES")
concerns = [
    "1. UNANSWERED CHALLENGE: Gulf states' $2T proves investment doesn't lead to capability - core commitment violated (Brandomian, confidence 0.92)",
    "2. EMPTY PLACEHOLDER: No one knows HOW to achieve sovereignty - procedural placeholder largely empty (Carey, confidence 0.90)",
    "3. SUBSTANTIALIST OBSTACLE: Treating capability as substance vs relational capacity - blocks understanding (Bachelardian, very_high entrenchment)",
    "4. VITAL NORM VIOLATION: Concept being used to justify increased dependency - self-undermining (Canguilhem, confidence 0.90)",
    "5. COHERENCE FAILURE: Components (autonomy, innovation) don't cohere - unstable structure (Carey/Deleuze, very_high severity)",
]
for concern in concerns:
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value=concern).alignment = wrap
    ws.row_dimensions[row].height = 30
    row += 1
row += 1

row = add_section_header(ws, row, "WHAT THIS MEANS")
row = add_explanation(ws, row,
    "The concept 'Technological Sovereignty' is STRUCTURALLY UNSTABLE. It has very high usage and policy adoption, but: "
    "(1) its core commitment (investment → capability → autonomy) is empirically falsified; (2) no one has worked out how to actually "
    "achieve it; (3) it's built from incompatible components; (4) it's increasingly used against its own vital norms. "
    "The concept is doing political work but may not survive serious empirical pressure.")

set_column_widths(ws, {'A': 15, 'B': 80, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15})

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_ANNOTATED_TechSovereignty.xlsx")
wb.save(output_path)
print(f"Annotated & populated schema saved to: {output_path}")
print(f"11 sheets with full explanations for each dimension")
