#!/usr/bin/env python3
"""
Generate DIMENSION-SPECIFIC Evidence Feedback Schema v2

Aligned with v5 expanded concept schema (47 tables).
Each dimension has feedback types derived from deep philosophical research
into Quine, Sellars, Brandom, Deleuze, Bachelard, Canguilhem, Hacking, Blumenberg, Carey.

TOTAL: 81 dimension-specific feedback types (up from 57 in v1)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Overview": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Hacking": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

def style_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def add_feedback_types(ws, start_row, dimension, types):
    """Add feedback types for a dimension."""
    row = start_row

    headers = ["Feedback Type", "What It Flags", "Tables Affected", "Severity", "Example"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for t in types:
        ws.cell(row=row, column=1, value=t["type"]).font = field_font
        ws.cell(row=row, column=2, value=t["flags"])
        ws.cell(row=row, column=3, value=t["tables"])
        ws.cell(row=row, column=4, value=t["severity"])
        ws.cell(row=row, column=5, value=t["example"])
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    return row + 1

def set_column_widths(ws):
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 55

# ============================================================================
# SHEET 1: Overview
# ============================================================================
ws = wb.active
ws.title = "1. Overview"

style_header(ws, 1, "DIMENSION-SPECIFIC EVIDENCE FEEDBACK SCHEMA v2", dimension_colors["Overview"])

overview = [
    "",
    "PURPOSE:",
    "When testing concept values against evidence clusters, each philosophical dimension",
    "has its OWN specific feedback types - derived from deep research into each thinker.",
    "",
    "This v2 schema is aligned with v5 expanded concept schema (47 tables).",
    "Each feedback type maps to specific tables and columns that may need revision.",
    "",
    "PHILOSOPHICAL GROUNDING:",
    "Each dimension's feedback types emerge from the thinker's core theoretical apparatus:",
    "- Quine: Web of belief, holism, ontological relativity, revisability costs",
    "- Sellars: Space of reasons, functional role semantics, myth of the given, manifest/scientific image",
    "- Brandom: Deontic scorekeeping, commitments/entitlements, game of giving and asking for reasons",
    "- Deleuze: Components, zones of indiscernibility, consistency, plane of immanence, personae",
    "- Bachelard: Obstacle taxonomy, three stages, psychoanalysis of knowledge, regional rationality",
    "- Canguilhem: Vital normativity, filiation, milieu, normal vs normative distinction",
    "- Hacking: Dynamic nominalism, looping effects, making up people, space of possibilities",
    "- Blumenberg: Absolute metaphors, nonconceptuality, lifeworld background, metakinetics",
    "- Carey: Core cognition, Quinian bootstrapping, placeholder structures, incommensurability",
    "",
    "FEEDBACK TYPES BY DIMENSION:",
    "",
    "QUINEAN (9 types): inference_broken, inference_missing, centrality_wrong,",
    "                   entrenchment_overestimated, web_inconsistency, holism_ripple,",
    "                   web_position_wrong, ontological_dependence_missed, revision_cost_miscalculated",
    "",
    "SELLARSIAN (9 types): false_given_exposed, hidden_commitment_revealed,",
    "                      givenness_marker_challenged, effect_reversed,",
    "                      space_of_reasons_violation, manifest_scientific_gap,",
    "                      functional_role_misspecified, inferential_articulation_wrong, image_tension_shifted",
    "",
    "BRANDOMIAN (7 types): commitment_violated, entitlement_exceeded, inference_defeated,",
    "                      incompatibility_missed, challenge_unanswered,",
    "                      perspectival_distortion, deontic_status_wrong",
    "",
    "DELEUZIAN (8 types): component_missing, component_extraneous, zone_mismapped,",
    "                     consistency_weaker, neighborhood_wrong, plane_misidentified,",
    "                     persona_different, problem_transformed",
    "",
    "BACHELARDIAN (9 types): new_obstacle_discovered, obstacle_deeper,",
    "                        blocked_understanding_unblocked, rupture_imminent,",
    "                        rupture_false, psychoanalytic_function_different,",
    "                        stage_mismatch, regional_rationality_wrong, obstacle_type_misidentified",
    "",
    "CANGUILHEM (9 types): evolution_different, normative_dimension_exposed,",
    "                      vitality_declining, vitality_reviving, filiation_broken,",
    "                      normal_normative_confusion, milieu_changed, vital_norm_violated,",
    "                      concept_migration_detected",
    "",
    "HACKING (9 types): style_mismatch, visibility_wrong, evidence_privileged_wrong,",
    "                   inference_pattern_different, new_style_emerging,",
    "                   objects_created_different, looping_effect_detected,",
    "                   kind_created_or_destroyed, possibility_space_shifted",
    "",
    "BLUMENBERG (9 types): metaphor_missed, metaphor_effect_different, metakinesis_different,",
    "                      conceptual_work_failing, nonconceptuality_revealed,",
    "                      absolutism_underestimated, lifeworld_tension_exposed,",
    "                      unbegrifflichkeit_encountered, background_metaphor_shift",
    "",
    "CAREY (12 types): hierarchy_wrong, component_missing, bootstrap_failed,",
    "                  incommensurability_revealed, mapping_different, core_cognition_wrong,",
    "                  placeholder_structure_wrong, core_system_engagement_missed,",
    "                  discontinuity_detected, representational_resource_insufficient,",
    "                  computational_constraint_violated, quinian_bootstrap_stalled",
    "",
    "TOTAL: 81 dimension-specific feedback types (up from 57 in v1)",
]

for row_idx, text in enumerate(overview, 3):
    ws.cell(row=row_idx, column=1, value=text)
    if text.startswith("PURPOSE") or text.startswith("HOW IT") or text.startswith("FEEDBACK") or text.startswith("PHILOSOPHICAL"):
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

set_column_widths(ws)

# ============================================================================
# SHEET 2: Quinean Feedback Types
# ============================================================================
ws = wb.create_sheet("2. Quinean")
style_header(ws, 1, "QUINEAN FEEDBACK TYPES - Web of Belief & Ontological Relativity", dimension_colors["Quinean"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Two Dogmas of Empiricism' - confirmation holism, no analytic/synthetic distinction, revisability, ontological relativity").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Quinean", [
    {
        "type": "inference_broken",
        "flags": "A claimed inferential connection doesn't hold - evidence shows 'If A then B' fails",
        "tables": "concept_inferences",
        "severity": "major",
        "example": "We claim 'tech sovereignty → indigenous development' but Gulf states have $2T and still can't do Series B"
    },
    {
        "type": "inference_missing",
        "flags": "Evidence reveals an important inferential connection we didn't capture",
        "tables": "concept_inferences",
        "severity": "minor",
        "example": "We missed 'tech sovereignty → need for protected procurement' which evidence shows is central"
    },
    {
        "type": "centrality_wrong",
        "flags": "Concept is more/less central to the web than we assessed",
        "tables": "concepts.centrality",
        "severity": "moderate",
        "example": "We marked 'technological sovereignty' as core but evidence shows it's intermediate - can be revised without major ripples"
    },
    {
        "type": "entrenchment_overestimated",
        "flags": "Concept is easier to revise than our entrenchment score suggested",
        "tables": "concepts.entrenchment_score",
        "severity": "moderate",
        "example": "Entrenchment was 0.9 but evidence shows actors readily abandon sovereignty claims when convenient"
    },
    {
        "type": "web_inconsistency",
        "flags": "The claimed inferential connections create contradictions within the web",
        "tables": "concept_web_tensions",
        "severity": "critical",
        "example": "We claim both 'sovereignty → autonomy' AND 'sovereignty → global integration' but these contradict"
    },
    {
        "type": "holism_ripple",
        "flags": "Change to this concept would force changes to more concepts than we predicted",
        "tables": "concept_inferences, concept_web_tensions",
        "severity": "major",
        "example": "Revising 'tech sovereignty' would force changes to 8 related concepts, not 3 as we thought"
    },
    {
        "type": "web_position_wrong",
        "flags": "Concept's position in belief web (bridge, hub, leaf, foundation) is misclassified",
        "tables": "concept_web_position",
        "severity": "major",
        "example": "We classified as 'hub' but evidence shows it's actually a 'bridge' connecting otherwise separate belief clusters"
    },
    {
        "type": "ontological_dependence_missed",
        "flags": "Concept's existence/identity depends on another entity we didn't identify (per Quine's ontological relativity)",
        "tables": "concept_ontological_dependence",
        "severity": "major",
        "example": "We missed that 'tech sovereignty' depends ontologically on 'the state' - the concept makes no sense without state actors"
    },
    {
        "type": "revision_cost_miscalculated",
        "flags": "The ramifications of revising this concept differ from our assessment (per Quine's pragmatic revisability)",
        "tables": "concept_revision_ramifications",
        "severity": "moderate",
        "example": "We said low revision cost but abandoning the concept would require reworking entire policy frameworks"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 3: Sellarsian Feedback Types
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
style_header(ws, 1, "SELLARSIAN FEEDBACK TYPES - Space of Reasons & Functional Role", dimension_colors["Sellarsian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Empiricism and Philosophy of Mind' - space of reasons, functional role semantics, myth of the given, manifest vs scientific image").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Sellarsian", [
    {
        "type": "false_given_exposed",
        "flags": "What we treated as foundational/self-evident is actually inferred from other premises",
        "tables": "concept_givenness",
        "severity": "critical",
        "example": "We marked 'sovereignty is achievable' as given, but evidence shows it's inferred from contested premises about state capacity"
    },
    {
        "type": "hidden_commitment_revealed",
        "flags": "Evidence exposes a baked-in assumption we didn't identify",
        "tables": "concept_hidden_commitments",
        "severity": "major",
        "example": "Evidence reveals hidden commitment: 'tech development is linear/predictable' which we didn't list"
    },
    {
        "type": "givenness_marker_challenged",
        "flags": "A rhetorical marker of givenness ('obviously', 'naturally') is directly contested by evidence",
        "tables": "concept_givenness_markers",
        "severity": "moderate",
        "example": "We noted 'naturally extends state power' but evidence shows tech often undermines state power"
    },
    {
        "type": "effect_reversed",
        "flags": "What we thought givenness enables is actually blocked, or vice versa",
        "tables": "concept_givenness_effects",
        "severity": "major",
        "example": "We said givenness blocks 'analysis of impossibility' but evidence shows such analysis IS happening"
    },
    {
        "type": "space_of_reasons_violation",
        "flags": "Concept is used outside proper justificatory context - not properly in the space of reasons",
        "tables": "concept_space_of_reasons",
        "severity": "moderate",
        "example": "Concept invoked without any inferential support - pure assertion masquerading as justified claim"
    },
    {
        "type": "manifest_scientific_gap",
        "flags": "Bigger gap between manifest (everyday) and scientific image than we identified",
        "tables": "concept_image_tension",
        "severity": "minor",
        "example": "Everyday use of 'sovereignty' diverges much more from technical IR usage than we noted"
    },
    {
        "type": "functional_role_misspecified",
        "flags": "The concept's role in the inferential network differs from what we specified (per Sellars' functional role semantics)",
        "tables": "concept_functional_role",
        "severity": "major",
        "example": "We specified 'enabling deployment premise' but concept actually functions as 'legitimation conclusion' in discourse"
    },
    {
        "type": "inferential_articulation_wrong",
        "flags": "The concept's articulation with other concepts in the space of reasons differs from what we mapped",
        "tables": "concept_space_of_reasons",
        "severity": "moderate",
        "example": "We said concept articulates with 'national security' but evidence shows primary articulation is with 'economic competitiveness'"
    },
    {
        "type": "image_tension_shifted",
        "flags": "The tension between manifest and scientific images of the concept has changed",
        "tables": "concept_image_tension",
        "severity": "moderate",
        "example": "Scientific image of tech development (complex, interdependent) is now bleeding into manifest image"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 4: Brandomian Feedback Types
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
style_header(ws, 1, "BRANDOMIAN FEEDBACK TYPES - Scorekeeping & Reasons", dimension_colors["Brandomian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Making It Explicit' - deontic scorekeeping, commitments/entitlements, game of giving and asking for reasons").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Brandomian", [
    {
        "type": "commitment_violated",
        "flags": "A claimed commitment is not honored in practice",
        "tables": "concept_commitments",
        "severity": "critical",
        "example": "Commitment: 'investment → autonomy' but Gulf $2T can't invest at Series B level - commitment violated"
    },
    {
        "type": "entitlement_exceeded",
        "flags": "Users claim more than they're entitled to based on the concept",
        "tables": "concept_commitments",
        "severity": "major",
        "example": "States claim entitlement to 'sovereignty success' based on rhetoric alone - exceeding inferential warrant"
    },
    {
        "type": "inference_defeated",
        "flags": "Counterexample defeats a claimed material inference (committive, permissive, or incompatibility)",
        "tables": "concept_inferential_roles",
        "severity": "critical",
        "example": "Committive inference 'tech sovereignty → right to exclude' defeated by EU Chips Act creating new dependencies"
    },
    {
        "type": "incompatibility_missed",
        "flags": "Evidence reveals an incompatibility relation we didn't capture",
        "tables": "concept_inferential_roles",
        "severity": "major",
        "example": "We missed: 'tech sovereignty' INCOMPATIBLE WITH 'participation in global supply chains'"
    },
    {
        "type": "challenge_unanswered",
        "flags": "A challenge to a commitment cannot be justified - no adequate reasons available",
        "tables": "concept_challenges, concept_justifications",
        "severity": "critical",
        "example": "Challenge: 'What entitles the claim investment → autonomy?' No justification in evidence cluster"
    },
    {
        "type": "perspectival_distortion",
        "flags": "De re translation doesn't preserve truth - our substitution changes meaning",
        "tables": "concept_perspectival_content",
        "severity": "moderate",
        "example": "Gulf says 'achieving sovereignty' but our translation 'dependency management' loses their self-understanding"
    },
    {
        "type": "deontic_status_wrong",
        "flags": "Something is attributed but not acknowledged, or undertaken but not attributed, etc.",
        "tables": "concept_commitments",
        "severity": "moderate",
        "example": "We marked commitment as 'acknowledged' but evidence shows actors only attribute it to others, not themselves"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 5: Deleuzian Feedback Types
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
style_header(ws, 1, "DELEUZIAN FEEDBACK TYPES - Concept Theory", dimension_colors["Deleuzian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'What is Philosophy?' - concepts as multiplicities, zones of indiscernibility, plane of immanence, conceptual personae").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Deleuzian", [
    {
        "type": "component_missing",
        "flags": "Evidence reveals a component of the concept we didn't identify",
        "tables": "concept_components",
        "severity": "major",
        "example": "We missed 'temporality' as component - sovereignty implies urgency/limited time window"
    },
    {
        "type": "component_extraneous",
        "flags": "A component we identified doesn't actually belong to this concept's multiplicity",
        "tables": "concept_components",
        "severity": "moderate",
        "example": "We listed 'innovation' as component but evidence shows it's separate from sovereignty proper"
    },
    {
        "type": "zone_mismapped",
        "flags": "Zone of indiscernibility between components doesn't function as we described",
        "tables": "concept_zones_of_indiscernibility",
        "severity": "moderate",
        "example": "Sovereignty-technology zone: we said tension, but evidence shows they're more integrated than we thought"
    },
    {
        "type": "consistency_weaker",
        "flags": "Endoconsistency (internal coherence) is weaker than we assessed",
        "tables": "concept_consistency",
        "severity": "major",
        "example": "We said 'moderate' consistency but evidence shows components are barely held together - 'unstable'"
    },
    {
        "type": "neighborhood_wrong",
        "flags": "Concept's relations to other concepts differ from what we mapped",
        "tables": "concept_neighborhood",
        "severity": "moderate",
        "example": "We said 'economic sovereignty' is closest neighbor but evidence shows 'digital sovereignty' is closer"
    },
    {
        "type": "plane_misidentified",
        "flags": "Concept operates on a different plane of immanence than we identified",
        "tables": "concept_plane_of_immanence",
        "severity": "critical",
        "example": "We said Westphalian plane but evidence shows concept now operates on 'techno-nationalist' plane"
    },
    {
        "type": "persona_different",
        "flags": "Different conceptual persona activates the concept than we identified",
        "tables": "concept_personae",
        "severity": "moderate",
        "example": "We said 'the sovereign' but evidence shows 'the technocrat' is the activating persona"
    },
    {
        "type": "problem_transformed",
        "flags": "The problem the concept addresses has transformed in ways we didn't capture",
        "tables": "concept_problems",
        "severity": "major",
        "example": "Problem shifted from 'how to achieve autonomy' to 'how to claim autonomy while managing dependency'"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 6: Bachelardian Feedback Types
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
style_header(ws, 1, "BACHELARDIAN FEEDBACK TYPES - Epistemological Obstacles", dimension_colors["Bachelardian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Formation of Scientific Mind' - obstacle taxonomy, three stages, psychoanalysis of knowledge, regional rationalism").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Bachelardian", [
    {
        "type": "new_obstacle_discovered",
        "flags": "Evidence reveals an epistemological obstacle function we didn't identify",
        "tables": "concept_obstacles",
        "severity": "major",
        "example": "We didn't identify 'techno-nationalism' as obstacle, but evidence shows it blocks recognition of interdependence"
    },
    {
        "type": "obstacle_deeper",
        "flags": "The obstacle is more entrenched than we assessed",
        "tables": "concept_obstacles",
        "severity": "major",
        "example": "We thought obstacle was 'verbal' (language-based) but evidence shows it's 'substantialist' - deeper resistance"
    },
    {
        "type": "blocked_understanding_unblocked",
        "flags": "Something we thought was blocked by the concept is actually being understood",
        "tables": "concept_obstacle_blocks",
        "severity": "moderate",
        "example": "We said concept blocks 'analysis of impossibility' but evidence shows scholars ARE doing this analysis"
    },
    {
        "type": "rupture_imminent",
        "flags": "Evidence suggests conditions for epistemological rupture are forming",
        "tables": "concept_obstacles",
        "severity": "major",
        "example": "Mounting failures (China chips, Russia sanctions) creating conditions for rupture with sovereignty concept"
    },
    {
        "type": "rupture_false",
        "flags": "A claimed rupture or breakthrough hasn't actually occurred",
        "tables": "concept_obstacles",
        "severity": "moderate",
        "example": "We noted a rupture point but evidence shows the old obstacle thinking persists"
    },
    {
        "type": "psychoanalytic_function_different",
        "flags": "The concept serves a different unconscious need than we identified",
        "tables": "concept_psychoanalytic_function",
        "severity": "moderate",
        "example": "We said need for 'control fantasy' but evidence suggests need for 'legitimation of spending'"
    },
    {
        "type": "stage_mismatch",
        "flags": "Concept is at a different cognitive stage (pre-scientific, proto-scientific, scientific) than we assessed",
        "tables": "concept_cognitive_stage",
        "severity": "major",
        "example": "We said 'scientific' but evidence shows concept operates with pre-scientific animistic thinking"
    },
    {
        "type": "regional_rationality_wrong",
        "flags": "Concept belongs to a different domain of rationality than we identified (per Bachelard's regional rationalism)",
        "tables": "concept_regional_rationality",
        "severity": "moderate",
        "example": "We assigned to 'geopolitical rationality' but concept actually operates within 'economic rationality'"
    },
    {
        "type": "obstacle_type_misidentified",
        "flags": "The type of obstacle (verbal, substantialist, animist, etc.) is different from what we classified",
        "tables": "concept_obstacles",
        "severity": "moderate",
        "example": "We classified as 'verbal obstacle' but evidence shows it's an 'obstacle of first experience' (immediate intuition)"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 7: Canguilhem Feedback Types
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
style_header(ws, 1, "CANGUILHEM FEEDBACK TYPES - Vital History of Concepts", dimension_colors["Canguilhem"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Normal and Pathological' - vital normativity, concept filiation, milieu, normal vs normative distinction").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Canguilhem", [
    {
        "type": "evolution_different",
        "flags": "Historical trajectory differs from the evolution we mapped",
        "tables": "concept_evolution",
        "severity": "moderate",
        "example": "We mapped linear evolution but evidence shows concept had a 'death' in 1990s before 2010s revival"
    },
    {
        "type": "normative_dimension_exposed",
        "flags": "Evidence reveals a hidden normative dimension (embedded value) we didn't identify",
        "tables": "concept_normative_dimensions",
        "severity": "major",
        "example": "We missed: 'sovereignty' embeds value that state action is inherently legitimate"
    },
    {
        "type": "vitality_declining",
        "flags": "Concept's health/vitality is worse than we assessed",
        "tables": "concept_vitality_indicators, concepts.health_status",
        "severity": "major",
        "example": "We said 'healthy' but evidence shows repeated failures indicating 'strained' status"
    },
    {
        "type": "vitality_reviving",
        "flags": "Concept shows unexpected signs of life/revival",
        "tables": "concept_vitality_indicators, concepts.health_status",
        "severity": "moderate",
        "example": "We said 'dying' but evidence shows new actors (EU, India) giving concept new life"
    },
    {
        "type": "filiation_broken",
        "flags": "Claimed conceptual lineage (filiation) doesn't hold",
        "tables": "concept_filiation",
        "severity": "moderate",
        "example": "We claimed descent from 'Westphalian sovereignty' but evidence shows different conceptual origin"
    },
    {
        "type": "normal_normative_confusion",
        "flags": "What we labeled as descriptive ('normal') is actually prescriptive ('normative')",
        "tables": "concept_normative_dimensions, concept_vital_norms",
        "severity": "major",
        "example": "We treated 'states pursue sovereignty' as descriptive fact but it's actually a normative prescription"
    },
    {
        "type": "milieu_changed",
        "flags": "The concept's operating milieu (institutional, intellectual, material environment) has shifted",
        "tables": "concept_milieu",
        "severity": "major",
        "example": "Concept's milieu shifted from 'nation-state geopolitics' to 'platform capitalism' environment"
    },
    {
        "type": "vital_norm_violated",
        "flags": "The concept is being used against its own vital norms (self-undermining deployment)",
        "tables": "concept_vital_norms",
        "severity": "critical",
        "example": "Using 'sovereignty' to justify becoming more dependent - violates concept's own vital norm of autonomy"
    },
    {
        "type": "concept_migration_detected",
        "flags": "Concept has migrated between fields/disciplines in ways we didn't track",
        "tables": "concept_filiation, concept_milieu",
        "severity": "moderate",
        "example": "We tracked IR origins but missed recent migration to 'technology ethics' discourse with transformed meaning"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 8: Hacking Feedback Types
# ============================================================================
ws = wb.create_sheet("8. Hacking")
style_header(ws, 1, "HACKING FEEDBACK TYPES - Dynamic Nominalism & Looping", dimension_colors["Hacking"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Historical Ontology' - dynamic nominalism, looping effects, making up people, styles of reasoning, space of possibilities").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Hacking", [
    {
        "type": "style_mismatch",
        "flags": "Concept requires a different reasoning style than we identified",
        "tables": "concept_reasoning_styles",
        "severity": "major",
        "example": "We said 'geopolitical strategic' style but evidence shows concept operates via 'economic modeling' style"
    },
    {
        "type": "visibility_wrong",
        "flags": "What we thought the style makes visible/invisible is reversed",
        "tables": "concept_style_visibility",
        "severity": "moderate",
        "example": "We said style makes 'corporate interests invisible' but evidence shows they're actually quite visible"
    },
    {
        "type": "evidence_privileged_wrong",
        "flags": "Different evidence types are privileged/marginalized than we identified",
        "tables": "concept_style_evidence",
        "severity": "moderate",
        "example": "We said 'case studies privileged' but evidence shows 'metrics/KPIs' are actually privileged"
    },
    {
        "type": "inference_pattern_different",
        "flags": "Actual reasoning moves/inference patterns differ from what we identified",
        "tables": "concept_style_inferences",
        "severity": "moderate",
        "example": "We said 'if competition then sovereignty needed' but actual inference is 'if others do X, we must too'"
    },
    {
        "type": "new_style_emerging",
        "flags": "Evidence shows emergence of a new reasoning style we didn't capture",
        "tables": "concept_reasoning_styles",
        "severity": "major",
        "example": "A new 'techno-civilizational' style is emerging that we didn't identify"
    },
    {
        "type": "objects_created_different",
        "flags": "The style creates different objects of inquiry than we identified",
        "tables": "concept_reasoning_styles, concept_kinds_created",
        "severity": "moderate",
        "example": "We said style creates 'strategic sectors' but evidence shows it creates 'technology gaps' as objects"
    },
    {
        "type": "looping_effect_detected",
        "flags": "Concept is looping back and changing the reality it describes (dynamic nominalism in action)",
        "tables": "concept_looping_effects",
        "severity": "major",
        "example": "Labeling countries as 'tech sovereign' is changing how those countries behave - classification creating reality"
    },
    {
        "type": "kind_created_or_destroyed",
        "flags": "The concept is 'making up' new kinds of things/people OR a kind we thought it created doesn't exist",
        "tables": "concept_kinds_created",
        "severity": "major",
        "example": "Concept has created new kind: 'tech-sovereign state' which didn't exist before - a Hacking-style creation"
    },
    {
        "type": "possibility_space_shifted",
        "flags": "The space of possibilities opened/closed by the concept differs from what we mapped",
        "tables": "concept_possibility_space",
        "severity": "major",
        "example": "We said concept opens 'industrial policy' space but evidence shows it actually closes 'international cooperation' space"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 9: Blumenberg Feedback Types
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
style_header(ws, 1, "BLUMENBERG FEEDBACK TYPES - Metaphorology & Nonconceptuality", dimension_colors["Blumenberg"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Paradigms for a Metaphorology' - absolute metaphors, nonconceptuality (Unbegrifflichkeit), lifeworld background, metakinetics").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Blumenberg", [
    {
        "type": "metaphor_missed",
        "flags": "Evidence reveals a root metaphor structuring the concept that we didn't identify",
        "tables": "concept_metaphors",
        "severity": "major",
        "example": "We missed 'sovereignty as immune system' metaphor - concept protecting from foreign 'infection'"
    },
    {
        "type": "metaphor_effect_different",
        "flags": "A metaphor enables/hides differently than we described",
        "tables": "concept_metaphor_effects",
        "severity": "moderate",
        "example": "We said territory metaphor 'hides networks' but evidence shows it also 'enables border thinking'"
    },
    {
        "type": "metakinesis_different",
        "flags": "Metaphor's historical transformation differs from what we mapped",
        "tables": "concept_metakinetics",
        "severity": "moderate",
        "example": "We mapped gradual shift but evidence shows sudden transformation after 2018 trade war"
    },
    {
        "type": "conceptual_work_failing",
        "flags": "The conceptual transformation work being attempted is not succeeding",
        "tables": "concept_work_in_progress",
        "severity": "major",
        "example": "Attempt to transform 'sovereignty' to encompass networks is failing - metaphor resists"
    },
    {
        "type": "nonconceptuality_revealed",
        "flags": "Evidence reveals an aspect of the concept that resists conceptualization (per Blumenberg's Unbegrifflichkeit)",
        "tables": "concept_nonconceptuality",
        "severity": "moderate",
        "example": "The 'feeling of autonomy' cannot be translated into policy terms - remains nonconceptual"
    },
    {
        "type": "absolutism_underestimated",
        "flags": "The metaphor is more 'absolute' (untranslatable to concepts) than we assessed",
        "tables": "concept_metaphors",
        "severity": "major",
        "example": "We said metaphor was 'instrumental' but evidence shows it's 'absolute' - can't be replaced"
    },
    {
        "type": "lifeworld_tension_exposed",
        "flags": "Tension between concept and underlying lifeworld experience is greater than we identified",
        "tables": "concept_lifeworld_connection",
        "severity": "moderate",
        "example": "Conceptual apparatus of 'sovereignty' is increasingly at odds with lived experience of global interdependence"
    },
    {
        "type": "unbegrifflichkeit_encountered",
        "flags": "There's an irreducible remainder that our conceptual schema cannot capture (per Blumenberg's nonconceptuality)",
        "tables": "concept_nonconceptuality",
        "severity": "major",
        "example": "Evidence points to something about 'tech sovereignty' that cannot be articulated in propositional form - pure affect/image"
    },
    {
        "type": "background_metaphor_shift",
        "flags": "The background metaphorics that make this concept thinkable have shifted",
        "tables": "concept_metaphors, concept_lifeworld_connection",
        "severity": "major",
        "example": "Background shifted from 'organic growth' metaphorics to 'immune defense' metaphorics - changes what sovereignty means"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 10: Carey Feedback Types
# ============================================================================
ws = wb.create_sheet("10. Carey")
style_header(ws, 1, "CAREY FEEDBACK TYPES - Bootstrapping & Conceptual Change", dimension_colors["Carey"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="Based on 'Origin of Concepts' - core cognition, Quinian bootstrapping, placeholder structures, discontinuity, computational constraints").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Carey", [
    {
        "type": "hierarchy_wrong",
        "flags": "Concept's hierarchy level is different than we assessed",
        "tables": "concept_hierarchy, concepts.hierarchy_level",
        "severity": "moderate",
        "example": "We said level 3 (complex) but evidence shows it's actually level 2 - simpler aggregation"
    },
    {
        "type": "component_missing",
        "flags": "A component concept we didn't identify is actually part of the bootstrap",
        "tables": "concept_built_from",
        "severity": "moderate",
        "example": "We missed 'legitimacy' as component concept needed to bootstrap 'tech sovereignty'"
    },
    {
        "type": "bootstrap_failed",
        "flags": "The combination of components doesn't achieve the claimed qualitative leap",
        "tables": "concept_hierarchy",
        "severity": "critical",
        "example": "Combining 'sovereignty' + 'technology' doesn't create new meaning - just awkward juxtaposition"
    },
    {
        "type": "incommensurability_revealed",
        "flags": "Concept cannot be reduced to or explained by its claimed components (Kuhnian incommensurability)",
        "tables": "concept_incommensurability",
        "severity": "major",
        "example": "'Tech sovereignty' has emergent properties that 'sovereignty' + 'technology' can't explain"
    },
    {
        "type": "mapping_different",
        "flags": "How the concept is acquired/learned differs from what we described",
        "tables": "concept_mapping_process",
        "severity": "moderate",
        "example": "We said 'extended mapping' but evidence shows 'fast mapping' from media - shallow uptake"
    },
    {
        "type": "core_cognition_wrong",
        "flags": "Our assessment of whether concept derives from core cognitive systems was wrong",
        "tables": "concept_core_cognition",
        "severity": "moderate",
        "example": "We said NO core cognition but 'in-group/out-group' core system may be involved"
    },
    {
        "type": "placeholder_structure_wrong",
        "flags": "The placeholder structure (partial concept awaiting elaboration) differs from what we identified",
        "tables": "concept_placeholder_structures",
        "severity": "moderate",
        "example": "We said 'slot for implementation' but evidence shows placeholder is 'slot for legitimate authority'"
    },
    {
        "type": "core_system_engagement_missed",
        "flags": "Concept engages with a core cognitive system we didn't identify",
        "tables": "concept_core_cognition",
        "severity": "major",
        "example": "We missed engagement with 'core object' system - sovereignty conceptualized as bounded thing"
    },
    {
        "type": "discontinuity_detected",
        "flags": "Evidence reveals conceptual discontinuity - qualitative leap we didn't map (per Carey's anti-continuity thesis)",
        "tables": "concept_incommensurability",
        "severity": "major",
        "example": "The way new actors use 'tech sovereignty' is genuinely discontinuous with older usage - not mere extension"
    },
    {
        "type": "representational_resource_insufficient",
        "flags": "The representational resources we attributed to concept learners are insufficient for this concept",
        "tables": "concept_placeholder_structures, concept_built_from",
        "severity": "moderate",
        "example": "Actors lack the background concepts needed to fully grasp 'tech sovereignty' - understanding is necessarily partial"
    },
    {
        "type": "computational_constraint_violated",
        "flags": "The concept requires cognitive operations that exceed specified constraints",
        "tables": "concept_bootstrapping_constraints",
        "severity": "moderate",
        "example": "Understanding this concept requires holding too many interdependencies in mind simultaneously - exceeds working memory"
    },
    {
        "type": "quinian_bootstrap_stalled",
        "flags": "The bootstrapping process (placeholder → elaboration → integration) has stalled at a stage",
        "tables": "concept_placeholder_structures, concept_bootstrapping_constraints",
        "severity": "major",
        "example": "Concept stuck at placeholder stage - everyone uses term but no one has elaborated internal structure"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 11: Feedback Database Schema
# ============================================================================
ws = wb.create_sheet("11. Database Schema")
style_header(ws, 1, "DIMENSION-SPECIFIC FEEDBACK - DATABASE SCHEMA v2", dimension_colors["Overview"])

row = 3
ws.cell(row=row, column=1, value="Master feedback table with dimension-specific type enum (81 types)").font = Font(bold=True, size=12)
row += 2

# Main table
headers = ["Field", "Type", "Description"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = table_header_fill
row += 1

fields = [
    ("id", "SERIAL PRIMARY KEY", ""),
    ("evidence_test_id", "INTEGER FK", "Link to evidence_tests"),
    ("dimension", "VARCHAR(20)", "quinean/sellarsian/brandomian/deleuzian/bachelardian/canguilhem/hacking/blumenberg/carey"),
    ("feedback_type", "VARCHAR(50)", "One of 81 dimension-specific types (see enum below)"),
    ("severity", "VARCHAR(20)", "critical/major/moderate/minor"),
    ("summary", "TEXT NOT NULL", "Brief description of finding"),
    ("details", "TEXT", "Full explanation with philosophical grounding"),
    ("affected_table", "VARCHAR(100)", "Which concept schema table is affected"),
    ("affected_row_id", "INTEGER", "ID of specific row if applicable"),
    ("evidence_basis", "TEXT", "What evidence supports this finding"),
    ("philosophical_significance", "TEXT", "Why this matters theoretically (per the dimension's framework)"),
    ("suggested_action", "TEXT", "What revision might address this"),
    ("confidence", "FLOAT", "LLM confidence 0-1"),
    ("status", "VARCHAR(20)", "pending/reviewed/actioned/dismissed"),
    ("source_type", "VARCHAR(30)", "llm_analysis/evidence_testing/internal_compute/user_input"),
    ("source_reference", "TEXT", "Reference to source (model, cluster id, etc.)"),
    ("created_at", "TIMESTAMPTZ", ""),
]

for f in fields:
    ws.cell(row=row, column=1, value=f[0]).font = field_font
    ws.cell(row=row, column=2, value=f[1]).fill = type_fill
    ws.cell(row=row, column=3, value=f[2])
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

# Enum definition
row += 2
ws.cell(row=row, column=1, value="FEEDBACK_TYPE ENUM (81 types by dimension):").font = Font(bold=True, size=12)
row += 2

enums = [
    ("quinean (9)", "inference_broken, inference_missing, centrality_wrong, entrenchment_overestimated, web_inconsistency, holism_ripple, web_position_wrong, ontological_dependence_missed, revision_cost_miscalculated"),
    ("sellarsian (9)", "false_given_exposed, hidden_commitment_revealed, givenness_marker_challenged, effect_reversed, space_of_reasons_violation, manifest_scientific_gap, functional_role_misspecified, inferential_articulation_wrong, image_tension_shifted"),
    ("brandomian (7)", "commitment_violated, entitlement_exceeded, inference_defeated, incompatibility_missed, challenge_unanswered, perspectival_distortion, deontic_status_wrong"),
    ("deleuzian (8)", "component_missing, component_extraneous, zone_mismapped, consistency_weaker, neighborhood_wrong, plane_misidentified, persona_different, problem_transformed"),
    ("bachelardian (9)", "new_obstacle_discovered, obstacle_deeper, blocked_understanding_unblocked, rupture_imminent, rupture_false, psychoanalytic_function_different, stage_mismatch, regional_rationality_wrong, obstacle_type_misidentified"),
    ("canguilhem (9)", "evolution_different, normative_dimension_exposed, vitality_declining, vitality_reviving, filiation_broken, normal_normative_confusion, milieu_changed, vital_norm_violated, concept_migration_detected"),
    ("hacking (9)", "style_mismatch, visibility_wrong, evidence_privileged_wrong, inference_pattern_different, new_style_emerging, objects_created_different, looping_effect_detected, kind_created_or_destroyed, possibility_space_shifted"),
    ("blumenberg (9)", "metaphor_missed, metaphor_effect_different, metakinesis_different, conceptual_work_failing, nonconceptuality_revealed, absolutism_underestimated, lifeworld_tension_exposed, unbegrifflichkeit_encountered, background_metaphor_shift"),
    ("carey (12)", "hierarchy_wrong, component_missing, bootstrap_failed, incommensurability_revealed, mapping_different, core_cognition_wrong, placeholder_structure_wrong, core_system_engagement_missed, discontinuity_detected, representational_resource_insufficient, computational_constraint_violated, quinian_bootstrap_stalled"),
]

for dim, types in enums:
    ws.cell(row=row, column=1, value=dim).font = Font(bold=True)
    ws.cell(row=row, column=2, value=types).alignment = wrap
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 100
ws.column_dimensions['C'].width = 50

# ============================================================================
# SHEET 12: Summary Statistics
# ============================================================================
ws = wb.create_sheet("12. Summary")
style_header(ws, 1, "FEEDBACK SCHEMA v2 SUMMARY", dimension_colors["Overview"])

row = 3
ws.cell(row=row, column=1, value="FEEDBACK TYPE COUNTS BY DIMENSION").font = Font(bold=True, size=12)
row += 2

headers = ["Dimension", "Type Count", "Key Focus Areas"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = table_header_fill
row += 1

summaries = [
    ("Quinean", 9, "Web position, ontological dependence, revision ramifications"),
    ("Sellarsian", 9, "Space of reasons, functional role, image tension"),
    ("Brandomian", 7, "Deontic scorekeeping, inferential roles, perspectival content"),
    ("Deleuzian", 8, "Components, consistency, plane, personae"),
    ("Bachelardian", 9, "Obstacles, stages, psychoanalytic function, regional rationality"),
    ("Canguilhem", 9, "Filiation, vital norms, milieu, normative dimensions"),
    ("Hacking", 9, "Looping effects, kinds created, possibility space, styles"),
    ("Blumenberg", 9, "Metaphors, nonconceptuality, lifeworld, metakinetics"),
    ("Carey", 12, "Core cognition, placeholders, bootstrapping, incommensurability"),
    ("TOTAL", 81, "Aligned with v5 concept schema (47 tables)"),
]

for dim, count, focus in summaries:
    ws.cell(row=row, column=1, value=dim)
    if dim == "TOTAL":
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2, value=count)
    ws.cell(row=row, column=3, value=focus)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 60

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Dimension_Feedback_Schema_v2.xlsx")
wb.save(output_path)
print(f"Dimension-specific feedback schema v2 saved to: {output_path}")
print(f"Total feedback types: 81 (aligned with v5 concept schema)")
