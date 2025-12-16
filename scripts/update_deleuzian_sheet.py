#!/usr/bin/env python3
"""
Update the Deleuzian sheet in the modular schema Excel
with corrected concept theory from "What is Philosophy?"
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Load existing workbook
wb_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D_Modular_v2.xlsx")
wb = load_workbook(wb_path)

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
new_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
removed_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
deleuzian_color = "1ABC9C"

# Delete old Deleuzian sheet and create new one
if "5. Deleuzian" in wb.sheetnames:
    del wb["5. Deleuzian"]

ws = wb.create_sheet("5. Deleuzian", 5)

# Title
ws.merge_cells('A1:D1')
cell = ws.cell(row=1, column=1, value="DELEUZIAN DIMENSION - Concept Theory (Revised)")
cell.font = Font(bold=True, size=14, color="FFFFFF")
cell.fill = PatternFill(start_color=deleuzian_color, end_color=deleuzian_color, fill_type="solid")

# Subtitle
ws.merge_cells('A2:D2')
ws.cell(row=2, column=1, value="Based on 'What is Philosophy?' (1991) - Deleuze's actual theory of concepts").font = Font(italic=True)

# Note about revision
ws.merge_cells('A4:D4')
note = ws.cell(row=4, column=1, value="NOTE: Previous version mixed in concepts from 'A Thousand Plateaus' (becomings, lines of flight, deterritorialization) which are about desire/social machines, not concepts per se. This revision focuses on Deleuze's theory OF concepts.")
note.font = Font(italic=True, color="C0392B")
note.alignment = wrap

def add_table(start_row, table_name, fields, note=None):
    row = start_row
    if note:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9)
        row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value=table_name).font = Font(bold=True, size=12)
    row += 1

    headers = ["Field", "Type", "Constraints", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for f in fields:
        ws.cell(row=row, column=1, value=f["name"]).font = field_font
        ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
        c3 = ws.cell(row=row, column=3, value=f.get("constraints", ""))
        if "FK" in f.get("constraints", ""):
            c3.fill = fk_fill
        ws.cell(row=row, column=4, value=f["desc"])
        if f.get("new"):
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = new_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1
    return row + 1

# Table 1: concept_components (NEW)
row = add_table(6, "concept_components (NEW - individual components of concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "component_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Name of component (e.g., 'doubting', 'thinking')", "new": True},
    {"name": "component_description", "type": "TEXT", "constraints": "", "desc": "What this component contributes to concept", "new": True},
    {"name": "is_intensive", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Is this an intensive variation?", "new": True},
    {"name": "order_index", "type": "INTEGER", "constraints": "", "desc": "Position in concept's structure", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From WIP: Every concept is a multiplicity of heterogeneous components")

# Example
ws.cell(row=row, column=1, value="Example for 'Technological Sovereignty':").font = Font(bold=True)
row += 1
examples = [
    "1. 'Sovereignty' - supreme authority, non-interference",
    "2. 'Technology' - productive capacity, innovation systems",
    "3. 'State' - political entity exercising control",
    "4. 'Territory' - bounded space of jurisdiction (problematic transfer)",
]
for ex in examples:
    ws.cell(row=row, column=1, value=ex)
    row += 1
row += 1

# Table 2: concept_zones_of_indiscernibility (NEW)
row = add_table(row, "concept_zones_of_indiscernibility (NEW - where components overlap)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "component_a_id", "type": "INTEGER", "constraints": "FK → concept_components", "desc": "First component", "new": True},
    {"name": "component_b_id", "type": "INTEGER", "constraints": "FK → concept_components", "desc": "Second component", "new": True},
    {"name": "zone_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What passes between components in this zone", "new": True},
    {"name": "what_becomes_possible", "type": "TEXT", "constraints": "", "desc": "What this zone enables", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From WIP: Components have zones of indiscernibility - loci of becoming within concepts")

# Table 3: concept_consistency (NEW)
row = add_table(row, "concept_consistency (NEW - endoconsistency analysis)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)", "new": True},
    {"name": "endoconsistency_description", "type": "TEXT", "constraints": "", "desc": "How components hold together internally", "new": True},
    {"name": "survey_point", "type": "TEXT", "constraints": "", "desc": "The 'point of absolute survey' that unifies", "new": True},
    {"name": "consistency_strength", "type": "VARCHAR(20)", "constraints": "strong/moderate/weak/unstable", "desc": "How well does concept cohere?", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From WIP: Endoconsistency = internal coherence of concept's components")

# Table 4: concept_neighborhood (NEW - exoconsistency)
row = add_table(row, "concept_neighborhood (NEW - exoconsistency / relations to other concepts)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "neighbor_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept (if in DB)", "new": True},
    {"name": "neighbor_description", "type": "TEXT", "constraints": "", "desc": "Description if not in DB", "new": True},
    {"name": "relation_type", "type": "VARCHAR(30)", "constraints": "", "desc": "bridge/resonance/interference/repulsion", "new": True},
    {"name": "neighborhood_order", "type": "INTEGER", "constraints": "", "desc": "How close (1=closest neighbor)", "new": True},
    {"name": "bridges_across_plane", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Does this bridge different planes?", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From WIP: Exoconsistency = concept's neighborhood, bridges to other concepts")

# Table 5: concept_plane_of_immanence (REVISED)
row = add_table(row, "concept_plane_of_immanence (REVISED - the ground on which concept operates)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "plane_name", "type": "VARCHAR(100)", "constraints": "", "desc": "Name/description of the plane"},
    {"name": "what_plane_presupposes", "type": "TEXT", "constraints": "", "desc": "Unthought assumptions of this plane", "new": True},
    {"name": "legitimate_problems", "type": "TEXT", "constraints": "", "desc": "What counts as a problem on this plane", "new": True},
    {"name": "excluded_problems", "type": "TEXT", "constraints": "", "desc": "What can't be asked on this plane", "new": True},
    {"name": "historical_emergence", "type": "VARCHAR(100)", "constraints": "", "desc": "When this plane emerged"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="From WIP: Plane of immanence = 'the absolute ground of philosophy, its earth'")

# Table 6: concept_personae (NEW)
row = add_table(row, "concept_personae (NEW - conceptual personae that activate the concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "persona_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Name of persona (e.g., 'the sovereign')", "new": True},
    {"name": "persona_description", "type": "TEXT", "constraints": "", "desc": "What this persona does/thinks", "new": True},
    {"name": "what_persona_enables", "type": "TEXT", "constraints": "", "desc": "What thinking this persona enables", "new": True},
    {"name": "historical_origin", "type": "TEXT", "constraints": "", "desc": "Where this persona comes from", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From WIP: Conceptual personae are the 'third element' of philosophy alongside plane and concepts")

# Table 7: concept_problems (REVISED)
row = add_table(row, "concept_problems (REVISED - problems the concept addresses)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "problem_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The problem being addressed"},
    {"name": "problem_type", "type": "VARCHAR(30)", "constraints": "", "desc": "political/epistemological/ontological/practical"},
    {"name": "triggering_event", "type": "TEXT", "constraints": "", "desc": "What event made this problem pressing", "new": True},
    {"name": "how_concept_responds", "type": "TEXT", "constraints": "", "desc": "How the concept addresses this problem", "new": True},
    {"name": "problem_transformed_to", "type": "TEXT", "constraints": "", "desc": "How problem changes through concept's use", "new": True},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="From WIP: Every event is a critical point in a problem; concepts are 'constellations of events to come'")

# Note about removed tables
row += 1
ws.merge_cells(f'A{row}:D{row}')
ws.cell(row=row, column=1, value="REMOVED (from A Thousand Plateaus, not concept theory):").font = Font(bold=True, color="C0392B")
row += 1

removed = [
    "concept_becomings - 'Becomings' is from desire theory, not concept theory",
    "concept_lines_of_flight - From rhizomatics/territorial analysis",
    "concept_creative_responses - Replaced by more precise 'how_concept_responds'",
    "concept_plane_assumptions - Merged into concept_plane_of_immanence.what_plane_presupposes",
    "concept_plane_effects - Merged into legitimate_problems/excluded_problems",
]
for item in removed:
    cell = ws.cell(row=row, column=1, value=f"  ✗ {item}")
    cell.fill = removed_fill
    row += 1

# Column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 55

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D_Modular_v3.xlsx")
wb.save(output_path)
print(f"Updated schema saved to: {output_path}")
