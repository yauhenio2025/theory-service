#!/usr/bin/env python3
"""
Generate DIMENSION-SPECIFIC Evidence Feedback Schema

Each dimension has its own feedback types derived from what that framework
would specifically flag when testing concepts against evidence.

This is NOT generic "tension/confirmation/extension" - these are
philosophically grounded feedback types specific to each thinker's framework.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Overview": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Davidson": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

def style_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def add_feedback_types(ws, start_row, dimension, types):
    """Add feedback types for a dimension."""
    row = start_row

    headers = ["Feedback Type", "What It Flags", "Tables Affected", "Severity", "Example"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for t in types:
        ws.cell(row=row, column=1, value=t["type"]).font = field_font
        ws.cell(row=row, column=2, value=t["flags"])
        ws.cell(row=row, column=3, value=t["tables"])
        ws.cell(row=row, column=4, value=t["severity"])
        ws.cell(row=row, column=5, value=t["example"])
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    return row + 1

def set_column_widths(ws):
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50

# ============================================================================
# SHEET 1: Overview
# ============================================================================
ws = wb.active
ws.title = "1. Overview"

style_header(ws, 1, "DIMENSION-SPECIFIC EVIDENCE FEEDBACK SCHEMA", dimension_colors["Overview"])

overview = [
    "",
    "PURPOSE:",
    "When testing concept values against evidence clusters, each philosophical dimension",
    "has its OWN specific feedback types - derived from what that framework would flag.",
    "",
    "NOT generic 'tension/confirmation/extension' - these are PHILOSOPHICALLY GROUNDED",
    "feedback types specific to each thinker's theoretical framework.",
    "",
    "HOW IT WORKS:",
    "1. Evidence clusters are tested against specific tables in specific dimensions",
    "2. LLM analyzes using that dimension's theoretical framework",
    "3. Feedback is categorized using that dimension's feedback types",
    "4. Editorial decisions are informed by the philosophical significance",
    "",
    "FEEDBACK TYPES BY DIMENSION:",
    "",
    "QUINEAN (6 types): inference_broken, inference_missing, centrality_wrong,",
    "                   entrenchment_overestimated, web_inconsistency, holism_ripple",
    "",
    "SELLARSIAN (6 types): false_given_exposed, hidden_commitment_revealed,",
    "                      givenness_marker_challenged, effect_reversed,",
    "                      space_of_reasons_violation, manifest_scientific_gap",
    "",
    "BRANDOMIAN (7 types): commitment_violated, entitlement_exceeded, inference_defeated,",
    "                      incompatibility_missed, challenge_unanswered,",
    "                      perspectival_distortion, deontic_status_wrong",
    "",
    "DELEUZIAN (8 types): component_missing, component_extraneous, zone_mismapped,",
    "                     consistency_weaker, neighborhood_wrong, plane_misidentified,",
    "                     persona_different, problem_transformed",
    "",
    "BACHELARDIAN (6 types): new_obstacle_discovered, obstacle_deeper,",
    "                        blocked_understanding_unblocked, rupture_imminent,",
    "                        rupture_false, psychoanalytic_function_different",
    "",
    "CANGUILHEM (6 types): evolution_different, normative_dimension_exposed,",
    "                      vitality_declining, vitality_reviving, filiation_broken,",
    "                      normal_normative_confusion",
    "",
    "DAVIDSON/HACKING (6 types): style_mismatch, visibility_wrong, evidence_privileged_wrong,",
    "                            inference_pattern_different, new_style_emerging,",
    "                            objects_created_different",
    "",
    "BLUMENBERG (6 types): metaphor_missed, metaphor_effect_different, metakinesis_different,",
    "                      conceptual_work_failing, nonconceptuality_revealed,",
    "                      absolutism_underestimated",
    "",
    "CAREY (6 types): hierarchy_wrong, component_missing, bootstrap_failed,",
    "                 incommensurability_revealed, mapping_different, core_cognition_wrong",
    "",
    "TOTAL: 57 dimension-specific feedback types",
]

for row_idx, text in enumerate(overview, 3):
    ws.cell(row=row_idx, column=1, value=text)
    if text.startswith("PURPOSE") or text.startswith("HOW IT") or text.startswith("FEEDBACK"):
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

set_column_widths(ws)

# ============================================================================
# SHEET 2: Quinean Feedback Types
# ============================================================================
ws = wb.create_sheet("2. Quinean")
style_header(ws, 1, "QUINEAN FEEDBACK TYPES - Web of Belief", dimension_colors["Quinean"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Quine's web of belief framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Quinean", [
    {
        "type": "inference_broken",
        "flags": "A claimed inferential connection doesn't hold - evidence shows 'If A then B' fails",
        "tables": "concept_inferences",
        "severity": "major",
        "example": "We claim 'tech sovereignty → indigenous development' but Gulf states have $2T and still can't do Series B"
    },
    {
        "type": "inference_missing",
        "flags": "Evidence reveals an important inferential connection we didn't capture",
        "tables": "concept_inferences",
        "severity": "minor",
        "example": "We missed 'tech sovereignty → need for protected procurement' which evidence shows is central"
    },
    {
        "type": "centrality_wrong",
        "flags": "Concept is more/less central to the web than we assessed",
        "tables": "concepts.centrality",
        "severity": "moderate",
        "example": "We marked 'technological sovereignty' as core but evidence shows it's intermediate - can be revised without major ripples"
    },
    {
        "type": "entrenchment_overestimated",
        "flags": "Concept is easier to revise than our entrenchment score suggested",
        "tables": "concepts.entrenchment_score",
        "severity": "moderate",
        "example": "Entrenchment was 0.9 but evidence shows actors readily abandon sovereignty claims when convenient"
    },
    {
        "type": "web_inconsistency",
        "flags": "The claimed inferential connections create contradictions within the web",
        "tables": "concept_web_tensions",
        "severity": "critical",
        "example": "We claim both 'sovereignty → autonomy' AND 'sovereignty → global integration' but these contradict"
    },
    {
        "type": "holism_ripple",
        "flags": "Change to this concept would force changes to more concepts than we predicted",
        "tables": "concept_inferences, concept_web_tensions",
        "severity": "major",
        "example": "Revising 'tech sovereignty' would force changes to 8 related concepts, not 3 as we thought"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 3: Sellarsian Feedback Types
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
style_header(ws, 1, "SELLARSIAN FEEDBACK TYPES - Myth of the Given", dimension_colors["Sellarsian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Sellars' critique of the given, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Sellarsian", [
    {
        "type": "false_given_exposed",
        "flags": "What we treated as foundational/self-evident is actually inferred from other premises",
        "tables": "concept_givenness",
        "severity": "critical",
        "example": "We marked 'sovereignty is achievable' as given, but evidence shows it's inferred from contested premises about state capacity"
    },
    {
        "type": "hidden_commitment_revealed",
        "flags": "Evidence exposes a baked-in assumption we didn't identify",
        "tables": "concept_hidden_commitments",
        "severity": "major",
        "example": "Evidence reveals hidden commitment: 'tech development is linear/predictable' which we didn't list"
    },
    {
        "type": "givenness_marker_challenged",
        "flags": "A rhetorical marker of givenness ('obviously', 'naturally') is directly contested by evidence",
        "tables": "concept_givenness_markers",
        "severity": "moderate",
        "example": "We noted 'naturally extends state power' but evidence shows tech often undermines state power"
    },
    {
        "type": "effect_reversed",
        "flags": "What we thought givenness enables is actually blocked, or vice versa",
        "tables": "concept_givenness_effects",
        "severity": "major",
        "example": "We said givenness blocks 'analysis of impossibility' but evidence shows such analysis IS happening"
    },
    {
        "type": "space_of_reasons_violation",
        "flags": "Concept is used outside proper justificatory context - not properly in the space of reasons",
        "tables": "concept_givenness",
        "severity": "moderate",
        "example": "Concept invoked without any inferential support - pure assertion masquerading as justified claim"
    },
    {
        "type": "manifest_scientific_gap",
        "flags": "Bigger gap between manifest (everyday) and scientific image than we identified",
        "tables": "concept_givenness",
        "severity": "minor",
        "example": "Everyday use of 'sovereignty' diverges much more from technical IR usage than we noted"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 4: Brandomian Feedback Types
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
style_header(ws, 1, "BRANDOMIAN FEEDBACK TYPES - Scorekeeping & Reasons", dimension_colors["Brandomian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Brandom's inferentialist framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Brandomian", [
    {
        "type": "commitment_violated",
        "flags": "A claimed commitment is not honored in practice",
        "tables": "concept_commitments",
        "severity": "critical",
        "example": "Commitment: 'investment → autonomy' but Gulf $2T can't invest at Series B level - commitment violated"
    },
    {
        "type": "entitlement_exceeded",
        "flags": "Users claim more than they're entitled to based on the concept",
        "tables": "concept_commitments",
        "severity": "major",
        "example": "States claim entitlement to 'sovereignty success' based on rhetoric alone - exceeding inferential warrant"
    },
    {
        "type": "inference_defeated",
        "flags": "Counterexample defeats a claimed material inference (committive, permissive, or incompatibility)",
        "tables": "concept_inferential_roles",
        "severity": "critical",
        "example": "Committive inference 'tech sovereignty → right to exclude' defeated by EU Chips Act creating new dependencies"
    },
    {
        "type": "incompatibility_missed",
        "flags": "Evidence reveals an incompatibility relation we didn't capture",
        "tables": "concept_inferential_roles",
        "severity": "major",
        "example": "We missed: 'tech sovereignty' INCOMPATIBLE WITH 'participation in global supply chains'"
    },
    {
        "type": "challenge_unanswered",
        "flags": "A challenge to a commitment cannot be justified - no adequate reasons available",
        "tables": "concept_challenges, concept_justifications",
        "severity": "critical",
        "example": "Challenge: 'What entitles the claim investment → autonomy?' No justification in evidence cluster"
    },
    {
        "type": "perspectival_distortion",
        "flags": "De re translation doesn't preserve truth - our substitution changes meaning",
        "tables": "concept_perspectival_content",
        "severity": "moderate",
        "example": "Gulf says 'achieving sovereignty' but our translation 'dependency management' loses their self-understanding"
    },
    {
        "type": "deontic_status_wrong",
        "flags": "Something is attributed but not acknowledged, or undertaken but not attributed, etc.",
        "tables": "concept_commitments",
        "severity": "moderate",
        "example": "We marked commitment as 'acknowledged' but evidence shows actors only attribute it to others, not themselves"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 5: Deleuzian Feedback Types
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
style_header(ws, 1, "DELEUZIAN FEEDBACK TYPES - Concept Theory", dimension_colors["Deleuzian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Deleuze's concept theory (from 'What is Philosophy?'), these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Deleuzian", [
    {
        "type": "component_missing",
        "flags": "Evidence reveals a component of the concept we didn't identify",
        "tables": "concept_components",
        "severity": "major",
        "example": "We missed 'temporality' as component - sovereignty implies urgency/limited time window"
    },
    {
        "type": "component_extraneous",
        "flags": "A component we identified doesn't actually belong to this concept's multiplicity",
        "tables": "concept_components",
        "severity": "moderate",
        "example": "We listed 'innovation' as component but evidence shows it's separate from sovereignty proper"
    },
    {
        "type": "zone_mismapped",
        "flags": "Zone of indiscernibility between components doesn't function as we described",
        "tables": "concept_zones_of_indiscernibility",
        "severity": "moderate",
        "example": "Sovereignty-technology zone: we said tension, but evidence shows they're more integrated than we thought"
    },
    {
        "type": "consistency_weaker",
        "flags": "Endoconsistency (internal coherence) is weaker than we assessed",
        "tables": "concept_consistency",
        "severity": "major",
        "example": "We said 'moderate' consistency but evidence shows components are barely held together - 'unstable'"
    },
    {
        "type": "neighborhood_wrong",
        "flags": "Concept's relations to other concepts differ from what we mapped",
        "tables": "concept_neighborhood",
        "severity": "moderate",
        "example": "We said 'economic sovereignty' is closest neighbor but evidence shows 'digital sovereignty' is closer"
    },
    {
        "type": "plane_misidentified",
        "flags": "Concept operates on a different plane of immanence than we identified",
        "tables": "concept_plane_of_immanence",
        "severity": "critical",
        "example": "We said Westphalian plane but evidence shows concept now operates on 'techno-nationalist' plane"
    },
    {
        "type": "persona_different",
        "flags": "Different conceptual persona activates the concept than we identified",
        "tables": "concept_personae",
        "severity": "moderate",
        "example": "We said 'the sovereign' but evidence shows 'the technocrat' is the activating persona"
    },
    {
        "type": "problem_transformed",
        "flags": "The problem the concept addresses has transformed in ways we didn't capture",
        "tables": "concept_problems",
        "severity": "major",
        "example": "Problem shifted from 'how to achieve autonomy' to 'how to claim autonomy while managing dependency'"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 6: Bachelardian Feedback Types
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
style_header(ws, 1, "BACHELARDIAN FEEDBACK TYPES - Epistemological Obstacles", dimension_colors["Bachelardian"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Bachelard's epistemological obstacle framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Bachelardian", [
    {
        "type": "new_obstacle_discovered",
        "flags": "Evidence reveals an epistemological obstacle function we didn't identify",
        "tables": "concept_obstacles",
        "severity": "major",
        "example": "We didn't identify 'techno-nationalism' as obstacle, but evidence shows it blocks recognition of interdependence"
    },
    {
        "type": "obstacle_deeper",
        "flags": "The obstacle is more entrenched than we assessed",
        "tables": "concept_obstacles",
        "severity": "major",
        "example": "We thought obstacle was 'verbal' (language-based) but evidence shows it's 'substantialist' - deeper resistance"
    },
    {
        "type": "blocked_understanding_unblocked",
        "flags": "Something we thought was blocked by the concept is actually being understood",
        "tables": "concept_obstacle_blocks",
        "severity": "moderate",
        "example": "We said concept blocks 'analysis of impossibility' but evidence shows scholars ARE doing this analysis"
    },
    {
        "type": "rupture_imminent",
        "flags": "Evidence suggests conditions for epistemological rupture are forming",
        "tables": "concept_obstacles",
        "severity": "major",
        "example": "Mounting failures (China chips, Russia sanctions) creating conditions for rupture with sovereignty concept"
    },
    {
        "type": "rupture_false",
        "flags": "A claimed rupture or breakthrough hasn't actually occurred",
        "tables": "concept_obstacles",
        "severity": "moderate",
        "example": "We noted a rupture point but evidence shows the old obstacle thinking persists"
    },
    {
        "type": "psychoanalytic_function_different",
        "flags": "The concept serves a different unconscious need than we identified",
        "tables": "concept_obstacles",
        "severity": "moderate",
        "example": "We said need for 'control fantasy' but evidence suggests need for 'legitimation of spending'"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 7: Canguilhem Feedback Types
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
style_header(ws, 1, "CANGUILHEM FEEDBACK TYPES - Life History of Concepts", dimension_colors["Canguilhem"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Canguilhem's concept history framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Canguilhem", [
    {
        "type": "evolution_different",
        "flags": "Historical trajectory differs from the evolution we mapped",
        "tables": "concept_evolution",
        "severity": "moderate",
        "example": "We mapped linear evolution but evidence shows concept had a 'death' in 1990s before 2010s revival"
    },
    {
        "type": "normative_dimension_exposed",
        "flags": "Evidence reveals a hidden normative dimension (embedded value) we didn't identify",
        "tables": "concept_normative_dimensions",
        "severity": "major",
        "example": "We missed: 'sovereignty' embeds value that state action is inherently legitimate"
    },
    {
        "type": "vitality_declining",
        "flags": "Concept's health/vitality is worse than we assessed",
        "tables": "concept_vitality_indicators, concepts.health_status",
        "severity": "major",
        "example": "We said 'healthy' but evidence shows repeated failures indicating 'strained' status"
    },
    {
        "type": "vitality_reviving",
        "flags": "Concept shows unexpected signs of life/revival",
        "tables": "concept_vitality_indicators, concepts.health_status",
        "severity": "moderate",
        "example": "We said 'dying' but evidence shows new actors (EU, India) giving concept new life"
    },
    {
        "type": "filiation_broken",
        "flags": "Claimed conceptual lineage (filiation) doesn't hold",
        "tables": "concept_evolution",
        "severity": "moderate",
        "example": "We claimed descent from 'Westphalian sovereignty' but evidence shows different conceptual origin"
    },
    {
        "type": "normal_normative_confusion",
        "flags": "What we labeled as descriptive ('normal') is actually prescriptive ('normative')",
        "tables": "concept_normative_dimensions",
        "severity": "major",
        "example": "We treated 'states pursue sovereignty' as descriptive fact but it's actually a normative prescription"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 8: Davidson/Hacking Feedback Types
# ============================================================================
ws = wb.create_sheet("8. Davidson-Hacking")
style_header(ws, 1, "DAVIDSON/HACKING FEEDBACK TYPES - Styles of Reasoning", dimension_colors["Davidson"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Hacking's styles of reasoning framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Davidson/Hacking", [
    {
        "type": "style_mismatch",
        "flags": "Concept requires a different reasoning style than we identified",
        "tables": "concept_reasoning_styles",
        "severity": "major",
        "example": "We said 'geopolitical strategic' style but evidence shows concept operates via 'economic modeling' style"
    },
    {
        "type": "visibility_wrong",
        "flags": "What we thought the style makes visible/invisible is reversed",
        "tables": "concept_style_visibility",
        "severity": "moderate",
        "example": "We said style makes 'corporate interests invisible' but evidence shows they're actually quite visible"
    },
    {
        "type": "evidence_privileged_wrong",
        "flags": "Different evidence types are privileged/marginalized than we identified",
        "tables": "concept_style_evidence",
        "severity": "moderate",
        "example": "We said 'case studies privileged' but evidence shows 'metrics/KPIs' are actually privileged"
    },
    {
        "type": "inference_pattern_different",
        "flags": "Actual reasoning moves/inference patterns differ from what we identified",
        "tables": "concept_style_inferences",
        "severity": "moderate",
        "example": "We said 'if competition then sovereignty needed' but actual inference is 'if others do X, we must too'"
    },
    {
        "type": "new_style_emerging",
        "flags": "Evidence shows emergence of a new reasoning style we didn't capture",
        "tables": "concept_reasoning_styles",
        "severity": "major",
        "example": "A new 'techno-civilizational' style is emerging that we didn't identify"
    },
    {
        "type": "objects_created_different",
        "flags": "The style creates different objects of inquiry than we identified",
        "tables": "concept_reasoning_styles",
        "severity": "moderate",
        "example": "We said style creates 'strategic sectors' but evidence shows it creates 'technology gaps' as objects"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 9: Blumenberg Feedback Types
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
style_header(ws, 1, "BLUMENBERG FEEDBACK TYPES - Metaphorology", dimension_colors["Blumenberg"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Blumenberg's metaphorology framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Blumenberg", [
    {
        "type": "metaphor_missed",
        "flags": "Evidence reveals a root metaphor structuring the concept that we didn't identify",
        "tables": "concept_metaphors",
        "severity": "major",
        "example": "We missed 'sovereignty as immune system' metaphor - concept protecting from foreign 'infection'"
    },
    {
        "type": "metaphor_effect_different",
        "flags": "A metaphor enables/hides differently than we described",
        "tables": "concept_metaphor_effects",
        "severity": "moderate",
        "example": "We said territory metaphor 'hides networks' but evidence shows it also 'enables border thinking'"
    },
    {
        "type": "metakinesis_different",
        "flags": "Metaphor's historical transformation differs from what we mapped",
        "tables": "concept_metakinetics",
        "severity": "moderate",
        "example": "We mapped gradual shift but evidence shows sudden transformation after 2018 trade war"
    },
    {
        "type": "conceptual_work_failing",
        "flags": "The conceptual transformation work being attempted is not succeeding",
        "tables": "concept_work_in_progress",
        "severity": "major",
        "example": "Attempt to transform 'sovereignty' to encompass networks is failing - metaphor resists"
    },
    {
        "type": "nonconceptuality_revealed",
        "flags": "Evidence reveals an aspect of the concept that resists conceptualization",
        "tables": "concept_metaphors",
        "severity": "moderate",
        "example": "The 'feeling of autonomy' cannot be translated into policy terms - remains nonconceptual"
    },
    {
        "type": "absolutism_underestimated",
        "flags": "The metaphor is more 'absolute' (untranslatable to concepts) than we assessed",
        "tables": "concept_metaphors",
        "severity": "major",
        "example": "We said metaphor was 'instrumental' but evidence shows it's 'absolute' - can't be replaced"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 10: Carey Feedback Types
# ============================================================================
ws = wb.create_sheet("10. Carey")
style_header(ws, 1, "CAREY FEEDBACK TYPES - Conceptual Bootstrapping", dimension_colors["Carey"])

ws.merge_cells('A3:E3')
ws.cell(row=3, column=1, value="When evidence tests concepts against Carey's bootstrapping framework, these are the specific issues it can flag:").font = Font(italic=True)

row = add_feedback_types(ws, 5, "Carey", [
    {
        "type": "hierarchy_wrong",
        "flags": "Concept's hierarchy level is different than we assessed",
        "tables": "concept_hierarchy, concepts.hierarchy_level",
        "severity": "moderate",
        "example": "We said level 3 (complex) but evidence shows it's actually level 2 - simpler aggregation"
    },
    {
        "type": "component_missing",
        "flags": "A component concept we didn't identify is actually part of the bootstrap",
        "tables": "concept_built_from",
        "severity": "moderate",
        "example": "We missed 'legitimacy' as component concept needed to bootstrap 'tech sovereignty'"
    },
    {
        "type": "bootstrap_failed",
        "flags": "The combination of components doesn't achieve the claimed qualitative leap",
        "tables": "concept_hierarchy",
        "severity": "critical",
        "example": "Combining 'sovereignty' + 'technology' doesn't create new meaning - just awkward juxtaposition"
    },
    {
        "type": "incommensurability_revealed",
        "flags": "Concept cannot be reduced to or explained by its claimed components",
        "tables": "concept_hierarchy",
        "severity": "major",
        "example": "'Tech sovereignty' has emergent properties that 'sovereignty' + 'technology' can't explain"
    },
    {
        "type": "mapping_different",
        "flags": "How the concept is acquired/learned differs from what we described",
        "tables": "concept_mapping_process",
        "severity": "moderate",
        "example": "We said 'extended mapping' but evidence shows 'fast mapping' from media - shallow uptake"
    },
    {
        "type": "core_cognition_wrong",
        "flags": "Our assessment of whether concept derives from core cognitive systems was wrong",
        "tables": "concepts.core_cognition_derived",
        "severity": "minor",
        "example": "We said NO core cognition but 'in-group/out-group' core system may be involved"
    },
])

set_column_widths(ws)

# ============================================================================
# SHEET 11: Feedback Database Schema
# ============================================================================
ws = wb.create_sheet("11. Database Schema")
style_header(ws, 1, "DIMENSION-SPECIFIC FEEDBACK - DATABASE SCHEMA", dimension_colors["Overview"])

row = 3
ws.cell(row=row, column=1, value="Master feedback table with dimension-specific type enum").font = Font(bold=True, size=12)
row += 2

# Main table
headers = ["Field", "Type", "Description"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = table_header_fill
row += 1

fields = [
    ("id", "SERIAL PRIMARY KEY", ""),
    ("evidence_test_id", "INTEGER FK", "Link to evidence_tests"),
    ("dimension", "VARCHAR(20)", "quinean/sellarsian/brandomian/deleuzian/bachelardian/canguilhem/davidson/blumenberg/carey"),
    ("feedback_type", "VARCHAR(40)", "One of 57 dimension-specific types (see enum below)"),
    ("severity", "VARCHAR(20)", "critical/major/moderate/minor"),
    ("summary", "TEXT NOT NULL", "Brief description of finding"),
    ("details", "TEXT", "Full explanation"),
    ("affected_table", "VARCHAR(100)", "Which concept schema table is affected"),
    ("affected_row_id", "INTEGER", "ID of specific row if applicable"),
    ("evidence_basis", "TEXT", "What evidence supports this finding"),
    ("philosophical_significance", "TEXT", "Why this matters theoretically"),
    ("suggested_action", "TEXT", "What revision might address this"),
    ("confidence", "FLOAT", "LLM confidence 0-1"),
    ("status", "VARCHAR(20)", "pending/reviewed/actioned/dismissed"),
    ("created_at", "TIMESTAMPTZ", ""),
]

for f in fields:
    ws.cell(row=row, column=1, value=f[0]).font = field_font
    ws.cell(row=row, column=2, value=f[1]).fill = type_fill
    ws.cell(row=row, column=3, value=f[2])
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

# Enum definition
row += 2
ws.cell(row=row, column=1, value="FEEDBACK_TYPE ENUM (by dimension):").font = Font(bold=True, size=12)
row += 2

enums = [
    ("quinean", "inference_broken, inference_missing, centrality_wrong, entrenchment_overestimated, web_inconsistency, holism_ripple"),
    ("sellarsian", "false_given_exposed, hidden_commitment_revealed, givenness_marker_challenged, effect_reversed, space_of_reasons_violation, manifest_scientific_gap"),
    ("brandomian", "commitment_violated, entitlement_exceeded, inference_defeated, incompatibility_missed, challenge_unanswered, perspectival_distortion, deontic_status_wrong"),
    ("deleuzian", "component_missing, component_extraneous, zone_mismapped, consistency_weaker, neighborhood_wrong, plane_misidentified, persona_different, problem_transformed"),
    ("bachelardian", "new_obstacle_discovered, obstacle_deeper, blocked_understanding_unblocked, rupture_imminent, rupture_false, psychoanalytic_function_different"),
    ("canguilhem", "evolution_different, normative_dimension_exposed, vitality_declining, vitality_reviving, filiation_broken, normal_normative_confusion"),
    ("davidson", "style_mismatch, visibility_wrong, evidence_privileged_wrong, inference_pattern_different, new_style_emerging, objects_created_different"),
    ("blumenberg", "metaphor_missed, metaphor_effect_different, metakinesis_different, conceptual_work_failing, nonconceptuality_revealed, absolutism_underestimated"),
    ("carey", "hierarchy_wrong, component_missing, bootstrap_failed, incommensurability_revealed, mapping_different, core_cognition_wrong"),
]

for dim, types in enums:
    ws.cell(row=row, column=1, value=dim).font = Font(bold=True)
    ws.cell(row=row, column=2, value=types)
    row += 1

set_column_widths(ws)

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Dimension_Feedback_Schema.xlsx")
wb.save(output_path)
print(f"Dimension-specific feedback schema saved to: {output_path}")
