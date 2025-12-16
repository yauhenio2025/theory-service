#!/usr/bin/env python3
"""
Generate Concept Schema v4:
1. Source tracking (source_type, source_reference) on ALL tables
2. Expanded Brandomian dimension based on full framework
3. All previous improvements (modular, revised Deleuzian, etc.)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
new_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
source_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")  # Green for source tracking
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Core": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Davidson": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

# Standard source tracking fields to add to every table
SOURCE_TRACKING_FIELDS = [
    {"name": "--- SOURCE TRACKING ---", "type": "", "constraints": "", "desc": "", "source": True},
    {"name": "source_type", "type": "VARCHAR(30)", "constraints": "", "desc": "llm_analysis/evidence_testing/internal_compute/user_input/import", "source": True},
    {"name": "source_reference", "type": "TEXT", "constraints": "", "desc": "Model ID, cluster_id, user_id, import batch, etc.", "source": True},
    {"name": "source_confidence", "type": "FLOAT", "constraints": "0-1", "desc": "Source's confidence in this data", "source": True},
]

def add_table(ws, start_row, table_name, fields, note=None, include_source=True):
    """Add a table schema section with source tracking."""
    row = start_row
    if note:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9)
        row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value=table_name).font = Font(bold=True, size=12)
    row += 1

    headers = ["Field", "Type", "Constraints", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    # Add regular fields
    for f in fields:
        ws.cell(row=row, column=1, value=f["name"]).font = field_font
        ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
        c3 = ws.cell(row=row, column=3, value=f.get("constraints", ""))
        if "FK" in f.get("constraints", ""):
            c3.fill = fk_fill
        ws.cell(row=row, column=4, value=f["desc"])
        if f.get("new"):
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = new_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    # Add source tracking fields
    if include_source:
        for f in SOURCE_TRACKING_FIELDS:
            ws.cell(row=row, column=1, value=f["name"]).font = field_font
            ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
            ws.cell(row=row, column=3, value=f.get("constraints", ""))
            ws.cell(row=row, column=4, value=f["desc"])
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = source_fill
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).alignment = wrap
            row += 1

    return row + 1

def style_dimension_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# ============================================================================
# Sheet 1: Overview
# ============================================================================
ws = wb.active
ws.title = "Overview"

overview = [
    ["ENRICHED CONCEPT SCHEMA v4"],
    [""],
    ["KEY FEATURES:"],
    ["1. SOURCE TRACKING on every table (who/what populated each row)"],
    ["2. EXPANDED BRANDOMIAN dimension (full inferentialist framework)"],
    ["3. MODULAR design (each claim is individually addressable)"],
    ["4. REVISED DELEUZIAN dimension (from 'What is Philosophy?')"],
    [""],
    ["SOURCE TRACKING FIELDS (on every table):"],
    ["Field", "Values", "Purpose"],
    ["source_type", "llm_analysis, evidence_testing, internal_compute, user_input, import", "Who/what created this row"],
    ["source_reference", "Model ID, cluster_id, user_id, batch_id, etc.", "Specific source identifier"],
    ["source_confidence", "0.0 - 1.0", "Source's confidence in this data"],
    [""],
    ["OPERATION TYPES:"],
    ["Type", "Source", "Example"],
    ["INTERNAL", "internal_compute", "Neighborhood computed from shared inferences"],
    ["EXTERNAL-LLM", "llm_analysis", "Claude analyzed concept components"],
    ["EXTERNAL-EVIDENCE", "evidence_testing", "Essay-flow cluster challenged claim"],
    ["EXTERNAL-USER", "user_input", "Expert corrected definition"],
    ["IMPORT", "import", "Batch imported from theory source"],
    [""],
    ["BRANDOMIAN EXPANSION (new in v4):"],
    ["- Three types of inferential consequences (committive, permissive, incompatibility)"],
    ["- Full scorekeeping model (acknowledge, attribute, undertake)"],
    ["- Challenge/justification tracking (game of giving and asking for reasons)"],
    ["- Perspectival analysis (de dicto vs de re ascriptions)"],
]

for row_idx, row_data in enumerate(overview, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif row_idx in [10, 16]:
            cell.font = header_font
            cell.fill = header_fill

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 45

# ============================================================================
# Sheet 2: Brandomian Dimension (EXPANDED)
# ============================================================================
ws = wb.create_sheet("Brandomian (Expanded)")
style_dimension_header(ws, 1, "BRANDOMIAN DIMENSION - Full Inferentialist Framework", dimension_colors["Brandomian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Robert Brandom's 'Making It Explicit' (1994) - Normative pragmatics, deontic scorekeeping, game of giving and asking for reasons").font = Font(italic=True)

row = 5

# Table 1: concept_inferential_roles
row = add_table(ws, row, "concept_inferential_roles (material inferences - NOT just logical)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "inference_direction", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "from_concept/to_concept (this concept as premise or conclusion)"},
    {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "committive/permissive/incompatibility", "new": True},
    {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The material inference"},
    {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept if in DB"},
    {"name": "is_material", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Content-dependent (vs purely formal)?", "new": True},
    {"name": "counterfactual_supporting", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Supports counterfactuals? (for incompatibility)", "new": True},
    {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Inference strength 0-1"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When evidence challenged this"},
], note="From Brandom: Three types - committive (deductive), permissive (inductive), incompatibility (modal)")

# Add explanation
ws.cell(row=row, column=1, value="INFERENCE TYPES:").font = Font(bold=True)
row += 1
types = [
    ("committive", "Commitment-preserving (deductive)", "'If X, then committed to Y'"),
    ("permissive", "Entitlement-preserving (inductive)", "'If entitled to X, entitled to Y'"),
    ("incompatibility", "Modal, counterfactual-supporting", "'X is incompatible with Y'"),
]
for t, desc, example in types:
    ws.cell(row=row, column=1, value=f"  {t}:")
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=example)
    row += 1
row += 1

# Table 2: concept_commitments (enhanced)
row = add_table(ws, row, "concept_commitments (deontic statuses)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "commitment_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "commitment/entitlement/incompatibility"},
    {"name": "statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "What you're committed/entitled to"},
    {"name": "deontic_status", "type": "VARCHAR(30)", "constraints": "", "desc": "acknowledged/attributed/undertaken", "new": True},
    {"name": "is_honored", "type": "BOOLEAN", "constraints": "", "desc": "Is commitment honored in practice?"},
    {"name": "violation_evidence", "type": "TEXT", "constraints": "", "desc": "If violated, where/how"},
    {"name": "inherited_from_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Inherited from parent concept?"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When last challenged"},
], note="From Brandom: Using a concept commits you to certain claims and entitles you to others")

# Table 3: concept_scorekeeping (enhanced)
row = add_table(ws, row, "concept_scorekeeping (deontic score changes over time)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "event_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "assertion/challenge/justification/withdrawal", "new": True},
    {"name": "event_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What happened"},
    {"name": "commitment_id", "type": "INTEGER", "constraints": "FK → concept_commitments", "desc": "Which commitment affected"},
    {"name": "score_change", "type": "VARCHAR(30)", "constraints": "", "desc": "commitment_gained/commitment_lost/entitlement_gained/entitlement_lost", "new": True},
    {"name": "attributed_to", "type": "TEXT", "constraints": "", "desc": "Who/what made this move", "new": True},
    {"name": "evidence_cluster_id", "type": "INTEGER", "constraints": "", "desc": "If from evidence testing"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="From Brandom: Discursive practice is deontic scorekeeping - tracking changes in commitments/entitlements")

# Table 4: concept_challenges (game of giving and asking for reasons) - NEW
row = add_table(ws, row, "concept_challenges (game of giving and asking for reasons) - NEW", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Concept being challenged", "new": True},
    {"name": "commitment_id", "type": "INTEGER", "constraints": "FK → concept_commitments", "desc": "Specific commitment challenged", "new": True},
    {"name": "challenge_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "request_reasons/incompatibility_claim/counterexample", "new": True},
    {"name": "challenge_content", "type": "TEXT", "constraints": "NOT NULL", "desc": "The challenge itself", "new": True},
    {"name": "challenger", "type": "TEXT", "constraints": "", "desc": "Who/what raised challenge (evidence cluster, user, etc.)", "new": True},
    {"name": "status", "type": "VARCHAR(20)", "constraints": "DEFAULT 'open'", "desc": "open/justified/withdrawn/unresolved", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
    {"name": "resolved_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "", "new": True},
], note="From Brandom: When challenged, you must give reasons or withdraw the claim")

# Table 5: concept_justifications (responses to challenges) - NEW
row = add_table(ws, row, "concept_justifications (responses to challenges) - NEW", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "challenge_id", "type": "INTEGER", "constraints": "FK → concept_challenges, NOT NULL", "desc": "Which challenge this responds to", "new": True},
    {"name": "justification_type", "type": "VARCHAR(30)", "constraints": "", "desc": "inference_chain/evidence_citation/authority/withdrawal", "new": True},
    {"name": "justification_content", "type": "TEXT", "constraints": "NOT NULL", "desc": "The justification given", "new": True},
    {"name": "supporting_inference_ids", "type": "INTEGER[]", "constraints": "", "desc": "Inferences used in justification", "new": True},
    {"name": "success", "type": "BOOLEAN", "constraints": "", "desc": "Did justification succeed?", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Brandom: Justification = providing reasons that preserve entitlement")

# Table 6: concept_perspectival_content (de dicto / de re) - NEW
row = add_table(ws, row, "concept_perspectival_content (de dicto vs de re ascriptions) - NEW", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "ascription_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "de_dicto/de_re", "new": True},
    {"name": "perspective_holder", "type": "TEXT", "constraints": "", "desc": "Whose perspective (for de dicto)", "new": True},
    {"name": "content_as_seen", "type": "TEXT", "constraints": "NOT NULL", "desc": "How concept appears from this perspective", "new": True},
    {"name": "differs_from_our_view", "type": "BOOLEAN", "constraints": "", "desc": "Does this differ from our (ascriber's) view?", "new": True},
    {"name": "our_translation", "type": "TEXT", "constraints": "", "desc": "How we (ascribers) would describe it (de re)", "new": True},
    {"name": "translation_preserves_truth", "type": "BOOLEAN", "constraints": "", "desc": "Does our translation preserve reference?", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Brandom: De re = hybrid perspective (ascriber + believer); enables objectivity")

# Add explanation of de dicto vs de re
ws.cell(row=row, column=1, value="DE DICTO vs DE RE:").font = Font(bold=True)
row += 1
explanations = [
    ("de_dicto", "From believer's perspective", "'X believes that the morning star is visible'"),
    ("de_re", "Hybrid (ascriber + believer)", "'X believes of Venus that it is visible' - we (ascribers) substitute our term"),
]
for atype, desc, example in explanations:
    ws.cell(row=row, column=1, value=f"  {atype}:")
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=example)
    row += 1

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 55

# ============================================================================
# Sheet 3: Source Tracking Reference
# ============================================================================
ws = wb.create_sheet("Source Tracking")
style_dimension_header(ws, 1, "SOURCE TRACKING - Applied to All Tables", "27AE60")

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Every row in every table has these three fields to track provenance").font = Font(italic=True)

row = 5
ws.cell(row=row, column=1, value="SOURCE_TYPE VALUES:").font = Font(bold=True)
row += 1

source_types = [
    ("llm_analysis", "Populated by LLM (Claude, etc.) analyzing concept", "Initial philosophical decomposition, component extraction"),
    ("evidence_testing", "Populated by evidence cluster analysis", "Challenge detected, violation evidence, inadequacy noted"),
    ("internal_compute", "Computed from other data in database", "Neighborhood order, centrality, hierarchy level"),
    ("user_input", "Entered by human user/expert", "Manual corrections, expert assessments"),
    ("import", "Batch imported from external source", "Theory source extraction, migration"),
]

headers = ["source_type", "Description", "Example Use"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
row += 1

for stype, desc, example in source_types:
    ws.cell(row=row, column=1, value=stype).font = field_font
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=example)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).alignment = wrap
    row += 1

row += 2
ws.cell(row=row, column=1, value="SOURCE_REFERENCE EXAMPLES:").font = Font(bold=True)
row += 1

refs = [
    ("llm_analysis", "claude-opus-4-20250514", "Model identifier"),
    ("evidence_testing", "cluster_id:1547", "Evidence cluster that triggered"),
    ("internal_compute", "job:recompute_neighborhoods:2025-01-15", "Batch job identifier"),
    ("user_input", "user:evgeny", "User who made edit"),
    ("import", "theory_source:morozov_guide:batch_001", "Import batch"),
]

headers = ["source_type", "source_reference example", "Notes"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
row += 1

for stype, ref, notes in refs:
    ws.cell(row=row, column=1, value=stype).font = field_font
    ws.cell(row=row, column=2, value=ref).font = field_font
    ws.cell(row=row, column=3, value=notes)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

row += 2
ws.cell(row=row, column=1, value="WHY THIS MATTERS:").font = Font(bold=True)
row += 1
reasons = [
    "1. PROVENANCE: Know where every claim came from",
    "2. RECOMPUTATION: Know what can be automatically recalculated",
    "3. TRUST: Weight claims differently based on source (LLM vs expert vs evidence)",
    "4. DEBUGGING: Trace why a concept has certain properties",
    "5. AUDIT: Track how concept changed over time and who/what changed it",
]
for reason in reasons:
    ws.cell(row=row, column=1, value=reason)
    row += 1

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 50

# ============================================================================
# Sheet 4: Table Summary (updated)
# ============================================================================
ws = wb.create_sheet("Table Summary")
style_dimension_header(ws, 1, "COMPLETE TABLE LIST - v4", "2C3E50")

ws.merge_cells('A3:F3')
ws.cell(row=3, column=1, value="All tables now include source_type, source_reference, source_confidence fields").font = Font(italic=True)

headers = ["#", "Table Name", "Dimension", "Relationship", "New in v4?", "Purpose"]
row = 5
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

tables = [
    # Core
    [1, "concepts", "Core", "1:concept", "", "Main concept record"],
    # Quinean
    [2, "concept_inferences", "Quinean", "Many:1", "", "Inferential connections"],
    [3, "concept_web_tensions", "Quinean", "Many:1", "", "Tensions with other concepts"],
    # Sellarsian
    [4, "concept_givenness", "Sellarsian", "1:1", "", "Myth of given analysis"],
    [5, "concept_givenness_markers", "Sellarsian", "Many:1", "", "Language markers"],
    [6, "concept_hidden_commitments", "Sellarsian", "Many:1", "", "Hidden assumptions"],
    [7, "concept_givenness_effects", "Sellarsian", "Many:1", "", "Enables/blocks"],
    # Brandomian (EXPANDED)
    [8, "concept_inferential_roles", "Brandomian", "Many:1", "EXPANDED", "Committive/permissive/incompatibility inferences"],
    [9, "concept_commitments", "Brandomian", "Many:1", "EXPANDED", "Deontic statuses with acknowledged/attributed/undertaken"],
    [10, "concept_scorekeeping", "Brandomian", "Many:1", "EXPANDED", "Score changes with event types"],
    [11, "concept_challenges", "Brandomian", "Many:1", "NEW", "Challenges in game of giving/asking reasons"],
    [12, "concept_justifications", "Brandomian", "Many:1", "NEW", "Responses to challenges"],
    [13, "concept_perspectival_content", "Brandomian", "Many:1", "NEW", "De dicto vs de re ascriptions"],
    # Deleuzian (REVISED)
    [14, "concept_components", "Deleuzian", "Many:1", "", "Concept components"],
    [15, "concept_zones_of_indiscernibility", "Deleuzian", "Many:1", "", "Where components overlap"],
    [16, "concept_consistency", "Deleuzian", "1:1", "", "Endoconsistency analysis"],
    [17, "concept_neighborhood", "Deleuzian", "Many:1", "", "Exoconsistency/relations"],
    [18, "concept_plane_of_immanence", "Deleuzian", "Many:1", "", "Background plane"],
    [19, "concept_personae", "Deleuzian", "Many:1", "", "Conceptual personae"],
    [20, "concept_problems", "Deleuzian", "Many:1", "", "Problems addressed"],
    # Bachelardian
    [21, "concept_obstacles", "Bachelardian", "1:1", "", "Obstacle analysis"],
    [22, "concept_obstacle_blocks", "Bachelardian", "Many:1", "", "What's blocked"],
    [23, "concept_inadequacy_evidence", "Bachelardian", "Many:1", "", "Empirical challenges"],
    # Canguilhem
    [24, "concept_evolution", "Canguilhem", "Many:1", "", "Historical transformations"],
    [25, "concept_normative_dimensions", "Canguilhem", "Many:1", "", "Embedded values"],
    [26, "concept_vitality_indicators", "Canguilhem", "Many:1", "", "Health indicators"],
    # Davidson/Hacking
    [27, "concept_reasoning_styles", "Davidson", "Many:1", "", "Required styles"],
    [28, "concept_style_visibility", "Davidson", "Many:1", "", "Visible/invisible"],
    [29, "concept_style_evidence", "Davidson", "Many:1", "", "Privileged/marginalized evidence"],
    [30, "concept_style_inferences", "Davidson", "Many:1", "", "Inference patterns"],
    # Blumenberg
    [31, "concept_metaphors", "Blumenberg", "Many:1", "", "Root metaphors"],
    [32, "concept_metaphor_effects", "Blumenberg", "Many:1", "", "Enables/hides"],
    [33, "concept_metakinetics", "Blumenberg", "Many:1", "", "Metaphor transformations"],
    [34, "concept_work_in_progress", "Blumenberg", "Many:1", "", "Conceptual work"],
    # Carey
    [35, "concept_hierarchy", "Carey", "1:1", "", "Bootstrap structure"],
    [36, "concept_built_from", "Carey", "Many:1", "", "Component concepts"],
    [37, "concept_mapping_process", "Carey", "Many:1", "", "How concept acquired"],
]

row = 6
for row_data in tables:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap
        if "NEW" in str(value) or "EXPANDED" in str(value):
            cell.fill = new_fill
    row += 1

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 45

# ============================================================================
# Sheet 5: Brandom Example (Technological Sovereignty)
# ============================================================================
ws = wb.create_sheet("Brandom Example")
style_dimension_header(ws, 1, "BRANDOMIAN ANALYSIS: Technological Sovereignty", dimension_colors["Brandomian"])

row = 3
ws.cell(row=row, column=1, value="INFERENTIAL ROLES:").font = Font(bold=True, size=12)
row += 1

headers = ["Type", "Direction", "Statement", "Counterfactual?"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
row += 1

inferences = [
    ("committive", "from", "If tech sovereign → committed to indigenous development possible", "No"),
    ("committive", "from", "If tech sovereign → committed to reduced dependency", "No"),
    ("permissive", "to", "High R&D investment → entitled to claim tech sovereignty", "No"),
    ("permissive", "to", "Protected markets → entitled to claim progress", "No"),
    ("incompatibility", "both", "Tech sovereignty INCOMPATIBLE WITH deep supply chain integration", "Yes"),
    ("incompatibility", "both", "Autonomy claims INCOMPATIBLE WITH ASML dependency", "Yes"),
]
for inf in inferences:
    for col, val in enumerate(inf, 1):
        ws.cell(row=row, column=col, value=val)
    row += 1

row += 2
ws.cell(row=row, column=1, value="COMMITMENTS & ENTITLEMENTS:").font = Font(bold=True, size=12)
row += 1

headers = ["Type", "Statement", "Deontic Status", "Honored?", "Violation"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
row += 1

commitments = [
    ("commitment", "Investment leads to autonomy", "acknowledged", "NO", "Gulf $2T can't invest at Series B"),
    ("commitment", "State can control tech trajectory", "acknowledged", "NO", "Depends on ASML, TSMC"),
    ("entitlement", "Can claim sovereignty progress", "attributed", "EXCEEDED", "Claims without substance"),
    ("incompatibility", "Sovereignty ∧ Global integration", "acknowledged", "-", "Logical tension"),
]
for c in commitments:
    for col, val in enumerate(c, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if val == "NO" or val == "EXCEEDED":
            cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    row += 1

row += 2
ws.cell(row=row, column=1, value="CHALLENGES (Game of Giving and Asking for Reasons):").font = Font(bold=True, size=12)
row += 1

headers = ["Challenge Type", "Content", "Challenger", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
row += 1

challenges = [
    ("request_reasons", "What entitles claim that investment → autonomy?", "Evidence cluster: Gulf SWF analysis", "unresolved"),
    ("counterexample", "China invested $100B+ but still depends on ASML", "Evidence cluster: Chip supply chain", "unresolved"),
    ("incompatibility_claim", "Can't have sovereignty AND participate in global supply chains", "Structural analysis", "open"),
]
for c in challenges:
    for col, val in enumerate(c, 1):
        ws.cell(row=row, column=col, value=val)
    row += 1

row += 2
ws.cell(row=row, column=1, value="PERSPECTIVAL CONTENT (De Dicto vs De Re):").font = Font(bold=True, size=12)
row += 1

headers = ["Perspective Holder", "De Dicto (their view)", "De Re (our translation)", "Preserves truth?"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
row += 1

perspectives = [
    ("Gulf states", "'We are achieving tech sovereignty'", "'They are achieving dependency management'", "Partially"),
    ("China", "'We will have chip sovereignty by 2030'", "'They are reducing but not eliminating ASML dependence'", "No"),
    ("EU", "'Digital sovereignty protects citizens'", "'Regulatory leverage within US tech ecosystem'", "Partially"),
]
for p in perspectives:
    for col, val in enumerate(p, 1):
        ws.cell(row=row, column=col, value=val)
    row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 35

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D_v4.xlsx")
wb.save(output_path)
print(f"Schema v4 saved to: {output_path}")
