#!/usr/bin/env python3
"""
Generate COMPLETE Concept Schema v4:
1. ALL 9 dimensions with full table definitions
2. Source tracking (source_type, source_reference, source_confidence) on EVERY table
3. Expanded Brandomian dimension (full inferentialist framework)
4. Revised Deleuzian dimension (from "What is Philosophy?")
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
new_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
source_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Core": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Davidson": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

# Standard source tracking fields
SOURCE_FIELDS = [
    {"name": "--- SOURCE TRACKING ---", "type": "", "constraints": "", "desc": "", "source": True},
    {"name": "source_type", "type": "VARCHAR(30)", "constraints": "", "desc": "llm_analysis/evidence_testing/internal_compute/user_input/import", "source": True},
    {"name": "source_reference", "type": "TEXT", "constraints": "", "desc": "Model ID, cluster_id, user_id, import batch, etc.", "source": True},
    {"name": "source_confidence", "type": "FLOAT", "constraints": "0-1", "desc": "Source's confidence in this data", "source": True},
]

def style_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def add_table(ws, start_row, table_name, fields, note=None, include_source=True):
    """Add table with source tracking fields."""
    row = start_row
    if note:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9)
        row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value=table_name).font = Font(bold=True, size=12)
    row += 1

    for col, h in enumerate(["Field", "Type", "Constraints", "Description"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    # Regular fields
    for f in fields:
        ws.cell(row=row, column=1, value=f["name"]).font = field_font
        ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
        c3 = ws.cell(row=row, column=3, value=f.get("constraints", ""))
        if "FK" in f.get("constraints", ""):
            c3.fill = fk_fill
        ws.cell(row=row, column=4, value=f["desc"])
        if f.get("new"):
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = new_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    # Source tracking fields
    if include_source:
        for f in SOURCE_FIELDS:
            ws.cell(row=row, column=1, value=f["name"]).font = field_font
            ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
            ws.cell(row=row, column=3, value=f.get("constraints", ""))
            ws.cell(row=row, column=4, value=f["desc"])
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = source_fill
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).alignment = wrap
            row += 1

    return row + 1

def set_column_widths(ws):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 55

# ============================================================================
# SHEET 1: Overview
# ============================================================================
ws = wb.active
ws.title = "1. Overview"

style_header(ws, 1, "ENRICHED CONCEPT SCHEMA v4 - COMPLETE", "2C3E50")

overview_text = [
    "",
    "KEY FEATURES OF v4:",
    "1. SOURCE TRACKING on every table - know who/what populated each row",
    "2. EXPANDED BRANDOMIAN - full inferentialist framework (6 tables)",
    "3. REVISED DELEUZIAN - based on 'What is Philosophy?' not 'A Thousand Plateaus'",
    "4. MODULAR design - each claim is individually addressable",
    "",
    "SOURCE TRACKING FIELDS (on every table):",
    "• source_type: llm_analysis | evidence_testing | internal_compute | user_input | import",
    "• source_reference: Specific identifier (model ID, cluster_id, user_id, etc.)",
    "• source_confidence: 0.0 - 1.0",
    "",
    "9 DIMENSIONS:",
    "1. QUINEAN (Web of Belief) - Inferential connections, centrality, entrenchment",
    "2. SELLARSIAN (Myth of Given) - Givenness, space of reasons, hidden commitments",
    "3. BRANDOMIAN (Scorekeeping) - Commitments, entitlements, challenges, perspectival content",
    "4. DELEUZIAN (Concept Theory) - Components, zones, consistency, plane, personae",
    "5. BACHELARDIAN (Obstacles) - Epistemological obstacles, ruptures, psychoanalysis",
    "6. CANGUILHEM (Life History) - Evolution, normative dimensions, vitality",
    "7. DAVIDSON/HACKING (Styles) - Reasoning styles, visibility, evidence types",
    "8. BLUMENBERG (Metaphorology) - Root metaphors, effects, metakinetics",
    "9. CAREY (Bootstrapping) - Hierarchy, components, mapping processes",
    "",
    "TOTAL TABLES: 37",
]

for row_idx, text in enumerate(overview_text, 3):
    ws.cell(row=row_idx, column=1, value=text)
    if text.startswith("KEY") or text.startswith("SOURCE") or text.startswith("9 DIM"):
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

set_column_widths(ws)

# ============================================================================
# SHEET 2: Core + Quinean
# ============================================================================
ws = wb.create_sheet("2. Quinean")
style_header(ws, 1, "QUINEAN DIMENSION - Web of Belief", dimension_colors["Quinean"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on W.V.O. Quine: Confirmation holism, web of belief, ontological relativity").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_inferences (individual inferential connections)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "forward/backward/lateral/contradiction"},
    {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The inference itself"},
    {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept (if in DB)"},
    {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Inference strength 0-1"},
    {"name": "defeasible", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Can be defeated by evidence?"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When last challenged"},
])

row = add_table(ws, row, "concept_web_tensions (theory holism tensions)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "tension_with_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Concept in tension with"},
    {"name": "tension_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Nature of the tension"},
    {"name": "resolution_cost", "type": "VARCHAR(20)", "constraints": "", "desc": "low/medium/high/catastrophic (Duhem-Quine)"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 3: Sellarsian
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
style_header(ws, 1, "SELLARSIAN DIMENSION - Myth of the Given", dimension_colors["Sellarsian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Wilfrid Sellars: Space of reasons, myth of the given, manifest vs scientific image").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_givenness (1:1 with concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "is_myth_of_given", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Falsely treated as foundational?"},
    {"name": "should_be_inferred_from", "type": "TEXT", "constraints": "", "desc": "What should actually support this"},
    {"name": "space_of_reasons_role", "type": "TEXT", "constraints": "", "desc": "Role in justification/inference"},
    {"name": "manifest_scientific_tension", "type": "TEXT", "constraints": "", "desc": "Tension between everyday and scientific views"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_givenness_markers (language markers)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "marker_text", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "'obviously', 'naturally', 'clearly', etc."},
    {"name": "example_usage", "type": "TEXT", "constraints": "", "desc": "Example in context"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_hidden_commitments (baked-in assumptions)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "commitment_text", "type": "TEXT", "constraints": "NOT NULL", "desc": "Hidden assumption"},
    {"name": "evidence_that_exposes", "type": "TEXT", "constraints": "", "desc": "What evidence would expose this"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_givenness_effects (enables/blocks)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "effect_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "enables/blocks"},
    {"name": "effect_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is enabled or blocked"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 4: Brandomian (EXPANDED)
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
style_header(ws, 1, "BRANDOMIAN DIMENSION - Full Inferentialist Framework (EXPANDED)", dimension_colors["Brandomian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Robert Brandom's 'Making It Explicit': Normative pragmatics, deontic scorekeeping, game of giving and asking for reasons").font = Font(italic=True)

row = 5

# Table 1: Inferential Roles (expanded with 3 types)
row = add_table(ws, row, "concept_inferential_roles (material inferences - NOT just logical)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "inference_direction", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "from_concept/to_concept"},
    {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "committive/permissive/incompatibility", "new": True},
    {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The material inference"},
    {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept if in DB"},
    {"name": "is_material", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Content-dependent (vs formal)?", "new": True},
    {"name": "counterfactual_supporting", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Supports counterfactuals? (for incompatibility)", "new": True},
    {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Inference strength 0-1"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When last challenged"},
], note="Three types: committive (commitment-preserving), permissive (entitlement-preserving), incompatibility (modal)")

# Table 2: Commitments (with deontic status)
row = add_table(ws, row, "concept_commitments (deontic statuses)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "commitment_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "commitment/entitlement/incompatibility"},
    {"name": "statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "What you're committed/entitled to"},
    {"name": "deontic_status", "type": "VARCHAR(30)", "constraints": "", "desc": "acknowledged/attributed/undertaken", "new": True},
    {"name": "is_honored", "type": "BOOLEAN", "constraints": "", "desc": "Is commitment honored in practice?"},
    {"name": "violation_evidence", "type": "TEXT", "constraints": "", "desc": "If violated, where/how"},
    {"name": "inherited_from_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Inherited from parent concept?"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": ""},
], note="Using a concept commits you to certain claims and entitles you to others")

# Table 3: Scorekeeping (track deontic score changes)
row = add_table(ws, row, "concept_scorekeeping (deontic score changes over time)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "event_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "assertion/challenge/justification/withdrawal", "new": True},
    {"name": "event_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What happened"},
    {"name": "commitment_id", "type": "INTEGER", "constraints": "FK → concept_commitments", "desc": "Which commitment affected"},
    {"name": "score_change", "type": "VARCHAR(30)", "constraints": "", "desc": "commitment_gained/lost/entitlement_gained/lost", "new": True},
    {"name": "attributed_to", "type": "TEXT", "constraints": "", "desc": "Who/what made this move", "new": True},
    {"name": "evidence_cluster_id", "type": "INTEGER", "constraints": "", "desc": "If from evidence testing"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Discursive practice is deontic scorekeeping - tracking changes in commitments/entitlements")

# Table 4: Challenges (game of giving and asking for reasons) - NEW
row = add_table(ws, row, "concept_challenges (game of giving and asking for reasons) - NEW", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Concept being challenged", "new": True},
    {"name": "commitment_id", "type": "INTEGER", "constraints": "FK → concept_commitments", "desc": "Specific commitment challenged", "new": True},
    {"name": "challenge_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "request_reasons/incompatibility_claim/counterexample", "new": True},
    {"name": "challenge_content", "type": "TEXT", "constraints": "NOT NULL", "desc": "The challenge itself", "new": True},
    {"name": "challenger", "type": "TEXT", "constraints": "", "desc": "Who/what raised challenge", "new": True},
    {"name": "status", "type": "VARCHAR(20)", "constraints": "DEFAULT 'open'", "desc": "open/justified/withdrawn/unresolved", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
    {"name": "resolved_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "", "new": True},
], note="When challenged, you must give reasons or withdraw the claim")

# Table 5: Justifications (responses to challenges) - NEW
row = add_table(ws, row, "concept_justifications (responses to challenges) - NEW", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "challenge_id", "type": "INTEGER", "constraints": "FK → concept_challenges, NOT NULL", "desc": "Which challenge this responds to", "new": True},
    {"name": "justification_type", "type": "VARCHAR(30)", "constraints": "", "desc": "inference_chain/evidence_citation/authority/withdrawal", "new": True},
    {"name": "justification_content", "type": "TEXT", "constraints": "NOT NULL", "desc": "The justification given", "new": True},
    {"name": "supporting_inference_ids", "type": "INTEGER[]", "constraints": "", "desc": "Inferences used in justification", "new": True},
    {"name": "success", "type": "BOOLEAN", "constraints": "", "desc": "Did justification succeed?", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="Justification = providing reasons that preserve entitlement")

# Table 6: Perspectival Content (de dicto / de re) - NEW
row = add_table(ws, row, "concept_perspectival_content (de dicto vs de re ascriptions) - NEW", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "ascription_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "de_dicto/de_re", "new": True},
    {"name": "perspective_holder", "type": "TEXT", "constraints": "", "desc": "Whose perspective (for de dicto)", "new": True},
    {"name": "content_as_seen", "type": "TEXT", "constraints": "NOT NULL", "desc": "How concept appears from this perspective", "new": True},
    {"name": "differs_from_our_view", "type": "BOOLEAN", "constraints": "", "desc": "Does this differ from our (ascriber's) view?", "new": True},
    {"name": "our_translation", "type": "TEXT", "constraints": "", "desc": "How we (ascribers) would describe it (de re)", "new": True},
    {"name": "translation_preserves_truth", "type": "BOOLEAN", "constraints": "", "desc": "Does our translation preserve reference?", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="De re = hybrid perspective (ascriber + believer); enables objectivity")

set_column_widths(ws)

# ============================================================================
# SHEET 5: Deleuzian (REVISED - from "What is Philosophy?")
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
style_header(ws, 1, "DELEUZIAN DIMENSION - Concept Theory (REVISED)", dimension_colors["Deleuzian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'What is Philosophy?' (1991) - Deleuze's actual theory OF concepts").font = Font(italic=True)

ws.merge_cells('A4:D4')
ws.cell(row=4, column=1, value="NOTE: Previous version mixed in concepts from 'A Thousand Plateaus' (becomings, lines of flight). This focuses on concept theory.").font = Font(italic=True, color="C0392B")

row = 6

# Table 1: Components
row = add_table(ws, row, "concept_components (individual components of concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "component_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Name of component (e.g., 'doubting')"},
    {"name": "component_description", "type": "TEXT", "constraints": "", "desc": "What this component contributes"},
    {"name": "is_intensive", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Is this an intensive variation?"},
    {"name": "order_index", "type": "INTEGER", "constraints": "", "desc": "Position in concept's structure"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Every concept is a multiplicity of heterogeneous components")

# Table 2: Zones of Indiscernibility
row = add_table(ws, row, "concept_zones_of_indiscernibility (where components overlap)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "component_a_id", "type": "INTEGER", "constraints": "FK → concept_components", "desc": "First component"},
    {"name": "component_b_id", "type": "INTEGER", "constraints": "FK → concept_components", "desc": "Second component"},
    {"name": "zone_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What passes between components in this zone"},
    {"name": "what_becomes_possible", "type": "TEXT", "constraints": "", "desc": "What this zone enables"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Components have zones of indiscernibility - loci of becoming within concepts")

# Table 3: Consistency (endo)
row = add_table(ws, row, "concept_consistency (endoconsistency analysis)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "endoconsistency_description", "type": "TEXT", "constraints": "", "desc": "How components hold together internally"},
    {"name": "survey_point", "type": "TEXT", "constraints": "", "desc": "The 'point of absolute survey' that unifies"},
    {"name": "consistency_strength", "type": "VARCHAR(20)", "constraints": "", "desc": "strong/moderate/weak/unstable"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Endoconsistency = internal coherence of concept's components")

# Table 4: Neighborhood (exoconsistency)
row = add_table(ws, row, "concept_neighborhood (exoconsistency / relations to other concepts)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "neighbor_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept (if in DB)"},
    {"name": "neighbor_description", "type": "TEXT", "constraints": "", "desc": "Description if not in DB"},
    {"name": "relation_type", "type": "VARCHAR(30)", "constraints": "", "desc": "bridge/resonance/interference/repulsion"},
    {"name": "neighborhood_order", "type": "INTEGER", "constraints": "", "desc": "How close (1=closest neighbor)"},
    {"name": "bridges_across_plane", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Does this bridge different planes?"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Exoconsistency = concept's neighborhood, bridges to other concepts")

# Table 5: Plane of Immanence
row = add_table(ws, row, "concept_plane_of_immanence (the ground on which concept operates)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "plane_name", "type": "VARCHAR(100)", "constraints": "", "desc": "Name/description of the plane"},
    {"name": "what_plane_presupposes", "type": "TEXT", "constraints": "", "desc": "Unthought assumptions of this plane"},
    {"name": "legitimate_problems", "type": "TEXT", "constraints": "", "desc": "What counts as a problem on this plane"},
    {"name": "excluded_problems", "type": "TEXT", "constraints": "", "desc": "What can't be asked on this plane"},
    {"name": "historical_emergence", "type": "VARCHAR(100)", "constraints": "", "desc": "When this plane emerged"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Plane of immanence = 'the absolute ground of philosophy, its earth'")

# Table 6: Conceptual Personae
row = add_table(ws, row, "concept_personae (conceptual personae that activate the concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "persona_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Name of persona (e.g., 'the sovereign')"},
    {"name": "persona_description", "type": "TEXT", "constraints": "", "desc": "What this persona does/thinks"},
    {"name": "what_persona_enables", "type": "TEXT", "constraints": "", "desc": "What thinking this persona enables"},
    {"name": "historical_origin", "type": "TEXT", "constraints": "", "desc": "Where this persona comes from"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Conceptual personae are the 'third element' of philosophy alongside plane and concepts")

# Table 7: Problems
row = add_table(ws, row, "concept_problems (problems the concept addresses)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "problem_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The problem being addressed"},
    {"name": "problem_type", "type": "VARCHAR(30)", "constraints": "", "desc": "political/epistemological/ontological/practical"},
    {"name": "triggering_event", "type": "TEXT", "constraints": "", "desc": "What event made this problem pressing"},
    {"name": "how_concept_responds", "type": "TEXT", "constraints": "", "desc": "How the concept addresses this problem"},
    {"name": "problem_transformed_to", "type": "TEXT", "constraints": "", "desc": "How problem changes through concept's use"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Concepts are 'constellations of events to come'")

set_column_widths(ws)

# ============================================================================
# SHEET 6: Bachelardian
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
style_header(ws, 1, "BACHELARDIAN DIMENSION - Epistemological Obstacles", dimension_colors["Bachelardian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Gaston Bachelard: Epistemological obstacles, rupture, psychoanalysis of knowledge").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_obstacles (1:1 - main obstacle analysis)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "is_obstacle", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Does concept block understanding?"},
    {"name": "obstacle_type", "type": "VARCHAR(30)", "constraints": "", "desc": "experience/verbal/pragmatic/quantitative/substantialist/animist"},
    {"name": "why_persists", "type": "TEXT", "constraints": "", "desc": "Ideological/class function"},
    {"name": "rupture_would_enable", "type": "TEXT", "constraints": "", "desc": "What becomes thinkable after rupture"},
    {"name": "rupture_trigger", "type": "TEXT", "constraints": "", "desc": "What would force abandonment"},
    {"name": "psychoanalytic_function", "type": "TEXT", "constraints": "", "desc": "Unconscious need it serves"},
    {"name": "epistemological_profile", "type": "TEXT", "constraints": "", "desc": "Where concept sits on progress scale"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_obstacle_blocks (what it blocks)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "blocked_understanding", "type": "TEXT", "constraints": "NOT NULL", "desc": "What understanding it prevents"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_inadequacy_evidence (empirical challenges)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "evidence_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Empirical challenge to the concept"},
    {"name": "source_cluster_id", "type": "INTEGER", "constraints": "", "desc": "Source evidence cluster"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 7: Canguilhem
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
style_header(ws, 1, "CANGUILHEM DIMENSION - Life History of Concepts", dimension_colors["Canguilhem"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Georges Canguilhem: Normal vs normative, concept life cycles, filiation of concepts").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_evolution (historical transformations)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "period", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Time period of transformation"},
    {"name": "transformation_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What changed"},
    {"name": "problem_driving", "type": "TEXT", "constraints": "", "desc": "What problem drove this change"},
    {"name": "who_transformed", "type": "TEXT", "constraints": "", "desc": "Intellectual tradition, actors"},
    {"name": "predecessor_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Concept filiation"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_normative_dimensions (embedded values)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "value_embedded", "type": "TEXT", "constraints": "NOT NULL", "desc": "What value is embedded"},
    {"name": "whose_values", "type": "TEXT", "constraints": "", "desc": "Whose interests this serves"},
    {"name": "what_excluded", "type": "TEXT", "constraints": "", "desc": "What's marked as abnormal/excluded"},
    {"name": "normal_vs_normative", "type": "VARCHAR(20)", "constraints": "", "desc": "normal/normative"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_vitality_indicators (tracking concept health)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "indicator_type", "type": "VARCHAR(30)", "constraints": "", "desc": "growth/strain/crisis/revival"},
    {"name": "indicator_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What shows the concept's health"},
    {"name": "observed_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": "When this was observed"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 8: Davidson/Hacking
# ============================================================================
ws = wb.create_sheet("8. Davidson-Hacking")
style_header(ws, 1, "DAVIDSON/HACKING DIMENSION - Styles of Reasoning", dimension_colors["Davidson"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Ian Hacking & Arnold Davidson: Historical epistemology, styles of reasoning, dynamic nominalism").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_reasoning_styles (styles required by concept)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "style_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Name of reasoning style"},
    {"name": "style_emerged", "type": "VARCHAR(100)", "constraints": "", "desc": "When this style emerged"},
    {"name": "objects_created", "type": "TEXT", "constraints": "", "desc": "What objects of inquiry style creates"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_style_visibility (what style makes visible/invisible)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": "Parent style"},
    {"name": "visibility_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "visible/invisible"},
    {"name": "what_affected", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is made visible or invisible"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_style_evidence (evidence types privileged/marginalized)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": "Parent style"},
    {"name": "evidence_status", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "privileged/marginalized"},
    {"name": "evidence_type", "type": "TEXT", "constraints": "NOT NULL", "desc": "Type of evidence"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_style_inferences (characteristic inference patterns)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": "Parent style"},
    {"name": "inference_pattern", "type": "TEXT", "constraints": "NOT NULL", "desc": "Characteristic reasoning move"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 9: Blumenberg
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
style_header(ws, 1, "BLUMENBERG DIMENSION - Metaphorology", dimension_colors["Blumenberg"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Hans Blumenberg: Absolute metaphors, nonconceptuality, metakinetics").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_metaphors (root metaphors)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "root_metaphor", "type": "TEXT", "constraints": "NOT NULL", "desc": "The metaphor itself"},
    {"name": "source_domain", "type": "TEXT", "constraints": "", "desc": "Where metaphor comes from"},
    {"name": "is_absolute", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Cannot be translated to concepts?"},
    {"name": "nonconceptuality_aspect", "type": "TEXT", "constraints": "", "desc": "What resists conceptualization?"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_metaphor_effects (enables/hides)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "metaphor_id", "type": "INTEGER", "constraints": "FK → concept_metaphors, NOT NULL", "desc": "Parent metaphor"},
    {"name": "effect_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "enables/hides"},
    {"name": "effect_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is enabled or hidden"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_metakinetics (how metaphors transform over time)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "metaphor_id", "type": "INTEGER", "constraints": "FK → concept_metaphors, NOT NULL", "desc": "Parent metaphor"},
    {"name": "period", "type": "VARCHAR(100)", "constraints": "", "desc": "Time period"},
    {"name": "transformation", "type": "TEXT", "constraints": "NOT NULL", "desc": "How metaphor meaning shifted"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_work_in_progress (conceptual work being done)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "original_meaning", "type": "TEXT", "constraints": "", "desc": "What concept originally meant"},
    {"name": "current_work", "type": "TEXT", "constraints": "NOT NULL", "desc": "What transformation is being attempted"},
    {"name": "who_doing_work", "type": "TEXT", "constraints": "", "desc": "Intellectual tradition, actors"},
    {"name": "work_status", "type": "VARCHAR(30)", "constraints": "", "desc": "succeeding/failing/ongoing"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 10: Carey
# ============================================================================
ws = wb.create_sheet("10. Carey")
style_header(ws, 1, "CAREY DIMENSION - Conceptual Bootstrapping", dimension_colors["Carey"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on Susan Carey: Core cognition, bootstrapping, conceptual discontinuities, incommensurability").font = Font(italic=True)

row = 5
row = add_table(ws, row, "concept_hierarchy (1:1 - bootstrap structure)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "combination_type", "type": "VARCHAR(30)", "constraints": "", "desc": "simple_aggregation/interactive/qualitative_leap"},
    {"name": "transparency", "type": "VARCHAR(20)", "constraints": "", "desc": "high/medium/low"},
    {"name": "bootstrap_failure_reason", "type": "TEXT", "constraints": "", "desc": "If failed, why"},
    {"name": "what_would_fix", "type": "TEXT", "constraints": "", "desc": "What would make bootstrap succeed"},
    {"name": "incommensurable_with", "type": "TEXT", "constraints": "", "desc": "Prior concepts this can't be reduced to"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_built_from (component concepts)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "component_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Component concept (if in DB)"},
    {"name": "component_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Description of component"},
    {"name": "component_level", "type": "INTEGER", "constraints": "", "desc": "Hierarchy level of component (0-3)"},
    {"name": "how_combined", "type": "TEXT", "constraints": "", "desc": "How this component is used"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_mapping_process (how concept is learned/acquired)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "mapping_type", "type": "VARCHAR(20)", "constraints": "", "desc": "fast/extended"},
    {"name": "mapping_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "How concept gets acquired"},
    {"name": "executive_function_needed", "type": "TEXT", "constraints": "", "desc": "What cognitive effort needed"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 11: Table Summary
# ============================================================================
ws = wb.create_sheet("11. Table Summary")
style_header(ws, 1, "COMPLETE TABLE LIST - v4", "2C3E50")

ws.merge_cells('A3:F3')
ws.cell(row=3, column=1, value="All tables include source_type, source_reference, source_confidence fields for provenance tracking").font = Font(italic=True)

headers = ["#", "Table Name", "Dimension", "Relationship", "New/Expanded in v4?", "Purpose"]
row = 5
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = table_header_fill

tables = [
    # Quinean
    [1, "concept_inferences", "Quinean", "Many:1", "", "Inferential connections"],
    [2, "concept_web_tensions", "Quinean", "Many:1", "", "Tensions with other concepts"],
    # Sellarsian
    [3, "concept_givenness", "Sellarsian", "1:1", "", "Myth of given analysis"],
    [4, "concept_givenness_markers", "Sellarsian", "Many:1", "", "Language markers"],
    [5, "concept_hidden_commitments", "Sellarsian", "Many:1", "", "Hidden assumptions"],
    [6, "concept_givenness_effects", "Sellarsian", "Many:1", "", "Enables/blocks"],
    # Brandomian (EXPANDED)
    [7, "concept_inferential_roles", "Brandomian", "Many:1", "EXPANDED", "Three inference types"],
    [8, "concept_commitments", "Brandomian", "Many:1", "EXPANDED", "With deontic status"],
    [9, "concept_scorekeeping", "Brandomian", "Many:1", "EXPANDED", "Score changes + events"],
    [10, "concept_challenges", "Brandomian", "Many:1", "NEW", "Game of giving/asking reasons"],
    [11, "concept_justifications", "Brandomian", "Many:1", "NEW", "Responses to challenges"],
    [12, "concept_perspectival_content", "Brandomian", "Many:1", "NEW", "De dicto vs de re"],
    # Deleuzian (REVISED)
    [13, "concept_components", "Deleuzian", "Many:1", "REVISED", "Concept components"],
    [14, "concept_zones_of_indiscernibility", "Deleuzian", "Many:1", "REVISED", "Where components overlap"],
    [15, "concept_consistency", "Deleuzian", "1:1", "REVISED", "Endoconsistency"],
    [16, "concept_neighborhood", "Deleuzian", "Many:1", "REVISED", "Exoconsistency"],
    [17, "concept_plane_of_immanence", "Deleuzian", "Many:1", "REVISED", "Background plane"],
    [18, "concept_personae", "Deleuzian", "Many:1", "REVISED", "Conceptual personae"],
    [19, "concept_problems", "Deleuzian", "Many:1", "REVISED", "Problems addressed"],
    # Bachelardian
    [20, "concept_obstacles", "Bachelardian", "1:1", "", "Obstacle analysis"],
    [21, "concept_obstacle_blocks", "Bachelardian", "Many:1", "", "What's blocked"],
    [22, "concept_inadequacy_evidence", "Bachelardian", "Many:1", "", "Empirical challenges"],
    # Canguilhem
    [23, "concept_evolution", "Canguilhem", "Many:1", "", "Historical transformations"],
    [24, "concept_normative_dimensions", "Canguilhem", "Many:1", "", "Embedded values"],
    [25, "concept_vitality_indicators", "Canguilhem", "Many:1", "", "Health indicators"],
    # Davidson/Hacking
    [26, "concept_reasoning_styles", "Davidson", "Many:1", "", "Required styles"],
    [27, "concept_style_visibility", "Davidson", "Many:1", "", "Visible/invisible"],
    [28, "concept_style_evidence", "Davidson", "Many:1", "", "Privileged/marginalized evidence"],
    [29, "concept_style_inferences", "Davidson", "Many:1", "", "Inference patterns"],
    # Blumenberg
    [30, "concept_metaphors", "Blumenberg", "Many:1", "", "Root metaphors"],
    [31, "concept_metaphor_effects", "Blumenberg", "Many:1", "", "Enables/hides"],
    [32, "concept_metakinetics", "Blumenberg", "Many:1", "", "Metaphor transformations"],
    [33, "concept_work_in_progress", "Blumenberg", "Many:1", "", "Conceptual work"],
    # Carey
    [34, "concept_hierarchy", "Carey", "1:1", "", "Bootstrap structure"],
    [35, "concept_built_from", "Carey", "Many:1", "", "Component concepts"],
    [36, "concept_mapping_process", "Carey", "Many:1", "", "How concept acquired"],
]

row = 6
for row_data in tables:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap
        if "NEW" in str(value) or "EXPANDED" in str(value) or "REVISED" in str(value):
            cell.fill = new_fill
    row += 1

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 40

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D_v4.xlsx")
wb.save(output_path)
print(f"Complete v4 schema saved to: {output_path}")
