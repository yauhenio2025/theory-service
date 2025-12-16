#!/usr/bin/env python3
"""
Generate Excel showing the CONCEPT SCHEMA - database structure for
capturing 9-dimensional concept depth, with example values for
"Technological Sovereignty".

This is the DATA MODEL for concepts, independent of evidence testing.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Create workbook
wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Core": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Davidson": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

def add_table_sheet(wb, sheet_name, dimension, table_name, description, fields, example_data):
    """Add a sheet showing table schema and example data."""
    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1, value=f"{dimension.upper()} - {table_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(
        start_color=dimension_colors.get(dimension, "2C3E50"),
        end_color=dimension_colors.get(dimension, "2C3E50"),
        fill_type="solid"
    )

    # Description
    ws.merge_cells('A2:E2')
    ws.cell(row=2, column=1, value=description)
    ws.cell(row=2, column=1).font = Font(italic=True)

    # Schema section
    ws.cell(row=4, column=1, value="DATABASE SCHEMA")
    ws.cell(row=4, column=1).font = Font(bold=True, size=12)

    # Field headers
    schema_headers = ["Field Name", "Type", "Constraints", "Description"]
    for col, header in enumerate(schema_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border

    # Fields
    row = 6
    for field in fields:
        ws.cell(row=row, column=1, value=field["name"]).font = field_font
        ws.cell(row=row, column=2, value=field["type"])
        ws.cell(row=row, column=2).fill = type_fill
        ws.cell(row=row, column=3, value=field.get("constraints", ""))
        if "FK" in field.get("constraints", ""):
            ws.cell(row=row, column=3).fill = fk_fill
        ws.cell(row=row, column=4, value=field["desc"])
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    # Example data section
    row += 2
    ws.cell(row=row, column=1, value="EXAMPLE DATA: Technological Sovereignty")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    # Example headers
    example_headers = ["Field", "Value"]
    for col, header in enumerate(example_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    # Example values
    for field_name, value in example_data.items():
        ws.cell(row=row, column=1, value=field_name).font = field_font
        ws.cell(row=row, column=2, value=str(value) if value else "(null)")
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15

# ============================================================================
# Sheet 1: Overview
# ============================================================================
ws = wb.active
ws.title = "Schema Overview"

overview_text = [
    ["ENRICHED CONCEPT SCHEMA - Database Structure"],
    [""],
    ["This schema captures concepts across 9 philosophical dimensions."],
    ["Each concept has ONE core record plus MULTIPLE related records in dimension tables."],
    [""],
    ["TABLE STRUCTURE"],
    [""],
    ["Table", "Dimension", "Relationship", "Purpose"],
    ["concepts", "Core + Extensions", "1 per concept", "Main concept record with extended fields"],
    ["concept_inferences", "Quinean", "Many per concept", "Inferential connections to other concepts"],
    ["concept_givenness", "Sellarsian", "1 per concept", "Analysis of 'myth of given' status"],
    ["concept_commitments", "Brandomian", "Many per concept", "Commitments and entitlements"],
    ["concept_problems", "Deleuzian", "Many per concept", "Problems the concept addresses"],
    ["concept_plane_assumptions", "Deleuzian", "Many per concept", "Background assumptions"],
    ["concept_obstacles", "Bachelardian", "1 per concept", "Epistemological obstacle analysis"],
    ["concept_evolution", "Canguilhem", "Many per concept", "Historical transformations"],
    ["concept_normative_dimensions", "Canguilhem", "Many per concept", "Embedded values"],
    ["concept_reasoning_styles", "Davidson", "Many per concept", "Required reasoning styles"],
    ["concept_metaphors", "Blumenberg", "Many per concept", "Root metaphors"],
    ["concept_work_in_progress", "Blumenberg", "Many per concept", "Conceptual work being done"],
    ["concept_hierarchy", "Carey", "1 per concept", "Bootstrap structure"],
    [""],
    ["RELATIONAL FIELDS"],
    [""],
    ["Some fields reference OTHER concepts in the database:"],
    ["- concept_inferences.target_concept_id → concepts.id"],
    ["- concept_hierarchy.built_from_concept_ids → concepts.id[]"],
    ["These enable mapping the web of relationships between concepts."],
]

for row_idx, row_data in enumerate(overview_text, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif row_idx == 8:
            cell.font = header_font
            cell.fill = header_fill
        elif row_idx > 8 and row_idx <= 21 and col_idx == 1:
            cell.font = field_font

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 50

# ============================================================================
# Sheet 2: Core concepts table (extended)
# ============================================================================
add_table_sheet(
    wb, "1. concepts (Core)", "Core", "concepts",
    "Main concept table with extended fields from Quinean, Canguilhem, and Carey dimensions",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "term", "type": "VARCHAR(255)", "constraints": "NOT NULL, INDEX", "desc": "Concept name/term"},
        {"name": "definition", "type": "TEXT", "constraints": "NOT NULL", "desc": "Core definition"},
        {"name": "category", "type": "VARCHAR(100)", "constraints": "", "desc": "Domain category"},
        {"name": "source_id", "type": "INTEGER", "constraints": "FK → theory_sources", "desc": "Where concept originated"},
        {"name": "status", "type": "ENUM", "constraints": "draft/active/deprecated/challenged", "desc": "Current status"},
        {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 1.0", "desc": "Confidence score 0-1"},
        {"name": "---", "type": "---", "constraints": "QUINEAN EXTENSIONS", "desc": "---"},
        {"name": "centrality", "type": "VARCHAR(20)", "constraints": "core/intermediate/peripheral", "desc": "Position in inferential web"},
        {"name": "web_coherence_impact", "type": "TEXT", "constraints": "", "desc": "How changing this affects other concepts"},
        {"name": "---", "type": "---", "constraints": "CANGUILHEM EXTENSIONS", "desc": "---"},
        {"name": "health_status", "type": "VARCHAR(20)", "constraints": "healthy/strained/dying/being_born", "desc": "Concept's current vitality"},
        {"name": "birth_period", "type": "VARCHAR(100)", "constraints": "", "desc": "When concept emerged"},
        {"name": "birth_problem", "type": "TEXT", "constraints": "", "desc": "Problem that birthed concept"},
        {"name": "---", "type": "---", "constraints": "CAREY EXTENSIONS", "desc": "---"},
        {"name": "hierarchy_level", "type": "INTEGER", "constraints": "0=primitive, 1-3+=complex", "desc": "Conceptual hierarchy level"},
        {"name": "bootstrap_status", "type": "VARCHAR(20)", "constraints": "successful/partial/failed", "desc": "Did bootstrap achieve leap?"},
        {"name": "---", "type": "---", "constraints": "TIMESTAMPS", "desc": "---"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
        {"name": "updated_at", "type": "TIMESTAMPTZ", "constraints": "ON UPDATE NOW()", "desc": "Last modification"},
    ],
    example_data={
        "id": 42,
        "term": "Technological Sovereignty",
        "definition": "A state's capacity to control its own technological development and infrastructure",
        "category": "sovereignty, technology, geopolitics",
        "source_id": 1,
        "status": "challenged",
        "confidence": 0.6,
        "centrality": "intermediate",
        "web_coherence_impact": "Changing affects 'sovereignty', 'autonomy', 'technology dependency' concepts",
        "health_status": "strained",
        "birth_period": "2010s (US-China tech tensions)",
        "birth_problem": "How to mobilize resources for tech independence",
        "hierarchy_level": 3,
        "bootstrap_status": "failed",
    }
)

# ============================================================================
# Sheet 3: concept_inferences (Quinean)
# ============================================================================
add_table_sheet(
    wb, "2. Inferences (Quinean)", "Quinean", "concept_inferences",
    "Inferential connections - what follows from/leads to this concept",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "forward/backward/lateral/contradiction"},
        {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The inference itself"},
        {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept (if in DB)"},
        {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Inference strength 0-1"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "inference_type": "forward",
        "inference_statement": "If tech sovereignty, then indigenous chip development possible",
        "target_concept_id": None,
        "strength": 0.7,
    }
)

# Add more example rows text
ws = wb["2. Inferences (Quinean)"]
row = ws.max_row + 2
ws.cell(row=row, column=1, value="ADDITIONAL EXAMPLE RECORDS:")
ws.cell(row=row, column=1).font = Font(bold=True)
examples = [
    ("forward", "If tech sovereignty, then reduced foreign dependency", 0.8),
    ("backward", "Tech sovereignty because sufficient R&D investment", 0.6),
    ("lateral", "Related to: economic sovereignty, data sovereignty", 0.9),
    ("contradiction", "Material dependency evidence contradicts autonomy claims", 0.85),
]
row += 1
for inf_type, statement, strength in examples:
    ws.cell(row=row, column=1, value=inf_type)
    ws.cell(row=row, column=2, value=statement)
    ws.cell(row=row, column=3, value=str(strength))
    row += 1

# ============================================================================
# Sheet 4: concept_givenness (Sellarsian)
# ============================================================================
add_table_sheet(
    wb, "3. Givenness (Sellarsian)", "Sellarsian", "concept_givenness",
    "Analysis of whether concept is falsely treated as foundational/self-evident",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
        {"name": "is_myth_of_given", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Is concept falsely treated as foundational?"},
        {"name": "givenness_markers", "type": "TEXT[]", "constraints": "", "desc": "Language markers: 'obviously', 'naturally', 'of course'"},
        {"name": "should_be_inferred_from", "type": "TEXT", "constraints": "", "desc": "What evidence/argument should actually support this"},
        {"name": "theoretical_commitments_embedded", "type": "TEXT[]", "constraints": "", "desc": "Hidden assumptions baked in"},
        {"name": "what_givenness_enables", "type": "TEXT", "constraints": "", "desc": "What does treating as given allow?"},
        {"name": "what_givenness_blocks", "type": "TEXT", "constraints": "", "desc": "What questions become unaskable?"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "is_myth_of_given": True,
        "givenness_markers": '["naturally", "clearly possible", "inevitable", "obvious strategic necessity"]',
        "should_be_inferred_from": "Analysis of actual material capacity, supply chains, knowledge dependencies",
        "theoretical_commitments_embedded": '["States can achieve tech autonomy through will+investment", "Technology behaves like territory"]',
        "what_givenness_enables": "Mobilizes resources, justifies protectionist policies, creates political legitimacy",
        "what_givenness_blocks": "Questions about structural impossibility, analysis of why sovereignty might be unachievable",
    }
)

# ============================================================================
# Sheet 5: concept_commitments (Brandomian)
# ============================================================================
add_table_sheet(
    wb, "4. Commitments (Brandomian)", "Brandomian", "concept_commitments",
    "Commitments and entitlements that using this concept involves",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "commitment_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "commitment/entitlement/incompatibility"},
        {"name": "statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "What you're committed to / entitled to"},
        {"name": "is_honored", "type": "BOOLEAN", "constraints": "", "desc": "Is this commitment actually honored? (null for incompatibilities)"},
        {"name": "violation_evidence", "type": "TEXT", "constraints": "", "desc": "If violated, where/how"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "commitment_type": "commitment",
        "statement": "Autonomy is achievable through investment",
        "is_honored": False,
        "violation_evidence": "Gulf $2T can't invest at Series B; China depends on ASML despite $100B+",
    }
)

ws = wb["4. Commitments (Brandomian)"]
row = ws.max_row + 2
ws.cell(row=row, column=1, value="ADDITIONAL EXAMPLE RECORDS:")
ws.cell(row=row, column=1).font = Font(bold=True)
examples = [
    ("commitment", "State can control tech development trajectory", False, "Dependent on foreign equipment/expertise"),
    ("entitlement", "Can claim sovereignty through rhetoric alone", None, "EXCEEDED - claims without substance"),
    ("incompatibility", "Sovereignty AND deep integration in global supply chains", None, "Logical tension unresolved"),
]
row += 1
for ctype, stmt, honored, evidence in examples:
    ws.cell(row=row, column=1, value=ctype)
    ws.cell(row=row, column=2, value=stmt)
    ws.cell(row=row, column=3, value=str(honored) if honored is not None else "n/a")
    ws.cell(row=row, column=4, value=evidence)
    row += 1

# ============================================================================
# Sheet 6: concept_problems (Deleuzian)
# ============================================================================
add_table_sheet(
    wb, "5. Problems (Deleuzian)", "Deleuzian", "concept_problems",
    "Problems/tensions the concept addresses and what transformations it enables/blocks",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "problem_addressed", "type": "TEXT", "constraints": "NOT NULL", "desc": "What tension/problem does concept navigate"},
        {"name": "tension_pole_a", "type": "TEXT", "constraints": "", "desc": "One side of the tension"},
        {"name": "tension_pole_b", "type": "TEXT", "constraints": "", "desc": "Other side of the tension"},
        {"name": "creative_responses", "type": "TEXT[]", "constraints": "", "desc": "How concept helps navigate problem"},
        {"name": "becomings_enabled", "type": "TEXT[]", "constraints": "", "desc": "What transformations does concept enable"},
        {"name": "becomings_blocked", "type": "TEXT[]", "constraints": "", "desc": "What transformations does concept prevent"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "problem_addressed": "How to claim autonomy while remaining materially dependent",
        "tension_pole_a": "Political need for sovereignty rhetoric (legitimacy, mobilization)",
        "tension_pole_b": "Material reality of structural dependency (supply chains, knowledge, equipment)",
        "creative_responses": '["Temporal displacement (future promise)", "Definitional shifting (partial sovereignty)", "Domain restriction (key areas only)"]',
        "becomings_enabled": '["Becoming-investor", "Becoming-protectionist", "Becoming-strategic"]',
        "becomings_blocked": '["Becoming-honest", "Becoming-interdependent", "Becoming-collaborative"]',
    }
)

# ============================================================================
# Sheet 7: concept_plane_assumptions (Deleuzian)
# ============================================================================
add_table_sheet(
    wb, "6. Plane (Deleuzian)", "Deleuzian", "concept_plane_assumptions",
    "Unquestioned background assumptions that make this concept possible",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "assumption", "type": "TEXT", "constraints": "NOT NULL", "desc": "The unquestioned assumption"},
        {"name": "makes_possible", "type": "TEXT[]", "constraints": "", "desc": "What this assumption enables"},
        {"name": "makes_impossible", "type": "TEXT[]", "constraints": "", "desc": "What this assumption forecloses"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "assumption": "States are the primary actors in technology development",
        "makes_possible": '["National tech policy", "State investment programs", "Security framing"]',
        "makes_impossible": '["Recognizing corporate primacy", "Network-based governance", "Post-state tech coordination"]',
    }
)

ws = wb["6. Plane (Deleuzian)"]
row = ws.max_row + 2
ws.cell(row=row, column=1, value="ADDITIONAL EXAMPLE RECORDS:")
ws.cell(row=row, column=1).font = Font(bold=True)
examples = [
    ("Technology can be 'owned' like territory", '["Sovereignty framing"]', '["Technology as relational"]'),
    ("Autonomy is achievable through investment", '["Resource mobilization"]', '["Structural critique"]'),
    ("Competition is the natural state", '["Strategic framings"]', '["Genuine cooperation models"]'),
]
row += 1
for assumption, possible, impossible in examples:
    ws.cell(row=row, column=1, value=assumption)
    ws.cell(row=row, column=2, value=possible)
    ws.cell(row=row, column=3, value=impossible)
    row += 1

# ============================================================================
# Sheet 8: concept_obstacles (Bachelardian)
# ============================================================================
add_table_sheet(
    wb, "7. Obstacles (Bachelardian)", "Bachelardian", "concept_obstacles",
    "Analysis of whether concept functions as an epistemological obstacle",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
        {"name": "is_obstacle", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Does concept block understanding?"},
        {"name": "obstacle_type", "type": "VARCHAR(30)", "constraints": "", "desc": "experience/verbal/pragmatic/quantitative/substantialist"},
        {"name": "what_it_blocks", "type": "TEXT[]", "constraints": "", "desc": "What understanding it prevents"},
        {"name": "evidence_of_inadequacy", "type": "TEXT[]", "constraints": "", "desc": "Empirical challenges to concept"},
        {"name": "why_persists", "type": "TEXT", "constraints": "", "desc": "Ideological/class function"},
        {"name": "rupture_would_enable", "type": "TEXT", "constraints": "", "desc": "What becomes thinkable after rupture"},
        {"name": "rupture_trigger", "type": "TEXT", "constraints": "", "desc": "What would force abandonment"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "is_obstacle": True,
        "obstacle_type": "verbal",
        "what_it_blocks": '["Recognition of structural dependency", "Analysis of power asymmetries", "Questions about achievability"]',
        "evidence_of_inadequacy": '["Gulf $2T but can\'t invest at Series B", "China depends on ASML", "Russia sanctions reveal dependencies"]',
        "why_persists": "Legitimizes state spending, creates protected markets, provides nationalist narrative",
        "rupture_would_enable": "Honest analysis of achievable autonomy, dependency management strategies",
        "rupture_trigger": "Sustained demonstration that investment doesn't produce autonomy",
    }
)

# ============================================================================
# Sheet 9: concept_evolution (Canguilhem)
# ============================================================================
add_table_sheet(
    wb, "8. Evolution (Canguilhem)", "Canguilhem", "concept_evolution",
    "Historical transformations of the concept over time",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "period", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "Time period of transformation"},
        {"name": "transformation_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What changed"},
        {"name": "problem_driving", "type": "TEXT", "constraints": "", "desc": "What problem drove this change"},
        {"name": "who_transformed", "type": "TEXT", "constraints": "", "desc": "Intellectual tradition, institution"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "period": "1648 (Westphalia)",
        "transformation_description": "Territorial sovereignty - supreme authority over bounded territory",
        "problem_driving": "Ending religious wars, establishing state system",
        "who_transformed": "European diplomats, emerging state system",
    }
)

ws = wb["8. Evolution (Canguilhem)"]
row = ws.max_row + 2
ws.cell(row=row, column=1, value="ADDITIONAL EXAMPLE RECORDS:")
ws.cell(row=row, column=1).font = Font(bold=True)
examples = [
    ("1950s-70s", "Economic sovereignty - control of resources and industry", "Decolonization", "Newly independent states"),
    ("1990s-2000s", "Data sovereignty - control over information flows", "Internet, privacy concerns", "EU regulators"),
    ("2010s-present", "Technological sovereignty - control over tech development", "US-China rivalry, Snowden", "Tech nationalists"),
]
row += 1
for period, desc, problem, who in examples:
    ws.cell(row=row, column=1, value=period)
    ws.cell(row=row, column=2, value=desc)
    ws.cell(row=row, column=3, value=problem)
    ws.cell(row=row, column=4, value=who)
    row += 1

# ============================================================================
# Sheet 10: concept_normative_dimensions (Canguilhem)
# ============================================================================
add_table_sheet(
    wb, "9. Normative (Canguilhem)", "Canguilhem", "concept_normative_dimensions",
    "Values embedded in the concept and whose interests they serve",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "value_embedded", "type": "TEXT", "constraints": "NOT NULL", "desc": "What value is embedded"},
        {"name": "whose_values", "type": "TEXT", "constraints": "", "desc": "Whose interests does this serve"},
        {"name": "what_excluded", "type": "TEXT", "constraints": "", "desc": "What's marked as 'abnormal' or excluded"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "value_embedded": "State control over technology is good",
        "whose_values": "State elites, security apparatus, domestic tech firms",
        "what_excluded": "Individual autonomy, corporate freedom, open collaboration",
    }
)

# ============================================================================
# Sheet 11: concept_reasoning_styles (Davidson)
# ============================================================================
add_table_sheet(
    wb, "10. Styles (Davidson)", "Davidson", "concept_reasoning_styles",
    "Reasoning styles required/enabled by the concept",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "style_required", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "What reasoning style is needed"},
        {"name": "what_visible", "type": "TEXT[]", "constraints": "", "desc": "What this style makes visible"},
        {"name": "what_invisible", "type": "TEXT[]", "constraints": "", "desc": "What's systematically hidden"},
        {"name": "evidence_types_privileged", "type": "TEXT[]", "constraints": "", "desc": "What counts as evidence"},
        {"name": "inference_patterns", "type": "TEXT[]", "constraints": "", "desc": "Characteristic reasoning moves"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "style_required": "geopolitical_realism",
        "what_visible": '["State competition", "Strategic sectors", "National interest", "Security vulnerabilities"]',
        "what_invisible": '["Class dynamics", "Corporate interests", "Impossibility of autonomy", "Labor flows"]',
        "evidence_types_privileged": '["State investments", "Trade statistics", "Military applications", "Government statements"]',
        "inference_patterns": '["If dependent, then vulnerable", "If strategic, must control", "If rival invests, we must too"]',
    }
)

# ============================================================================
# Sheet 12: concept_metaphors (Blumenberg)
# ============================================================================
add_table_sheet(
    wb, "11. Metaphors (Blumenberg)", "Blumenberg", "concept_metaphors",
    "Root metaphors that structure the concept",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "root_metaphor", "type": "TEXT", "constraints": "NOT NULL", "desc": "The metaphor itself"},
        {"name": "source_domain", "type": "TEXT", "constraints": "", "desc": "Where metaphor comes from"},
        {"name": "what_metaphor_enables", "type": "TEXT[]", "constraints": "", "desc": "What thinking it makes possible"},
        {"name": "what_metaphor_hides", "type": "TEXT[]", "constraints": "", "desc": "What it obscures"},
        {"name": "resists_conceptualization", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Is this an 'absolute metaphor'?"},
        {"name": "why_resists", "type": "TEXT", "constraints": "", "desc": "Why can't it be made precise"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "root_metaphor": "Sovereignty as territorial possession",
        "source_domain": "Territorial/military control",
        "what_metaphor_enables": '["Thinking tech can be owned like territory", "Border concepts for data", "Exclusion as defense"]',
        "what_metaphor_hides": '["Tech is networked not bounded", "Requires ongoing relationships", "Knowledge flows through people"]',
        "resists_conceptualization": True,
        "why_resists": "Territory and technology have fundamentally different autonomy conditions",
    }
)

# ============================================================================
# Sheet 13: concept_work_in_progress (Blumenberg)
# ============================================================================
add_table_sheet(
    wb, "12. Work (Blumenberg)", "Blumenberg", "concept_work_in_progress",
    "Conceptual work being done on the concept - attempts to transform it",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
        {"name": "original_meaning", "type": "TEXT", "constraints": "", "desc": "What concept originally meant"},
        {"name": "current_work", "type": "TEXT", "constraints": "NOT NULL", "desc": "What transformation is being attempted"},
        {"name": "who_doing_work", "type": "TEXT", "constraints": "", "desc": "Intellectual tradition, actors"},
        {"name": "work_status", "type": "VARCHAR(30)", "constraints": "", "desc": "succeeding/failing/ongoing"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "original_meaning": "Sovereignty: supreme authority over territory",
        "current_work": "Extending territorial sovereignty logic to technological domain",
        "who_doing_work": "Tech nationalists, security establishment, some academics",
        "work_status": "failing",
    }
)

# ============================================================================
# Sheet 14: concept_hierarchy (Carey)
# ============================================================================
add_table_sheet(
    wb, "13. Hierarchy (Carey)", "Carey", "concept_hierarchy",
    "Bootstrap structure - what primitives/concepts this is built from",
    fields=[
        {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "Unique identifier"},
        {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
        {"name": "built_from_concept_ids", "type": "INTEGER[]", "constraints": "FK[] → concepts", "desc": "Concepts this is bootstrapped from"},
        {"name": "built_from_descriptions", "type": "TEXT[]", "constraints": "", "desc": "How components combine"},
        {"name": "combination_type", "type": "VARCHAR(30)", "constraints": "", "desc": "simple_aggregation/interactive/qualitative_leap"},
        {"name": "transparency", "type": "VARCHAR(20)", "constraints": "high/medium/low", "desc": "How visible are components"},
        {"name": "bootstrap_failure_reason", "type": "TEXT", "constraints": "", "desc": "If failed, why"},
        {"name": "what_would_fix", "type": "TEXT", "constraints": "", "desc": "What would make bootstrap succeed"},
        {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "Creation timestamp"},
    ],
    example_data={
        "id": 1,
        "concept_id": 42,
        "built_from_concept_ids": "[15, 23]",  # sovereignty, technology
        "built_from_descriptions": '["Sovereignty (state + territory + authority)", "Technology domain application"]',
        "combination_type": "qualitative_leap",
        "transparency": "low",
        "bootstrap_failure_reason": "Territorial sovereignty logic doesn't transfer to technology domain",
        "what_would_fix": "New concept: 'Strategic positioning' built from network position + leverage + influence",
    }
)

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D.xlsx")
wb.save(output_path)
print(f"Concept schema Excel saved to: {output_path}")
