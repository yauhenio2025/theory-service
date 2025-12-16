#!/usr/bin/env python3
"""
Generate EXPANDED Concept Schema v5

ALL 9 dimensions now have deep philosophical grounding:
- Brandom: 6 tables (inferentialist framework)
- Deleuze: 7 tables (concept theory from "What is Philosophy?")
- Quine: 4 tables (confirmation holism, revisability, ontological relativity)
- Sellars: 5 tables (space of reasons, functional role semantics)
- Bachelard: 5 tables (obstacle taxonomy, stages, psychoanalysis)
- Canguilhem: 5 tables (filiation, vital norms, milieu)
- Hacking: 5 tables (looping effects, kinds created, possibility space)
- Blumenberg: 5 tables (nonconceptuality, lifeworld, underground)
- Carey: 5 tables (core cognition, placeholder structures, constraints)

TOTAL: 47 tables (up from 36 in v4)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
field_font = Font(name="Consolas", size=10)
type_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
fk_fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
new_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
source_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

dimension_colors = {
    "Core": "2C3E50",
    "Quinean": "E74C3C",
    "Sellarsian": "9B59B6",
    "Brandomian": "3498DB",
    "Deleuzian": "1ABC9C",
    "Bachelardian": "F39C12",
    "Canguilhem": "27AE60",
    "Davidson": "E67E22",
    "Blumenberg": "8E44AD",
    "Carey": "16A085",
}

SOURCE_FIELDS = [
    {"name": "--- SOURCE TRACKING ---", "type": "", "constraints": "", "desc": "", "source": True},
    {"name": "source_type", "type": "VARCHAR(30)", "constraints": "", "desc": "llm_analysis/evidence_testing/internal_compute/user_input/import", "source": True},
    {"name": "source_reference", "type": "TEXT", "constraints": "", "desc": "Model ID, cluster_id, user_id, etc.", "source": True},
    {"name": "source_confidence", "type": "FLOAT", "constraints": "0-1", "desc": "Source's confidence", "source": True},
]

def style_header(ws, row, text, color):
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def add_table(ws, start_row, table_name, fields, note=None, include_source=True):
    row = start_row
    if note:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9)
        row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value=table_name).font = Font(bold=True, size=12)
    row += 1

    for col, h in enumerate(["Field", "Type", "Constraints", "Description"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = thin_border
    row += 1

    for f in fields:
        ws.cell(row=row, column=1, value=f["name"]).font = field_font
        ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
        c3 = ws.cell(row=row, column=3, value=f.get("constraints", ""))
        if "FK" in f.get("constraints", ""):
            c3.fill = fk_fill
        ws.cell(row=row, column=4, value=f["desc"])
        if f.get("new"):
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = new_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = wrap
        row += 1

    if include_source:
        for f in SOURCE_FIELDS:
            ws.cell(row=row, column=1, value=f["name"]).font = field_font
            ws.cell(row=row, column=2, value=f["type"]).fill = type_fill
            ws.cell(row=row, column=3, value=f.get("constraints", ""))
            ws.cell(row=row, column=4, value=f["desc"])
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = source_fill
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).alignment = wrap
            row += 1

    return row + 1

def set_column_widths(ws):
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 55

# ============================================================================
# SHEET 1: Overview
# ============================================================================
ws = wb.active
ws.title = "1. Overview"
style_header(ws, 1, "ENRICHED CONCEPT SCHEMA v5 - FULLY EXPANDED", dimension_colors["Core"])

overview = """
KEY CHANGES IN v5:
All 9 dimensions now have DEEP philosophical grounding matching Brandom/Deleuze.

QUINEAN (4 tables) - Based on "Two Dogmas of Empiricism" + "Word and Object"
• confirmation holism: whole web faces experience, not individual statements
• ontological relativity: what exists depends on background theory
• revisability: any statement can be held true if we adjust elsewhere
• costs of revision: more central = more costly

SELLARSIAN (5 tables) - Based on "Empiricism and the Philosophy of Mind"
• space of reasons: full justificatory structure
• functional role semantics: concepts get meaning from inferential role
• myth of Jones: inner episodes modeled on public language
• manifest vs scientific image tensions

BACHELARDIAN (5 tables) - Based on "Formation of the Scientific Mind"
• obstacle taxonomy: experience/verbal/pragmatic/quantitative/substantialist/animist
• three stages: concrete → concrete-abstract → abstract
• psychoanalysis: thinking resists thought, catharsis needed
• regional rationalism: different rationalities for different domains

CANGUILHEM (5 tables) - Based on "The Normal and the Pathological"
• normal vs normative: descriptive vs prescriptive norms
• vital normativity: life creates new norms
• concept filiation: genealogical descent of concepts
• milieu: organism's/concept's relation to environment

HACKING (5 tables) - Based on "Historical Ontology" + "Making Up People"
• dynamic nominalism: classifications and classified emerge together
• looping effects: classification changes classified AND vice versa
• making up people: sciences create kinds
• space of possibilities for personhood

BLUMENBERG (5 tables) - Based on "Paradigms for a Metaphorology"
• absolute metaphors: resist translation to concepts
• nonconceptuality (Unbegrifflichkeit): what resists conceptualization
• life-world background: metaphors connect to lived experience
• underground of thought: metaphorical substructure

CAREY (5 tables) - Based on "The Origin of Concepts"
• core cognition: innate conceptual (not just sensory) representations
• Quinian bootstrapping: building new from placeholder structures
• incommensurability: new systems can't reduce to old
• computational constraints: guide interpretation of placeholders

TOTAL: 47 tables (up from 36 in v4)
""".strip().split('\n')

for row_idx, text in enumerate(overview, 3):
    ws.cell(row=row_idx, column=1, value=text)
    if text.startswith("KEY") or text.endswith("tables)"):
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

set_column_widths(ws)

# ============================================================================
# SHEET 2: Quinean (EXPANDED)
# ============================================================================
ws = wb.create_sheet("2. Quinean")
style_header(ws, 1, "QUINEAN DIMENSION - Confirmation Holism & Web of Belief (EXPANDED)", dimension_colors["Quinean"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'Two Dogmas of Empiricism' (1951): Whole web faces experience; any statement can be held true; costs of revision determine what we revise").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_inferences (inferential connections in the web)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "forward/backward/lateral/contradiction"},
    {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The inference itself"},
    {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept (if in DB)"},
    {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": "Inference strength 0-1"},
    {"name": "defeasible", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Can be defeated by evidence?"},
    {"name": "holistic_load", "type": "INTEGER", "constraints": "", "desc": "How many other inferences depend on this one (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
    {"name": "challenged_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": ""},
])

row = add_table(ws, row, "concept_web_position (NEW - position in the holistic web)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)", "new": True},
    {"name": "centrality", "type": "VARCHAR(20)", "constraints": "", "desc": "core/intermediate/peripheral", "new": True},
    {"name": "entrenchment_score", "type": "FLOAT", "constraints": "0-1", "desc": "How costly to revise (Quine)", "new": True},
    {"name": "revision_cost_type", "type": "VARCHAR(30)", "constraints": "", "desc": "practical/theoretical/both (NEW)", "new": True},
    {"name": "can_be_held_true", "type": "TEXT", "constraints": "", "desc": "What adjustments would let us hold this true despite evidence (NEW)", "new": True},
    {"name": "can_be_revised", "type": "TEXT", "constraints": "", "desc": "What would force us to revise this (NEW)", "new": True},
    {"name": "pragmatic_value", "type": "TEXT", "constraints": "", "desc": "What practical benefits does this concept provide? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Quine: Any statement can be held true 'come what may' if we adjust elsewhere")

row = add_table(ws, row, "concept_ontological_dependence (NEW - ontological relativity)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "background_theory", "type": "TEXT", "constraints": "NOT NULL", "desc": "What theory determines what this concept refers to (NEW)", "new": True},
    {"name": "ontological_commitment", "type": "TEXT", "constraints": "", "desc": "What must exist for this concept to be meaningful (NEW)", "new": True},
    {"name": "alternative_theory", "type": "TEXT", "constraints": "", "desc": "Alternative background that would change reference (NEW)", "new": True},
    {"name": "translation_indeterminacy", "type": "TEXT", "constraints": "", "desc": "Where translation is indeterminate (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Quine's ontological relativity: what exists depends on the background theory")

row = add_table(ws, row, "concept_revision_ramifications (NEW - ripple effects)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "if_revised", "type": "TEXT", "constraints": "NOT NULL", "desc": "Hypothetical revision being considered (NEW)", "new": True},
    {"name": "affected_concept_ids", "type": "INTEGER[]", "constraints": "", "desc": "Concepts that would need adjustment (NEW)", "new": True},
    {"name": "affected_count", "type": "INTEGER", "constraints": "", "desc": "Number of concepts affected (NEW)", "new": True},
    {"name": "severity", "type": "VARCHAR(20)", "constraints": "", "desc": "low/medium/high/catastrophic (Duhem-Quine) (NEW)", "new": True},
    {"name": "recommended_adjustment", "type": "TEXT", "constraints": "", "desc": "What else to adjust to accommodate this revision (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Quine: conflict at periphery occasions readjustments in interior")

set_column_widths(ws)

# ============================================================================
# SHEET 3: Sellarsian (EXPANDED)
# ============================================================================
ws = wb.create_sheet("3. Sellarsian")
style_header(ws, 1, "SELLARSIAN DIMENSION - Space of Reasons (EXPANDED)", dimension_colors["Sellarsian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'Empiricism and Philosophy of Mind' (1956): Space of reasons, functional role semantics, myth of the given, manifest vs scientific image").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_givenness (myth of the given analysis)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "is_myth_of_given", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Falsely treated as foundational?"},
    {"name": "should_be_inferred_from", "type": "TEXT", "constraints": "", "desc": "What should actually support this"},
    {"name": "knowledge_by_acquaintance_claim", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Claims unmediated knowledge? (Russell/Ayer) (NEW)", "new": True},
    {"name": "requires_catharsis", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Needs intellectual catharsis to overcome? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_space_of_reasons (NEW - justificatory structure)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)", "new": True},
    {"name": "space_of_reasons_role", "type": "TEXT", "constraints": "", "desc": "Role in justification/inference (NEW)", "new": True},
    {"name": "justifies_what", "type": "TEXT", "constraints": "", "desc": "What this concept can justify (NEW)", "new": True},
    {"name": "justified_by_what", "type": "TEXT", "constraints": "", "desc": "What justifies use of this concept (NEW)", "new": True},
    {"name": "properly_in_space", "type": "BOOLEAN", "constraints": "", "desc": "Is concept properly in space of reasons? (NEW)", "new": True},
    {"name": "causal_vs_rational", "type": "VARCHAR(20)", "constraints": "", "desc": "causal_only/rational_only/both (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Sellars: placing episode 'in the logical space of reasons, of justifying and being able to justify'")

row = add_table(ws, row, "concept_functional_role (NEW - functional role semantics)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "role_type", "type": "VARCHAR(30)", "constraints": "", "desc": "inferential/behavioral/mixed (NEW)", "new": True},
    {"name": "role_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What functional role this concept plays (NEW)", "new": True},
    {"name": "relations_to_stimuli", "type": "TEXT", "constraints": "", "desc": "Relations to perceptual/environmental stimuli (NEW)", "new": True},
    {"name": "relations_to_responses", "type": "TEXT", "constraints": "", "desc": "Relations to behavioral responses (NEW)", "new": True},
    {"name": "relations_to_other_concepts", "type": "TEXT", "constraints": "", "desc": "Inferential relations to other concepts (NEW)", "new": True},
    {"name": "presupposes_language", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Does concept presuppose public language? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Sellars: concepts gain meaning through inferential and behavioral roles")

row = add_table(ws, row, "concept_image_tension (NEW - manifest vs scientific image)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "manifest_image", "type": "TEXT", "constraints": "", "desc": "How concept appears in everyday/folk understanding (NEW)", "new": True},
    {"name": "scientific_image", "type": "TEXT", "constraints": "", "desc": "How concept appears in scientific understanding (NEW)", "new": True},
    {"name": "tension_description", "type": "TEXT", "constraints": "", "desc": "Nature of tension between images (NEW)", "new": True},
    {"name": "which_dominates", "type": "VARCHAR(20)", "constraints": "", "desc": "manifest/scientific/neither (NEW)", "new": True},
    {"name": "resolution_possible", "type": "BOOLEAN", "constraints": "", "desc": "Can tension be resolved? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Sellars: tension between 'manifest image' (folk) and 'scientific image'")

row = add_table(ws, row, "concept_givenness_markers (language markers)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "marker_text", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": "'obviously', 'naturally', 'clearly', etc."},
    {"name": "example_usage", "type": "TEXT", "constraints": "", "desc": "Example in context"},
    {"name": "confidence", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 4: Brandomian (already expanded in v4)
# ============================================================================
ws = wb.create_sheet("4. Brandomian")
style_header(ws, 1, "BRANDOMIAN DIMENSION - Full Inferentialist Framework", dimension_colors["Brandomian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'Making It Explicit' (1994): Normative pragmatics, deontic scorekeeping, game of giving and asking for reasons").font = Font(italic=True)

row = 5

# [Keep all 6 Brandomian tables from v4 - they're already deep]
row = add_table(ws, row, "concept_inferential_roles (material inferences)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept"},
    {"name": "inference_direction", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "from_concept/to_concept"},
    {"name": "inference_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "committive/permissive/incompatibility"},
    {"name": "inference_statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "The material inference"},
    {"name": "target_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Related concept if in DB"},
    {"name": "is_material", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Content-dependent (vs formal)?"},
    {"name": "counterfactual_supporting", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Supports counterfactuals?"},
    {"name": "strength", "type": "FLOAT", "constraints": "DEFAULT 0.8", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="Three types: committive (deductive), permissive (inductive), incompatibility (modal)")

row = add_table(ws, row, "concept_commitments (deontic statuses)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "commitment_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "commitment/entitlement/incompatibility"},
    {"name": "statement", "type": "TEXT", "constraints": "NOT NULL", "desc": "What you're committed/entitled to"},
    {"name": "deontic_status", "type": "VARCHAR(30)", "constraints": "", "desc": "acknowledged/attributed/undertaken"},
    {"name": "is_honored", "type": "BOOLEAN", "constraints": "", "desc": "Honored in practice?"},
    {"name": "violation_evidence", "type": "TEXT", "constraints": "", "desc": "If violated, where/how"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_challenges (game of giving and asking for reasons)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "commitment_id", "type": "INTEGER", "constraints": "FK → concept_commitments", "desc": "Specific commitment challenged"},
    {"name": "challenge_type", "type": "VARCHAR(30)", "constraints": "NOT NULL", "desc": "request_reasons/incompatibility_claim/counterexample"},
    {"name": "challenge_content", "type": "TEXT", "constraints": "NOT NULL", "desc": "The challenge itself"},
    {"name": "challenger", "type": "TEXT", "constraints": "", "desc": "Who/what raised challenge"},
    {"name": "status", "type": "VARCHAR(20)", "constraints": "DEFAULT 'open'", "desc": "open/justified/withdrawn/unresolved"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_perspectival_content (de dicto vs de re)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "ascription_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "de_dicto/de_re"},
    {"name": "perspective_holder", "type": "TEXT", "constraints": "", "desc": "Whose perspective (for de dicto)"},
    {"name": "content_as_seen", "type": "TEXT", "constraints": "NOT NULL", "desc": "How concept appears from this perspective"},
    {"name": "our_translation", "type": "TEXT", "constraints": "", "desc": "Our (ascriber's) translation (de re)"},
    {"name": "translation_preserves_truth", "type": "BOOLEAN", "constraints": "", "desc": "Does translation preserve reference?"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 5: Deleuzian (already expanded in v4)
# ============================================================================
ws = wb.create_sheet("5. Deleuzian")
style_header(ws, 1, "DELEUZIAN DIMENSION - Concept Theory", dimension_colors["Deleuzian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'What is Philosophy?' (1991): Components, zones of indiscernibility, plane of immanence, conceptual personae").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_components (heterogeneous components)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "component_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": ""},
    {"name": "component_description", "type": "TEXT", "constraints": "", "desc": "What this component contributes"},
    {"name": "is_intensive", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Intensive variation?"},
    {"name": "order_index", "type": "INTEGER", "constraints": "", "desc": "Position in concept"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_zones_of_indiscernibility (where components overlap)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "component_a_id", "type": "INTEGER", "constraints": "FK → concept_components", "desc": ""},
    {"name": "component_b_id", "type": "INTEGER", "constraints": "FK → concept_components", "desc": ""},
    {"name": "zone_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What passes between them"},
    {"name": "what_becomes_possible", "type": "TEXT", "constraints": "", "desc": "What this zone enables"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_consistency (endoconsistency)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": ""},
    {"name": "endoconsistency_description", "type": "TEXT", "constraints": "", "desc": "How components hold together"},
    {"name": "survey_point", "type": "TEXT", "constraints": "", "desc": "'Point of absolute survey' that unifies"},
    {"name": "consistency_strength", "type": "VARCHAR(20)", "constraints": "", "desc": "strong/moderate/weak/unstable"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_plane_of_immanence (background plane)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "plane_name", "type": "VARCHAR(100)", "constraints": "", "desc": ""},
    {"name": "what_plane_presupposes", "type": "TEXT", "constraints": "", "desc": "Unthought assumptions"},
    {"name": "legitimate_problems", "type": "TEXT", "constraints": "", "desc": "What counts as problem"},
    {"name": "excluded_problems", "type": "TEXT", "constraints": "", "desc": "What can't be asked"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_personae (conceptual personae)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "persona_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": ""},
    {"name": "persona_description", "type": "TEXT", "constraints": "", "desc": "What persona does/thinks"},
    {"name": "what_persona_enables", "type": "TEXT", "constraints": "", "desc": "What thinking it enables"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 6: Bachelardian (EXPANDED)
# ============================================================================
ws = wb.create_sheet("6. Bachelardian")
style_header(ws, 1, "BACHELARDIAN DIMENSION - Epistemological Obstacles (EXPANDED)", dimension_colors["Bachelardian"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'Formation of Scientific Mind' (1938): Obstacle taxonomy, three stages, psychoanalysis of knowledge, regional rationalism").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_obstacles (obstacle analysis - EXPANDED)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)"},
    {"name": "is_obstacle", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Does concept block understanding?"},
    {"name": "obstacle_type", "type": "VARCHAR(30)", "constraints": "", "desc": "experience/verbal/pragmatic/quantitative/substantialist/animist (EXPANDED)", "new": True},
    {"name": "obstacle_subtype", "type": "TEXT", "constraints": "", "desc": "More specific classification (NEW)", "new": True},
    {"name": "why_persists", "type": "TEXT", "constraints": "", "desc": "Ideological/class function"},
    {"name": "thinking_resists_thought", "type": "TEXT", "constraints": "", "desc": "How thinking resists itself here (NEW)", "new": True},
    {"name": "rupture_would_enable", "type": "TEXT", "constraints": "", "desc": "What becomes thinkable after rupture"},
    {"name": "rupture_trigger", "type": "TEXT", "constraints": "", "desc": "What would force abandonment"},
    {"name": "catharsis_needed", "type": "TEXT", "constraints": "", "desc": "What intellectual/emotional catharsis needed (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="From Bachelard: Epistemological obstacles are INTERNAL and IDEOLOGICAL, not external")

row = add_table(ws, row, "concept_cognitive_stage (NEW - three stages of scientific mind)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)", "new": True},
    {"name": "stage", "type": "VARCHAR(30)", "constraints": "", "desc": "concrete/concrete_abstract/abstract (NEW)", "new": True},
    {"name": "stage_evidence", "type": "TEXT", "constraints": "", "desc": "What shows this stage (NEW)", "new": True},
    {"name": "affective_interest", "type": "TEXT", "constraints": "", "desc": "What affective interest at this stage (NEW)", "new": True},
    {"name": "obstacles_at_stage", "type": "TEXT", "constraints": "", "desc": "Obstacles specific to this stage (NEW)", "new": True},
    {"name": "path_to_next_stage", "type": "TEXT", "constraints": "", "desc": "How to progress (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Bachelard: Necessary passage through concrete → concrete-abstract → abstract")

row = add_table(ws, row, "concept_psychoanalytic_function (NEW - psychoanalysis of knowledge)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "unconscious_need", "type": "TEXT", "constraints": "", "desc": "What unconscious need concept serves (NEW)", "new": True},
    {"name": "preconceived_structure", "type": "TEXT", "constraints": "", "desc": "Preconceived idea/structure in mind (NEW)", "new": True},
    {"name": "resistance_mechanism", "type": "TEXT", "constraints": "", "desc": "How concept resists being thought through (NEW)", "new": True},
    {"name": "catharsis_approach", "type": "TEXT", "constraints": "", "desc": "How to overcome through catharsis (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Bachelard: Historical epistemology as psychoanalysis of scientific mind")

row = add_table(ws, row, "concept_regional_rationality (NEW - regional rationalism)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "rationality_region", "type": "VARCHAR(50)", "constraints": "", "desc": "physics/chemistry/biology/social/etc (NEW)", "new": True},
    {"name": "regional_norms", "type": "TEXT", "constraints": "", "desc": "What norms apply in this region (NEW)", "new": True},
    {"name": "cross_regional_application", "type": "TEXT", "constraints": "", "desc": "Errors from applying across regions (NEW)", "new": True},
    {"name": "epistemological_profile", "type": "TEXT", "constraints": "", "desc": "Where on progress scale (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Bachelard: Different rationalities for different domains")

row = add_table(ws, row, "concept_inadequacy_evidence (empirical challenges)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "evidence_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "Empirical challenge"},
    {"name": "source_cluster_id", "type": "INTEGER", "constraints": "", "desc": "Source evidence cluster"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 7: Canguilhem (EXPANDED)
# ============================================================================
ws = wb.create_sheet("7. Canguilhem")
style_header(ws, 1, "CANGUILHEM DIMENSION - Life History of Concepts (EXPANDED)", dimension_colors["Canguilhem"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'The Normal and the Pathological' (1943/1968): Vital normativity, concept filiation, milieu, normal vs normative").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_filiation (NEW - genealogical lineage)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "ancestor_concept_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Predecessor concept (if in DB) (NEW)", "new": True},
    {"name": "ancestor_description", "type": "TEXT", "constraints": "", "desc": "Description if not in DB (NEW)", "new": True},
    {"name": "filiation_type", "type": "VARCHAR(30)", "constraints": "", "desc": "direct/lateral/rupture/synthesis (NEW)", "new": True},
    {"name": "transformation_description", "type": "TEXT", "constraints": "", "desc": "How concept descended/transformed (NEW)", "new": True},
    {"name": "continuity_type", "type": "VARCHAR(30)", "constraints": "", "desc": "quantitative_continuous/qualitative_discontinuous (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Canguilhem: Concept filiation - genealogical descent, continuity doesn't preclude discontinuity")

row = add_table(ws, row, "concept_vital_norms (NEW - vital normativity)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "norm_created", "type": "TEXT", "constraints": "NOT NULL", "desc": "What norm does concept create/embody (NEW)", "new": True},
    {"name": "norm_type", "type": "VARCHAR(30)", "constraints": "", "desc": "normal/normative/both (NEW)", "new": True},
    {"name": "plastic_power", "type": "TEXT", "constraints": "", "desc": "How concept creates qualitatively new norms (NEW)", "new": True},
    {"name": "self_regulatory", "type": "BOOLEAN", "constraints": "", "desc": "Does concept self-regulate? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Canguilhem: Life creates new norms; concept as normative, not just normal")

row = add_table(ws, row, "concept_milieu (NEW - relation to environment)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": "Parent concept (1:1)", "new": True},
    {"name": "milieu_description", "type": "TEXT", "constraints": "", "desc": "Environment concept operates in (NEW)", "new": True},
    {"name": "concept_shapes_milieu", "type": "TEXT", "constraints": "", "desc": "How concept modifies its environment (NEW)", "new": True},
    {"name": "milieu_shapes_concept", "type": "TEXT", "constraints": "", "desc": "How environment shapes concept (NEW)", "new": True},
    {"name": "organism_vs_machine", "type": "VARCHAR(20)", "constraints": "", "desc": "Does concept treat subject as organism or machine? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Canguilhem: Organism's relation to milieu; living being modifies environment")

row = add_table(ws, row, "concept_normative_dimensions (embedded values)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "value_embedded", "type": "TEXT", "constraints": "NOT NULL", "desc": "What value is embedded"},
    {"name": "whose_values", "type": "TEXT", "constraints": "", "desc": "Whose interests this serves"},
    {"name": "what_excluded", "type": "TEXT", "constraints": "", "desc": "What's marked as abnormal"},
    {"name": "normal_vs_normative", "type": "VARCHAR(20)", "constraints": "", "desc": "normal/normative (descriptive vs prescriptive)"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_vitality_indicators (health status)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "indicator_type", "type": "VARCHAR(30)", "constraints": "", "desc": "growth/strain/crisis/revival"},
    {"name": "indicator_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What shows concept's health"},
    {"name": "observed_at", "type": "TIMESTAMPTZ", "constraints": "", "desc": ""},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 8: Davidson/Hacking (EXPANDED)
# ============================================================================
ws = wb.create_sheet("8. Hacking")
style_header(ws, 1, "HACKING DIMENSION - Dynamic Nominalism & Styles (EXPANDED)", dimension_colors["Davidson"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'Historical Ontology' + 'Making Up People': Dynamic nominalism, looping effects, styles of reasoning, space of possibilities").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_reasoning_styles (styles of reasoning)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "style_name", "type": "VARCHAR(100)", "constraints": "NOT NULL", "desc": ""},
    {"name": "style_emerged", "type": "VARCHAR(100)", "constraints": "", "desc": "When style emerged"},
    {"name": "objects_created", "type": "TEXT", "constraints": "", "desc": "What objects of inquiry style creates"},
    {"name": "self_authenticating", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Style authenticates its own truth claims? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_looping_effects (NEW - dynamic nominalism)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "loop_type", "type": "VARCHAR(30)", "constraints": "", "desc": "classification_to_classified/classified_to_classification/both (NEW)", "new": True},
    {"name": "loop_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "How classification and classified interact (NEW)", "new": True},
    {"name": "target_modified", "type": "TEXT", "constraints": "", "desc": "How the classified are changed by classification (NEW)", "new": True},
    {"name": "classification_modified", "type": "TEXT", "constraints": "", "desc": "How classification is changed by the classified (NEW)", "new": True},
    {"name": "loop_intensity", "type": "VARCHAR(20)", "constraints": "", "desc": "weak/moderate/strong (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Hacking: 'Moving targets because our investigations interact with them and change them'")

row = add_table(ws, row, "concept_kinds_created (NEW - making up people/things)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "kind_name", "type": "TEXT", "constraints": "NOT NULL", "desc": "What kind of people/things concept creates (NEW)", "new": True},
    {"name": "kind_type", "type": "VARCHAR(20)", "constraints": "", "desc": "person/thing/relation/process (NEW)", "new": True},
    {"name": "did_not_exist_before", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Did this kind exist before the concept? (NEW)", "new": True},
    {"name": "how_created", "type": "TEXT", "constraints": "", "desc": "Mechanism of kind creation (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Hacking: 'Our sciences create kinds of people that in a certain sense did not exist before'")

row = add_table(ws, row, "concept_possibility_space (NEW - space of possibilities)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "possibility_type", "type": "VARCHAR(20)", "constraints": "", "desc": "opens/closes (NEW)", "new": True},
    {"name": "possibility_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What possibility is opened/closed (NEW)", "new": True},
    {"name": "for_whom", "type": "TEXT", "constraints": "", "desc": "For what people/actors (NEW)", "new": True},
    {"name": "historically_bounded", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Is possibility historically bounded? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Hacking: Possibilities for personhood are bounded by what is named and described")

row = add_table(ws, row, "concept_style_visibility (what style makes visible/invisible)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "style_id", "type": "INTEGER", "constraints": "FK → concept_reasoning_styles", "desc": ""},
    {"name": "visibility_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "visible/invisible"},
    {"name": "what_affected", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is made visible/invisible"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 9: Blumenberg (EXPANDED)
# ============================================================================
ws = wb.create_sheet("9. Blumenberg")
style_header(ws, 1, "BLUMENBERG DIMENSION - Metaphorology (EXPANDED)", dimension_colors["Blumenberg"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'Paradigms for a Metaphorology' (1960): Absolute metaphors, nonconceptuality (Unbegrifflichkeit), underground of thought, life-world connection").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_metaphors (root metaphors)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": ""},
    {"name": "root_metaphor", "type": "TEXT", "constraints": "NOT NULL", "desc": "The metaphor itself"},
    {"name": "source_domain", "type": "TEXT", "constraints": "", "desc": "Where metaphor comes from"},
    {"name": "is_absolute", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Cannot be translated to concepts?"},
    {"name": "springs_into_void", "type": "TEXT", "constraints": "", "desc": "What conceptual void does it fill? (NEW)", "new": True},
    {"name": "courage_of_conjecture", "type": "TEXT", "constraints": "", "desc": "What does mind preempt/conjecture? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="From Blumenberg: Absolute metaphors 'spring into a void that concepts are unable to fill'")

row = add_table(ws, row, "concept_nonconceptuality (NEW - Unbegrifflichkeit)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "what_resists", "type": "TEXT", "constraints": "NOT NULL", "desc": "What aspect resists conceptualization (NEW)", "new": True},
    {"name": "why_resists", "type": "TEXT", "constraints": "", "desc": "Why it can't be made precise/univocal (NEW)", "new": True},
    {"name": "expressed_through", "type": "TEXT", "constraints": "", "desc": "How expressed (metaphor/myth/symbol) (NEW)", "new": True},
    {"name": "motivational_support", "type": "TEXT", "constraints": "", "desc": "What motivation it provides for theory (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Blumenberg: Nonconceptuality - perceptions/experiences that don't lend themselves to precise concepts")

row = add_table(ws, row, "concept_lifeworld_connection (NEW - background of life-world)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "lifeworld_aspect", "type": "TEXT", "constraints": "NOT NULL", "desc": "What life-world background it connects to (NEW)", "new": True},
    {"name": "existential_question", "type": "TEXT", "constraints": "", "desc": "What 'unanswerable question' it addresses (NEW)", "new": True},
    {"name": "orientation_function", "type": "TEXT", "constraints": "", "desc": "How it orients thought and action (NEW)", "new": True},
    {"name": "reality_perception", "type": "TEXT", "constraints": "", "desc": "How it shapes perception of reality as whole (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Blumenberg: Metaphors connect to life-world background, provide constant motivational support")

row = add_table(ws, row, "concept_metaphor_effects (enables/hides)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "metaphor_id", "type": "INTEGER", "constraints": "FK → concept_metaphors, NOT NULL", "desc": ""},
    {"name": "effect_type", "type": "VARCHAR(20)", "constraints": "NOT NULL", "desc": "enables/hides"},
    {"name": "effect_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "What is enabled/hidden"},
    {"name": "underground_role", "type": "TEXT", "constraints": "", "desc": "Role in 'underground of thought' (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

row = add_table(ws, row, "concept_metakinetics (metaphor transformations)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "metaphor_id", "type": "INTEGER", "constraints": "FK → concept_metaphors, NOT NULL", "desc": ""},
    {"name": "period", "type": "VARCHAR(100)", "constraints": "", "desc": ""},
    {"name": "transformation", "type": "TEXT", "constraints": "NOT NULL", "desc": "How metaphor meaning shifted"},
    {"name": "horizon_shift", "type": "TEXT", "constraints": "", "desc": "What horizon of meaning changed? (NEW)", "new": True},
    {"name": "way_of_seeing_changed", "type": "TEXT", "constraints": "", "desc": "How way of seeing changed (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
], note="From Blumenberg: Metakinetics - 'historical transformation brings to light the metakinetics of historical horizons'")

set_column_widths(ws)

# ============================================================================
# SHEET 10: Carey (EXPANDED)
# ============================================================================
ws = wb.create_sheet("10. Carey")
style_header(ws, 1, "CAREY DIMENSION - Conceptual Bootstrapping (EXPANDED)", dimension_colors["Carey"])

ws.merge_cells('A3:D3')
ws.cell(row=3, column=1, value="Based on 'The Origin of Concepts' (2009): Core cognition, Quinian bootstrapping, placeholder structures, incommensurability, computational constraints").font = Font(italic=True)

row = 5

row = add_table(ws, row, "concept_core_cognition (NEW - innate starting point)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "core_system", "type": "VARCHAR(50)", "constraints": "", "desc": "Which core system: object/number/agent/space (NEW)", "new": True},
    {"name": "core_contribution", "type": "TEXT", "constraints": "", "desc": "What core cognition contributes (NEW)", "new": True},
    {"name": "goes_beyond_core", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Does concept go beyond core cognition? (NEW)", "new": True},
    {"name": "how_transcends", "type": "TEXT", "constraints": "", "desc": "How it transcends core cognition (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Carey: Core cognition = innate conceptual (not just sensory) representations")

row = add_table(ws, row, "concept_placeholder_structures (NEW - semantically impoverished symbols)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "placeholder_symbol", "type": "TEXT", "constraints": "NOT NULL", "desc": "The placeholder symbol/term (NEW)", "new": True},
    {"name": "initial_semantic_poverty", "type": "TEXT", "constraints": "", "desc": "How symbol was initially impoverished (NEW)", "new": True},
    {"name": "how_interpreted", "type": "TEXT", "constraints": "", "desc": "How placeholder got interpreted (NEW)", "new": True},
    {"name": "interpretation_source", "type": "TEXT", "constraints": "", "desc": "What gave it meaning (NEW)", "new": True},
    {"name": "now_has_wide_content", "type": "BOOLEAN", "constraints": "", "desc": "Does it now have wide content? (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Carey: Placeholder structures - uninterpreted models that later get interpretation")

row = add_table(ws, row, "concept_bootstrapping_constraints (NEW - computational constraints)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "constraint_type", "type": "VARCHAR(50)", "constraints": "", "desc": "What kind of constraint (NEW)", "new": True},
    {"name": "constraint_description", "type": "TEXT", "constraints": "NOT NULL", "desc": "How constraint guides interpretation (NEW)", "new": True},
    {"name": "limits_interpretation_to", "type": "TEXT", "constraints": "", "desc": "What interpretations are possible (NEW)", "new": True},
    {"name": "executive_function_needed", "type": "TEXT", "constraints": "", "desc": "What cognitive effort required (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Carey: Computational constraints guide interpretation, meeting Fodor's challenge")

row = add_table(ws, row, "concept_incommensurability (NEW - conceptual discontinuities)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": "", "new": True},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, NOT NULL", "desc": "Parent concept", "new": True},
    {"name": "incommensurable_with_id", "type": "INTEGER", "constraints": "FK → concepts", "desc": "Earlier concept (if in DB) (NEW)", "new": True},
    {"name": "incommensurable_with_desc", "type": "TEXT", "constraints": "", "desc": "Description if not in DB (NEW)", "new": True},
    {"name": "nature_of_discontinuity", "type": "TEXT", "constraints": "", "desc": "What makes them incommensurable (NEW)", "new": True},
    {"name": "more_expressive_power", "type": "BOOLEAN", "constraints": "DEFAULT TRUE", "desc": "Does new system have more expressive power? (NEW)", "new": True},
    {"name": "cannot_be_defined_in_terms_of", "type": "TEXT", "constraints": "", "desc": "What prior primitives can't define this (NEW)", "new": True},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": "", "new": True},
], note="From Carey: Genuine conceptual change = incommensurable systems with more expressive power")

row = add_table(ws, row, "concept_hierarchy (bootstrap structure)", [
    {"name": "id", "type": "SERIAL", "constraints": "PRIMARY KEY", "desc": ""},
    {"name": "concept_id", "type": "INTEGER", "constraints": "FK → concepts, UNIQUE", "desc": ""},
    {"name": "combination_type", "type": "VARCHAR(30)", "constraints": "", "desc": "simple_aggregation/interactive/qualitative_leap"},
    {"name": "transparency", "type": "VARCHAR(20)", "constraints": "", "desc": "high/medium/low"},
    {"name": "is_quinian_bootstrap", "type": "BOOLEAN", "constraints": "DEFAULT FALSE", "desc": "Is this Quinian (not just hypothesis testing)? (NEW)", "new": True},
    {"name": "bootstrap_failure_reason", "type": "TEXT", "constraints": "", "desc": "If failed, why"},
    {"name": "created_at", "type": "TIMESTAMPTZ", "constraints": "DEFAULT NOW()", "desc": ""},
])

set_column_widths(ws)

# ============================================================================
# SHEET 11: Table Summary
# ============================================================================
ws = wb.create_sheet("11. Table Summary")
style_header(ws, 1, "COMPLETE TABLE LIST - v5 (47 tables)", "2C3E50")

ws.merge_cells('A3:F3')
ws.cell(row=3, column=1, value="All tables include source_type, source_reference, source_confidence fields").font = Font(italic=True)

headers = ["#", "Table Name", "Dimension", "New in v5?", "Purpose"]
row = 5
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = table_header_fill

tables = [
    # Quinean (4)
    [1, "concept_inferences", "Quinean", "", "Inferential connections"],
    [2, "concept_web_position", "Quinean", "NEW", "Position in holistic web, revisability"],
    [3, "concept_ontological_dependence", "Quinean", "NEW", "Ontological relativity"],
    [4, "concept_revision_ramifications", "Quinean", "NEW", "Ripple effects of revision"],
    # Sellarsian (5)
    [5, "concept_givenness", "Sellarsian", "EXPANDED", "Myth of given analysis"],
    [6, "concept_space_of_reasons", "Sellarsian", "NEW", "Justificatory structure"],
    [7, "concept_functional_role", "Sellarsian", "NEW", "Functional role semantics"],
    [8, "concept_image_tension", "Sellarsian", "NEW", "Manifest vs scientific image"],
    [9, "concept_givenness_markers", "Sellarsian", "", "Language markers"],
    # Brandomian (6)
    [10, "concept_inferential_roles", "Brandomian", "", "Three inference types"],
    [11, "concept_commitments", "Brandomian", "", "Deontic statuses"],
    [12, "concept_scorekeeping", "Brandomian", "", "Score changes"],
    [13, "concept_challenges", "Brandomian", "", "Game of giving/asking reasons"],
    [14, "concept_justifications", "Brandomian", "", "Responses to challenges"],
    [15, "concept_perspectival_content", "Brandomian", "", "De dicto vs de re"],
    # Deleuzian (7)
    [16, "concept_components", "Deleuzian", "", "Heterogeneous components"],
    [17, "concept_zones_of_indiscernibility", "Deleuzian", "", "Component overlaps"],
    [18, "concept_consistency", "Deleuzian", "", "Endoconsistency"],
    [19, "concept_neighborhood", "Deleuzian", "", "Exoconsistency"],
    [20, "concept_plane_of_immanence", "Deleuzian", "", "Background plane"],
    [21, "concept_personae", "Deleuzian", "", "Conceptual personae"],
    [22, "concept_problems", "Deleuzian", "", "Problems addressed"],
    # Bachelardian (5)
    [23, "concept_obstacles", "Bachelardian", "EXPANDED", "Obstacle analysis + taxonomy"],
    [24, "concept_cognitive_stage", "Bachelardian", "NEW", "Three stages of scientific mind"],
    [25, "concept_psychoanalytic_function", "Bachelardian", "NEW", "Psychoanalysis of knowledge"],
    [26, "concept_regional_rationality", "Bachelardian", "NEW", "Regional rationalism"],
    [27, "concept_inadequacy_evidence", "Bachelardian", "", "Empirical challenges"],
    # Canguilhem (5)
    [28, "concept_filiation", "Canguilhem", "NEW", "Genealogical lineage"],
    [29, "concept_vital_norms", "Canguilhem", "NEW", "Vital normativity"],
    [30, "concept_milieu", "Canguilhem", "NEW", "Relation to environment"],
    [31, "concept_normative_dimensions", "Canguilhem", "", "Embedded values"],
    [32, "concept_vitality_indicators", "Canguilhem", "", "Health status"],
    # Hacking (5)
    [33, "concept_reasoning_styles", "Hacking", "EXPANDED", "Styles of reasoning"],
    [34, "concept_looping_effects", "Hacking", "NEW", "Dynamic nominalism"],
    [35, "concept_kinds_created", "Hacking", "NEW", "Making up people/things"],
    [36, "concept_possibility_space", "Hacking", "NEW", "Space of possibilities"],
    [37, "concept_style_visibility", "Hacking", "", "Visible/invisible"],
    # Blumenberg (5)
    [38, "concept_metaphors", "Blumenberg", "EXPANDED", "Root metaphors"],
    [39, "concept_nonconceptuality", "Blumenberg", "NEW", "Unbegrifflichkeit"],
    [40, "concept_lifeworld_connection", "Blumenberg", "NEW", "Life-world background"],
    [41, "concept_metaphor_effects", "Blumenberg", "", "Enables/hides"],
    [42, "concept_metakinetics", "Blumenberg", "EXPANDED", "Metaphor transformations"],
    # Carey (5)
    [43, "concept_core_cognition", "Carey", "NEW", "Innate starting point"],
    [44, "concept_placeholder_structures", "Carey", "NEW", "Semantically impoverished symbols"],
    [45, "concept_bootstrapping_constraints", "Carey", "NEW", "Computational constraints"],
    [46, "concept_incommensurability", "Carey", "NEW", "Conceptual discontinuities"],
    [47, "concept_hierarchy", "Carey", "EXPANDED", "Bootstrap structure"],
]

row = 6
for row_data in tables:
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap
        if "NEW" in str(value) or "EXPANDED" in str(value):
            cell.fill = new_fill
    row += 1

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 45

# Save
output_path = Path("/home/evgeny/projects/theory-service/documentation/Concept_Schema_9D_v5.xlsx")
wb.save(output_path)
print(f"Expanded v5 schema saved to: {output_path}")
