"""
Generate 12-Dimension Concept Schema Spreadsheets

Creates two spreadsheets:
1. Schema Explanation - All 12 dimensions with field explanations
2. Tech Sovereignty Example - Mock data for "Technological Sovereignty" concept
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Color palette for dimensions
COLORS = {
    'quinean': 'E8F0FE',       # Light blue
    'sellarsian': 'FCE8E6',    # Light red/pink
    'brandomian': 'F3E8FD',    # Light purple
    'deleuzian': 'E6F4EA',     # Light green
    'bachelardian': 'FEF7E0',  # Light orange/amber
    'canguilhem': 'FDE7F3',    # Light pink
    'hacking': 'E0F7FA',       # Light cyan
    'blumenberg': 'FFF8E1',    # Light yellow
    'carey': 'EFEBE9',         # Light brown/gray
    'kuhnian': 'E8EAF6',       # Light indigo (NEW)
    'pragmatist': 'E0F2F1',    # Light teal (NEW)
    'foucauldian': 'FCE4EC',   # Light rose (NEW)
}

DIMENSION_INFO = {
    '1. Quinean': {
        'title': 'QUINEAN DIMENSION - Web of Belief',
        'thinker': 'W.V.O. Quine',
        'key_works': '"Two Dogmas of Empiricism" (1951), "Word and Object" (1960)',
        'core_insight': 'Beliefs face evidence as a corporate body (confirmation holism). No belief is immune to revision. Concepts form an interconnected web where revising one affects others.',
        'what_it_captures': 'How the concept connects to other beliefs - its INFERENTIAL POSITION in the web. What follows from it? What contradicts it? How costly is it to revise?',
        'key_questions': [
            'What other beliefs must you hold if you accept this concept?',
            'What would contradict or undermine this concept?',
            'How central is this concept to your overall framework?',
            'What would be the cost of abandoning this concept?'
        ],
        'tables': [
            {
                'name': 'concept_inferences',
                'description': 'Inferential connections to/from this concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('inference_type', 'ENUM', 'forward/backward/lateral/contradiction'),
                    ('inference_statement', 'TEXT', 'The inference itself'),
                    ('target_concept_id', 'INTEGER FK', 'Related concept if in DB'),
                    ('strength', 'FLOAT 0-1', 'How strong is this inference'),
                    ('defeasible', 'BOOLEAN', 'Can be defeated by evidence?'),
                    ('source_type', 'ENUM', 'user_input/llm_analysis/evidence'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_web_tensions',
                'description': 'Tensions with other parts of the belief web',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('tension_with', 'TEXT', 'What it is in tension with'),
                    ('tension_description', 'TEXT', 'Nature of the tension'),
                    ('resolution_cost', 'ENUM', 'low/medium/high/catastrophic'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'quinean'
    },
    '2. Sellarsian': {
        'title': 'SELLARSIAN DIMENSION - Myth of the Given',
        'thinker': 'Wilfrid Sellars',
        'key_works': '"Empiricism and the Philosophy of Mind" (1956)',
        'core_insight': 'Nothing is epistemically "given" - even perception requires conceptual capacities. Concepts that seem self-evident often smuggle in hidden assumptions.',
        'what_it_captures': 'What the concept treats as SELF-EVIDENT that actually requires justification. The hidden assumptions baked into the concept. The manifest vs scientific image tension.',
        'key_questions': [
            'What does this concept treat as obvious or natural?',
            'What hidden assumptions are embedded in this concept?',
            'What justification is actually needed but being skipped?',
            'Is there a tension between everyday and scientific views?'
        ],
        'tables': [
            {
                'name': 'concept_givenness',
                'description': 'Main givenness analysis (1:1 with concept)',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK UNIQUE', 'Parent concept (1:1)'),
                    ('is_myth_of_given', 'BOOLEAN', 'Is concept falsely foundational?'),
                    ('should_be_inferred_from', 'TEXT', 'What evidence should support this'),
                    ('space_of_reasons_role', 'TEXT', 'Role in justification/inference'),
                    ('manifest_scientific_tension', 'TEXT', 'Everyday vs scientific view tension'),
                ]
            },
            {
                'name': 'concept_givenness_markers',
                'description': 'Language markers of false self-evidence',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('marker_text', 'VARCHAR(100)', 'Language marker: "obviously", "naturally"'),
                    ('example_usage', 'TEXT', 'Example of this marker in use'),
                    ('confidence', 'FLOAT 0-1', 'Confidence this is a marker'),
                ]
            },
            {
                'name': 'concept_hidden_commitments',
                'description': 'Hidden assumptions baked into the concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('commitment_text', 'TEXT', 'The hidden assumption'),
                    ('evidence_that_exposes', 'TEXT', 'What could expose this'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'sellarsian'
    },
    '3. Brandomian': {
        'title': 'BRANDOMIAN DIMENSION - Normative Commitments',
        'thinker': 'Robert Brandom',
        'key_works': '"Making It Explicit" (1994), "Articulating Reasons" (2000)',
        'core_insight': 'Concepts are bundles of normative commitments and entitlements. Using a concept commits you to certain claims and entitles you to others. Meaning = inferential role.',
        'what_it_captures': 'What NORMATIVE COMMITMENTS using this concept involves. What you are obligated to accept, what you are entitled to assert, and what is materially incompatible.',
        'key_questions': [
            'What claims are you COMMITTED to if you use this concept?',
            'What claims are you ENTITLED to make?',
            'What is MATERIALLY INCOMPATIBLE with this concept?',
            'What inheritance chains does this concept participate in?'
        ],
        'tables': [
            {
                'name': 'concept_commitments',
                'description': 'What using the concept commits you to',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('commitment_type', 'ENUM', 'deontic_commitment/deontic_entitlement'),
                    ('commitment_statement', 'TEXT', 'What you must/may accept'),
                    ('derived_from', 'TEXT', 'Where this commitment comes from'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_incompatibilities',
                'description': 'Material incompatibilities - what the concept rules out',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('incompatible_with', 'TEXT', 'What is ruled out'),
                    ('why_incompatible', 'TEXT', 'Why these cannot coexist'),
                    ('incompatibility_type', 'ENUM', 'logical/material/practical'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_scorekeeping',
                'description': 'Deontic scorekeeping - tracking who owes what to whom',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('score_keeper', 'TEXT', 'Who is tracking commitments'),
                    ('score_type', 'ENUM', 'commitment_score/entitlement_score'),
                    ('score_description', 'TEXT', 'What is being tracked'),
                ]
            }
        ],
        'color': 'brandomian'
    },
    '4. Deleuzian': {
        'title': 'DELEUZIAN DIMENSION - Problems & Becomings',
        'thinker': 'Gilles Deleuze (with Félix Guattari)',
        'key_works': '"What Is Philosophy?" (1991), "A Thousand Plateaus" (1980)',
        'core_insight': 'Concepts are created to solve problems. They enable "lines of flight" - movements of deterritorialization that escape existing constraints. Focus on becoming, not being.',
        'what_it_captures': 'What TRANSFORMATIONS and BECOMINGS this concept enables. What "lines of flight" does it open? What does it deterritorialize? What new problems does it create?',
        'key_questions': [
            'What problem does this concept solve or address?',
            'What "lines of flight" does it enable - what escapes become possible?',
            'What does it deterritorialize - what fixed identities does it dissolve?',
            'What new becomings does it make thinkable?'
        ],
        'tables': [
            {
                'name': 'concept_problems',
                'description': 'The problems this concept addresses',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('problem_statement', 'TEXT', 'The problem being addressed'),
                    ('problem_origin', 'TEXT', 'Where this problem comes from'),
                    ('how_concept_addresses', 'TEXT', 'How concept tackles the problem'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_lines_of_flight',
                'description': 'Lines of flight - escapes the concept enables',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('flight_description', 'TEXT', 'What escape/transformation is enabled'),
                    ('from_what', 'TEXT', 'What is being escaped from'),
                    ('to_what', 'TEXT', 'What is being moved toward'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_deterritorializations',
                'description': 'What the concept deterritorializes/dissolves',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('what_deterritorialized', 'TEXT', 'What fixed identity is dissolved'),
                    ('new_territory', 'TEXT', 'What replaces it'),
                    ('intensity', 'ENUM', 'relative/absolute'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'deleuzian'
    },
    '5. Bachelardian': {
        'title': 'BACHELARDIAN DIMENSION - Epistemological Obstacles',
        'thinker': 'Gaston Bachelard',
        'key_works': '"The Formation of the Scientific Mind" (1938), "The New Scientific Spirit" (1934)',
        'core_insight': 'Concepts can become epistemological obstacles - blocking rather than enabling understanding. Progress requires RUPTURE with old concepts, not gradual evolution.',
        'what_it_captures': 'How this concept might BLOCK understanding. What epistemological obstacles it creates or overcomes. What RUPTURE would be required to transcend it.',
        'key_questions': [
            'Does this concept block certain kinds of understanding?',
            'What existing framework is this concept breaking from?',
            'What would a rupture with this concept enable?',
            'What unconscious needs does this concept serve?'
        ],
        'tables': [
            {
                'name': 'concept_obstacles',
                'description': 'Epistemological obstacle analysis',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK UNIQUE', 'Parent concept (1:1)'),
                    ('is_obstacle', 'BOOLEAN', 'Does concept block understanding?'),
                    ('obstacle_type', 'ENUM', 'experience/verbal/pragmatic/quantitative/substantialist/animist'),
                    ('what_it_blocks', 'TEXT', 'What understanding it prevents'),
                    ('why_persists', 'TEXT', 'Ideological/class function'),
                    ('rupture_would_enable', 'TEXT', 'What becomes thinkable after rupture'),
                    ('psychoanalytic_function', 'TEXT', 'Unconscious need it serves'),
                ]
            },
            {
                'name': 'concept_ruptures',
                'description': 'Ruptures this concept creates or requires',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('rupture_from', 'TEXT', 'What framework is broken from'),
                    ('rupture_trigger', 'TEXT', 'What forced the rupture'),
                    ('what_became_possible', 'TEXT', 'What became thinkable'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'bachelardian'
    },
    '6. Canguilhem': {
        'title': 'CANGUILHEM DIMENSION - Vital Norms',
        'thinker': 'Georges Canguilhem',
        'key_works': '"The Normal and the Pathological" (1943/1966)',
        'core_insight': 'Concepts have life histories - birth, health, decay. They embody NORMS that distinguish normal from pathological. Normativity precedes normalization.',
        'what_it_captures': 'The LIFE HISTORY of the concept - its birth, health, evolution. What NORMS it embeds. What it marks as normal vs pathological. Who sets these norms.',
        'key_questions': [
            'What problem gave birth to this concept?',
            'Is the concept healthy, strained, or dying?',
            'What does the concept mark as normal vs abnormal/pathological?',
            'Whose interests do these norms serve?'
        ],
        'tables': [
            {
                'name': 'concept_vitality',
                'description': 'Life history and health of the concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK UNIQUE', 'Parent concept (1:1)'),
                    ('health_status', 'ENUM', 'healthy/strained/dying/being_born'),
                    ('birth_period', 'VARCHAR(100)', 'When concept emerged'),
                    ('birth_problem', 'TEXT', 'Problem that birthed concept'),
                    ('death_signs', 'TEXT', 'Signs of conceptual decay'),
                    ('vitality_indicators', 'TEXT', 'How we know it is healthy'),
                ]
            },
            {
                'name': 'concept_norms',
                'description': 'Norms the concept embeds',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('norm_statement', 'TEXT', 'The norm being embedded'),
                    ('what_is_normal', 'TEXT', 'What counts as normal'),
                    ('what_is_pathological', 'TEXT', 'What counts as abnormal'),
                    ('whose_interests', 'TEXT', 'Who benefits from this norm'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'canguilhem'
    },
    '7. Hacking': {
        'title': 'HACKING DIMENSION - Dynamic Nominalism',
        'thinker': 'Ian Hacking',
        'key_works': '"The Social Construction of What?" (1999), "Historical Ontology" (2002)',
        'core_insight': 'Concepts CREATE what they describe through "looping effects". Naming things changes them. Styles of reasoning create new kinds of objects and new kinds of people.',
        'what_it_captures': 'How this concept CREATES what it describes. What "looping effects" occur - how does the concept change what it names? What new objects or kinds does it bring into being?',
        'key_questions': [
            'What does this concept create or bring into being?',
            'What "looping effects" occur - how does naming change the named?',
            'What style of reasoning does this concept embody or enable?',
            'What new kinds of objects or people does it create?'
        ],
        'tables': [
            {
                'name': 'concept_creations',
                'description': 'What the concept creates through naming',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('what_created', 'TEXT', 'Object/kind brought into being'),
                    ('how_created', 'TEXT', 'Mechanism of creation'),
                    ('existed_before', 'BOOLEAN', 'Did it exist before naming?'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_looping_effects',
                'description': 'Looping effects - how naming changes the named',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('what_is_named', 'TEXT', 'What is being named/classified'),
                    ('how_naming_changes_it', 'TEXT', 'How the named responds to classification'),
                    ('loop_direction', 'ENUM', 'reinforcing/undermining/transforming'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_styles_of_reasoning',
                'description': 'Styles of reasoning the concept embodies',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('style_name', 'TEXT', 'Name of reasoning style'),
                    ('what_becomes_possible', 'TEXT', 'What this style makes thinkable'),
                    ('what_counts_as_evidence', 'TEXT', 'What counts as evidence in this style'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'hacking'
    },
    '8. Blumenberg': {
        'title': 'BLUMENBERG DIMENSION - Absolute Metaphors',
        'thinker': 'Hans Blumenberg',
        'key_works': '"Paradigms for a Metaphorology" (1960), "Work on Myth" (1979)',
        'core_insight': 'Some metaphors are "absolute" - they cannot be cashed out in literal terms. These background metaphors structure thinking without being noticed. They carry historical sediment.',
        'what_it_captures': 'The METAPHORS that structure this concept, especially those that cannot be translated into literal terms. What historical sediment do they carry? What do they reveal vs hide?',
        'key_questions': [
            'What root metaphor underlies this concept?',
            'Can this metaphor be translated into literal terms?',
            'What does the metaphor reveal? What does it hide?',
            'What historical associations does the metaphor carry?'
        ],
        'tables': [
            {
                'name': 'concept_metaphors',
                'description': 'Root metaphors underlying the concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('metaphor', 'TEXT', 'The metaphor itself'),
                    ('source_domain', 'TEXT', 'Where metaphor comes from'),
                    ('target_domain', 'TEXT', 'What it is applied to'),
                    ('is_absolute', 'BOOLEAN', 'Cannot be literalized?'),
                    ('what_reveals', 'TEXT', 'What metaphor makes visible'),
                    ('what_hides', 'TEXT', 'What metaphor obscures'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_metaphor_history',
                'description': 'Historical sediment in metaphors',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('historical_layer', 'TEXT', 'Historical period/context'),
                    ('associations_from_period', 'TEXT', 'What associations this period adds'),
                    ('still_active', 'BOOLEAN', 'Are these associations still active?'),
                ]
            }
        ],
        'color': 'blumenberg'
    },
    '9. Carey': {
        'title': 'CAREY DIMENSION - Conceptual Bootstrapping',
        'thinker': 'Susan Carey',
        'key_works': '"The Origin of Concepts" (2009)',
        'core_insight': 'Complex concepts are "bootstrapped" from simpler ones through processes like placeholder structures, analogy, and conceptual combination. Some concepts require restructuring of core cognition.',
        'what_it_captures': 'How this concept was BUILT from simpler components. What cognitive primitives does it draw on? Did the concept require genuine conceptual change or just elaboration?',
        'key_questions': [
            'What simpler concepts combine to make this concept?',
            'What core cognitive systems does it draw on?',
            'Did formation require genuine conceptual change?',
            'What makes this concept hard or easy to learn?'
        ],
        'tables': [
            {
                'name': 'concept_components',
                'description': 'Simpler concepts that combine to form this one',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('component_concept', 'TEXT', 'Simpler concept used'),
                    ('role_in_combination', 'TEXT', 'How it contributes'),
                    ('combination_type', 'ENUM', 'conjunction/disjunction/placeholder/analogy'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_core_cognition',
                'description': 'Core cognitive systems the concept draws on',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('core_system', 'ENUM', 'object/agent/number/space/causation/social'),
                    ('how_used', 'TEXT', 'How this core system is utilized'),
                    ('extended_beyond', 'BOOLEAN', 'Does concept extend beyond core?'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_learning_constraints',
                'description': 'What makes concept hard/easy to learn',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('constraint_type', 'ENUM', 'facilitating/blocking'),
                    ('constraint_description', 'TEXT', 'What helps or hinders learning'),
                    ('target_population', 'TEXT', 'Who faces this constraint'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'carey'
    },
    '10. Kuhnian': {
        'title': 'KUHNIAN DIMENSION - Paradigm Structure [NEW]',
        'thinker': 'Thomas Kuhn',
        'key_works': '"The Structure of Scientific Revolutions" (1962)',
        'core_insight': 'Concepts operate within PARADIGMS - shared frameworks of assumptions, exemplars, and practices. Normal work refines concepts within a paradigm; revolutionary work replaces paradigms.',
        'what_it_captures': 'What PARADIGM this concept belongs to. What counts as an exemplar? What would be an anomaly? Is this concept normal science or revolutionary?',
        'key_questions': [
            'What paradigm does this concept belong to?',
            'What are the exemplary instances of this concept?',
            'What would count as an anomaly - something the concept cannot handle?',
            'Is this concept normal science or potentially revolutionary?'
        ],
        'tables': [
            {
                'name': 'concept_paradigm',
                'description': 'Paradigm context for the concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK UNIQUE', 'Parent concept (1:1)'),
                    ('paradigm_name', 'TEXT', 'Name of the paradigm'),
                    ('paradigm_assumptions', 'TEXT', 'Core assumptions of the paradigm'),
                    ('disciplinary_matrix', 'TEXT', 'Shared commitments of practitioners'),
                    ('is_revolutionary', 'BOOLEAN', 'Does concept challenge paradigm?'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_exemplars',
                'description': 'Exemplary instances that define the concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('exemplar_description', 'TEXT', 'The exemplary case'),
                    ('why_paradigmatic', 'TEXT', 'Why this is a paradigm case'),
                    ('what_it_teaches', 'TEXT', 'What practitioners learn from it'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_anomalies',
                'description': 'Anomalies - what the concept cannot handle',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('anomaly_description', 'TEXT', 'The anomalous case'),
                    ('why_anomalous', 'TEXT', 'Why the concept cannot handle this'),
                    ('seriousness', 'ENUM', 'minor/moderate/crisis_inducing'),
                    ('resolution_attempts', 'TEXT', 'How anomaly has been addressed'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'kuhnian'
    },
    '11. Pragmatist': {
        'title': 'PRAGMATIST DIMENSION - Performative Consequences [NEW]',
        'thinker': 'William James, John Dewey, Richard Rorty',
        'key_works': '"Pragmatism" (James, 1907), "Experience and Nature" (Dewey, 1925)',
        'core_insight': 'The meaning of a concept lies in its PRACTICAL CONSEQUENCES. Concepts are tools - what matters is what they enable us to DO. Truth is what works.',
        'what_it_captures': 'What USING this concept enables you to DO. What practical difference does adopting it make? What actions, interventions, or proposals become possible?',
        'key_questions': [
            'What does using this concept enable you to DO?',
            'What practical difference does adopting it make?',
            'What interventions or actions become possible?',
            'What proposals can you make with this concept that you could not make without it?'
        ],
        'tables': [
            {
                'name': 'concept_practical_effects',
                'description': 'What the concept enables practically',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('effect_type', 'ENUM', 'enables_action/enables_proposal/enables_intervention'),
                    ('effect_description', 'TEXT', 'What becomes possible'),
                    ('without_concept', 'TEXT', 'What was not possible before'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_cash_value',
                'description': '"Cash value" - practical difference the concept makes',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('domain', 'TEXT', 'Domain where difference is felt'),
                    ('what_changes', 'TEXT', 'What changes when concept is adopted'),
                    ('for_whom', 'TEXT', 'Who experiences the difference'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_tool_uses',
                'description': 'How the concept functions as a tool',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('tool_function', 'TEXT', 'What the concept does as a tool'),
                    ('problem_addressed', 'TEXT', 'What problem the tool addresses'),
                    ('effectiveness', 'ENUM', 'highly_effective/moderately_effective/questionable'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'pragmatist'
    },
    '12. Foucauldian': {
        'title': 'FOUCAULDIAN DIMENSION - Power-Knowledge Relations [NEW]',
        'thinker': 'Michel Foucault',
        'key_works': '"Discipline and Punish" (1975), "History of Sexuality Vol. 1" (1976)',
        'core_insight': 'Knowledge and power are intertwined. Concepts make phenomena GOVERNABLE. They naturalize certain power relations while making others unthinkable.',
        'what_it_captures': 'What POWER RELATIONS this concept naturalizes or makes invisible. What does it make governable? Whose authority does it legitimize? What populations does it bring under scrutiny?',
        'key_questions': [
            'What power relations does this concept naturalize or make invisible?',
            'What does this concept make governable or manageable?',
            'Whose authority does this concept legitimize?',
            'What populations or phenomena does it bring under scrutiny?'
        ],
        'tables': [
            {
                'name': 'concept_power_relations',
                'description': 'Power relations embedded in the concept',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('power_relation', 'TEXT', 'The power relation'),
                    ('how_naturalized', 'TEXT', 'How concept makes it seem natural'),
                    ('who_benefits', 'TEXT', 'Who benefits from this arrangement'),
                    ('who_is_disadvantaged', 'TEXT', 'Who is disadvantaged'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_governmentality',
                'description': 'What the concept makes governable',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('what_governed', 'TEXT', 'Population/phenomenon made governable'),
                    ('governance_technique', 'TEXT', 'How governance is exercised'),
                    ('knowledge_produced', 'TEXT', 'What knowledge is generated'),
                    ('resistance_possible', 'TEXT', 'Where resistance might emerge'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            },
            {
                'name': 'concept_discourse_effects',
                'description': 'Effects of the concept in discourse',
                'fields': [
                    ('id', 'SERIAL PRIMARY KEY', 'Unique identifier'),
                    ('concept_id', 'INTEGER FK', 'Parent concept'),
                    ('what_speakable', 'TEXT', 'What the concept makes sayable'),
                    ('what_unspeakable', 'TEXT', 'What it makes difficult to say'),
                    ('subject_positions_created', 'TEXT', 'New subject positions enabled'),
                    ('confidence', 'FLOAT 0-1', 'Confidence in this claim'),
                ]
            }
        ],
        'color': 'foucauldian'
    }
}

# Tech Sovereignty example data
TECH_SOVEREIGNTY_DATA = {
    'core': {
        'term': 'Technological Sovereignty',
        'definition': 'The capacity of a state or political entity to control, develop, and govern critical technologies within its jurisdiction, reducing dependency on foreign actors for essential technological capabilities.',
        'category': 'Political Economy of Technology',
        'status': 'active',
        'confidence': 0.9,
        'centrality': 'core',
        'health_status': 'healthy',
        'birth_period': '2010s-2020s (intensified post-2018 US-China tech war)',
        'hierarchy_level': 3
    },
    'quinean': {
        'inferences': [
            {'type': 'forward', 'statement': 'If tech sovereignty → indigenous semiconductor manufacturing capability required', 'strength': 0.85},
            {'type': 'forward', 'statement': 'If tech sovereignty → control over critical supply chains necessary', 'strength': 0.9},
            {'type': 'forward', 'statement': 'If tech sovereignty → domestic AI/ML research ecosystem needed', 'strength': 0.8},
            {'type': 'backward', 'statement': 'Tech sovereignty ← sufficient R&D investment capacity', 'strength': 0.7},
            {'type': 'contradiction', 'statement': 'Material dependence on foreign chip foundries contradicts sovereignty claims', 'strength': 0.85},
            {'type': 'contradiction', 'statement': 'Global just-in-time supply chains are incompatible with full sovereignty', 'strength': 0.75},
        ],
        'tensions': [
            {'with': 'Economic efficiency via specialization', 'description': 'Sovereignty may require inefficient domestic production', 'cost': 'high'},
            {'with': 'Open science norms', 'description': 'Sovereignty may restrict research collaboration', 'cost': 'medium'},
            {'with': 'Liberal trade principles', 'description': 'Requires protectionist industrial policy', 'cost': 'high'},
        ]
    },
    'sellarsian': {
        'givenness': {
            'is_myth': True,
            'should_be_inferred_from': 'Empirical analysis of actual state capacity, comparative advantage, and historical success of autarky',
            'manifest_scientific_tension': 'Everyday view assumes sovereignty is achievable through will; scientific analysis shows structural constraints'
        },
        'markers': [
            {'marker': 'naturally', 'example': '"Technology naturally extends state power"'},
            {'marker': 'obviously', 'example': '"Obviously a matter of national security"'},
            {'marker': 'clearly possible', 'example': '"Sovereignty is clearly possible with sufficient investment"'},
            {'marker': 'inherently', 'example': '"States are inherently capable of technological independence"'},
        ],
        'hidden_commitments': [
            {'commitment': 'Nation-states are the appropriate unit of technological governance', 'exposure': 'Evidence of successful regional/supranational tech governance'},
            {'commitment': 'Technological autarky is achievable without prohibitive costs', 'exposure': 'Cost-benefit analysis of domestic chip production'},
            {'commitment': 'Security concerns outweigh efficiency gains from specialization', 'exposure': 'Historical analysis of actual security incidents vs trade benefits'},
        ]
    },
    'brandomian': {
        'commitments': [
            {'type': 'deontic_commitment', 'statement': 'If you claim tech sovereignty, you are committed to funding domestic R&D'},
            {'type': 'deontic_commitment', 'statement': 'If you claim tech sovereignty, you are committed to restricting foreign access to critical IP'},
            {'type': 'deontic_entitlement', 'statement': 'Tech sovereignty entitles states to regulate technology exports'},
            {'type': 'deontic_entitlement', 'statement': 'Tech sovereignty entitles states to screen foreign investment in tech sectors'},
        ],
        'incompatibilities': [
            {'with': 'Radical free trade', 'why': 'Sovereignty requires protecting domestic industries from foreign competition'},
            {'with': 'Technology as global public good', 'why': 'Sovereignty implies exclusive national control'},
            {'with': 'Borderless innovation networks', 'why': 'Sovereignty requires national boundaries on knowledge flows'},
        ]
    },
    'deleuzian': {
        'problems': [
            {'statement': 'Dependence on foreign technology creates strategic vulnerability', 'origin': 'US sanctions on Huawei, chip export controls'},
            {'statement': 'Tech colonialism - extraction of data without local value creation', 'origin': 'Big Tech platform dominance in Global South'},
        ],
        'lines_of_flight': [
            {'description': 'Escape from unilateral US tech hegemony', 'from': 'US-dominated tech stack', 'to': 'Multipolar tech ecosystem'},
            {'description': 'Creation of alternative technical standards', 'from': 'Western-defined standards bodies', 'to': 'Competing standard-setting'},
        ],
        'deterritorializations': [
            {'what': 'Tech as neutral/apolitical domain', 'new_territory': 'Tech as site of geopolitical contestation', 'intensity': 'relative'},
            {'what': 'Single global internet', 'new_territory': 'Fragmented "splinternet"', 'intensity': 'relative'},
        ]
    },
    'bachelardian': {
        'obstacles': {
            'is_obstacle': True,
            'type': 'pragmatic',
            'what_blocks': 'Recognition of structural interdependence and benefits of specialization',
            'why_persists': 'Serves nationalist political projects; mobilizes resources for industrial policy',
            'rupture_would_enable': 'Thinking about technology governance beyond the nation-state',
            'psychoanalytic': 'Satisfies desire for control in face of technological complexity'
        },
        'ruptures': [
            {'from': 'Neoliberal "end of history" globalization', 'trigger': 'US-China tech war, COVID supply chain disruptions', 'enabled': 'Strategic industrial policy became thinkable again'},
        ]
    },
    'canguilhem': {
        'vitality': {
            'health': 'healthy',
            'birth': '2010s, accelerated 2018+',
            'birth_problem': 'Recognition that technological dependence creates geopolitical vulnerability',
            'vitality_signs': 'Appearing in major policy documents (EU, China, India); billions in investment'
        },
        'norms': [
            {'norm': 'States should control critical technology', 'normal': 'Domestic tech champions', 'pathological': 'Foreign platform dominance', 'whose_interests': 'National security apparatus, domestic tech industry'},
            {'norm': 'Technology should serve national development', 'normal': 'Indigenous innovation', 'pathological': 'Dependence on foreign solutions', 'whose_interests': 'State planners, local entrepreneurs'},
        ]
    },
    'hacking': {
        'creations': [
            {'what': '"Sovereign cloud" as a category', 'how': 'Naming creates demand for compliant infrastructure', 'existed_before': False},
            {'what': '"Critical technology" as governance object', 'how': 'Classification creates export control regimes', 'existed_before': False},
        ],
        'looping_effects': [
            {'named': 'Tech companies as national champions', 'change': 'Companies adopt nationalist framing, lobby for protection', 'direction': 'reinforcing'},
            {'named': 'Foreign tech as threat', 'change': 'Increased suspicion leads to actual restrictions, creating retaliation cycle', 'direction': 'reinforcing'},
        ],
        'styles': [
            {'style': 'Security reasoning applied to technology', 'enables': 'Treating tech choices as national security matters', 'evidence': 'Intelligence assessments, threat classifications'},
        ]
    },
    'blumenberg': {
        'metaphors': [
            {'metaphor': 'SOVEREIGNTY AS TERRITORY', 'source': 'Westphalian state sovereignty', 'target': 'Technology governance', 'absolute': True, 'reveals': 'Control, boundaries, exclusion rights', 'hides': 'Inherently networked nature of modern tech'},
            {'metaphor': 'TECHNOLOGY AS INFRASTRUCTURE', 'source': 'Physical infrastructure (roads, ports)', 'target': 'Digital systems', 'absolute': False, 'reveals': 'Public good aspects, investment needs', 'hides': 'Dynamic, rapidly evolving nature'},
        ],
        'history': [
            {'period': 'Cold War', 'associations': 'Military-industrial complex, space race, nuclear deterrence', 'still_active': True},
            {'period': 'Post-colonial period', 'associations': 'Development economics, import substitution, non-alignment', 'still_active': True},
        ]
    },
    'carey': {
        'components': [
            {'concept': 'Sovereignty (political)', 'role': 'Provides framework of exclusive control', 'type': 'analogy'},
            {'concept': 'Critical infrastructure', 'role': 'Defines what requires protection', 'type': 'placeholder'},
            {'concept': 'Industrial policy', 'role': 'Provides intervention mechanisms', 'type': 'conjunction'},
            {'concept': 'National security', 'role': 'Provides legitimation frame', 'type': 'conjunction'},
        ],
        'core_cognition': [
            {'system': 'agent', 'how': 'States treated as agents with intentions and capabilities', 'extended': True},
            {'system': 'object', 'how': 'Technology treated as bounded objects that can be possessed', 'extended': True},
        ],
        'learning_constraints': [
            {'type': 'facilitating', 'description': 'Intuitive analogy to territorial sovereignty', 'population': 'Policy makers'},
            {'type': 'blocking', 'description': 'Complexity of actual tech supply chains', 'population': 'General public'},
        ]
    },
    'kuhnian': {
        'paradigm': {
            'name': 'Techno-nationalist political economy',
            'assumptions': 'States are primary tech governance actors; security concerns dominate; autarky is possible and desirable',
            'matrix': 'Industrial policy tools, strategic trade theory, security studies',
            'revolutionary': False  # Works within existing paradigm of state-centric IR
        },
        'exemplars': [
            {'description': 'China\'s semiconductor self-sufficiency drive post-2018', 'why': 'Most comprehensive sovereignty effort', 'teaches': 'Scale of investment required; limits of autarky'},
            {'description': 'EU\'s GDPR and data sovereignty initiatives', 'why': 'Regulatory rather than industrial approach', 'teaches': 'Alternative sovereignty via rules not production'},
        ],
        'anomalies': [
            {'description': 'Taiwan\'s success as small open economy chip leader', 'why': 'Disproves need for sovereignty for tech leadership', 'seriousness': 'moderate'},
            {'description': 'Failure of import substitution in developing countries', 'why': 'Historical precedent for sovereignty-focused tech policy failure', 'seriousness': 'moderate'},
        ]
    },
    'pragmatist': {
        'practical_effects': [
            {'type': 'enables_action', 'description': 'Justify massive public investment in domestic chip fabs', 'without': 'Such investment seen as market distortion'},
            {'type': 'enables_proposal', 'description': 'Propose technology export controls as legitimate policy', 'without': 'Seen as protectionism'},
            {'type': 'enables_intervention', 'description': 'Screen and block foreign tech investments', 'without': 'Seen as violating free market principles'},
        ],
        'cash_value': [
            {'domain': 'Industrial policy', 'what_changes': 'State intervention in tech markets becomes legitimate', 'for_whom': 'Governments, domestic tech firms'},
            {'domain': 'International trade', 'what_changes': 'Tech export restrictions become acceptable', 'for_whom': 'Security establishments'},
        ],
        'tool_uses': [
            {'function': 'Mobilize resources for industrial policy', 'problem': 'Need to justify non-market state intervention', 'effectiveness': 'highly_effective'},
            {'function': 'Delegitimize foreign tech presence', 'problem': 'Competition from more advanced foreign firms', 'effectiveness': 'moderately_effective'},
        ]
    },
    'foucauldian': {
        'power_relations': [
            {'relation': 'State authority over tech sector', 'naturalized': 'Framing tech as national security makes state control seem obvious', 'benefits': 'State security apparatus, national champions', 'disadvantaged': 'Foreign tech firms, consumers wanting choice'},
            {'relation': 'Hierarchy between tech-sovereign and dependent nations', 'naturalized': 'Sovereignty discourse creates have/have-not binary', 'benefits': 'Tech-leading nations', 'disadvantaged': 'Global South, smaller states'},
        ],
        'governmentality': [
            {'governed': 'Technology firms (domestic and foreign)', 'technique': 'Licensing, security reviews, investment screening', 'knowledge': 'Tech audits, supply chain mapping', 'resistance': 'Regulatory arbitrage, offshore operations'},
            {'governed': 'Citizens as data subjects', 'technique': 'Data localization requirements', 'knowledge': 'Data flow monitoring, compliance audits', 'resistance': 'VPNs, encrypted communications'},
        ],
        'discourse_effects': [
            {'speakable': 'Tech as geopolitical weapon, national asset', 'unspeakable': 'Tech as global commons, shared human heritage', 'subject_positions': 'Citizen as beneficiary of sovereign tech; foreign firm as potential threat'},
        ]
    }
}


def create_schema_explanation_spreadsheet():
    """Create spreadsheet explaining all 12 dimensions with fields."""
    wb = Workbook()

    # Overview sheet
    ws = wb.active
    ws.title = 'Overview'

    # Title
    ws['A1'] = 'CONCEPT SCHEMA v6 - 12-DIMENSIONAL FRAMEWORK'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A3'] = 'WHAT IS THIS SCHEMA?'
    ws['A3'].font = Font(bold=True, size=12)

    ws['A4'] = '''This schema analyzes concepts through 12 philosophical dimensions. Each dimension reveals different aspects of how a concept works, what it presupposes, what it commits users to, and how it evolves over time. The framework has been extended from 9 to 12 dimensions to better capture complex research programs.'''
    ws.merge_cells('A4:E4')

    ws['A6'] = 'THE 12 DIMENSIONS'
    ws['A6'].font = Font(bold=True, size=12)

    row = 7
    ws[f'A{row}'] = 'Dimension'
    ws[f'B{row}'] = 'Thinker'
    ws[f'C{row}'] = 'Core Question'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)

    dimensions_summary = [
        ('Quinean', 'Quine', 'How does this connect to other beliefs in the web?'),
        ('Sellarsian', 'Sellars', 'What is falsely treated as self-evident?'),
        ('Brandomian', 'Brandom', 'What normative commitments does using this involve?'),
        ('Deleuzian', 'Deleuze', 'What transformations and becomings does this enable?'),
        ('Bachelardian', 'Bachelard', 'Does this block understanding? What rupture is needed?'),
        ('Canguilhem', 'Canguilhem', 'What norms does this embed? What is marked as pathological?'),
        ('Hacking', 'Hacking', 'What does this concept create through naming?'),
        ('Blumenberg', 'Blumenberg', 'What metaphors structure this concept?'),
        ('Carey', 'Carey', 'How was this bootstrapped from simpler concepts?'),
        ('Kuhnian [NEW]', 'Kuhn', 'What paradigm does this belong to? What are its anomalies?'),
        ('Pragmatist [NEW]', 'James/Dewey', 'What does USING this concept enable you to DO?'),
        ('Foucauldian [NEW]', 'Foucault', 'What power relations does this naturalize?'),
    ]

    for dim, thinker, question in dimensions_summary:
        row += 1
        ws[f'A{row}'] = dim
        ws[f'B{row}'] = thinker
        ws[f'C{row}'] = question

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60

    # Create individual dimension sheets
    for dim_key, dim_info in DIMENSION_INFO.items():
        ws = wb.create_sheet(title=dim_key[:31])  # Sheet names limited to 31 chars

        # Header
        ws['A1'] = dim_info['title']
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws['A3'] = f"Thinker: {dim_info['thinker']}"
        ws['A4'] = f"Key Works: {dim_info['key_works']}"
        ws.merge_cells('A4:F4')

        ws['A6'] = 'CORE INSIGHT'
        ws['A6'].font = Font(bold=True)
        ws['A7'] = dim_info['core_insight']
        ws.merge_cells('A7:F7')

        ws['A9'] = 'WHAT THIS DIMENSION CAPTURES'
        ws['A9'].font = Font(bold=True)
        ws['A10'] = dim_info['what_it_captures']
        ws.merge_cells('A10:F10')

        ws['A12'] = 'KEY QUESTIONS'
        ws['A12'].font = Font(bold=True)
        row = 13
        for q in dim_info['key_questions']:
            ws[f'A{row}'] = f"• {q}"
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

        row += 1
        ws[f'A{row}'] = 'DATABASE TABLES'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for table in dim_info['tables']:
            row += 1
            ws[f'A{row}'] = table['name']
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = table['description']
            ws.merge_cells(f'B{row}:F{row}')
            row += 1

            # Field headers
            ws[f'A{row}'] = 'Field'
            ws[f'B{row}'] = 'Type'
            ws[f'C{row}'] = 'Description'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
            row += 1

            for field_name, field_type, field_desc in table['fields']:
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = field_type
                ws[f'C{row}'] = field_desc
                row += 1

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

    # Save
    output_path = '/home/evgeny/Downloads/Concept_Schema_12D_Explanation.xlsx'
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_tech_sovereignty_spreadsheet():
    """Create spreadsheet with Tech Sovereignty example data."""
    wb = Workbook()

    # Overview sheet
    ws = wb.active
    ws.title = 'Overview'

    core = TECH_SOVEREIGNTY_DATA['core']

    ws['A1'] = 'TECHNOLOGICAL SOVEREIGNTY - 12-DIMENSIONAL ANALYSIS'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    ws['A3'] = 'CONCEPT CORE DATA'
    ws['A3'].font = Font(bold=True, size=12)

    data = [
        ('Term', core['term']),
        ('Definition', core['definition']),
        ('Category', core['category']),
        ('Status', core['status']),
        ('Confidence', core['confidence']),
        ('Centrality (Quinean)', core['centrality']),
        ('Health Status (Canguilhem)', core['health_status']),
        ('Birth Period', core['birth_period']),
        ('Hierarchy Level (Carey)', core['hierarchy_level']),
    ]

    row = 4
    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = str(value)
        ws.merge_cells(f'B{row}:D{row}')
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80

    # Quinean sheet
    ws = wb.create_sheet('1. Quinean')
    ws['A1'] = 'QUINEAN DIMENSION - Web of Belief'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'INFERENCES'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Type'
    ws['B4'] = 'Statement'
    ws['C4'] = 'Strength'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)

    row = 5
    for inf in TECH_SOVEREIGNTY_DATA['quinean']['inferences']:
        ws[f'A{row}'] = inf['type']
        ws[f'B{row}'] = inf['statement']
        ws[f'C{row}'] = inf['strength']
        row += 1

    row += 1
    ws[f'A{row}'] = 'WEB TENSIONS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'In Tension With'
    ws[f'B{row}'] = 'Description'
    ws[f'C{row}'] = 'Resolution Cost'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for t in TECH_SOVEREIGNTY_DATA['quinean']['tensions']:
        ws[f'A{row}'] = t['with']
        ws[f'B{row}'] = t['description']
        ws[f'C{row}'] = t['cost']
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15

    # Sellarsian sheet
    ws = wb.create_sheet('2. Sellarsian')
    ws['A1'] = 'SELLARSIAN DIMENSION - Myth of the Given'
    ws['A1'].font = Font(bold=True, size=14)

    giv = TECH_SOVEREIGNTY_DATA['sellarsian']['givenness']
    ws['A3'] = 'GIVENNESS ANALYSIS'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Is Myth of Given: {giv['is_myth']}"
    ws['A5'] = f"Should Be Inferred From: {giv['should_be_inferred_from']}"
    ws['A6'] = f"Manifest/Scientific Tension: {giv['manifest_scientific_tension']}"

    ws['A8'] = 'GIVENNESS MARKERS'
    ws['A8'].font = Font(bold=True)
    ws['A9'] = 'Marker'
    ws['B9'] = 'Example Usage'
    ws['A9'].font = Font(bold=True)
    ws['B9'].font = Font(bold=True)

    row = 10
    for m in TECH_SOVEREIGNTY_DATA['sellarsian']['markers']:
        ws[f'A{row}'] = m['marker']
        ws[f'B{row}'] = m['example']
        row += 1

    row += 1
    ws[f'A{row}'] = 'HIDDEN COMMITMENTS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Commitment'
    ws[f'B{row}'] = 'Evidence That Exposes'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1

    for c in TECH_SOVEREIGNTY_DATA['sellarsian']['hidden_commitments']:
        ws[f'A{row}'] = c['commitment']
        ws[f'B{row}'] = c['exposure']
        row += 1

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 55

    # Brandomian sheet
    ws = wb.create_sheet('3. Brandomian')
    ws['A1'] = 'BRANDOMIAN DIMENSION - Normative Commitments'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'COMMITMENTS & ENTITLEMENTS'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Type'
    ws['B4'] = 'Statement'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)

    row = 5
    for c in TECH_SOVEREIGNTY_DATA['brandomian']['commitments']:
        ws[f'A{row}'] = c['type']
        ws[f'B{row}'] = c['statement']
        row += 1

    row += 1
    ws[f'A{row}'] = 'INCOMPATIBILITIES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Incompatible With'
    ws[f'B{row}'] = 'Why Incompatible'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1

    for i in TECH_SOVEREIGNTY_DATA['brandomian']['incompatibilities']:
        ws[f'A{row}'] = i['with']
        ws[f'B{row}'] = i['why']
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 70

    # Deleuzian sheet
    ws = wb.create_sheet('4. Deleuzian')
    ws['A1'] = 'DELEUZIAN DIMENSION - Problems & Becomings'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'PROBLEMS ADDRESSED'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Problem'
    ws['B4'] = 'Origin'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)

    row = 5
    for p in TECH_SOVEREIGNTY_DATA['deleuzian']['problems']:
        ws[f'A{row}'] = p['statement']
        ws[f'B{row}'] = p['origin']
        row += 1

    row += 1
    ws[f'A{row}'] = 'LINES OF FLIGHT'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Description'
    ws[f'B{row}'] = 'From → To'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1

    for l in TECH_SOVEREIGNTY_DATA['deleuzian']['lines_of_flight']:
        ws[f'A{row}'] = l['description']
        ws[f'B{row}'] = f"{l['from']} → {l['to']}"
        row += 1

    row += 1
    ws[f'A{row}'] = 'DETERRITORIALIZATIONS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'What Dissolved'
    ws[f'B{row}'] = 'New Territory'
    ws[f'C{row}'] = 'Intensity'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for d in TECH_SOVEREIGNTY_DATA['deleuzian']['deterritorializations']:
        ws[f'A{row}'] = d['what']
        ws[f'B{row}'] = d['new_territory']
        ws[f'C{row}'] = d['intensity']
        row += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15

    # Bachelardian sheet
    ws = wb.create_sheet('5. Bachelardian')
    ws['A1'] = 'BACHELARDIAN DIMENSION - Epistemological Obstacles'
    ws['A1'].font = Font(bold=True, size=14)

    obs = TECH_SOVEREIGNTY_DATA['bachelardian']['obstacles']
    ws['A3'] = 'OBSTACLE ANALYSIS'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Is Obstacle: {obs['is_obstacle']}"
    ws['A5'] = f"Obstacle Type: {obs['type']}"
    ws['A6'] = f"What It Blocks: {obs['what_blocks']}"
    ws['A7'] = f"Why It Persists: {obs['why_persists']}"
    ws['A8'] = f"Rupture Would Enable: {obs['rupture_would_enable']}"
    ws['A9'] = f"Psychoanalytic Function: {obs['psychoanalytic']}"

    ws['A11'] = 'RUPTURES'
    ws['A11'].font = Font(bold=True)
    ws['A12'] = 'From'
    ws['B12'] = 'Trigger'
    ws['C12'] = 'What Became Possible'
    ws['A12'].font = Font(bold=True)
    ws['B12'].font = Font(bold=True)
    ws['C12'].font = Font(bold=True)

    row = 13
    for r in TECH_SOVEREIGNTY_DATA['bachelardian']['ruptures']:
        ws[f'A{row}'] = r['from']
        ws[f'B{row}'] = r['trigger']
        ws[f'C{row}'] = r['enabled']
        row += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 45

    # Canguilhem sheet
    ws = wb.create_sheet('6. Canguilhem')
    ws['A1'] = 'CANGUILHEM DIMENSION - Vital Norms'
    ws['A1'].font = Font(bold=True, size=14)

    vit = TECH_SOVEREIGNTY_DATA['canguilhem']['vitality']
    ws['A3'] = 'CONCEPT VITALITY'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Health Status: {vit['health']}"
    ws['A5'] = f"Birth Period: {vit['birth']}"
    ws['A6'] = f"Birth Problem: {vit['birth_problem']}"
    ws['A7'] = f"Vitality Signs: {vit['vitality_signs']}"

    ws['A9'] = 'NORMS EMBEDDED'
    ws['A9'].font = Font(bold=True)
    ws['A10'] = 'Norm'
    ws['B10'] = 'What Is Normal'
    ws['C10'] = 'What Is Pathological'
    ws['D10'] = 'Whose Interests'
    ws['A10'].font = Font(bold=True)
    ws['B10'].font = Font(bold=True)
    ws['C10'].font = Font(bold=True)
    ws['D10'].font = Font(bold=True)

    row = 11
    for n in TECH_SOVEREIGNTY_DATA['canguilhem']['norms']:
        ws[f'A{row}'] = n['norm']
        ws[f'B{row}'] = n['normal']
        ws[f'C{row}'] = n['pathological']
        ws[f'D{row}'] = n['whose_interests']
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 35

    # Hacking sheet
    ws = wb.create_sheet('7. Hacking')
    ws['A1'] = 'HACKING DIMENSION - Dynamic Nominalism'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'WHAT CONCEPT CREATES'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'What Created'
    ws['B4'] = 'How Created'
    ws['C4'] = 'Existed Before?'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)

    row = 5
    for c in TECH_SOVEREIGNTY_DATA['hacking']['creations']:
        ws[f'A{row}'] = c['what']
        ws[f'B{row}'] = c['how']
        ws[f'C{row}'] = 'Yes' if c['existed_before'] else 'No'
        row += 1

    row += 1
    ws[f'A{row}'] = 'LOOPING EFFECTS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'What Is Named'
    ws[f'B{row}'] = 'How Naming Changes It'
    ws[f'C{row}'] = 'Direction'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for l in TECH_SOVEREIGNTY_DATA['hacking']['looping_effects']:
        ws[f'A{row}'] = l['named']
        ws[f'B{row}'] = l['change']
        ws[f'C{row}'] = l['direction']
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 15

    # Blumenberg sheet
    ws = wb.create_sheet('8. Blumenberg')
    ws['A1'] = 'BLUMENBERG DIMENSION - Absolute Metaphors'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'ROOT METAPHORS'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Metaphor'
    ws['B4'] = 'Source → Target'
    ws['C4'] = 'Absolute?'
    ws['D4'] = 'Reveals'
    ws['E4'] = 'Hides'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws['E4'].font = Font(bold=True)

    row = 5
    for m in TECH_SOVEREIGNTY_DATA['blumenberg']['metaphors']:
        ws[f'A{row}'] = m['metaphor']
        ws[f'B{row}'] = f"{m['source']} → {m['target']}"
        ws[f'C{row}'] = 'Yes' if m['absolute'] else 'No'
        ws[f'D{row}'] = m['reveals']
        ws[f'E{row}'] = m['hides']
        row += 1

    row += 1
    ws[f'A{row}'] = 'HISTORICAL SEDIMENT'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Period'
    ws[f'B{row}'] = 'Associations'
    ws[f'C{row}'] = 'Still Active?'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for h in TECH_SOVEREIGNTY_DATA['blumenberg']['history']:
        ws[f'A{row}'] = h['period']
        ws[f'B{row}'] = h['associations']
        ws[f'C{row}'] = 'Yes' if h['still_active'] else 'No'
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 40

    # Carey sheet
    ws = wb.create_sheet('9. Carey')
    ws['A1'] = 'CAREY DIMENSION - Conceptual Bootstrapping'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'COMPONENT CONCEPTS'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Component'
    ws['B4'] = 'Role'
    ws['C4'] = 'Combination Type'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)

    row = 5
    for c in TECH_SOVEREIGNTY_DATA['carey']['components']:
        ws[f'A{row}'] = c['concept']
        ws[f'B{row}'] = c['role']
        ws[f'C{row}'] = c['type']
        row += 1

    row += 1
    ws[f'A{row}'] = 'CORE COGNITION SYSTEMS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'System'
    ws[f'B{row}'] = 'How Used'
    ws[f'C{row}'] = 'Extended Beyond?'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for c in TECH_SOVEREIGNTY_DATA['carey']['core_cognition']:
        ws[f'A{row}'] = c['system']
        ws[f'B{row}'] = c['how']
        ws[f'C{row}'] = 'Yes' if c['extended'] else 'No'
        row += 1

    row += 1
    ws[f'A{row}'] = 'LEARNING CONSTRAINTS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Type'
    ws[f'B{row}'] = 'Description'
    ws[f'C{row}'] = 'Population'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for c in TECH_SOVEREIGNTY_DATA['carey']['learning_constraints']:
        ws[f'A{row}'] = c['type']
        ws[f'B{row}'] = c['description']
        ws[f'C{row}'] = c['population']
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20

    # Kuhnian sheet (NEW)
    ws = wb.create_sheet('10. Kuhnian NEW')
    ws['A1'] = 'KUHNIAN DIMENSION - Paradigm Structure [NEW]'
    ws['A1'].font = Font(bold=True, size=14)

    par = TECH_SOVEREIGNTY_DATA['kuhnian']['paradigm']
    ws['A3'] = 'PARADIGM CONTEXT'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Paradigm: {par['name']}"
    ws['A5'] = f"Core Assumptions: {par['assumptions']}"
    ws['A6'] = f"Disciplinary Matrix: {par['matrix']}"
    ws['A7'] = f"Is Revolutionary: {'Yes' if par['revolutionary'] else 'No'}"

    ws['A9'] = 'EXEMPLARS'
    ws['A9'].font = Font(bold=True)
    ws['A10'] = 'Exemplar'
    ws['B10'] = 'Why Paradigmatic'
    ws['C10'] = 'What It Teaches'
    ws['A10'].font = Font(bold=True)
    ws['B10'].font = Font(bold=True)
    ws['C10'].font = Font(bold=True)

    row = 11
    for e in TECH_SOVEREIGNTY_DATA['kuhnian']['exemplars']:
        ws[f'A{row}'] = e['description']
        ws[f'B{row}'] = e['why']
        ws[f'C{row}'] = e['teaches']
        row += 1

    row += 1
    ws[f'A{row}'] = 'ANOMALIES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Anomaly'
    ws[f'B{row}'] = 'Why Anomalous'
    ws[f'C{row}'] = 'Seriousness'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for a in TECH_SOVEREIGNTY_DATA['kuhnian']['anomalies']:
        ws[f'A{row}'] = a['description']
        ws[f'B{row}'] = a['why']
        ws[f'C{row}'] = a['seriousness']
        row += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # Pragmatist sheet (NEW)
    ws = wb.create_sheet('11. Pragmatist NEW')
    ws['A1'] = 'PRAGMATIST DIMENSION - Performative Consequences [NEW]'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'PRACTICAL EFFECTS'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Type'
    ws['B4'] = 'Effect Description'
    ws['C4'] = 'Without Concept'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)

    row = 5
    for e in TECH_SOVEREIGNTY_DATA['pragmatist']['practical_effects']:
        ws[f'A{row}'] = e['type']
        ws[f'B{row}'] = e['description']
        ws[f'C{row}'] = e['without']
        row += 1

    row += 1
    ws[f'A{row}'] = 'CASH VALUE'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Domain'
    ws[f'B{row}'] = 'What Changes'
    ws[f'C{row}'] = 'For Whom'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for c in TECH_SOVEREIGNTY_DATA['pragmatist']['cash_value']:
        ws[f'A{row}'] = c['domain']
        ws[f'B{row}'] = c['what_changes']
        ws[f'C{row}'] = c['for_whom']
        row += 1

    row += 1
    ws[f'A{row}'] = 'TOOL USES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'Function'
    ws[f'B{row}'] = 'Problem Addressed'
    ws[f'C{row}'] = 'Effectiveness'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for t in TECH_SOVEREIGNTY_DATA['pragmatist']['tool_uses']:
        ws[f'A{row}'] = t['function']
        ws[f'B{row}'] = t['problem']
        ws[f'C{row}'] = t['effectiveness']
        row += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 35

    # Foucauldian sheet (NEW)
    ws = wb.create_sheet('12. Foucauldian NEW')
    ws['A1'] = 'FOUCAULDIAN DIMENSION - Power-Knowledge Relations [NEW]'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'POWER RELATIONS NATURALIZED'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Power Relation'
    ws['B4'] = 'How Naturalized'
    ws['C4'] = 'Who Benefits'
    ws['D4'] = 'Who Is Disadvantaged'
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)

    row = 5
    for p in TECH_SOVEREIGNTY_DATA['foucauldian']['power_relations']:
        ws[f'A{row}'] = p['relation']
        ws[f'B{row}'] = p['naturalized']
        ws[f'C{row}'] = p['benefits']
        ws[f'D{row}'] = p['disadvantaged']
        row += 1

    row += 1
    ws[f'A{row}'] = 'GOVERNMENTALITY'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'What Governed'
    ws[f'B{row}'] = 'Technique'
    ws[f'C{row}'] = 'Knowledge Produced'
    ws[f'D{row}'] = 'Resistance Possible'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'].font = Font(bold=True)
    row += 1

    for g in TECH_SOVEREIGNTY_DATA['foucauldian']['governmentality']:
        ws[f'A{row}'] = g['governed']
        ws[f'B{row}'] = g['technique']
        ws[f'C{row}'] = g['knowledge']
        ws[f'D{row}'] = g['resistance']
        row += 1

    row += 1
    ws[f'A{row}'] = 'DISCOURSE EFFECTS'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = 'What Speakable'
    ws[f'B{row}'] = 'What Unspeakable'
    ws[f'C{row}'] = 'Subject Positions Created'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    for d in TECH_SOVEREIGNTY_DATA['foucauldian']['discourse_effects']:
        ws[f'A{row}'] = d['speakable']
        ws[f'B{row}'] = d['unspeakable']
        ws[f'C{row}'] = d['subject_positions']
        row += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 35

    # Save
    output_path = '/home/evgeny/Downloads/TechSovereignty_12D_Analysis.xlsx'
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


if __name__ == '__main__':
    print("Generating 12-Dimension Schema Spreadsheets...")
    create_schema_explanation_spreadsheet()
    create_tech_sovereignty_spreadsheet()
    print("Done!")
